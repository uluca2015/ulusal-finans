#!/usr/bin/env python3

# -*- coding: utf-8 -*-

"""

NAKİT AKIŞ YÖNETİM SİSTEMİ  v3.1 (Web Edition)

"""

import sys, os, json, threading, webbrowser, io, base64, traceback, hashlib, secrets, time

from http.server import HTTPServer, BaseHTTPRequestHandler

from urllib.parse import urlparse, parse_qs

import pandas as pd

from datetime import datetime, timedelta



BANKA_KART_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "nakit_akis_banka_kart.json")

BANKA_KART_DATA = []



def load_banka_kart():

    global BANKA_KART_DATA

    if os.path.exists(BANKA_KART_FILE):

        try:

            with open(BANKA_KART_FILE, 'r', encoding='utf-8') as f:

                BANKA_KART_DATA = json.load(f)

            print(f"  {len(BANKA_KART_DATA)} banka karti yuklendi")

        except: BANKA_KART_DATA = []



def save_banka_kart():

    try:

        with open(BANKA_KART_FILE, 'w', encoding='utf-8') as f:

            json.dump(BANKA_KART_DATA, f, ensure_ascii=False, indent=2)

    except Exception as e:

        print(f"Banka karti kayit hatasi: {e}")



IPOTEK_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "nakit_akis_ipotek.json")

IPOTEK_DATA = []



def load_ipotek():

    global IPOTEK_DATA

    if os.path.exists(IPOTEK_FILE):

        try:

            with open(IPOTEK_FILE, 'r', encoding='utf-8') as f:

                IPOTEK_DATA = json.load(f)

            print(f"  {len(IPOTEK_DATA)} ipotek kaydi yuklendi")

        except: IPOTEK_DATA = []



def save_ipotek():

    try:

        with open(IPOTEK_FILE, 'w', encoding='utf-8') as f:

            json.dump(IPOTEK_DATA, f, ensure_ascii=False, indent=2)

    except Exception as e:

        print(f"Ipotek kayit hatasi: {e}")



USERS_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "nakit_akis_users.json")

SESSIONS   = {}
SESSION_TTL = 28800  # 8 saat (sliding window)
SESSIONS_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "nakit_akis_sessions.json")

def load_sessions():
    global SESSIONS
    if os.path.exists(SESSIONS_FILE):
        try:
            with open(SESSIONS_FILE, 'r') as f:
                data = json.load(f)
            now = time.time()
            SESSIONS = {k:v for k,v in data.items() if v.get('expires',0) > now}
            print(f"  {len(SESSIONS)} aktif oturum yuklendi")
        except: SESSIONS = {}

def save_sessions():
    try:
        now = time.time()
        aktif = {k:v for k,v in SESSIONS.items() if v.get('expires',0) > now}
        with open(SESSIONS_FILE, 'w') as f:
            json.dump(aktif, f)
    except: pass



def hash_password(pw):

    return hashlib.sha256(pw.encode('utf-8')).hexdigest()



def load_users():

    if os.path.exists(USERS_FILE):

        try:

            with open(USERS_FILE, 'r', encoding='utf-8') as f:

                return json.load(f)

        except: pass

    default = {"admin": {"password": hash_password("admin123"), "ad": "Yonetici", "rol": "admin", "created_at": datetime.now().strftime('%d.%m.%Y %H:%M')}}

    save_users(default)

    return default



def save_users(users):

    try:

        with open(USERS_FILE, 'w', encoding='utf-8') as f:

            json.dump(users, f, ensure_ascii=False, indent=2)

    except: pass



def create_session(username):
    token = secrets.token_hex(32)
    SESSIONS[token] = {'username': username, 'expires': time.time() + SESSION_TTL}
    save_sessions()
    return token



def get_session(token):

    if not token: return None

    s = SESSIONS.get(token)

    if not s: return None

    if time.time() > s['expires']:

        del SESSIONS[token]; return None

    # Sliding window: her aktivitede süreyi uzat
    s['expires'] = time.time() + SESSION_TTL
    save_sessions()
    return s



def get_token_from_request(handler):

    for part in handler.headers.get('Cookie','').split(';'):

        part = part.strip()

        if part.startswith('na_token='):

            return part[9:]

    return None



def check_auth(handler):

    # Cookie'den token

    token = get_token_from_request(handler)

    if token:

        s = get_session(token)

        if s: return s

    # X-Token header (JS api wrapper'dan)

    xtoken = handler.headers.get('X-Token','')

    if xtoken:

        s = get_session(xtoken)

        if s: return s

    # Authorization header (multipart için)

    auth = handler.headers.get('Authorization','')

    if auth.startswith('Bearer '):

        return get_session(auth[7:].strip())

    return None



GITHUB_TOKEN  = os.environ.get('GITHUB_TOKEN', '')

GITHUB_REPO   = os.environ.get('GITHUB_REPO', 'uluca2015/ulusal-finans')

GITHUB_BRANCH = 'main'

BACKUP_PREFIX = 'yedek/'



def github_api(method, endpoint, data=None):

    import urllib.request

    if not GITHUB_TOKEN: return None

    url = f'https://api.github.com{endpoint}'

    headers = {'Authorization': f'token {GITHUB_TOKEN}', 'Accept': 'application/vnd.github.v3+json',

                'Content-Type': 'application/json', 'User-Agent': 'nakit-akis-backup'}

    try:

        body = json.dumps(data).encode('utf-8') if data else None

        req = urllib.request.Request(url, data=body, headers=headers, method=method)

        with urllib.request.urlopen(req, timeout=30) as r:

            return json.loads(r.read())

    except Exception as e:

        print(f"GitHub API hatasi: {e}"); return None



def github_dosya_yukle(dosya_yolu, icerik_str, mesaj):

    if not GITHUB_TOKEN: return False

    try:

        icerik_b64 = base64.b64encode(icerik_str.encode('utf-8')).decode('utf-8')

        mevcut = github_api('GET', f'/repos/{GITHUB_REPO}/contents/{dosya_yolu}')

        sha = mevcut.get('sha') if mevcut else None

        veri = {'message': mesaj, 'content': icerik_b64, 'branch': GITHUB_BRANCH}

        if sha: veri['sha'] = sha

        return github_api('PUT', f'/repos/{GITHUB_REPO}/contents/{dosya_yolu}', veri) is not None

    except: return False



def github_dosya_oku(dosya_yolu):

    if not GITHUB_TOKEN: return None

    try:

        sonuc = github_api('GET', f'/repos/{GITHUB_REPO}/contents/{dosya_yolu}')

        if sonuc and 'content' in sonuc:

            return json.loads(base64.b64decode(sonuc['content']).decode('utf-8'))

    except: return None



def github_yedek_al():

    if not GITHUB_TOKEN: return False

    ts = datetime.now().strftime('%d.%m.%Y %H:%M')

    mesaj = f"Otomatik yedek - {ts}"

    dosyalar = {

        f'{BACKUP_PREFIX}nakit_akis_data.json':       json.dumps(DATA, ensure_ascii=False, indent=2),

        f'{BACKUP_PREFIX}nakit_akis_gelir.json':      json.dumps(GELIR_DATA, ensure_ascii=False, indent=2),

        f'{BACKUP_PREFIX}nakit_akis_banka.json':      json.dumps(BANKA_DATA, ensure_ascii=False, indent=2),

        f'{BACKUP_PREFIX}nakit_akis_kredi.json':      json.dumps(KREDI_DATA, ensure_ascii=False, indent=2),

        f'{BACKUP_PREFIX}nakit_akis_cari.json':       json.dumps(CARI_DATA, ensure_ascii=False, indent=2),

        f'{BACKUP_PREFIX}nakit_akis_notlar.json':     json.dumps(NOTLAR, ensure_ascii=False, indent=2),

        f'{BACKUP_PREFIX}nakit_akis_ipotek.json':     json.dumps(IPOTEK_DATA, ensure_ascii=False, indent=2),

    }

    basarili = sum(1 for yol,icerik in dosyalar.items() if github_dosya_yukle(yol,icerik,mesaj))

    print(f"  GitHub yedek: {basarili}/{len(dosyalar)} dosya [{ts}]")

    return basarili > 0



def github_yedekten_yukle():

    global DATA, GELIR_DATA, BANKA_DATA, KREDI_DATA, CARI_DATA, NOTLAR, IPOTEK_DATA

    if not GITHUB_TOKEN: return False

    dosya_map = {

        f'{BACKUP_PREFIX}nakit_akis_data.json':   'gider',

        f'{BACKUP_PREFIX}nakit_akis_gelir.json':  'gelir',

        f'{BACKUP_PREFIX}nakit_akis_banka.json':  'banka',

        f'{BACKUP_PREFIX}nakit_akis_kredi.json':  'kredi',

        f'{BACKUP_PREFIX}nakit_akis_cari.json':   'cari',

        f'{BACKUP_PREFIX}nakit_akis_notlar.json': 'notlar',

        f'{BACKUP_PREFIX}nakit_akis_ipotek.json': 'ipotek',

    }

    yuklu = 0

    for yol, tur in dosya_map.items():

        veri = github_dosya_oku(yol)

        if veri is None: continue

        if tur=='gider':  DATA=veri; save_data()

        elif tur=='gelir': GELIR_DATA=veri; save_gelir()

        elif tur=='banka': BANKA_DATA=veri; save_banka()

        elif tur=='kredi': KREDI_DATA=veri; save_kredi()

        elif tur=='cari':  CARI_DATA=veri; save_cari()

        elif tur=='notlar': NOTLAR=veri; save_notlar()

        elif tur=='ipotek': IPOTEK_DATA=veri; save_ipotek()

        yuklu += 1

    print(f"  GitHub yedeginden {yuklu} dosya yuklendi")

    return yuklu > 0



def yedek_zamanlayici():

    import time as _t

    while True:

        simdi = datetime.now()

        hedef = simdi.replace(hour=3, minute=0, second=0, microsecond=0)

        if simdi >= hedef: hedef = hedef + timedelta(days=1)

        _t.sleep((hedef - simdi).total_seconds())

        github_yedek_al()



LOGIN_HTML = '''<!DOCTYPE html>

<html lang="tr"><head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">

<title>Nakit Akış — Giriş</title>

<style>*{box-sizing:border-box;margin:0;padding:0}body{background:linear-gradient(135deg,#1e3a5f,#0f2027);min-height:100vh;display:flex;align-items:center;justify-content:center;font-family:"Segoe UI",sans-serif}.card{background:#fff;border-radius:16px;padding:40px 36px;width:360px;box-shadow:0 20px 60px rgba(0,0,0,.4)}.logo{text-align:center;margin-bottom:28px}.logo h1{color:#1e3a5f;font-size:22px;font-weight:700;margin-top:8px}.logo p{color:#6b7280;font-size:13px;margin-top:4px}label{display:block;font-size:13px;font-weight:600;color:#374151;margin-bottom:6px}input{width:100%;padding:11px 14px;border:1.5px solid #d1d5db;border-radius:8px;font-size:14px;outline:none}.field{margin-bottom:18px}button{width:100%;padding:12px;background:#1e3a5f;color:#fff;border:none;border-radius:8px;font-size:15px;font-weight:600;cursor:pointer}.err{color:#dc2626;font-size:13px;background:#fee2e2;padding:10px 14px;border-radius:8px;margin-bottom:16px;display:none}</style></head>

<body><div class="card"><div class="logo"><div style="font-size:40px">💰</div><h1>Nakit Akış Yönetimi</h1><p>Sisteme giriş yapın</p></div>

<div class="err" id="err"></div>

<div class="field"><label>Kullanıcı Adı</label><input type="text" id="u" autocomplete="username"></div>

<div class="field"><label>Şifre</label><input type="password" id="p" autocomplete="current-password" onkeydown="if(event.key==='Enter')login()"></div>

<button onclick="login()">Giriş Yap</button>

<p style="text-align:center;margin-top:16px;font-size:12px;color:#9ca3af">Varsayılan: admin / admin123</p></div>

<script>async function login(){const u=document.getElementById('u').value.trim(),p=document.getElementById('p').value,e=document.getElementById('err');e.style.display='none';if(!u||!p){e.textContent='Kullanıcı adı ve şifre gerekli.';e.style.display='block';return;}const r=await fetch('/api/login',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({username:u,password:p})});const d=await r.json();if(d.ok){window.location.href='/';}else{e.textContent=d.msg||'Giriş başarısız.';e.style.display='block';}}</script></body></html>'''





# ── PATHS ─────────────────────────────────────────────────────────────────────

APP_DIR     = os.path.dirname(os.path.abspath(__file__))

SAVE_FILE        = os.path.join(APP_DIR, "nakit_akis_data.json")

GELIR_SAVE_FILE  = os.path.join(APP_DIR, "nakit_akis_gelir.json")

BANKA_SAVE_FILE  = os.path.join(APP_DIR, "nakit_akis_banka.json")

KREDI_SAVE_FILE  = os.path.join(APP_DIR, "nakit_akis_kredi.json")

CARI_SAVE_FILE   = os.path.join(APP_DIR, "nakit_akis_cari.json")

API_SAVE_FILE    = os.path.join(APP_DIR, "nakit_akis_api.json")

LOGO_CONFIG_FILE = os.path.join(APP_DIR, "nakit_akis_logo.json")

NOTLAR_FILE      = os.path.join(APP_DIR, "nakit_akis_notlar.json")

GMAIL_FILE       = os.path.join(APP_DIR, "nakit_akis_gmail.json")

GMAIL_TOKEN_FILE = os.path.join(APP_DIR, "nakit_akis_gmail_token.json")

AI_CONFIG_FILE   = os.path.join(APP_DIR, "nakit_akis_ai.json")

OTEL_DATA_FILE   = os.path.join(APP_DIR, "nakit_akis_otel.json")

OTEL_CONFIG_FILE = os.path.join(APP_DIR, "nakit_akis_otel_config.json")

HATIRLATMA_FILE  = os.path.join(APP_DIR, "nakit_akis_hatirlatma.json")

BANKA_SAVE_FILE  = os.path.join(APP_DIR, "nakit_akis_banka.json")

KREDI_SAVE_FILE  = os.path.join(APP_DIR, "nakit_akis_kredi.json")

CONFIG_FILE = os.path.join(APP_DIR, "nakit_akis_config.json")

AYLAR = ['Ocak','Şubat','Mart','Nisan','Mayıs','Haziran','Temmuz',

         'Ağustos','Eylül','Ekim','Kasım','Aralık']



# ── GLOBAL STATE ──────────────────────────────────────────────────────────────

DATA          = []        # Gider/ödeme kayıtları

GELIR_DATA    = []        # Gelir kayıtları

BANKA_DATA    = []        # Banka bakiyeleri

KREDI_DATA    = []        # Kredi limitleri

BANKA_DATA    = []        # Banka bakiyeleri (manuel giriş)

KREDI_DATA    = []        # Kredi limitleri (ROTATİF/KMH/VİNOV/K.Kartı)

CARI_DATA     = []        # Cari hesap kayıtları (alacak/borç)

API_CONFIG    = []        # Banka API ayarları

KUR_CACHE     = {}        # TCMB kur cache {tarih: {USD:..., EUR:...}}

LOGO_CONFIG   = {}        # Logo SQL bağlantı ayarları

FATURA_CACHE  = {}        # Fatura önbellek

NOTLAR        = []        # Notlar listesi

GMAIL_CONFIG  = {}        # Gmail OAuth ayarları

AI_CONFIG     = {}        # OpenAI API ayarları

AI_CONV_FILE  = os.path.join(APP_DIR, 'nakit_akis_ai_conv.json')

OTEL_DATA     = {'otel1': [], 'otel2': []}  # Otel rezervasyonları

OTEL_CONFIG   = {}                           # Otel ayarları

HATIRLATMALAR = []                           # Hatırlatıcılar

WHAPI_CONFIG  = {}                           # Whapi.Cloud API ayarları

DOCS_META     = []                           # Doküman meta verileri

WHAPI_FILE    = os.path.join(APP_DIR, "nakit_akis_whapi.json")

DOCS_META_FILE = os.path.join(APP_DIR, "nakit_akis_docs.json")

DOCS_DIR       = os.path.join(APP_DIR, "belgeler")

WHAPI_MSG_FILE = os.path.join(APP_DIR, "nakit_akis_wa_mesajlar.json")

WHAPI_MSGS    = {}   # {chat_id: {ad, mesajlar: [...]}}

SOURCE_FILE   = None   # Kullanıcının tanımladığı Excel yolu

SOURCE_SHEET  = None   # Sayfa adı (None = otomatik)

SOURCE_MTIME  = None   # Son değişiklik zamanı

AUTO_RELOAD   = True

LAST_SYNC_MSG = "Henüz senkronize edilmedi"

LAST_SYNC_OK  = False



# ── CONFIG ────────────────────────────────────────────────────────────────────

def load_config():

    global SOURCE_SHEET

    global SOURCE_FILE, SOURCE_SHEET, AUTO_RELOAD

    if os.path.exists(CONFIG_FILE):

        try:

            with open(CONFIG_FILE, 'r', encoding='utf-8') as f:

                cfg = json.load(f)

            SOURCE_FILE  = cfg.get('source_file')

            SOURCE_SHEET = None  # Sayfa adi her seferinde otomatik bulunur

            AUTO_RELOAD  = cfg.get('auto_reload', True)

            if SOURCE_FILE:

                print(f"  Kaynak dosya: {SOURCE_FILE}")

        except Exception as e:

            print(f"Config okunamadı: {e}")



def save_config():

    try:

        with open(CONFIG_FILE, 'w', encoding='utf-8') as f:

            json.dump({

                'source_file':  SOURCE_FILE,

                'source_sheet': None,  # Her acilista otomatik bulunur

                'auto_reload':  AUTO_RELOAD,

            }, f, ensure_ascii=False, indent=2)

    except Exception as e:

        print(f"Config kayıt hatası: {e}")



# ── DATA ──────────────────────────────────────────────────────────────────────

def load_data():

    global DATA, GELIR_DATA, BANKA_DATA, KREDI_DATA, CARI_DATA, CARI_DATA, API_CONFIG, LOGO_CONFIG, NOTLAR, GMAIL_CONFIG, AI_CONFIG, OTEL_DATA, OTEL_CONFIG, HATIRLATMALAR, WHAPI_CONFIG, WHAPI_MSGS, DOCS_META

    if os.path.exists(SAVE_FILE):

        try:

            with open(SAVE_FILE, 'r', encoding='utf-8') as f:

                DATA = json.load(f)

            print(f"  {len(DATA)} gider kaydı yüklendi")

        except Exception as e:

            print(f"Kayıt okunamadı: {e}")

            DATA = []

    if os.path.exists(GELIR_SAVE_FILE):

        try:

            with open(GELIR_SAVE_FILE, 'r', encoding='utf-8') as f:

                GELIR_DATA = json.load(f)

            print(f"  {len(GELIR_DATA)} gelir kaydı yüklendi")

        except Exception as e:

            print(f"Gelir kaydı okunamadı: {e}")

            GELIR_DATA = []

    if os.path.exists(BANKA_SAVE_FILE):

        try:

            with open(BANKA_SAVE_FILE, 'r', encoding='utf-8') as f:

                BANKA_DATA = json.load(f)

            print(f"  {len(BANKA_DATA)} banka kaydı yüklendi")

        except Exception as e:

            print(f"Banka kaydı okunamadı: {e}")

            BANKA_DATA = []

    if os.path.exists(KREDI_SAVE_FILE):

        try:

            with open(KREDI_SAVE_FILE, 'r', encoding='utf-8') as f:

                KREDI_DATA = json.load(f)

            print(f"  {len(KREDI_DATA)} kredi kaydı yüklendi")

        except Exception as e:

            print(f"Kredi kaydı okunamadı: {e}")

            KREDI_DATA = []

    if os.path.exists(BANKA_SAVE_FILE):

        try:

            with open(BANKA_SAVE_FILE, 'r', encoding='utf-8') as f:

                BANKA_DATA = json.load(f)

            print(f"  {len(BANKA_DATA)} banka bakiyesi yüklendi")

        except:

            BANKA_DATA = []

    if os.path.exists(KREDI_SAVE_FILE):

        try:

            with open(KREDI_SAVE_FILE, 'r', encoding='utf-8') as f:

                KREDI_DATA = json.load(f)

            print(f"  {len(KREDI_DATA)} kredi limiti yüklendi")

        except:

            KREDI_DATA = []

    if os.path.exists(CARI_SAVE_FILE):

        try:

            with open(CARI_SAVE_FILE, 'r', encoding='utf-8') as f:

                CARI_DATA = json.load(f)

            print(f"  {len(CARI_DATA)} cari kayıt yüklendi")

        except:

            CARI_DATA = []

    if os.path.exists(API_SAVE_FILE):

        try:

            with open(API_SAVE_FILE, 'r', encoding='utf-8') as f:

                API_CONFIG = json.load(f)

            print(f"  {len(API_CONFIG)} banka API ayarı yüklendi")

        except:

            API_CONFIG = []

    if os.path.exists(LOGO_CONFIG_FILE):

        try:

            with open(LOGO_CONFIG_FILE, 'r', encoding='utf-8') as f:

                LOGO_CONFIG = json.load(f)

            if LOGO_CONFIG.get('server'):

                print(f"  Logo SQL config yüklendi: {LOGO_CONFIG.get('server')}")

        except:

            LOGO_CONFIG = {}

    if os.path.exists(NOTLAR_FILE):

        try:

            with open(NOTLAR_FILE, 'r', encoding='utf-8') as f:

                NOTLAR = json.load(f)

            print(f"  {len(NOTLAR)} not yüklendi")

        except:

            NOTLAR = []

    if os.path.exists(GMAIL_FILE):

        try:

            with open(GMAIL_FILE, 'r', encoding='utf-8') as f:

                GMAIL_CONFIG = json.load(f)

        except:

            GMAIL_CONFIG = {}

    if os.path.exists(WHAPI_FILE):

        try:

            with open(WHAPI_FILE, 'r', encoding='utf-8') as f:

                WHAPI_CONFIG = json.load(f)

        except:

            WHAPI_CONFIG = {}

    global DOCS_META

    if os.path.exists(DOCS_META_FILE):

        try:

            with open(DOCS_META_FILE, 'r', encoding='utf-8') as f:

                DOCS_META = json.load(f)

        except:

            DOCS_META = []

    # Belgeler klasörü oluştur

    os.makedirs(DOCS_DIR, exist_ok=True)

    if os.path.exists(WHAPI_MSG_FILE):

        try:

            with open(WHAPI_MSG_FILE, 'r', encoding='utf-8') as f:

                WHAPI_MSGS = json.load(f)

        except:

            WHAPI_MSGS = {}



def save_data():

    try:

        with open(SAVE_FILE, 'w', encoding='utf-8') as f:

            json.dump(DATA, f, ensure_ascii=False, indent=2)

    except Exception as e:

        print(f"Kayıt hatası: {e}")



def save_gelir():

    try:

        with open(GELIR_SAVE_FILE, 'w', encoding='utf-8') as f:

            json.dump(GELIR_DATA, f, ensure_ascii=False, indent=2)

    except Exception as e:

        print(f"Gelir kayıt hatası: {e}")



def save_banka():

    try:

        with open(BANKA_SAVE_FILE, 'w', encoding='utf-8') as f:

            json.dump(BANKA_DATA, f, ensure_ascii=False, indent=2)

    except Exception as e:

        print(f"Banka kayıt hatası: {e}")



def save_kredi():

    try:

        with open(KREDI_SAVE_FILE, 'w', encoding='utf-8') as f:

            json.dump(KREDI_DATA, f, ensure_ascii=False, indent=2)

    except Exception as e:

        print(f"Kredi kayıt hatası: {e}")



def save_cari():

    try:

        with open(CARI_SAVE_FILE, 'w', encoding='utf-8') as f:

            json.dump(CARI_DATA, f, ensure_ascii=False, indent=2)

    except Exception as e:

        print(f"Cari kayıt hatası: {e}")



def save_api_config():

    try:

        with open(API_SAVE_FILE, 'w', encoding='utf-8') as f:

            json.dump(API_CONFIG, f, ensure_ascii=False, indent=2)

    except Exception as e:

        print(f"API config kayıt hatası: {e}")



def save_logo_config():

    try:

        with open(LOGO_CONFIG_FILE, 'w', encoding='utf-8') as f:

            json.dump(LOGO_CONFIG, f, ensure_ascii=False, indent=2)

    except Exception as e:

        print(f"Logo config kayıt hatası: {e}")



def save_notlar():

    try:

        with open(NOTLAR_FILE, 'w', encoding='utf-8') as f:

            json.dump(NOTLAR, f, ensure_ascii=False, indent=2)

    except Exception as e:

        print(f"Notlar kayıt hatası: {e}")



def save_gmail_config():

    try:

        with open(GMAIL_FILE, 'w', encoding='utf-8') as f:

            json.dump(GMAIL_CONFIG, f, ensure_ascii=False, indent=2)

    except Exception as e:

        print(f"Gmail config kayıt hatası: {e}")



def save_ai_config():

    try:

        with open(AI_CONFIG_FILE, 'w', encoding='utf-8') as f:

            json.dump(AI_CONFIG, f, ensure_ascii=False, indent=2)

    except Exception as e:

        print(f"AI config kayıt hatası: {e}")



def save_otel_data():

    try:

        with open(OTEL_DATA_FILE, 'w', encoding='utf-8') as f:

            json.dump(OTEL_DATA, f, ensure_ascii=False, indent=2)

    except Exception as e:

        print(f"Otel data kayıt hatası: {e}")



def save_otel_config():

    try:

        with open(OTEL_CONFIG_FILE, 'w', encoding='utf-8') as f:

            json.dump(OTEL_CONFIG, f, ensure_ascii=False, indent=2)

    except Exception as e:

        print(f"Otel config kayıt hatası: {e}")



def save_hatirlatmalar():

    try:

        with open(HATIRLATMA_FILE, 'w', encoding='utf-8') as f:

            json.dump(HATIRLATMALAR, f, ensure_ascii=False, indent=2)

    except Exception as e:

        print(f"Hatırlatma kayıt hatası: {e}")



def save_banka():

    try:

        with open(BANKA_SAVE_FILE, 'w', encoding='utf-8') as f:

            json.dump(BANKA_DATA, f, ensure_ascii=False, indent=2)

    except Exception as e:

        print(f"Banka kayıt hatası: {e}")



def save_kredi():

    try:

        with open(KREDI_SAVE_FILE, 'w', encoding='utf-8') as f:

            json.dump(KREDI_DATA, f, ensure_ascii=False, indent=2)

    except Exception as e:

        print(f"Kredi kayıt hatası: {e}")



# ── EXCEL IMPORT ──────────────────────────────────────────────────────────────

def _norm_sheet(s):

    import unicodedata as _ud

    r = str(s).upper()

    r = r.replace(chr(304),"I").replace(chr(350),"S").replace(chr(286),"G")

    r = r.replace(chr(220),"U").replace(chr(214),"O").replace(chr(199),"C")

    r2 = _ud.normalize("NFD", r)

    return "".join(c for c in r2 if _ud.category(c) != "Mn")

def import_excel(filepath, sheet_hint=None):

    try:

        import warnings; warnings.filterwarnings('ignore')

        xl = pd.ExcelFile(filepath)

        # Sayfa sec

        target = None

        if sheet_hint:

            for s in xl.sheet_names:

                if s == sheet_hint: target = s; break

            if not target:

                for s in xl.sheet_names:

                    if s.upper() == sheet_hint.upper(): target = s; break

            if not target:

                hn = _norm_sheet(sheet_hint)

                for s in xl.sheet_names:

                    if _norm_sheet(s) == hn: target = s; break

        if not target:

            for s in xl.sheet_names:

                sn = _norm_sheet(s)

                if ("NAKIT" in sn and "AKIS" in sn and "V2" in sn) or "NAKITAKISV2" in sn:

                    target = s; break

        if not target:

            for s in xl.sheet_names:

                sn = _norm_sheet(s)

                if ("NAKIT" in sn and "AKIS" in sn) or "NAKITAKIS" in sn:

                    target = s; break

        if not target:

            for s in xl.sheet_names:

                sn = _norm_sheet(s)

                if "NAKIT" in sn or "AKIS" in sn:

                    target = s; break

        if not target:

            target = xl.sheet_names[0]

        print("  Secilen sayfa:", repr(target))



        df_raw = pd.read_excel(filepath, sheet_name=target, header=None)



        # Başlık satırını bul — en çok anahtar kelime içeren satırı seç

        header_row = -1

        best_score = 0

        HEADER_KEYS = ['GRUP','ODEME','FIRMA','GIDER','TUTAR','VADE','TARIH','DURUM','ACIKLAMA','FAIZ','BSMV','ANAPARA']

        for i in range(min(30, len(df_raw))):

            row = df_raw.iloc[i].fillna('').astype(str)

            vals = _norm_sheet(' '.join(row.values))

            score = sum(1 for k in HEADER_KEYS if k in vals)

            if score > best_score:

                best_score = score

                header_row = i

        if header_row == -1 or best_score == 0:

            header_row = 0

            print(f"  Header bulunamadi, 0. satir kullaniliyor")

        else:

            print(f"  Header satiri bulundu: {header_row}. satir (skor:{best_score})")



        # df_raw'dan header set et - ikinci read_excel yok

        df = df_raw.iloc[header_row+1:].copy()

        df.columns = df_raw.iloc[header_row].fillna('').astype(str).str.strip().values

        df = df.reset_index(drop=True)

        print(f"  Sutunlar: {list(df.columns)[:6]}")

        print(f"  Satir sayisi: {len(df)}")



        col_map = {

            'firma':    ['GRUP FİRMASI','GRUP FIRMASI','FİRMA','FIRMA','GRUP'],

            'tur':      ['ÖDEME TÜRÜ','ODEME TURU','TÜR','TUR','TYPE','ÖDEME TİPİ'],

            'yer':      ['ÖDEME YAPILACAK YER','BANKA','YER','PLACE','KURUM'],

            'tarih':    ['VADE - ÖDEME TARİHİ','VADE - ÖDEME TARIHI','TARİH','TARIH','DATE','VADE','ÖDEME TARİHİ'],

            'ay':       ['ÖDEME AYI','ODEME AYI','AY','MONTH'],

            'durum':    ['DURUMU','DURUM','STATUS'],

            'anapara':  ['ÖDEME ANAPARA','ODEME ANAPARA','ANAPARA','PRINCIPAL'],

            'faiz':     ['ÖDEME FAİZ','ODEME FAIZ','FAİZ','FAIZ','INTEREST'],

            'bsmv':     ['ÖDEME BSMV','ODEME BSMV','BSMV'],

            'tutar':    ['ÖDEME TUTARI','ODEME TUTARI','TUTAR','AMOUNT','TOPLAM'],

            'aciklama': ['AÇIKLAMA','ACIKLAMA','NOTE','NOTES','DESCRIPTION'],

        }



        found = {}

        for field, aliases in col_map.items():

            for col in df.columns:

                cu = _norm_sheet(str(col).strip())

                for alias in aliases:

                    an = _norm_sheet(alias)

                    if an in cu or cu in an:

                        found[field] = col

                        break

                if field in found:

                    break

        print(f"  Bulunan sutunlar: {found}")



        def parse_date(v):

            if pd.isna(v) or v == '':

                return ''

            if isinstance(v, (pd.Timestamp, datetime)):

                return v.strftime('%d.%m.%Y')

            s = str(v).strip()

            for fmt in ['%Y-%m-%d','%d.%m.%Y','%d/%m/%Y','%m/%d/%Y']:

                try:

                    return datetime.strptime(s.split(' ')[0].split('T')[0], fmt).strftime('%d.%m.%Y')

                except:

                    pass

            try:

                n = float(s)

                d = datetime(1899, 12, 30) + timedelta(days=int(n))

                return d.strftime('%d.%m.%Y')

            except:

                pass

            return s[:10] if len(s) >= 10 else s



        def parse_ay(v, tarih=''):

            if v and not pd.isna(v):

                s = str(v).strip()

                if s in AYLAR:

                    return s

                for m in AYLAR:

                    if m[:3].upper() in s.upper():

                        return m

            if tarih and len(tarih) == 10:

                try:

                    idx = int(tarih[3:5]) - 1

                    if 0 <= idx < 12:

                        return AYLAR[idx]

                except:

                    pass

            return ''



        records = []

        for _, row in df.iterrows():

            def g(field):

                if field in found:

                    v = row.get(found[field])

                    return v if not (isinstance(v, float) and pd.isna(v)) else ''

                return ''



            firma = str(g('firma')).strip() if g('firma') != '' else ''

            try:

                tutar = float(g('tutar') or 0)

            except:

                tutar = 0

            if not firma or firma in ['nan','None',''] or tutar == 0:

                continue



            tarih = parse_date(g('tarih'))

            ay = parse_ay(g('ay'), tarih)



            durum = str(g('durum')).strip() if g('durum') != '' else ''

            if 'DENDİ' in durum.upper() and 'MEDİ' not in durum.upper():

                durum = 'ÖDENDİ'

            else:

                durum = durum if durum else 'ÖDENMEDİ'



            def safe_float(v):

                try: return round(float(v or 0), 2)

                except: return 0.0



            records.append({

                'firma':    firma,

                'tur':      str(g('tur')).strip() if g('tur') != '' else '',

                'yer':      str(g('yer')).strip() if g('yer') != '' else '',

                'tarih':    tarih,

                'ay':       ay,

                'durum':    durum,

                'anapara':  safe_float(g('anapara')),

                'faiz':     safe_float(g('faiz')),

                'bsmv':     safe_float(g('bsmv')),

                'tutar':    round(tutar, 2),

                'aciklama': str(g('aciklama')).strip() if g('aciklama') not in ['','nan','None',None] else '',

            })



        if not records:

            # Hangi sütunların bulunduğunu logla

            print(f"  Uyari: Kayıt bulunamadı. Sayfa:{target}, Sutunlar:{list(df.columns)[:8]}, Baslik:{header_row}")

        print(f"  import_excel: {len(records)} kayit, sayfa={target}, found_cols={list(found.keys())}")

        return records, len(records), target



    except Exception as e:

        traceback.print_exc()

        return [], 0, str(e)



# ── KAYNAK DOSYA OTOMATIK YENİLEME ───────────────────────────────────────────

def reload_from_source(force=False):

    """Kaynak Excel dosyasını okur.

    OneDrive/kilitli dosyalar için geçici kopyaya okuma yapar."""

    global DATA, SOURCE_MTIME, LAST_SYNC_MSG, LAST_SYNC_OK

    if not SOURCE_FILE:

        return False

    if not os.path.exists(SOURCE_FILE):

        LAST_SYNC_MSG = f"Dosya bulunamadı: {SOURCE_FILE}"

        LAST_SYNC_OK  = False

        return False



    import shutil, tempfile

    tmp_path = None

    try:

        mtime = os.path.getmtime(SOURCE_FILE)

        if not force and mtime == SOURCE_MTIME:

            return False  # Değişmemiş, atla



        # Dosyayı geçici konuma kopyala (OneDrive/kilitli dosyalar için)

        ext = os.path.splitext(SOURCE_FILE)[1]

        ext = os.path.splitext(SOURCE_FILE)[1]

        tmp_fd, tmp_path = tempfile.mkstemp(suffix=ext)

        os.close(tmp_fd)

        # Dosyayi gecici konuma kopyala

        import shutil as _shutil

        try:

            _shutil.copy2(SOURCE_FILE, tmp_path)

        except Exception:

            # Kilitliyse openpyxl ile oku

            import openpyxl as _opx

            _wb = _opx.load_workbook(SOURCE_FILE, read_only=True, data_only=True)

            _wb.save(tmp_path); _wb.close()

        records, count, sheet = import_excel(tmp_path, SOURCE_SHEET)

        # import_excel hata dönüyorsa (count=0, sheet=hata_str) bunu yakala

        if count == 0 and sheet and len(sheet) > 50:  # Uzun string = hata mesajı

            LAST_SYNC_MSG = f'Dosya okuma hatasi: {sheet[:200]}'

            LAST_SYNC_OK = False

            return False



        if count > 0:

            DATA = records

            save_data()

            SOURCE_MTIME = mtime

            ts = datetime.now().strftime('%d.%m.%Y %H:%M:%S')

            LAST_SYNC_MSG = f"{ts} — {count} kayıt okundu (sayfa: {sheet})"

            try: db_sync_all()  # SQLite güncelle

            except Exception as _dbe: print(f"DB sync hatası: {_dbe}")

            LAST_SYNC_OK  = True

            print(f"  Kaynak güncellendi: {count} kayıt [{ts}]")

            return True

        else:

            # Sayfa listesini göster

            try:

                import openpyxl

                wb_tmp = openpyxl.load_workbook(tmp_path, read_only=True, data_only=True)

                sheets_str = ', '.join(wb_tmp.sheetnames)

                wb_tmp.close()

            except:

                try:

                    xl2 = pd.ExcelFile(tmp_path)

                    sheets_str = ', '.join(xl2.sheet_names)

                except:

                    sheets_str = '?'

            LAST_SYNC_MSG = (

                f"Kayıt bulunamadı — Okunan sayfa: '{sheet}'\n"

                f"Dosya: {os.path.basename(SOURCE_FILE)}\n"

                f"Mevcut sayfalar: {sheets_str}\n"

                f"\nDİKKAT: Doğru dosyayı seçtiğinizden emin olun!\n"

                f"0-ULUSAL_ABC_FİNANSMAN_MODEL_ÇALIŞMA.xlsx gibi ana dosyayı seçin.\n"

                f"Sayfa Adı kutusunu BOŞ bırakın — sistem otomatik bulacak."

            )

            LAST_SYNC_OK = False



    except PermissionError:

        LAST_SYNC_MSG = "Dosya kilitli — Excel uygulamasını kapatıp tekrar deneyin"

        LAST_SYNC_OK  = False

        print("Kaynak dosya kilitli (PermissionError)")

    except FileNotFoundError:

        LAST_SYNC_MSG = f"Dosya bulunamadı: {SOURCE_FILE}"

        LAST_SYNC_OK  = False

    except Exception as e:

        err_type = type(e).__name__

        LAST_SYNC_MSG = f"Hata ({err_type}): {str(e)}"

        LAST_SYNC_OK  = False

        print(f"Kaynak dosya hatası [{err_type}]: {e}")

        traceback.print_exc()

    finally:

        # Geçici dosyayı temizle

        if tmp_path and os.path.exists(tmp_path):

            try: os.unlink(tmp_path)

            except: pass



    return False



def source_watcher():

    pass

    import time

    time.sleep(15)  # Başlangıçta bekle

    while True:

        if SOURCE_FILE and AUTO_RELOAD:

            reload_from_source(force=False)

        time.sleep(60)  # 60 saniyede bir kontrol (30 yerine)



# ── EXPORT ────────────────────────────────────────────────────────────────────

def export_excel_data(mode='all', filter_params=None):

    import openpyxl

    from openpyxl.styles import Font, PatternFill, Alignment



    wb = openpyxl.Workbook()



    if mode == 'summary':

        ws = wb.active

        ws.title = 'Aylık Özet'

        hdr_fill = PatternFill("solid", fgColor="1E3A5F")

        hdr_font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)

        headers = ['Ay','Toplam (₺)','Ödendi (₺)','Bekleyen (₺)','Kayıt Sayısı']

        for col, h in enumerate(headers, 1):

            cell = ws.cell(row=1, column=col, value=h)

            cell.fill = hdr_fill; cell.font = hdr_font

            cell.alignment = Alignment(horizontal='center')

        for ay in AYLAR:

            rows = [d for d in DATA if d.get('ay') == ay]

            if not rows: continue

            ws.append([ay,

                round(sum(d['tutar'] for d in rows)),

                round(sum(d['tutar'] for d in rows if d.get('durum')=='ÖDENDİ')),

                round(sum(d['tutar'] for d in rows if d.get('durum')!='ÖDENDİ')),

                len(rows)])

        for c,w in zip(['A','B','C','D','E'],[14,20,20,20,12]):

            ws.column_dimensions[c].width = w

    else:

        src = DATA

        if mode == 'filtered' and filter_params:

            src = apply_filter(filter_params)

        ws = wb.active; ws.title = 'Nakit Akış'

        hdr_fill = PatternFill("solid", fgColor="1E3A5F")

        hdr_font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)

        headers = ['Firma','Ödeme Türü','Ödeme Yeri','Tarih','Ay','Durum',

                   'Anapara (₺)','Faiz (₺)','BSMV (₺)','Ödeme Tutarı (₺)','Açıklama']

        for col, h in enumerate(headers, 1):

            cell = ws.cell(row=1, column=col, value=h)

            cell.fill = hdr_fill; cell.font = hdr_font

            cell.alignment = Alignment(horizontal='center')

        gf = PatternFill("solid", fgColor="D5F5E3")

        rf = PatternFill("solid", fgColor="FADBD8")

        for r, d in enumerate(src, 2):

            vals = [d.get(k,'') for k in ['firma','tur','yer','tarih','ay','durum',

                                           'anapara','faiz','bsmv','tutar','aciklama']]

            for col, val in enumerate(vals, 1):

                cell = ws.cell(row=r, column=col, value=val)

                cell.fill = gf if d.get('durum')=='ÖDENDİ' else rf

                if col >= 7: cell.number_format = '#,##0'

        for c,w in zip(['A','B','C','D','E','F','G','H','I','J','K'],

                        [10,25,28,12,10,12,16,14,12,16,35]):

            ws.column_dimensions[c].width = w

        ws.auto_filter.ref = ws.dimensions

        ws.freeze_panes = 'A2'



    out = io.BytesIO()

    wb.save(out)

    return out.getvalue()



def get_yil(d):

    """Kayıttan yıl bilgisini çıkar."""

    t = d.get('tarih','')

    if t and len(t)==10:

        try: return t[6:10]

        except: pass

    return ''



def apply_filter(params):

    result = DATA

    if params.get('yil'):   result = [d for d in result if get_yil(d) == params['yil']]

    if params.get('ay'):    result = [d for d in result if d.get('ay')    == params['ay']]

    if params.get('firma'): result = [d for d in result if d.get('firma') == params['firma']]

    if params.get('tur'):   result = [d for d in result if d.get('tur')   == params['tur']]

    if params.get('durum'): result = [d for d in result if d.get('durum') == params['durum']]

    if params.get('yer'):   result = [d for d in result if d.get('yer','') == params['yer']]

    if params.get('q'):

        q = params['q'].lower()

        result = [d for d in result if q in

                  (d.get('yer','') + d.get('firma','') + d.get('tur','') + d.get('aciklama','')).lower()]

    # Tarih aralığı filtresi - input type=date 'YYYY-MM-DD' formatında geliyor

    if params.get('tarih_bas') or params.get('tarih_bit'):

        def tarih_to_iso(t):

            if not t or len(t) < 8: return ''

            if '-' in t: return t  # Zaten ISO

            p = t.split('.')

            if len(p) == 3: return f'{p[2]}-{p[1]}-{p[0]}'

            return ''

        bas = params.get('tarih_bas','')

        bit = params.get('tarih_bit','')

        filtered = []

        for d in result:

            dt = tarih_to_iso(d.get('tarih',''))

            if not dt: continue

            if bas and dt < bas: continue

            if bit and dt > bit: continue

            filtered.append(d)

        result = filtered

    return result



# ── CHARTS ────────────────────────────────────────────────────────────────────

def make_chart(chart_type):

    import matplotlib

    matplotlib.use('Agg')

    import matplotlib.pyplot as plt

    import matplotlib.ticker as mticker



    plt.rcParams.update({

        'figure.facecolor':'#161b27','axes.facecolor':'#1e2535',

        'axes.edgecolor':'#2a3348','axes.labelcolor':'#8892aa',

        'xtick.color':'#8892aa','ytick.color':'#8892aa',

        'text.color':'#e8eaf2','grid.color':'#2a3348',

        'font.family':'DejaVu Sans','font.size':9,

    })



    fig, ax = plt.subplots(figsize=(7,3.2) if chart_type!='pie' else (5,3.5))

    fig.patch.set_facecolor('#161b27')



    # Yıl+Ay bazlı gruplama

    month_keys = {}   # 'YYYY-Ay' -> tutar

    month_paid = {}

    month_pend = {}

    by_type    = {}

    for d in DATA:

        ay  = d.get('ay','')

        yil = get_yil(d) or '????'

        key = f"{yil}-{ay}"

        if ay in AYLAR:

            month_keys[key] = month_keys.get(key, 0) + d.get('tutar',0)

            if d.get('durum')=='ÖDENDİ': month_paid[key] = month_paid.get(key,0) + d.get('tutar',0)

            else:                          month_pend[key] = month_pend.get(key,0) + d.get('tutar',0)

        by_type[d.get('tur','Diğer')] = by_type.get(d.get('tur','Diğer'),0) + d.get('tutar',0)



    # Kronolojik sırala

    def sort_key(k):

        parts = k.split('-', 1)

        return (parts[0], AYLAR.index(parts[1]) if parts[1] in AYLAR else 99)

    months = sorted(month_keys.keys(), key=sort_key)

    # Etiket: Yıl değişince yılı göster

    def make_label(keys):

        labels = []

        prev_yil = None

        for k in keys:

            yil, ay = k.split('-', 1)

            if yil != prev_yil:

                labels.append(f"{ay}\n{yil}")

                prev_yil = yil

            else:

                labels.append(ay)

        return labels

    month_labels = make_label(months)

    # compat aliases

    by_month = month_keys

    by_paid  = month_paid

    by_pend  = month_pend



    if chart_type == 'monthly':

        bars = ax.bar(range(len(months)),[by_month[m]/1e6 for m in months],

                      color='#3b82f6',alpha=.85,width=.6,zorder=2)

        ax.set_xticks(range(len(months)))

        ax.set_xticklabels(month_labels,rotation=45,ha='right',fontsize=8)

        ax.set_title('Aylık Toplam Ödeme (Milyon ₺)',color='#e8eaf2',fontsize=10,pad=8)

        ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda x,_:f'{x:.1f}M'))

        ax.grid(axis='y',alpha=.4,zorder=1)

        ax.set_axisbelow(True)

        for bar in bars:

            h = bar.get_height()

            ax.text(bar.get_x()+bar.get_width()/2, h+.02, f'{h:.1f}M',

                    ha='center',va='bottom',fontsize=7.5,color='#8892aa')



    elif chart_type == 'status':

        x = range(len(months)); w = .38

        ax.bar([i-w/2 for i in x],[by_paid[m]/1e6 for m in months],w,

               label='Ödendi',color='#10b981',alpha=.85,zorder=2)

        ax.bar([i+w/2 for i in x],[by_pend[m]/1e6 for m in months],w,

               label='Bekleyen',color='#ef4444',alpha=.75,zorder=2)

        ax.set_xticks(list(x)); ax.set_xticklabels(month_labels,rotation=45,ha='right',fontsize=8)

        ax.set_title('Ödendi / Bekleyen (Milyon ₺)',color='#e8eaf2',fontsize=10,pad=8)

        ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda x,_:f'{x:.1f}M'))

        ax.legend(facecolor='#252d40',edgecolor='#2a3348',labelcolor='#8892aa')

        ax.grid(axis='y',alpha=.4,zorder=1); ax.set_axisbelow(True)



    elif chart_type == 'pie':

        labels = list(by_type.keys()); vals = [by_type[k] for k in labels]

        colors = ['#3b82f6','#6366f1','#10b981','#f59e0b','#ef4444','#8b5cf6','#06b6d4']

        short  = [l[:22]+'…' if len(l)>22 else l for l in labels]

        wedges, texts, atexts = ax.pie(vals, labels=short, colors=colors[:len(vals)],

            autopct='%1.1f%%', startangle=90,

            textprops={'color':'#8892aa','fontsize':8},

            wedgeprops={'linewidth':1,'edgecolor':'#161b27'})

        for at in atexts: at.set_color('#e8eaf2'); at.set_fontsize(8)

        ax.set_title('Tür Bazlı Dağılım',color='#e8eaf2',fontsize=10,pad=8)



    plt.tight_layout(pad=1.2)

    buf = io.BytesIO()

    plt.savefig(buf,format='png',dpi=130,bbox_inches='tight',facecolor='#161b27')

    plt.close()

    buf.seek(0)

    return base64.b64encode(buf.read()).decode()



# ── EXCEL GERİ YAZMA ─────────────────────────────────────────────────────────

def write_back_to_excel(data_idx):

    """

    Kaynak Excel dosyasındaki ilgili satırın DURUMU sütununu günceller.

    Kaydı, tarih + firma + tutar üçlüsüyle eşleştirir.

    """

    if not SOURCE_FILE or not os.path.exists(SOURCE_FILE):

        return False

    try:

        import openpyxl

        rec = DATA[data_idx]

        wb  = openpyxl.load_workbook(SOURCE_FILE)



        # Doğru sayfayı bul

        target_sheet = SOURCE_SHEET

        if not target_sheet:

            for sn in wb.sheetnames:

                if any(k in sn.upper() for k in ['NAKİT','NAKIT','AKIŞ','AKIS']):

                    target_sheet = sn

                    break

        if not target_sheet:

            target_sheet = wb.sheetnames[0]

        ws = wb[target_sheet]



        # Başlık satırını ve DURUMU sütununu bul

        header_row  = None

        durum_col   = None

        firma_col   = None

        tarih_col   = None

        tutar_col   = None



        for row in ws.iter_rows():

            for cell in row:

                val = str(cell.value or '').strip().upper()

                if 'GRUP' in val and 'FİRMA' in val or val in ('FİRMA','FIRMA','GRUP FİRMASI','GRUP FIRMASI'):

                    firma_col  = cell.column

                    header_row = cell.row

                if 'DURUMU' in val or val == 'DURUM':

                    durum_col  = cell.column

                if 'VADE' in val or 'TARİH' in val or 'TARIH' in val:

                    tarih_col  = cell.column

                if 'ÖDEME TUTARI' in val or 'ODEME TUTARI' in val or val == 'TUTAR':

                    tutar_col  = cell.column

            if header_row:

                break



        if not header_row or not durum_col:

            print(f"  Excel geri yazma: DURUMU sütunu bulunamadı")

            return False



        # Hedef satırı bul: tarih + firma + tutar eşleşmesi

        target_row = None

        for row in ws.iter_rows(min_row=header_row+1):

            # Firma eşleşmesi

            firma_val = str(row[firma_col-1].value or '').strip() if firma_col else ''

            if firma_val != rec.get('firma',''):

                continue

            # Tutar eşleşmesi

            if tutar_col:

                try:

                    tutar_val = float(row[tutar_col-1].value or 0)

                    if abs(tutar_val - rec.get('tutar',0)) > 0.5:

                        continue

                except:

                    pass

            # Tarih eşleşmesi

            if tarih_col:

                cell_tarih = row[tarih_col-1].value

                if cell_tarih:

                    import re

                    from datetime import datetime

                    cell_str = ''

                    if hasattr(cell_tarih, 'strftime'):

                        cell_str = cell_tarih.strftime('%d.%m.%Y')

                    else:

                        s = str(cell_tarih).strip()

                        m = re.search(r'(\d{4})-(\d{2})-(\d{2})', s)

                        if m: cell_str = f"{m.group(3)}.{m.group(2)}.{m.group(1)}"

                        elif len(s)>=10 and s[2]=='.' and s[5]=='.': cell_str = s[:10]

                    if cell_str and cell_str != rec.get('tarih',''):

                        continue

            target_row = row

            break



        if not target_row:

            # Tarih bulunamazsa sadece firma+tutar ile dene

            for row in ws.iter_rows(min_row=header_row+1):

                firma_val = str(row[firma_col-1].value or '').strip() if firma_col else ''

                if firma_val != rec.get('firma',''): continue

                if tutar_col:

                    try:

                        tutar_val = float(row[tutar_col-1].value or 0)

                        if abs(tutar_val - rec.get('tutar',0)) < 0.5:

                            target_row = row

                            break

                    except: pass



        if not target_row:

            print(f"  Excel geri yazma: Satır bulunamadı ({rec.get('firma')} / {rec.get('tarih')} / {rec.get('tutar')})")

            return False



        # DURUMU güncelle

        target_row[durum_col-1].value = rec['durum']

        wb.save(SOURCE_FILE)

        print(f"  Excel güncellendi: {rec.get('firma')} / {rec.get('tarih')} → {rec['durum']}")

        return True



    except Exception as e:

        print(f"  Excel geri yazma hatası: {e}")

        return False



# ── BANKA EKSTRE PARSE ───────────────────────────────────────────────────────

def parse_banka_ekstre(filepath, ext):

    """

    Banka ekstresi dosyasını otomatik tanır ve parse eder.

    Desteklenen: Vakıfbank, Garanti, İş Bankası, YapıKredi, Akbank, Halkbank, Ziraat

    """

    import warnings

    warnings.filterwarnings('ignore')



    try:

        if ext in ['.csv', '.txt']:

            # CSV dene - farklı encoding'ler

            df = None

            for enc in ['utf-8-sig', 'utf-8', 'cp1254', 'latin-1']:

                try:

                    import io

                    raw = open(filepath, 'rb').read()

                    text = raw.decode(enc)

                    # Separator tahmin et

                    sep = ';' if text.count(';') > text.count(',') else ','

                    df = pd.read_csv(io.StringIO(text), sep=sep, header=None, dtype=str)

                    break

                except: pass

            if df is None:

                return {'ok': False, 'msg': 'CSV okunamadı'}

        else:

            df = pd.read_excel(filepath, header=None, dtype=str)



        # Banka formatını tanı

        all_text = ' '.join(df.fillna('').astype(str).values.flatten()).upper()



        banka = 'BILINMIYOR'

        if 'VAKIF' in all_text: banka = 'VAKIFBANK'

        elif 'GARANTİ' in all_text or 'GARANTI' in all_text or 'BBVA' in all_text: banka = 'GARANTİ BBVA'

        elif 'İŞ BANKASI' in all_text or 'IS BANKASI' in all_text: banka = 'İŞ BANKASI'

        elif 'YAPI KREDİ' in all_text or 'YAPIKREDI' in all_text: banka = 'YAPİ KREDİ'

        elif 'AKBANK' in all_text: banka = 'AKBANK'

        elif 'HALKBANK' in all_text or 'HALK BANK' in all_text: banka = 'HALKBANK'

        elif 'ZİRAAT' in all_text or 'ZIRAAT' in all_text: banka = 'ZİRAAT BANKASI'

        elif 'TEB' in all_text: banka = 'TEB'

        elif 'FİNANS' in all_text or 'FINANS' in all_text: banka = 'QNB FİNANSBANK'

        elif 'DENİZ' in all_text or 'DENIZ' in all_text: banka = 'DENİZBANK'



        # Başlık satırını bul

        def norm(s): return str(s).upper().replace('İ','I').replace('Ş','S').replace('Ğ','G').replace('Ü','U').replace('Ö','O').replace('Ç','C')



        header_row = 0

        col_tarih = col_aciklama = col_borc = col_alacak = col_bakiye = None



        for i in range(min(20, len(df))):

            row_norm = [norm(str(v)) for v in df.iloc[i].fillna('')]

            row_str = ' '.join(row_norm)

            if any(k in row_str for k in ['TARIH','ISLEM','TUTAR','BORC','ALACAK','BAKIYE','DATE','AMOUNT']):

                header_row = i

                for j, v in enumerate(row_norm):

                    if 'TARIH' in v or 'DATE' in v: col_tarih = col_tarih or j

                    if 'ACIKLAMA' in v or 'ISLEM' in v or 'DESCRIPTION' in v or 'DETAY' in v: col_aciklama = col_aciklama or j

                    if 'BORC' in v or 'CIKIS' in v or 'DEBIT' in v: col_borc = col_borc or j

                    if 'ALACAK' in v or 'GIRIS' in v or 'CREDIT' in v: col_alacak = col_alacak or j

                    if 'BAKIYE' in v or 'BALANCE' in v or 'KALAN' in v: col_bakiye = col_bakiye or j

                break



        # Veri satırlarını parse et

        records = []

        bakiye_son = None



        for i in range(header_row + 1, len(df)):

            row = df.iloc[i]

            def gv(col):

                if col is None: return ''

                v = str(row.iloc[col]).strip() if col < len(row) else ''

                return v if v not in ['nan','None','NaN',''] else ''



            tarih     = gv(col_tarih)

            aciklama  = gv(col_aciklama)

            borc_str  = gv(col_borc)

            alacak_str= gv(col_alacak)

            bakiye_str= gv(col_bakiye)



            if not tarih and not borc_str and not alacak_str: continue



            def parse_num(s):

                if not s: return None

                s = s.replace('.','').replace(',','.').replace(' ','').replace('TL','').strip()

                try: return float(s)

                except: return None



            borc   = parse_num(borc_str)

            alacak = parse_num(alacak_str)

            bakiye = parse_num(bakiye_str)

            if bakiye is not None: bakiye_son = bakiye



            if borc is None and alacak is None: continue



            # Tarih normalize

            tarih_fmt = ''

            for fmt in ['%d.%m.%Y','%d/%m/%Y','%Y-%m-%d','%d-%m-%Y','%d.%m.%y']:

                try:

                    from datetime import datetime

                    tarih_fmt = datetime.strptime(tarih.split(' ')[0], fmt).strftime('%d.%m.%Y')

                    break

                except: pass

            if not tarih_fmt: tarih_fmt = tarih[:10] if len(tarih) >= 10 else tarih



            records.append({

                'tarih':     tarih_fmt,

                'aciklama':  aciklama[:80] if aciklama else '',

                'borc':      round(borc, 2) if borc else 0,

                'alacak':    round(alacak, 2) if alacak else 0,

                'bakiye':    round(bakiye, 2) if bakiye else None,

                'banka':     banka,

            })



        return {

            'ok':      True,

            'banka':   banka,

            'records': records,

            'count':   len(records),

            'bakiye_son': round(bakiye_son, 2) if bakiye_son else None,

            'cols': {'tarih': col_tarih, 'aciklama': col_aciklama,

                     'borc': col_borc, 'alacak': col_alacak, 'bakiye': col_bakiye},

        }



    except Exception as e:

        traceback.print_exc()

        return {'ok': False, 'msg': str(e)}



# ── TCMB KUR SERVİSİ ─────────────────────────────────────────────────────────

def tcmb_kur_cek(tarih_str=None):

    """

    TCMB'den kur çeker.

    tarih_str: 'DD.MM.YYYY' formatında, None ise bugün.

    Döndürür: {'USD': {'alis':..,'satis':..,'isim':..}, 'EUR': {...}, ...}

    """

    import urllib.request

    import xml.etree.ElementTree as ET



    try:

        if tarih_str:

            # DD.MM.YYYY → DDMMYYYY ve YYYYMM

            parts = tarih_str.split('.')

            if len(parts) == 3:

                gun, ay, yil = parts[0].zfill(2), parts[1].zfill(2), parts[2]

                dosya = f"{gun}{ay}{yil}"

                klasor = f"{yil}{ay}"

                url = f"https://www.tcmb.gov.tr/kurlar/{klasor}/{dosya}.xml"

            else:

                url = "https://www.tcmb.gov.tr/kurlar/today.xml"

        else:

            url = "https://www.tcmb.gov.tr/kurlar/today.xml"



        req = urllib.request.Request(url, headers={

            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'

        })

        with urllib.request.urlopen(req, timeout=15) as r:

            data = r.read()



        root = ET.fromstring(data)

        tarih_xml = root.attrib.get('Tarih', tarih_str or '')



        kurlar = {}

        for cur in root.findall('Currency'):

            kod  = cur.attrib.get('Kod', cur.attrib.get('CurrencyCode', ''))

            isim = cur.findtext('Isim') or cur.findtext('CurrencyName') or ''

            isim_en = cur.findtext('CurrencyName') or isim



            def parse_kur(tag):

                v = cur.findtext(tag, '')

                try: return round(float(v.replace(',','.')), 4) if v else None

                except: return None



            alis   = parse_kur('ForexBuying')

            satis  = parse_kur('ForexSelling')

            efektif_alis  = parse_kur('BanknoteBuying')

            efektif_satis = parse_kur('BanknoteSelling')



            if kod and (alis or satis):

                kurlar[kod] = {

                    'kod': kod,

                    'isim': isim,

                    'isim_en': isim_en,

                    'alis': alis,

                    'satis': satis,

                    'efektif_alis': efektif_alis,

                    'efektif_satis': efektif_satis,

                    'birim': int(cur.findtext('Unit') or cur.findtext('Birim') or 1),

                }



        return {'ok': True, 'tarih': tarih_xml, 'kurlar': kurlar, 'url': url}



    except Exception as e:

        return {'ok': False, 'hata': str(e), 'tarih': tarih_str or ''}



# ── LOGO TIGER SQL BAĞLANTISI ────────────────────────────────────────────────

def logo_connect():

    """Logo Tiger SQL Server bağlantısı. pyodbc veya pymssql kullanır."""

    if not LOGO_CONFIG.get('server'):

        raise Exception("Logo SQL bağlantı ayarları tanımlanmamış")



    server   = LOGO_CONFIG['server']

    database = LOGO_CONFIG['database']

    username = LOGO_CONFIG.get('username', '')

    password = LOGO_CONFIG.get('password', '')

    port     = LOGO_CONFIG.get('port', 1433)



    # pyodbc ile bağlan (SQL Server ODBC Driver)

    try:

        import pyodbc

        # Farklı driver versiyonlarını dene

        drivers = [

            'ODBC Driver 18 for SQL Server',

            'ODBC Driver 17 for SQL Server',

            'ODBC Driver 13 for SQL Server',

            'SQL Server Native Client 11.0',

            'SQL Server',

        ]

        conn = None

        last_err = None

        for driver in drivers:

            try:

                if username:

                    conn_str = (

                        f"DRIVER={{{driver}}};"

                        f"SERVER={server},{port};"

                        f"DATABASE={database};"

                        f"UID={username};PWD={password};"

                        f"TrustServerCertificate=yes;"

                        f"Encrypt=no;"

                    )

                else:

                    conn_str = (

                        f"DRIVER={{{driver}}};"

                        f"SERVER={server},{port};"

                        f"DATABASE={database};"

                        f"Trusted_Connection=yes;"

                        f"TrustServerCertificate=yes;"

                    )

                conn = pyodbc.connect(conn_str, timeout=10)

                print(f"  Logo bağlantısı OK ({driver})")

                return conn

            except Exception as e:

                last_err = e

                continue

        raise Exception(f"ODBC bağlantısı başarısız: {last_err}")

    except ImportError:

        pass



    # pymssql fallback

    try:

        import pymssql

        conn = pymssql.connect(

            server=server, port=str(port),

            user=username, password=password,

            database=database, timeout=10,

            charset='UTF-8'

        )

        print("  Logo bağlantısı OK (pymssql)")

        return conn

    except ImportError:

        raise Exception(

            "SQL Server bağlantısı için pyodbc veya pymssql kurulu değil. "

            "CMD'de şunu çalıştırın: pip install pyodbc"

        )





def logo_firma_no():

    """LOGO_CONFIG'den varsayılan firma numarası al."""

    return str(LOGO_CONFIG.get('firma_no', '001')).zfill(3)



def logo_firma_listesi():

    """

    LOGO_CONFIG'deki tüm tanımlı firmaları döndür.

    Format: [{'no': '001', 'ad': 'Firma A'}, ...]

    """

    firmalar = LOGO_CONFIG.get('firmalar', [])

    if not firmalar:

        # Eski tekli firma yapısından oku

        no = logo_firma_no()

        ad = LOGO_CONFIG.get('firma_adi', f'Firma {no}')

        return [{'no': no, 'ad': ad}]

    return firmalar



def logo_fatura_cek(tip='satis', limit=500, offset=0, filters=None):

    """

    Logo Tiger'dan çok firma + çok dönem fatura çeker.

    filters:

      - firmalar: ['001','002'] veya None (tümü)

      - yillar:   [2023, 2024] veya None (tümü)

      - baslangic: 'YYYY-MM-DD'

      - bitis:     'YYYY-MM-DD'

      - cari:      cari adı/kodu araması

      - tip_kodlar: [1,2,7,8] özel filtre

    """

    if tip == 'satis':

        tip_kod_default = "INV.TRCODE IN (7,8)"

        tip_label = 'SATIS'

    elif tip == 'alis':

        tip_kod_default = "INV.TRCODE IN (1,2)"

        tip_label = 'ALIS'

    else:

        tip_kod_default = "INV.TRCODE IN (1,2,7,8)"

        tip_label = 'TUMU'



    # Özel tip kodu filtresi

    if filters and filters.get('tip_kodlar'):

        kodlar = ','.join(str(k) for k in filters['tip_kodlar'])

        tip_kod_default = f"INV.TRCODE IN ({kodlar})"



    # Hangi firmalar sorgulanacak

    tum_firmalar = logo_firma_listesi()

    if filters and filters.get('firmalar'):

        sec_firmalar = [f for f in tum_firmalar if f['no'] in filters['firmalar']]

    else:

        sec_firmalar = tum_firmalar



    if not sec_firmalar:

        sec_firmalar = [{'no': logo_firma_no(), 'ad': 'Firma'}]



    # Her firma için ayrı sorgu, sonuçları birleştir

    tum_rows = []



    try:

        conn = logo_connect()



        for firma_cfg in sec_firmalar:

            firma = str(firma_cfg['no']).zfill(3)

            firma_adi = firma_cfg.get('ad', firma)



            where_parts = [tip_kod_default, "INV.CANCELLED = 0"]



            if filters:

                if filters.get('baslangic'):

                    where_parts.append(f"CAST(INV.DATE_ AS DATE) >= '{filters['baslangic']}'")

                if filters.get('bitis'):

                    where_parts.append(f"CAST(INV.DATE_ AS DATE) <= '{filters['bitis']}'")

                if filters.get('yillar'):

                    yil_list = ','.join(str(y) for y in filters['yillar'])

                    where_parts.append(f"YEAR(INV.DATE_) IN ({yil_list})")

                if filters.get('cari'):

                    cari_q = filters['cari'].replace("'", "''")

                    where_parts.append(

                        f"(CL.DEFINITION_ LIKE '%{cari_q}%' OR CL.CODE LIKE '%{cari_q}%')"

                    )

                if filters.get('min_tutar'):

                    where_parts.append(f"INV.NETTOTAL >= {float(filters['min_tutar'])}")

                if filters.get('max_tutar'):

                    where_parts.append(f"INV.NETTOTAL <= {float(filters['max_tutar'])}")



            where = ' AND '.join(where_parts)

            firma_limit = limit if len(sec_firmalar) == 1 else min(limit, limit // len(sec_firmalar) + 200)



            sql = f"""

            SELECT TOP {firma_limit}

                INV.FICHENO        AS fatura_no,

                CONVERT(VARCHAR,INV.DATE_,104) AS tarih,

                YEAR(INV.DATE_)    AS yil,

                MONTH(INV.DATE_)   AS ay,

                CL.DEFINITION_     AS cari_unvan,

                CL.CODE            AS cari_kod,

                INV.GROSSTOTAL     AS brut_toplam,

                INV.TOTALDISCOUNTS AS toplam_iskonto,

                INV.TOTALVAT       AS toplam_kdv,

                INV.NETTOTAL       AS net_toplam,

                INV.CURRSEL        AS doviz_turu,

                INV.TOTNETFC       AS doviz_tutar,

                INV.TRCODE         AS tip_kodu,

                CASE INV.TRCODE

                    WHEN 7 THEN 'Satış Faturası'

                    WHEN 8 THEN 'Satış İade'

                    WHEN 1 THEN 'Alış Faturası'

                    WHEN 2 THEN 'Alış İade'

                    ELSE 'Diğer'

                END                AS tip_adi,

                INV.LOGICALREF     AS id,

                '{firma_adi}'    AS logo_firma_adi,

                '{firma}'        AS logo_firma_no

            FROM LG_{firma}_01_INVOICE INV

            LEFT JOIN LG_{firma}_CLCARD CL ON INV.CLIENTREF = CL.LOGICALREF

            WHERE {where}

            ORDER BY INV.DATE_ DESC, INV.FICHENO DESC

            OFFSET {offset} ROWS FETCH NEXT {firma_limit} ROWS ONLY

            """



            try:

                cursor = conn.cursor()

                cursor.execute(sql)

                cols = [desc[0] for desc in cursor.description]

                for row in cursor.fetchall():

                    d = {}

                    for i2, col in enumerate(cols):

                        v = row[i2]

                        if hasattr(v, 'isoformat'): v = str(v)

                        elif v is None: v = ''

                        d[col] = v

                    tum_rows.append(d)

                cursor.close()

            except Exception as e2:

                print(f"  Firma {firma} fatura hatası: {e2}")



        conn.close()



        # Tarihe göre sırala

        tum_rows.sort(key=lambda x: (x.get('tarih',''), x.get('fatura_no','')), reverse=True)

        if len(tum_rows) > limit:

            tum_rows = tum_rows[:limit]



        return {'ok': True, 'records': tum_rows, 'count': len(tum_rows), 'tip': tip_label}



    except Exception as e:

        traceback.print_exc()

        return {'ok': False, 'msg': str(e), 'records': []}





def logo_stok_cek(q='', limit=300, offset=0, firmalar=None, yillar=None, grup=None, aktif_hareket=False, baslangic=None, bitis=None):

    """Cok firma stok listesi — yil filtresi + donem satis ozeti."""

    tum_firmalar = logo_firma_listesi()

    sec = [f for f in tum_firmalar if f['no'] in firmalar] if firmalar else tum_firmalar

    if not sec: sec = [{'no': logo_firma_no(), 'ad': 'Firma'}]



    tum_rows = []

    try:

        conn = logo_connect()

        for firma_cfg in sec:

            firma     = str(firma_cfg['no']).zfill(3)

            firma_adi = firma_cfg.get('ad', firma)



            where = "IT.ACTIVE = 0"

            if q:

                q2 = q.replace("'","''")

                where += f" AND (IT.CODE LIKE '%{q2}%' OR IT.DEFINITION_ LIKE '%{q2}%')"

            if grup:

                where += f" AND IT.STGRPCODE = '{grup.replace(chr(39),chr(39)*2)}'"



            yil_filter = ""

            tarih_parts = []

            if yillar:

                yil_list = ','.join(str(y) for y in yillar)

                tarih_parts.append(f"YEAR(INV_S.DATE_) IN ({yil_list})")

            if baslangic:

                tarih_parts.append(f"CAST(INV_S.DATE_ AS DATE) >= '{baslangic}'")

            if bitis:

                tarih_parts.append(f"CAST(INV_S.DATE_ AS DATE) <= '{bitis}'")

            if tarih_parts:

                yil_filter = "AND " + " AND ".join(tarih_parts)

            if aktif_hareket or baslangic or bitis:

                extra_cond = f"YEAR(INV2.DATE_) IN ({','.join(str(y) for y in yillar)})" if yillar else "1=1"

                bas_c = f"AND CAST(INV2.DATE_ AS DATE) >= '{baslangic}'" if baslangic else ""

                bit_c = f"AND CAST(INV2.DATE_ AS DATE) <= '{bitis}'" if bitis else ""

                where += (f" AND EXISTS ("

                    f"SELECT 1 FROM LG_{firma}_01_STLINE SL2 "

                    f"INNER JOIN LG_{firma}_01_INVOICE INV2 ON SL2.INVOICEREF=INV2.LOGICALREF "

                    f"WHERE SL2.STOCKREF=IT.LOGICALREF AND ({extra_cond}) {bas_c} {bit_c})")



            q_str = (

                f"SELECT TOP {limit} "

                f"IT.CODE AS stok_kodu, IT.DEFINITION_ AS stok_adi, "

                f"IT.STGRPCODE AS stok_grubu, IT.UNITSETCODE AS birim, IT.SPECODE AS ozel_kod, "

                f"IT.LOGICALREF AS id, IT.LASTPURCHPRICE AS son_alis_fiyati, "

                f"IT.LASTSELLINGPRICE AS son_satis_fiyati, "

                f"ISNULL((SELECT SUM(SL_S.AMOUNT) FROM LG_{firma}_01_STLINE SL_S "

                f"INNER JOIN LG_{firma}_01_INVOICE INV_S ON SL_S.INVOICEREF=INV_S.LOGICALREF "

                f"WHERE SL_S.STOCKREF=IT.LOGICALREF AND INV_S.TRCODE IN (7,8) "

                f"AND INV_S.CANCELLED=0 {yil_filter}),0) AS donem_satis_adet, "

                f"ISNULL((SELECT SUM(SL_S.LINENET) FROM LG_{firma}_01_STLINE SL_S "

                f"INNER JOIN LG_{firma}_01_INVOICE INV_S ON SL_S.INVOICEREF=INV_S.LOGICALREF "

                f"WHERE SL_S.STOCKREF=IT.LOGICALREF AND INV_S.TRCODE IN (7,8) "

                f"AND INV_S.CANCELLED=0 {yil_filter}),0) AS donem_satis_tutar, "

                f"ISNULL((SELECT SUM(SL_S.AMOUNT) FROM LG_{firma}_01_STLINE SL_S "

                f"INNER JOIN LG_{firma}_01_INVOICE INV_S ON SL_S.INVOICEREF=INV_S.LOGICALREF "

                f"WHERE SL_S.STOCKREF=IT.LOGICALREF AND INV_S.TRCODE IN (1,2) "

                f"AND INV_S.CANCELLED=0 {yil_filter}),0) AS donem_alis_adet, "

                f"ISNULL((SELECT SUM(CASE WHEN SL_C.IOCODE IN (1,3) THEN SL_C.AMOUNT ELSE -SL_C.AMOUNT END) "

                f"FROM LG_{firma}_01_STLINE SL_C WHERE SL_C.STOCKREF=IT.LOGICALREF "

                f"AND SL_C.LINETYPE=0),0) AS stok_bakiye, "

                f"'{firma_adi}' AS logo_firma_adi, '{firma}' AS logo_firma_no "

                f"FROM LG_{firma}_ITEMS IT WHERE {where} "

                f"ORDER BY IT.CODE OFFSET {offset} ROWS FETCH NEXT {limit} ROWS ONLY"

            )

            try:

                cursor = conn.cursor(); cursor.execute(q_str)

                cols = [d[0] for d in cursor.description]

                for row in cursor.fetchall():

                    d = {}

                    for i2, col in enumerate(cols):

                        v = row[i2]

                        if v is None: v = ''

                        elif isinstance(v, float): v = round(v, 4)

                        d[col] = v

                    tum_rows.append(d)

                cursor.close()

            except Exception as e2:

                print(f"  Firma {firma} stok hatasi: {e2}")

        conn.close()

        tum_rows.sort(key=lambda x: (str(x.get('logo_firma_no','')), str(x.get('stok_kodu',''))))

        return {'ok': True, 'records': tum_rows[:limit], 'count': len(tum_rows)}

    except Exception as e:

        traceback.print_exc()

        return {'ok': False, 'msg': str(e), 'records': []}



def logo_cari_cek(q='', limit=500, offset=0, tip='', firmalar=None, yillar=None, baslangic=None, bitis=None):

    """

    Cok firma cari listesi — yil filtresi + donem ciro ozeti.

    yillar: o yillarda faturasi olan cariler + donem cirolarini gosterir.

    """

    tum_firmalar = logo_firma_listesi()

    sec = [f for f in tum_firmalar if f['no'] in firmalar] if firmalar else tum_firmalar

    if not sec: sec = [{'no': logo_firma_no(), 'ad': 'Firma'}]



    tum_rows = []

    try:

        conn = logo_connect()

        for firma_cfg in sec:

            firma     = str(firma_cfg['no']).zfill(3)

            firma_adi = firma_cfg.get('ad', firma)



            where = "CL.ACTIVE = 0"

            if q:

                q2 = q.replace("'","''")

                where += f" AND (CL.CODE LIKE '%{q2}%' OR CL.DEFINITION_ LIKE '%{q2}%' OR CL.TAXNR LIKE '%{q2}%')"

            if tip == 'musteri':    where += " AND CL.ACCCLASS IN (2,3)"

            elif tip == 'tedarikci': where += " AND CL.ACCCLASS IN (1,3)"



            yil_filter = ""

            tarih_parts_c = []

            if yillar:

                yil_list = ','.join(str(y) for y in yillar)

                tarih_parts_c.append(f"YEAR(INV_C.DATE_) IN ({yil_list})")

            if baslangic:

                tarih_parts_c.append(f"CAST(INV_C.DATE_ AS DATE) >= '{baslangic}'")

            if bitis:

                tarih_parts_c.append(f"CAST(INV_C.DATE_ AS DATE) <= '{bitis}'")

            if tarih_parts_c:

                yil_filter = "AND " + " AND ".join(tarih_parts_c)

            if yillar or baslangic or bitis:

                bas_c2 = f"AND CAST(INV_F.DATE_ AS DATE) >= '{baslangic}'" if baslangic else ""

                bit_c2 = f"AND CAST(INV_F.DATE_ AS DATE) <= '{bitis}'" if bitis else ""

                yil_c2 = f"AND YEAR(INV_F.DATE_) IN ({','.join(str(y) for y in yillar)})" if yillar else ""

                where += (f" AND EXISTS (SELECT 1 FROM LG_{firma}_01_INVOICE INV_F "

                          f"WHERE INV_F.CLIENTREF=CL.LOGICALREF AND INV_F.CANCELLED=0 "

                          f"{yil_c2} {bas_c2} {bit_c2})")



            q_str = (

                f"SELECT TOP {limit} "

                f"CL.CODE AS cari_kodu, CL.DEFINITION_ AS cari_adi, "

                f"CL.TAXNR AS vergi_no, CL.TAXOFFICE AS vergi_dairesi, "

                f"CL.ADDR1 AS adres, CL.CITY AS sehir, "

                f"CL.TELNRS1 AS telefon, CL.EMAILADDR AS email, "

                f"CASE CL.ACCCLASS WHEN 1 THEN 'Tedarikci' WHEN 2 THEN 'Musteri' "

                f"WHEN 3 THEN 'Hem Musteri Hem Tedarikci' ELSE 'Diger' END AS cari_tipi, "

                f"CL.LOGICALREF AS id, "

                # Dönem satış ciro

                f"ISNULL((SELECT SUM(INV_C.NETTOTAL) FROM LG_{firma}_01_INVOICE INV_C "

                f"WHERE INV_C.CLIENTREF=CL.LOGICALREF AND INV_C.TRCODE IN (7,8) "

                f"AND INV_C.CANCELLED=0 {yil_filter}),0) AS donem_satis_ciro, "

                # Dönem alış ciro

                f"ISNULL((SELECT SUM(INV_C.NETTOTAL) FROM LG_{firma}_01_INVOICE INV_C "

                f"WHERE INV_C.CLIENTREF=CL.LOGICALREF AND INV_C.TRCODE IN (1,2) "

                f"AND INV_C.CANCELLED=0 {yil_filter}),0) AS donem_alis_ciro, "

                # Dönem fatura sayısı

                f"ISNULL((SELECT COUNT(*) FROM LG_{firma}_01_INVOICE INV_C "

                f"WHERE INV_C.CLIENTREF=CL.LOGICALREF AND INV_C.CANCELLED=0 "

                f"{yil_filter}),0) AS donem_fatura_sayisi, "

                # Güncel bakiye

                f"ISNULL((SELECT SUM(CASE WHEN CF.SIGN=0 THEN CF.AMOUNT ELSE -CF.AMOUNT END) "

                f"FROM LG_{firma}_01_CLFLINE CF WHERE CF.CLIENTREF=CL.LOGICALREF),0) AS cari_bakiye, "

                f"'{firma_adi}' AS logo_firma_adi, '{firma}' AS logo_firma_no "

                f"FROM LG_{firma}_CLCARD CL "

                f"WHERE {where} "

                f"ORDER BY CL.CODE "

                f"OFFSET {offset} ROWS FETCH NEXT {limit} ROWS ONLY"

            )

            try:

                cursor = conn.cursor(); cursor.execute(q_str)

                cols = [d[0] for d in cursor.description]

                for row in cursor.fetchall():

                    d = {}

                    for i2, col in enumerate(cols):

                        v = row[i2]

                        if v is None: v = ''

                        elif isinstance(v, float): v = round(v, 2)

                        d[col] = v

                    tum_rows.append(d)

                cursor.close()

            except Exception as e2:

                print(f"  Firma {firma} cari hatasi: {e2}")

        conn.close()

        tum_rows.sort(key=lambda x: (str(x.get('logo_firma_no','')), str(x.get('cari_kodu',''))))

        return {'ok': True, 'records': tum_rows[:limit], 'count': len(tum_rows)}

    except Exception as e:

        traceback.print_exc()

        return {'ok': False, 'msg': str(e), 'records': []}



def logo_cari_hesap_ozet(cari_kod):

    """Cari hesap bakiye ve hareket özeti."""

    firma = logo_firma_no()

    sql = f"""

    SELECT TOP 1

        CL.CODE, CL.DEFINITION_,

        (SELECT ISNULL(SUM(CASE WHEN SIGN=0 THEN AMOUNT ELSE -AMOUNT END),0)

         FROM LG_{firma}_01_CLFLINE

         WHERE CLIENTREF=CL.LOGICALREF) AS BAKIYE,

        (SELECT COUNT(*) FROM LG_{firma}_01_INVOICE

         WHERE CLIENTREF=CL.LOGICALREF AND CANCELLED=0) AS FATURA_ADET

    FROM LG_{firma}_CLCARD CL

    WHERE CL.CODE = '{cari_kod.replace("'","''")}'

    """

    try:

        conn = logo_connect(); cur = conn.cursor(); cur.execute(sql)

        row = cur.fetchone()

        cur.close(); conn.close()

        if row:

            return {'ok':True,'bakiye':float(row[2] or 0),'fatura_adet':int(row[3] or 0)}

        return {'ok':False,'msg':'Cari bulunamadı'}

    except Exception as e:

        return {'ok':False,'msg':str(e)}





# ── GMAIL API ─────────────────────────────────────────────────────────────────

def gmail_get_service():

    """Gmail API servisini döndürür. Token yoksa None döner."""

    try:

        from google.oauth2.credentials import Credentials

        from googleapiclient.discovery import build



        if not os.path.exists(GMAIL_TOKEN_FILE):

            return None, "Token bulunamadı — önce OAuth ile giriş yapın"



        with open(GMAIL_TOKEN_FILE, 'r') as f:

            token_data = json.load(f)



        creds = Credentials(

            token=token_data.get('token'),

            refresh_token=token_data.get('refresh_token'),

            token_uri='https://oauth2.googleapis.com/token',

            client_id=GMAIL_CONFIG.get('client_id'),

            client_secret=GMAIL_CONFIG.get('client_secret'),

        )

        service = build('gmail', 'v1', credentials=creds)

        return service, None

    except ImportError:

        return None, "google-api-python-client kurulu değil. CMD'de: pip install google-api-python-client google-auth-oauthlib"

    except Exception as e:

        return None, str(e)





def gmail_mail_listesi(max_results=50, query=''):

    """Gelen kutusu maillerini listeler."""

    service, err = gmail_get_service()

    if not service:

        return {'ok': False, 'msg': err}

    try:

        q = f"in:inbox {query}".strip()

        result = service.users().messages().list(

            userId='me', maxResults=max_results, q=q

        ).execute()

        messages = result.get('messages', [])



        mails = []

        for msg in messages[:max_results]:

            m = service.users().messages().get(

                userId='me', id=msg['id'],

                format='metadata',

                metadataHeaders=['From','To','Subject','Date']

            ).execute()

            headers = {h['name']: h['value'] for h in m.get('payload',{}).get('headers',[])}

            snippet = m.get('snippet','')

            labels  = m.get('labelIds',[])

            mails.append({

                'id':       msg['id'],

                'kimden':   headers.get('From',''),

                'kime':     headers.get('To',''),

                'konu':     headers.get('Subject',''),

                'tarih':    headers.get('Date',''),

                'snippet':  snippet[:150],

                'okunmadi': 'UNREAD' in labels,

            })

        return {'ok': True, 'mails': mails, 'count': len(mails)}

    except Exception as e:

        return {'ok': False, 'msg': str(e)}





def gmail_mail_oku(mail_id):

    """Mail içeriğini okur."""

    service, err = gmail_get_service()

    if not service:

        return {'ok': False, 'msg': err}

    try:

        import base64

        m = service.users().messages().get(

            userId='me', id=mail_id, format='full'

        ).execute()

        headers = {h['name']: h['value']

                   for h in m.get('payload',{}).get('headers',[])}



        # Body çıkar

        def get_body(payload):

            if payload.get('body',{}).get('data'):

                return base64.urlsafe_b64decode(

                    payload['body']['data'] + '=='

                ).decode('utf-8', errors='replace')

            for part in payload.get('parts', []):

                if part.get('mimeType') == 'text/plain':

                    data = part.get('body',{}).get('data','')

                    if data:

                        return base64.urlsafe_b64decode(

                            data + '=='

                        ).decode('utf-8', errors='replace')

            for part in payload.get('parts', []):

                if part.get('mimeType') == 'text/html':

                    data = part.get('body',{}).get('data','')

                    if data:

                        return base64.urlsafe_b64decode(

                            data + '=='

                        ).decode('utf-8', errors='replace')

            return ''



        body = get_body(m.get('payload', {}))



        # Okundu işaretle

        service.users().messages().modify(

            userId='me', id=mail_id,

            body={'removeLabelIds': ['UNREAD']}

        ).execute()



        return {

            'ok':     True,

            'kimden': headers.get('From',''),

            'kime':   headers.get('To',''),

            'konu':   headers.get('Subject',''),

            'tarih':  headers.get('Date',''),

            'body':   body[:8000],

        }

    except Exception as e:

        return {'ok': False, 'msg': str(e)}





def gmail_mail_gonder(kime, konu, body):

    """Mail gönderir."""

    service, err = gmail_get_service()

    if not service:

        return {'ok': False, 'msg': err}

    try:

        import base64

        from email.mime.text import MIMEText

        msg = MIMEText(body, 'plain', 'utf-8')

        msg['to']      = kime

        msg['from']    = GMAIL_CONFIG.get('email', 'me')

        msg['subject'] = konu

        raw = base64.urlsafe_b64encode(msg.as_bytes()).decode()

        service.users().messages().send(

            userId='me', body={'raw': raw}

        ).execute()

        return {'ok': True}

    except Exception as e:

        return {'ok': False, 'msg': str(e)}





def gmail_oauth_url():

    """OAuth onay URL'si üretir."""

    try:

        from google_auth_oauthlib.flow import Flow

        if not GMAIL_CONFIG.get('client_id') or not GMAIL_CONFIG.get('client_secret'):

            return {'ok': False, 'msg': 'Client ID ve Secret girilmemiş'}



        flow = Flow.from_client_config(

            {'web': {

                'client_id':     GMAIL_CONFIG['client_id'],

                'client_secret': GMAIL_CONFIG['client_secret'],

                'auth_uri':  'https://accounts.google.com/o/oauth2/auth',

                'token_uri': 'https://oauth2.googleapis.com/token',

                'redirect_uris': ['http://localhost:5678/api/gmail/oauth/callback'],

            }},

            scopes=['https://www.googleapis.com/auth/gmail.modify']

        )

        flow.redirect_uri = 'http://localhost:5678/api/gmail/oauth/callback'

        url, state = flow.authorization_url(

            access_type='offline', include_granted_scopes='true', prompt='consent'

        )

        return {'ok': True, 'url': url}

    except ImportError:

        return {'ok': False, 'msg': 'google-auth-oauthlib kurulu değil. CMD: pip install google-auth-oauthlib'}

    except Exception as e:

        return {'ok': False, 'msg': str(e)}





def gmail_oauth_callback(code):

    """OAuth callback — token kaydeder."""

    try:

        from google_auth_oauthlib.flow import Flow

        flow = Flow.from_client_config(

            {'web': {

                'client_id':     GMAIL_CONFIG['client_id'],

                'client_secret': GMAIL_CONFIG['client_secret'],

                'auth_uri':  'https://accounts.google.com/o/oauth2/auth',

                'token_uri': 'https://oauth2.googleapis.com/token',

                'redirect_uris': ['http://localhost:5678/api/gmail/oauth/callback'],

            }},

            scopes=['https://www.googleapis.com/auth/gmail.modify']

        )

        flow.redirect_uri = 'http://localhost:5678/api/gmail/oauth/callback'

        flow.fetch_token(code=code)

        creds = flow.credentials

        token_data = {

            'token':         creds.token,

            'refresh_token': creds.refresh_token,

            'token_uri':     creds.token_uri,

            'client_id':     creds.client_id,

            'client_secret': creds.client_secret,

        }

        with open(GMAIL_TOKEN_FILE, 'w') as f:

            json.dump(token_data, f)

        return {'ok': True}

    except Exception as e:

        return {'ok': False, 'msg': str(e)}



# ── OPENAI / ChatGPT ENTEGRASYoNU ────────────────────────────────────────────

def ai_chat(messages, model=None):

    """

    OpenAI API ile chat. messages: [{role, content}]

    Finansal verilerle zenginleştirilmiş sistem prompt kullanır.

    """

    import urllib.request, urllib.error



    api_key = AI_CONFIG.get('api_key', '')

    if not api_key:

        return {'ok': False, 'msg': 'API key girilmemiş — AI Asistan ayarlarından OpenAI API key girin'}



    model = model or AI_CONFIG.get('model', 'gpt-4o-mini')



    # Sistem prompt: nakit akış bağlamı

    system_prompt = """Sen Nakit Akış Yönetim Sistemi'nin yapay zeka asistanısın.

Kullanıcı bir finans yöneticisi. Şirket grupları: ULUSAL, ABC, BRG, BRK.

Görevin:

- Nakit akış, finansman, likidite konularında uzmanca yardım etmek

- Fatura, ödeme, kredi limit sorularını yanıtlamak

- Finansal karar desteği sunmak

- Türkçe olarak net ve profesyonel cevaplar vermek

- Gerektiğinde finansal formüller ve hesaplamalar yapmak

Cevaplarında pratik ve uygulanabilir öneriler ver."""



    # Mevcut finansal bağlamı ekle

    try:

        toplam_gider  = sum(d.get('tutar', 0) for d in DATA)

        odenmemis     = sum(d.get('tutar', 0) for d in DATA if d.get('durum') != 'ÖDENDİ')

        toplam_gelir  = sum(d.get('tutar', 0) for d in GELIR_DATA)

        toplam_banka  = sum(float(b.get('bakiye', 0) or 0) for b in BANKA_DATA)

        kalan_kredi   = sum(float(k.get('kalan_limit', 0) or 0) for k in KREDI_DATA if k.get('aktif', True))

        system_prompt += f"""



Mevcut finansal durum (anlık):

- Toplam gider kaydı: {len(DATA)} adet, {toplam_gider:,.0f} ₺

- Ödenmemiş gider: {odenmemis:,.0f} ₺

- Toplam gelir: {toplam_gelir:,.0f} ₺

- Banka bakiyesi toplamı: {toplam_banka:,.0f} ₺

- Kullanılabilir kredi limiti: {kalan_kredi:,.0f} ₺

- Net pozisyon: {(toplam_gelir - odenmemis):,.0f} ₺"""

    except:

        pass



    payload = {

        'model': model,

        'messages': [{'role': 'system', 'content': system_prompt}] + messages,

        'max_tokens': 2000,

        'temperature': 0.7,

    }



    try:

        req = urllib.request.Request(

            'https://api.openai.com/v1/chat/completions',

            data=json.dumps(payload).encode('utf-8'),

            headers={

                'Authorization': f'Bearer {api_key}',

                'Content-Type': 'application/json',

            },

            method='POST'

        )

        with urllib.request.urlopen(req, timeout=30) as resp:

            result = json.loads(resp.read().decode('utf-8'))

        content_text = result['choices'][0]['message']['content']

        return {'ok': True, 'content': content_text, 'model': model}

    except urllib.error.HTTPError as e:

        err_body = e.read().decode('utf-8', errors='replace')

        try:

            err_json = json.loads(err_body)

            err_msg = err_json.get('error', {}).get('message', err_body[:200])

        except:

            err_msg = err_body[:200]

        return {'ok': False, 'msg': f'OpenAI API hatası ({e.code}): {err_msg}'}

    except Exception as e:

        return {'ok': False, 'msg': str(e)}



# ── SQLite VERİTABANI ─────────────────────────────────────────────────────────

import sqlite3 as _sqlite3



DB_FILE = os.path.join(APP_DIR, "nakit_akis.db")



def db_connect():

    conn = _sqlite3.connect(DB_FILE, timeout=10)

    conn.row_factory = _sqlite3.Row

    conn.execute("PRAGMA journal_mode=WAL")

    conn.execute("PRAGMA foreign_keys=ON")

    return conn



def db_init():

    """Tabloları oluştur (ilk çalıştırmada)."""

    conn = db_connect()

    c = conn.cursor()

    c.executescript("""

        CREATE TABLE IF NOT EXISTS giderler (

            id          INTEGER PRIMARY KEY AUTOINCREMENT,

            firma       TEXT,

            tur         TEXT,

            yer         TEXT,

            tarih       TEXT,

            ay          TEXT,

            yil         TEXT,

            durum       TEXT DEFAULT 'ÖDENMEDİ',

            anapara     REAL DEFAULT 0,

            faiz        REAL DEFAULT 0,

            bsmv        REAL DEFAULT 0,

            tutar       REAL DEFAULT 0,

            aciklama    TEXT,

            kaynak      TEXT DEFAULT 'manuel',

            created_at  TEXT DEFAULT (datetime('now','localtime')),

            updated_at  TEXT DEFAULT (datetime('now','localtime'))

        );

        CREATE TABLE IF NOT EXISTS gelirler (

            id          INTEGER PRIMARY KEY AUTOINCREMENT,

            firma       TEXT,

            tur         TEXT,

            aciklama    TEXT,

            tarih       TEXT,

            ay          TEXT,

            durum       TEXT DEFAULT 'BEKLENİYOR',

            tutar       REAL DEFAULT 0,

            created_at  TEXT DEFAULT (datetime('now','localtime')),

            updated_at  TEXT DEFAULT (datetime('now','localtime'))

        );

        CREATE TABLE IF NOT EXISTS banka_bakiyeleri (

            id          INTEGER PRIMARY KEY AUTOINCREMENT,

            firma       TEXT,

            banka       TEXT,

            hesap_turu  TEXT,

            bakiye      REAL DEFAULT 0,

            tarih       TEXT,

            kaynak      TEXT DEFAULT 'manuel',

            updated_at  TEXT DEFAULT (datetime('now','localtime'))

        );

        CREATE TABLE IF NOT EXISTS kredi_limitleri (

            id          INTEGER PRIMARY KEY AUTOINCREMENT,

            firma       TEXT,

            banka       TEXT,

            tur         TEXT,

            limit_tutar REAL DEFAULT 0,

            kullanilan  REAL DEFAULT 0,

            kalan       REAL DEFAULT 0,

            faiz_yillik REAL DEFAULT 0,

            bitis_tarihi TEXT,

            aktif       INTEGER DEFAULT 1,

            updated_at  TEXT DEFAULT (datetime('now','localtime'))

        );

        CREATE TABLE IF NOT EXISTS cari_hareketler (

            id          INTEGER PRIMARY KEY AUTOINCREMENT,

            cari_firma  TEXT,

            grup_firma  TEXT,

            tur         TEXT,

            belge_no    TEXT,

            tarih       TEXT,

            vade        TEXT,

            tutar       REAL DEFAULT 0,

            durum       TEXT DEFAULT 'AÇIK',

            aciklama    TEXT,

            created_at  TEXT DEFAULT (datetime('now','localtime')),

            updated_at  TEXT DEFAULT (datetime('now','localtime'))

        );

        CREATE TABLE IF NOT EXISTS notlar (

            id          INTEGER PRIMARY KEY AUTOINCREMENT,

            baslik      TEXT,

            metin       TEXT,

            renk        TEXT DEFAULT 'default',

            created_at  TEXT DEFAULT (datetime('now','localtime'))

        );

        CREATE TABLE IF NOT EXISTS sync_log (

            id          INTEGER PRIMARY KEY AUTOINCREMENT,

            tablo       TEXT,

            islem       TEXT,

            kayit_sayisi INTEGER,

            durum       TEXT,

            mesaj       TEXT,

            created_at  TEXT DEFAULT (datetime('now','localtime'))

        );

    """)

    conn.commit()

    conn.close()

    print(f"  SQLite DB hazır: {DB_FILE}")



def db_sync_all():

    """Tüm JSON verilerini SQLite'a senkronize eder."""

    results = {}



    conn = db_connect()

    c = conn.cursor()



    # Giderler

    try:

        c.execute("DELETE FROM giderler")

        for rec in DATA:

            c.execute("""INSERT INTO giderler

                (firma,tur,yer,tarih,ay,yil,durum,anapara,faiz,bsmv,tutar,aciklama,kaynak)

                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)""", (

                rec.get('firma',''), rec.get('tur',''), rec.get('yer',''),

                rec.get('tarih',''), rec.get('ay',''),

                rec.get('tarih','')[6:10] if len(rec.get('tarih',''))>=10 else '',

                rec.get('durum','ÖDENMEDİ'),

                float(rec.get('anapara',0) or 0), float(rec.get('faiz',0) or 0),

                float(rec.get('bsmv',0) or 0), float(rec.get('tutar',0) or 0),

                rec.get('aciklama',''), 'excel'

            ))

        results['giderler'] = len(DATA)

    except Exception as e:

        results['giderler_hata'] = str(e)



    # Gelirler

    try:

        c.execute("DELETE FROM gelirler")

        for rec in GELIR_DATA:

            c.execute("""INSERT INTO gelirler

                (firma,tur,aciklama,tarih,ay,durum,tutar)

                VALUES (?,?,?,?,?,?,?)""", (

                rec.get('firma',''), rec.get('tur',''), rec.get('aciklama',''),

                rec.get('tarih',''), rec.get('ay',''),

                rec.get('durum','BEKLENİYOR'), float(rec.get('tutar',0) or 0)

            ))

        results['gelirler'] = len(GELIR_DATA)

    except Exception as e:

        results['gelirler_hata'] = str(e)



    # Banka

    try:

        c.execute("DELETE FROM banka_bakiyeleri")

        for rec in BANKA_DATA:

            c.execute("""INSERT INTO banka_bakiyeleri

                (firma,banka,hesap_turu,bakiye,tarih,kaynak)

                VALUES (?,?,?,?,?,?)""", (

                rec.get('firma',''), rec.get('banka',''),

                rec.get('hesap_turu',''), float(rec.get('bakiye',0) or 0),

                rec.get('tarih',''), rec.get('kaynak','manuel')

            ))

        results['banka'] = len(BANKA_DATA)

    except Exception as e:

        results['banka_hata'] = str(e)



    # Kredi

    try:

        c.execute("DELETE FROM kredi_limitleri")

        for rec in KREDI_DATA:

            c.execute("""INSERT INTO kredi_limitleri

                (firma,banka,tur,limit_tutar,kullanilan,kalan,faiz_yillik,bitis_tarihi,aktif)

                VALUES (?,?,?,?,?,?,?,?,?)""", (

                rec.get('firma',''), rec.get('banka',''), rec.get('tur',''),

                float(rec.get('limit',0) or 0), float(rec.get('kullanilan',0) or 0),

                float(rec.get('kalan_limit',0) or 0), float(rec.get('faiz_yillik',0) or 0),

                rec.get('bitis',''), 1 if rec.get('aktif',True) else 0

            ))

        results['kredi'] = len(KREDI_DATA)

    except Exception as e:

        results['kredi_hata'] = str(e)



    # Cari

    try:

        c.execute("DELETE FROM cari_hareketler")

        for rec in CARI_DATA:

            c.execute("""INSERT INTO cari_hareketler

                (cari_firma,grup_firma,tur,belge_no,tarih,vade,tutar,durum,aciklama)

                VALUES (?,?,?,?,?,?,?,?,?)""", (

                rec.get('cari_firma',''), rec.get('grup_firma',''), rec.get('tur',''),

                rec.get('belge_no',''), rec.get('tarih',''), rec.get('vade',''),

                float(rec.get('tutar',0) or 0), rec.get('durum','AÇIK'), rec.get('aciklama','')

            ))

        results['cari'] = len(CARI_DATA)

    except Exception as e:

        results['cari_hata'] = str(e)



    # Notlar

    try:

        c.execute("DELETE FROM notlar")

        for rec in NOTLAR:

            c.execute("INSERT INTO notlar (id,baslik,metin,renk,created_at) VALUES (?,?,?,?,?)", (

                rec.get('id'), rec.get('baslik',''), rec.get('metin',''),

                rec.get('renk','default'), rec.get('tarih','')

            ))

        results['notlar'] = len(NOTLAR)

    except Exception as e:

        results['notlar_hata'] = str(e)



    # Log

    c.execute("""INSERT INTO sync_log (tablo,islem,kayit_sayisi,durum,mesaj)

        VALUES ('all','full_sync',?,?,?)""", (

        sum(v for v in results.values() if isinstance(v,int)),

        'OK', str(results)

    ))



    conn.commit()

    conn.close()



    from datetime import datetime as _dt

    print(f"  DB sync: {results} [{_dt.now().strftime('%H:%M:%S')}]")

    return results



def db_query(sql, params=(), fetchall=True):

    """Genel SELECT sorgusu."""

    conn = db_connect()

    c = conn.cursor()

    c.execute(sql, params)

    if fetchall:

        rows = [dict(r) for r in c.fetchall()]

    else:

        row = c.fetchone()

        rows = dict(row) if row else {}

    conn.close()

    return rows



# ── OTEL YÖNETİMİ ────────────────────────────────────────────────────────────

def otel_istatistik(otel_id, baslangic=None, bitis=None):

    """Belirtilen otel için istatistik hesaplar."""

    from datetime import datetime as _dt, timedelta as _td

    rezervasyonlar = OTEL_DATA.get(otel_id, [])

    cfg = OTEL_CONFIG.get(otel_id, {})

    oda_sayisi = int(cfg.get('oda_sayisi', 50))



    bugun = _dt.now().date()

    if baslangic:

        try: bas = _dt.strptime(baslangic, '%Y-%m-%d').date()

        except: bas = bugun - _td(days=30)

    else:

        bas = bugun - _td(days=30)

    if bitis:

        try: bit = _dt.strptime(bitis, '%Y-%m-%d').date()

        except: bit = bugun + _td(days=30)

    else:

        bit = bugun + _td(days=30)



    toplam_gelir = 0

    toplam_gece  = 0

    kanal_map    = {}

    durum_map    = {'checkin':0,'checkout':0,'konakliyor':0,'iptal':0,'rezerve':0}

    bugun_checkin  = []

    bugun_checkout = []

    aktif_rezerve  = []



    for r in rezervasyonlar:

        try:

            giris = _dt.strptime(r.get('giris_tarihi',''), '%Y-%m-%d').date()

            cikis = _dt.strptime(r.get('cikis_tarihi',''), '%Y-%m-%d').date()

        except:

            continue



        gece  = max(1, (cikis - giris).days)

        tutar = float(r.get('tutar', 0) or 0)

        kanal = r.get('kanal', 'Direkt') or 'Direkt'

        durum = r.get('durum', 'Rezerve') or 'Rezerve'

        dk    = durum.lower().replace(' ','').replace('-','')



        if giris == bugun: bugun_checkin.append(r)

        if cikis == bugun: bugun_checkout.append(r)

        if giris <= bugun < cikis and 'iptal' not in dk:

            aktif_rezerve.append(r)



        if 'iptal' in dk: durum_map['iptal'] += 1

        elif giris == bugun: durum_map['checkin'] += 1

        elif cikis == bugun: durum_map['checkout'] += 1

        elif giris <= bugun < cikis: durum_map['konakliyor'] += 1

        else: durum_map['rezerve'] += 1



        if not (giris <= bit and cikis >= bas): continue

        if 'iptal' in dk: continue



        toplam_gelir += tutar

        toplam_gece  += gece

        kanal_map[kanal] = kanal_map.get(kanal, 0) + tutar



    dolu_oda = len(aktif_rezerve)

    doluluk  = round(dolu_oda / oda_sayisi * 100, 1) if oda_sayisi else 0



    # Günlük doluluk takvimi (geçmiş 7 + gelecek 53 gün)

    takvim = []

    for i in range(60):

        gun = bugun + _td(days=i-7)

        gun_str = gun.strftime('%Y-%m-%d')

        dolu = 0

        for r in rezervasyonlar:

            if 'iptal' in (r.get('durum','').lower()):

                continue

            try:

                _g = _dt.strptime(r.get('giris_tarihi',''), '%Y-%m-%d').date()

                _c = _dt.strptime(r.get('cikis_tarihi',''), '%Y-%m-%d').date()

                if _g <= gun < _c:

                    dolu += 1

            except:

                pass

        takvim.append({

            'tarih': gun_str,

            'dolu':  dolu,

            'bos':   max(0, oda_sayisi - dolu),

            'oran':  round(dolu/oda_sayisi*100,1) if oda_sayisi else 0,

        })



    return {

        'toplam_gelir':    round(toplam_gelir, 2),

        'toplam_gece':     toplam_gece,

        'toplam_rezerv':   len(rezervasyonlar),

        'aktif_konak':     len(aktif_rezerve),

        'doluluk_oran':    doluluk,

        'oda_sayisi':      oda_sayisi,

        'bugun_checkin':   len(bugun_checkin),

        'bugun_checkout':  len(bugun_checkout),

        'kanal_dagilim':   kanal_map,

        'durum_dagilim':   durum_map,

        'bugun_checkin_list':  bugun_checkin[:20],

        'bugun_checkout_list': bugun_checkout[:20],

        'takvim':          takvim,

    }



# ── SATIŞ ANALİZİ ────────────────────────────────────────────────────────────

def logo_satis_analiz(filters=None):

    """

    Logo'dan satış analizi için fatura kalemleri çeker.

    Çok firma + çok yıl destekli.

    """

    tum_firmalar = logo_firma_listesi()

    if filters and filters.get('firmalar'):

        sec = [f for f in tum_firmalar if f['no'] in filters['firmalar']]

    else:

        sec = tum_firmalar

    if not sec:

        sec = [{'no': logo_firma_no(), 'ad': 'Firma'}]



    where_extra = []

    if filters:

        if filters.get('baslangic'):

            where_extra.append(f"CAST(INV.DATE_ AS DATE) >= '{filters['baslangic']}'")

        if filters.get('bitis'):

            where_extra.append(f"CAST(INV.DATE_ AS DATE) <= '{filters['bitis']}'")

        if filters.get('yillar'):

            yil_list = ','.join(str(y) for y in filters['yillar'])

            where_extra.append(f"YEAR(INV.DATE_) IN ({yil_list})")

        if filters.get('stok_q'):

            q2 = filters['stok_q'].replace("'","''")

            where_extra.append(f"(IT.CODE LIKE '%{q2}%' OR IT.DEFINITION_ LIKE '%{q2}%')")



    extra_where = (' AND ' + ' AND '.join(where_extra)) if where_extra else ''



    tum_rows = []

    try:

        conn = logo_connect()

        for firma_cfg in sec:

            firma     = str(firma_cfg['no']).zfill(3)

            firma_adi = firma_cfg.get('ad', firma)

            sql = f"""

            SELECT

                IT.CODE            AS stok_kodu,

                IT.DEFINITION_     AS stok_adi,

                IT.STGRPCODE       AS stok_grubu,

                ST.UNITSETCODE     AS birim,

                ST.AMOUNT          AS miktar,

                ST.PRICE           AS satis_fiyati,

                ST.TOTAL           AS satir_toplam,

                ST.VATRATE         AS kdv_orani,

                ST.DISCPER         AS iskonto_oran,

                ST.LINENET         AS net_tutar,

                INV.TRCODE         AS fatura_tipi,

                YEAR(INV.DATE_)    AS yil,

                MONTH(INV.DATE_)   AS ay,

                CONVERT(VARCHAR,INV.DATE_,104) AS tarih,

                INV.FICHENO        AS fatura_no,

                CL.DEFINITION_     AS cari_adi,

                CL.CODE            AS cari_kod,

                '{firma_adi}'     AS logo_firma_adi,

                '{firma}'         AS logo_firma_no

            FROM LG_{firma}_01_STLINE ST

            INNER JOIN LG_{firma}_01_INVOICE INV ON ST.INVOICEREF = INV.LOGICALREF

            LEFT  JOIN LG_{firma}_ITEMS IT       ON ST.STOCKREF   = IT.LOGICALREF

            LEFT  JOIN LG_{firma}_CLCARD CL      ON INV.CLIENTREF = CL.LOGICALREF

            WHERE INV.CANCELLED = 0

              AND ST.LINETYPE   = 0

              AND IT.CODE IS NOT NULL

              AND INV.TRCODE IN (7,8)   -- Sadece satış faturaları

              {extra_where}

            ORDER BY INV.DATE_ DESC

            """

            try:

                cursor = conn.cursor()

                cursor.execute(sql)

                cols = [d[0] for d in cursor.description]

                for row in cursor.fetchall():

                    d = {}

                    for i2, col in enumerate(cols):

                        v = row[i2]

                        if hasattr(v, 'isoformat'): v = str(v)

                        elif v is None: v = ''

                        d[col] = v

                    tum_rows.append(d)

                cursor.close()

            except Exception as e2:

                print(f"  Firma {firma} satış analiz hatası: {e2}")

        conn.close()

    except Exception as e:

        traceback.print_exc()

        return {'ok': False, 'msg': str(e), 'rows': []}



    return {'ok': True, 'rows': tum_rows, 'count': len(tum_rows)}





def logo_stok_bakiye(filters=None):

    """Stok bakiyelerini çeker — negatif stok tespiti için."""

    tum_firmalar = logo_firma_listesi()

    if filters and filters.get('firmalar'):

        sec = [f for f in tum_firmalar if f['no'] in filters['firmalar']]

    else:

        sec = tum_firmalar

    if not sec:

        sec = [{'no': logo_firma_no(), 'ad': 'Firma'}]



    tum_rows = []

    try:

        conn = logo_connect()

        for firma_cfg in sec:

            firma     = str(firma_cfg['no']).zfill(3)

            firma_adi = firma_cfg.get('ad', firma)

            sql = f"""

            SELECT

                IT.CODE        AS stok_kodu,

                IT.DEFINITION_ AS stok_adi,

                IT.STGRPCODE   AS stok_grubu,

                ISNULL(

                    (SELECT SUM(CASE WHEN IOCODE IN (1,3) THEN AMOUNT ELSE -AMOUNT END)

                     FROM LG_{firma}_01_STLINE SL

                     WHERE SL.STOCKREF = IT.LOGICALREF AND SL.LINETYPE = 0),

                0) AS bakiye,

                IT.LASTPURCHPRICE  AS son_alis_fiyati,

                IT.LASTSELLINGPRICE AS son_satis_fiyati,

                '{firma_adi}' AS logo_firma_adi,

                '{firma}'    AS logo_firma_no

            FROM LG_{firma}_ITEMS IT

            WHERE IT.ACTIVE = 0

            ORDER BY bakiye ASC

            """

            try:

                cursor = conn.cursor()

                cursor.execute(sql)

                cols = [d[0] for d in cursor.description]

                for row in cursor.fetchall():

                    d = dict(zip(cols,[''  if v is None else (float(v) if isinstance(v,(int,float)) else str(v)) for v in row]))

                    tum_rows.append(d)

                cursor.close()

            except Exception as e2:

                print(f"  Firma {firma} bakiye hatası: {e2}")

        conn.close()

    except Exception as e:

        traceback.print_exc()

        return {'ok': False, 'msg': str(e), 'rows': []}



    return {'ok': True, 'rows': tum_rows}



# ── HTTP HANDLER ──────────────────────────────────────────────────────────────

class AppHandler(BaseHTTPRequestHandler):

    def log_message(self, format, *args): pass



    def send_json(self, data, status=200):
        body = json.dumps(data, ensure_ascii=False).encode('utf-8')
        self.send_response(status)
        self.send_header('Content-Type','application/json; charset=utf-8')
        self.send_header('Content-Length', len(body))
        # Cookie'yi yenile (sliding window - oturum uzasın)
        try:
            token = get_token_from_request(self) or self.headers.get('X-Token','')
            if token and get_session(token):
                self.send_header('Set-Cookie',
                    f'na_token={token}; Path=/; HttpOnly; SameSite=Lax; Max-Age={SESSION_TTL}')
        except: pass
        self.end_headers()

        self.wfile.write(body)



    def send_html(self, html):

        body = html.encode('utf-8')

        self.send_response(200)

        self.send_header('Content-Type','text/html; charset=utf-8')

        self.send_header('Content-Length',len(body))

        self.end_headers()

        self.wfile.write(body)



    def send_file(self, data, filename, mime):

        self.send_response(200)

        self.send_header('Content-Type', mime)

        self.send_header('Content-Disposition',f'attachment; filename="{filename}"')

        self.send_header('Content-Length',len(data))

        self.end_headers()

        self.wfile.write(data)



    def do_GET(self):

        global DOCS_META, DOCS_DIR

        parsed = urlparse(self.path)

        path   = parsed.path

        qs     = parse_qs(parsed.query)



        if path == '/login':

            self.send_html(LOGIN_HTML); return

        if path == '/api/logout':

            token = get_token_from_request(self)

            if token and token in SESSIONS: del SESSIONS[token]

            self.send_response(302); self.send_header('Location','/login')

            self.send_header('Set-Cookie','na_token=; Path=/; HttpOnly; SameSite=Lax; Max-Age=0'); self.end_headers(); return

        if path == '/api/me':

            sess = check_auth(self)

            if not sess: self.send_json({'ok':False},401); return

            users = load_users(); u = users.get(sess['username'],{})

            self.send_json({'ok':True,'username':sess['username'],'ad':u.get('ad',sess['username']),'rol':u.get('rol','user')}); return

        if path not in ('/api/login',):

            sess = check_auth(self)

            if not sess:

                if path.startswith('/api/'):

                    self.send_json({'ok':False,'msg':'Oturum doldu','redirect':'/login'},401)

                else:

                    self.send_response(302); self.send_header('Location','/login'); self.end_headers()

                return



        if path == '/api/ipotek':

            sess = check_auth(self)

            if not sess: self.send_json({'ok':False},401); return

            self.send_json({'ok':True,'records':IPOTEK_DATA}); return



        if path == '/api/kredi/sablon':

            sess = check_auth(self)

            if not sess: self.send_json({'ok':False},401); return

            try:

                import openpyxl

                from openpyxl.styles import Font, PatternFill, Alignment

                wb = openpyxl.Workbook(); ws = wb.active; ws.title = 'Kredi Limitleri'

                hdr_fill = PatternFill("solid", fgColor="1E3A5F")

                hdr_font = Font(bold=True, color="FFFFFF", size=10)

                headers = ['FİRMA','BANKA','KREDİ TÜRÜ','KREDİ/KART NO','LİMİT (₺)','KULLANILAN (₺)','YILLIK FAİZ (%)','BİTİŞ TARİHİ','AD/AÇIKLAMA']

                widths  = [10, 18, 14, 20, 16, 16, 14, 14, 20]

                for col,(h,w) in enumerate(zip(headers,widths),1):

                    cell = ws.cell(row=1,column=col,value=h)

                    cell.fill=hdr_fill; cell.font=hdr_font

                    cell.alignment=Alignment(horizontal='center',wrap_text=True)

                    ws.column_dimensions[ws.cell(row=1,column=col).column_letter].width=w

                ws.row_dimensions[1].height=35

                ornek_rows = [

                    ['ULUSAL','VAKIFBANK','ROTATİF','1234567890','10000000','3500000','38.5','31.12.2026','Rotatif Kredi'],

                    ['ULUSAL','YAPIKREDI','KREDİ KARTI','4111 1111 1111 1111','500000','120000','45.0','','Kurumsal Kart'],

                    ['ABC','AKBANK','KMH','9876543210','5000000','0','36.0','30.06.2026','KMH Hesabı'],

                ]

                for ri,row in enumerate(ornek_rows,2):

                    for col,val in enumerate(row,1):

                        cell = ws.cell(row=ri,column=col,value=val)

                        cell.fill=PatternFill("solid",fgColor="E8F5E9")

                        cell.font=Font(italic=True,color="555555",size=9)

                ws.cell(row=5,column=1,value="* 2-4. satırlar örnektir, silebilirsiniz. Faiz % olarak girin (38.5 = %38.5)")

                ws.merge_cells(f'A5:{ws.cell(row=5,column=len(headers)).column_letter}5')

                ws.freeze_panes='A2'

                import io as _io2; out=_io2.BytesIO(); wb.save(out)

                self.send_file(out.getvalue(),'kredi_limitleri_sablon.xlsx','application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

            except Exception as e: self.send_json({'ok':False,'msg':str(e)})

            return



        if path == '/api/ipotek/sablon':

            sess = check_auth(self)

            if not sess: self.send_json({'ok':False},401); return

            try:

                import openpyxl

                from openpyxl.styles import Font, PatternFill, Alignment

                wb = openpyxl.Workbook(); ws = wb.active; ws.title = 'Ipotek'

                hdr_fill = PatternFill("solid", fgColor="1E3A5F")

                hdr_font = Font(bold=True, color="FFFFFF", size=10)

                headers = ['S.NO','TASINMAZ KOD','SAHİBİ','İLİ','İLÇE','ADRES','ADA','PARSEL','B.B','KAT','BLOK','NİTELİK','İPOTEK','TUTAR','İPOTEK TARİHİ','EKSPERTİZ DEĞERİ','EKSPERTİZ TARİHİ','TAŞINMAZ TİPİ','DASK POLİÇE NO','DASK VADE','KONUT POLİÇE NO','KONUT VADE']

                widths = [6,12,20,12,14,20,8,8,8,8,8,18,14,14,14,16,14,16,16,12,16,12]

                for col,(h,w) in enumerate(zip(headers,widths),1):

                    cell = ws.cell(row=1,column=col,value=h)

                    cell.fill=hdr_fill; cell.font=hdr_font

                    cell.alignment=Alignment(horizontal='center',wrap_text=True)

                    ws.column_dimensions[ws.cell(row=1,column=col).column_letter].width=w

                ws.row_dimensions[1].height=35

                ornek = ['1','IPT-001','Sahin Erkin','Istanbul','Sisli','Eskisehir Mah.','1316','76','5','1','','Mesken','Akbank','5000000','01.01.2024','8000000','15.06.2023','Kat Mulkiyeti','76416712','20.10.2025','0001-123456','20.10.2025']

                for col,val in enumerate(ornek,1):

                    cell = ws.cell(row=2,column=col,value=val)

                    cell.fill=PatternFill("solid",fgColor="E8F5E9")

                    cell.font=Font(italic=True,color="555555",size=9)

                ws.cell(row=3,column=1,value="* 2. satir ornektir, silebilirsiniz")

                ws.merge_cells(f'A3:{ws.cell(row=3,column=len(headers)).column_letter}3')

                ws.freeze_panes='A2'

                out=io.BytesIO(); wb.save(out)

                self.send_file(out.getvalue(),'ipotek_sablon.xlsx','application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

            except Exception as e: self.send_json({'ok':False,'msg':str(e)})

            return



        if path == '/api/ipotek/excel':

            sess = check_auth(self)

            if not sess: self.send_json({'ok':False},401); return

            try:

                import openpyxl

                from openpyxl.styles import Font, PatternFill, Alignment

                wb = openpyxl.Workbook(); ws = wb.active; ws.title = 'Ipotek Takibi'

                hdr_fill=PatternFill("solid",fgColor="1E3A5F"); hdr_font=Font(bold=True,color="FFFFFF",size=10)

                headers=['S.No','Kod','Sahibi','Il','Ilce','Adres','Ada','Parsel','B.B','Kat','Blok','Nitelik','Ipotek Bankasi','Tutar','Ipotek Tarihi','Ekspertiz Degeri','Ekspertiz Tarihi','Tasınmaz Tipi','DASK Police','DASK Vade','Konut Police','Konut Vade']

                for col,h in enumerate(headers,1):

                    cell=ws.cell(row=1,column=col,value=h)

                    cell.fill=hdr_fill; cell.font=hdr_font

                    cell.alignment=Alignment(horizontal='center')

                for ri,r in enumerate(IPOTEK_DATA,2):

                    vals=[r.get('sno',''),r.get('kod',''),r.get('sahip',''),r.get('il',''),r.get('ilce',''),r.get('adres',''),r.get('ada',''),r.get('parsel',''),r.get('bb',''),r.get('kat',''),r.get('blok',''),r.get('nitelik',''),r.get('banka',''),r.get('tutar',0),r.get('ipotek_tarih',''),r.get('exp_deger',0),r.get('exp_tarih',''),r.get('tbl_tip',''),r.get('dask_no',''),r.get('dask_vade',''),r.get('konut_no',''),r.get('konut_vade','')]

                    for col,val in enumerate(vals,1): ws.cell(row=ri,column=col,value=val)

                ws.auto_filter.ref=ws.dimensions; ws.freeze_panes='A2'

                out=io.BytesIO(); wb.save(out)

                self.send_file(out.getvalue(),f"ipotek_{datetime.now().strftime('%Y%m%d')}.xlsx",'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

            except Exception as e: self.send_json({'ok':False,'msg':str(e)})

            return



        if path == '/api/banka_kart':

            sess = check_auth(self)

            if not sess: self.send_json({'ok':False},401); return

            self.send_json({'ok':True,'records':BANKA_KART_DATA}); return



        if path == '/api/banka_kart/ipotek_listesi':

            # Bu bankaya ait ipotekleri getir

            sess = check_auth(self)

            if not sess: self.send_json({'ok':False},401); return

            banka_adi = qs.get('banka',[''])[0]

            ilgili = [r for r in IPOTEK_DATA if (r.get('banka') or '').lower() == banka_adi.lower()]

            self.send_json({'ok':True,'records':ilgili}); return



        if path == '/api/yedek/github':

            sess = check_auth(self)

            if not sess: self.send_json({'ok':False},401); return

            ok = github_yedek_al()

            self.send_json({'ok':ok,'msg':'GitHub yedegi tamamlandi' if ok else 'Hata - token kontrol edin'}); return



        if path == '/api/shutdown':

            import threading

            def _shutdown():

                import time; time.sleep(0.5)

                self.server.shutdown()

            threading.Thread(target=_shutdown, daemon=True).start()

            self.send_response(200)

            self.send_header('Content-Type','text/plain')

            self.end_headers()

            self.wfile.write(b'shutting down')

            return



        if path == '/':

            page = qs.get('page',[''])[0]

            self.send_html(get_html(page))



        elif path == '/api/data':

            params = {k: v[0] for k,v in qs.items()}

            result = apply_filter(params)

            # Her kayda gerçek DATA index'ini ekle

            for rec in result:

                try:

                    rec['_idx'] = DATA.index(rec)

                except ValueError:

                    # Obje referansı eşleşmiyorsa tarih+firma+tutar ile bul

                    rec['_idx'] = next(

                        (i for i,d in enumerate(DATA)

                         if d.get('firma')==rec.get('firma') and

                            d.get('tarih')==rec.get('tarih') and

                            abs(d.get('tutar',0)-rec.get('tutar',0))<0.5),

                        -1

                    )

            self.send_json({'records': result, 'total': len(DATA)})



        elif path == '/api/stats':

            total   = sum(d['tutar'] for d in DATA)

            odendi  = [d for d in DATA if d.get('durum')=='ÖDENDİ']

            bekl    = [d for d in DATA if d.get('durum')!='ÖDENDİ']

            now     = datetime.now(); in30 = now + timedelta(days=30)

            soon    = []

            for d in bekl:

                t = d.get('tarih','')

                if t and len(t)==10:

                    try:

                        dt = datetime.strptime(t,'%d.%m.%Y')

                        if now <= dt <= in30: soon.append(d)

                    except: pass

            self.send_json({

                'total':total,'total_count':len(DATA),

                'odendi':sum(d['tutar'] for d in odendi),'odendi_count':len(odendi),

                'bekleyen':sum(d['tutar'] for d in bekl),'bekleyen_count':len(bekl),

                'soon':sum(d['tutar'] for d in soon),'soon_count':len(soon),

                'firmas':sorted(set(d['firma'] for d in DATA)),

                'turler':sorted(set(d['tur'] for d in DATA)),

                'aylar':[m for m in AYLAR if any(d['ay']==m for d in DATA)],

                'yillar':sorted(set(get_yil(d) for d in DATA if get_yil(d))),

            })



        elif path == '/api/chart':

            ct = qs.get('type',['monthly'])[0]

            try:    self.send_json({'img': make_chart(ct)})

            except Exception as e: self.send_json({'error':str(e)},500)



        elif path == '/api/source/status':

            # Kaynak dosya durumu

            self.send_json({

                'file':     SOURCE_FILE or '',

                'sheet':    SOURCE_SHEET or '',

                'auto':     AUTO_RELOAD,

                'msg':      LAST_SYNC_MSG,

                'ok':       LAST_SYNC_OK,

                'exists':   bool(SOURCE_FILE and os.path.exists(SOURCE_FILE)),

            })



        elif path == '/api/source/reload':

            ok = reload_from_source(force=True)

            self.send_json({'ok':ok,'msg':LAST_SYNC_MSG,'count':len(DATA)})



        elif path == '/api/source/browse':

            # Windows dosya secici diyalogu ac

            try:

                import tkinter as tk

                from tkinter import filedialog

                root = tk.Tk()

                root.withdraw()

                root.attributes('-topmost', True)

                filepath = filedialog.askopenfilename(

                    title='Excel Dosyasi Sec',

                    filetypes=[('Excel Dosyalari', '*.xlsx *.xls'), ('Tum Dosyalar', '*.*')]

                )

                root.destroy()

                if filepath:

                    filepath = os.path.normpath(filepath)

                    self.send_json({'ok': True, 'path': filepath})

                else:

                    self.send_json({'ok': False, 'cancelled': True})

            except Exception as e:

                self.send_json({'ok': False, 'error': str(e)})



        elif path == '/api/cari':

            # Cari hesap listesi - firma bazlı özet

            params = {k: v[0] for k,v in qs.items()}

            firma_f  = params.get('firma','')

            tur_f    = params.get('tur','')   # ALACAK / BORÇ

            durum_f  = params.get('durum','') # ACIK / KAPALI

            q_f      = params.get('q','').lower()



            result = CARI_DATA

            if firma_f:  result = [d for d in result if d.get('cari_firma','') == firma_f]

            if tur_f:    result = [d for d in result if d.get('tur','') == tur_f]

            if durum_f:  result = [d for d in result if d.get('durum','') == durum_f]

            if q_f:      result = [d for d in result if q_f in (

                d.get('cari_firma','') + d.get('aciklama','') + d.get('belge_no','')

            ).lower()]



            # İndeks ekle

            all_idx = {id(d): i for i,d in enumerate(CARI_DATA)}

            result_with_idx = [dict(d, _idx=all_idx.get(id(d), i))

                               for i, d in enumerate(result)]



            # Özet: toplam alacak/borç

            toplam_alacak = sum(d.get('tutar',0) for d in CARI_DATA if d.get('tur')=='ALACAK' and d.get('durum')=='AÇIK')

            toplam_borc   = sum(d.get('tutar',0) for d in CARI_DATA if d.get('tur')=='BORÇ'   and d.get('durum')=='AÇIK')

            firmalar      = sorted(set(d.get('cari_firma','') for d in CARI_DATA if d.get('cari_firma')))



            self.send_json({

                'records': result_with_idx,

                'total': len(CARI_DATA),

                'toplam_alacak': round(toplam_alacak, 2),

                'toplam_borc':   round(toplam_borc, 2),

                'net': round(toplam_alacak - toplam_borc, 2),

                'firmalar': firmalar,

            })



        elif path == '/api/cari/ozet':

            # Firma bazlı cari hesap özeti

            firma_map = {}

            for d in CARI_DATA:

                fn = d.get('cari_firma', 'Diğer')

                if fn not in firma_map:

                    firma_map[fn] = {

                        'cari_firma': fn,

                        'grup_firma': d.get('grup_firma',''),

                        'alacak': 0, 'borc': 0,

                        'alacak_vadeli': 0, 'borc_vadeli': 0,

                        'hareket_sayisi': 0,

                        'son_hareket': '',

                    }

                tutar = d.get('tutar', 0)

                if d.get('tur') == 'ALACAK':

                    firma_map[fn]['alacak'] += tutar

                    if d.get('durum') == 'VADEDE': firma_map[fn]['alacak_vadeli'] += tutar

                else:

                    firma_map[fn]['borc'] += tutar

                    if d.get('durum') == 'VADEDE': firma_map[fn]['borc_vadeli'] += tutar

                firma_map[fn]['hareket_sayisi'] += 1

                tarih = d.get('tarih','')

                if tarih > firma_map[fn]['son_hareket']:

                    firma_map[fn]['son_hareket'] = tarih



            ozet = sorted([

                {**v, 'net': round(v['alacak']-v['borc'],2),

                 'alacak': round(v['alacak'],2), 'borc': round(v['borc'],2)}

                for v in firma_map.values()

            ], key=lambda x: abs(x['net']), reverse=True)



            self.send_json({'ozet': ozet})



        elif path == '/api/banka':

            self.send_json({'records': BANKA_DATA})



        elif path == '/api/logo/stok':

            q        = qs.get('q',[''])[0]

            limit    = int(qs.get('limit',['300'])[0])

            offset   = int(qs.get('offset',['0'])[0])

            firmalar = qs.get('firmalar',[''])[0].split(',') if qs.get('firmalar') else None

            yillar   = None

            if qs.get('yillar'):

                try: yillar = [int(y) for y in qs['yillar'][0].split(',') if y]

                except: pass

            grup     = qs.get('grup',[''])[0] or None

            aktif    = qs.get('aktif_hareket',['0'])[0] == '1'

            bas_s    = qs.get('baslangic',[''])[0] or None

            bit_s    = qs.get('bitis',[''])[0] or None

            result   = logo_stok_cek(q, limit, offset, firmalar, yillar, grup, aktif, bas_s, bit_s)

            result['firmalar_list'] = sorted(set(r.get('logo_firma_adi','') for r in result.get('records',[])))

            result['gruplar'] = sorted(set(r.get('stok_grubu','') for r in result.get('records',[]) if r.get('stok_grubu')))

            self.send_json(result)



        elif path == '/api/logo/cari':

            q        = qs.get('q',[''])[0]

            limit    = int(qs.get('limit',['500'])[0])

            offset   = int(qs.get('offset',['0'])[0])

            tip      = qs.get('tip',[''])[0]

            firmalar = qs.get('firmalar',[''])[0].split(',') if qs.get('firmalar') else None

            yillar   = None

            if qs.get('yillar'):

                try: yillar = [int(y) for y in qs['yillar'][0].split(',') if y]

                except: pass

            bas_c    = qs.get('baslangic',[''])[0] or None

            bit_c    = qs.get('bitis',[''])[0] or None

            result   = logo_cari_cek(q, limit, offset, tip, firmalar, yillar, bas_c, bit_c)

            result['firmalar_list'] = sorted(set(r.get('logo_firma_adi','') for r in result.get('records',[])))

            self.send_json(result)



        elif path == '/api/logo/cari/ozet':

            kod = qs.get('kod',[''])[0]

            self.send_json(logo_cari_hesap_ozet(kod))



        elif path == '/api/hatirlatma':

            from datetime import datetime as _dt, date as _date

            bugun = _dt.now().strftime('%Y-%m-%d')

            simdi = _dt.now().strftime('%H:%M')

            filtre = qs.get('filtre',['tumu'])[0]

            result = []

            for h in HATIRLATMALAR:

                tarih = h.get('tarih','')

                if filtre == 'bugun' and tarih != bugun: continue

                if filtre == 'bekleyen' and (tarih < bugun or h.get('tamamlandi')): continue

                if filtre == 'tamamlandi' and not h.get('tamamlandi'): continue

                gecti = tarih < bugun and not h.get('tamamlandi')

                result.append({**h, 'gecti': gecti})

            result.sort(key=lambda x: (x.get('tamamlandi',False), x.get('tarih',''), x.get('saat','')))

            bugun_bekleyen = sum(1 for h in HATIRLATMALAR

                if h.get('tarih','') == bugun and not h.get('tamamlandi'))

            acil = sum(1 for h in HATIRLATMALAR

                if h.get('tarih','') <= bugun and not h.get('tamamlandi'))

            self.send_json({'records': result, 'bugun_bekleyen': bugun_bekleyen, 'acil': acil})



        elif path == '/api/otel/config':

            self.send_json({'config': OTEL_CONFIG})



        elif path == '/api/otel/rezervasyonlar':

            otel_id  = qs.get('otel',['otel1'])[0]

            durum_f  = qs.get('durum',[''])[0]

            kanal_f  = qs.get('kanal',[''])[0]

            q_f      = qs.get('q',[''])[0].lower()

            bas_f    = qs.get('baslangic',[''])[0]

            bit_f    = qs.get('bitis',[''])[0]

            from datetime import datetime as _dt

            rezervler = OTEL_DATA.get(otel_id, [])

            if durum_f:

                rezervler = [r for r in rezervler if r.get('durum','').lower() == durum_f.lower()]

            if kanal_f:

                rezervler = [r for r in rezervler if r.get('kanal','') == kanal_f]

            if q_f:

                rezervler = [r for r in rezervler if q_f in (

                    r.get('misafir_adi','') + r.get('telefon','') +

                    r.get('rezervasyon_no','') + r.get('aciklama','')

                ).lower()]

            if bas_f:

                rezervler = [r for r in rezervler if r.get('giris_tarihi','') >= bas_f]

            if bit_f:

                rezervler = [r for r in rezervler if r.get('giris_tarihi','') <= bit_f]

            rezervler_sorted = sorted(rezervler, key=lambda x: x.get('giris_tarihi',''), reverse=True)

            kanallar = sorted(set(r.get('kanal','Direkt') for r in OTEL_DATA.get(otel_id,[])))

            self.send_json({'records': rezervler_sorted, 'count': len(rezervler_sorted), 'kanallar': kanallar})



        elif path == '/api/otel/istatistik':

            otel_id   = qs.get('otel',['otel1'])[0]

            baslangic = qs.get('baslangic',[''])[0] or None

            bitis     = qs.get('bitis',[''])[0] or None

            istat = otel_istatistik(otel_id, baslangic, bitis)

            self.send_json(istat)



        elif path == '/api/db/status':

            try:

                info = {}

                conn = db_connect()

                c = conn.cursor()

                for tablo in ['giderler','gelirler','banka_bakiyeleri','kredi_limitleri','cari_hareketler','notlar']:

                    c.execute(f"SELECT COUNT(*) FROM {tablo}")

                    info[tablo] = c.fetchone()[0]

                c.execute("SELECT name FROM sqlite_master WHERE type='table'")

                info['tablolar'] = [r[0] for r in c.fetchall()]

                c.execute("SELECT page_count * page_size FROM pragma_page_count(), pragma_page_size()")

                row = c.fetchone()

                info['db_boyut_kb'] = round((row[0] if row else 0) / 1024, 1)

                info['db_dosya'] = DB_FILE

                c.execute("SELECT * FROM sync_log ORDER BY id DESC LIMIT 1")

                last = c.fetchone()

                info['son_sync'] = dict(last) if last else None

                conn.close()

                self.send_json({'ok': True, 'info': info})

            except Exception as e:

                self.send_json({'ok': False, 'msg': str(e)})



        elif path == '/api/db/sorgu':

            # Özel SQL sorgusu çalıştır (sadece SELECT)

            sql = qs.get('sql',[''])[0]

            if not sql.strip().upper().startswith('SELECT'):

                self.send_json({'ok': False, 'msg': 'Sadece SELECT sorguları desteklenir'})

                return

            try:

                rows = db_query(sql)

                self.send_json({'ok': True, 'rows': rows[:500], 'count': len(rows)})

            except Exception as e:

                self.send_json({'ok': False, 'msg': str(e)})



        elif path == '/api/ai/config':

            safe = {k:v for k,v in AI_CONFIG.items() if k != 'api_key'}

            safe['has_key'] = bool(AI_CONFIG.get('api_key'))

            self.send_json({'config': safe})



        elif path == '/api/notlar':

            self.send_json({'notlar': NOTLAR})



        elif path == '/api/gmail/config':

            safe = {k:v for k,v in GMAIL_CONFIG.items() if k != 'client_secret'}

            connected = os.path.exists(GMAIL_TOKEN_FILE)

            self.send_json({'config': safe, 'connected': connected})



        elif path == '/api/gmail/mails':

            q          = qs.get('q',[''])[0]

            max_r      = int(qs.get('max',[50])[0])

            self.send_json(gmail_mail_listesi(max_r, q))



        elif path == '/api/gmail/mail':

            mail_id = qs.get('id',[''])[0]

            self.send_json(gmail_mail_oku(mail_id))



        elif path == '/api/gmail/oauth/url':

            self.send_json(gmail_oauth_url())



        elif path == '/api/gmail/oauth/callback':

            code = qs.get('code',[''])[0]

            result = gmail_oauth_callback(code)

            # Tarayıcıda güzel mesaj göster

            if result['ok']:

                html = '<html><body style="font-family:sans-serif;background:#0d1117;color:#e6edf3;padding:40px"><h2 style="color:#10b981">Gmail baglantisi basarili!</h2><p>Bu sekmeyi kapatip uygulamaya donebilirsiniz.</p></body></html>'.encode('utf-8')

            else:

                html = f'<html><body style="font-family:sans-serif;background:#0d1117;color:#e6edf3;padding:40px"><h2 style="color:#ef4444">&#10007; Hata</h2><p>{result["msg"]}</p></body></html>'.encode()

            self.send_response(200)

            self.send_header('Content-Type','text/html; charset=utf-8')

            self.end_headers()

            self.wfile.write(html)

            return



        elif path == '/api/logo/config':

            self.send_json({'config': {k:v for k,v in LOGO_CONFIG.items() if k != 'password'}})



        elif path == '/api/logo/test':

            result = logo_test_connection()

            self.send_json(result)



        elif path == '/api/logo/fatura':

            tip    = qs.get('tip',['satis'])[0]

            limit  = int(qs.get('limit',['500'])[0])

            offset = int(qs.get('offset',['0'])[0])

            filters = {}

            if qs.get('baslangic'):  filters['baslangic']  = qs['baslangic'][0]

            if qs.get('bitis'):      filters['bitis']      = qs['bitis'][0]

            if qs.get('cari'):       filters['cari']       = qs['cari'][0]

            if qs.get('min_tutar'):  filters['min_tutar']  = qs['min_tutar'][0]

            if qs.get('max_tutar'):  filters['max_tutar']  = qs['max_tutar'][0]

            # Çok firma

            if qs.get('firmalar'):

                filters['firmalar'] = qs['firmalar'][0].split(',')

            # Çok yıl

            if qs.get('yillar'):

                try: filters['yillar'] = [int(y) for y in qs['yillar'][0].split(',') if y]

                except: pass

            result = logo_fatura_cek(tip, limit, offset, filters)

            # Yıl listesi ekle

            yillar = sorted(set(str(r.get('yil','')) for r in result.get('records',[]) if r.get('yil')), reverse=True)

            firmalar_list = sorted(set(r.get('logo_firma_adi','') for r in result.get('records',[])))

            cariler = sorted(set(r.get('cari_unvan','') for r in result.get('records',[]) if r.get('cari_unvan')))[:100]

            result['yillar'] = yillar

            result['firmalar_list'] = firmalar_list

            result['cariler'] = cariler

            self.send_json(result)



        elif path == '/api/logo/satis_analiz':

            filters = {}

            if qs.get('firmalar'):  filters['firmalar']  = qs['firmalar'][0].split(',')

            if qs.get('yillar'):

                try: filters['yillar'] = [int(y) for y in qs['yillar'][0].split(',') if y]

                except: pass

            if qs.get('baslangic'): filters['baslangic'] = qs['baslangic'][0]

            if qs.get('bitis'):     filters['bitis']     = qs['bitis'][0]

            if qs.get('stok_q'):    filters['stok_q']    = qs['stok_q'][0]

            result = logo_satis_analiz(filters)

            self.send_json(result)



        elif path == '/api/logo/stok_bakiye':

            filters = {}

            if qs.get('firmalar'): filters['firmalar'] = qs['firmalar'][0].split(',')

            result = logo_stok_bakiye(filters)

            self.send_json(result)



        elif path == '/api/logo/firma_listesi':

            firmalar = logo_firma_listesi()

            self.send_json({'firmalar': firmalar})



        elif path == '/api/logo/fatura/detay':

            fatura_id = qs.get('id',['0'])[0]

            result = logo_fatura_detay(fatura_id)

            self.send_json(result)



        elif path == '/api/kur':

            # TCMB kur çek — tarih parametresi: ?tarih=DD.MM.YYYY

            tarih = qs.get('tarih', [None])[0]

            cache_key = tarih or 'bugun'



            # Cache kontrolü (aynı gün için tekrar çekme)

            if cache_key in KUR_CACHE:

                cached = KUR_CACHE[cache_key]

                cached['from_cache'] = True

                self.send_json(cached)

                return



            result = tcmb_kur_cek(tarih)



            if result['ok']:

                KUR_CACHE[cache_key] = result  # Cache'e kaydet

                # Cache boyutunu sınırla (son 30 gün)

                if len(KUR_CACHE) > 30:

                    oldest = next(iter(KUR_CACHE))

                    del KUR_CACHE[oldest]



            self.send_json(result)



        elif path == '/api/banka_api/config':

            self.send_json({'config': API_CONFIG})



        elif path == '/api/banka_api/test':

            # API bağlantı testi (ileride gerçek banka API'si buraya)

            banka_id = qs.get('id',[''])[0]

            cfg = next((c for c in API_CONFIG if c.get('id')==banka_id), None)

            if not cfg:

                self.send_json({'ok':False,'msg':'Konfigürasyon bulunamadı'})

                return

            if not cfg.get('aktif'):

                self.send_json({'ok':False,'msg':'Bu API aktif değil'})

                return

            # Şimdilik mock test - ileride gerçek HTTP isteği

            self.send_json({

                'ok': True,

                'msg': f"Bağlantı ayarları kaydedildi. API anahtarı alındığında aktif edilebilir.",

                'banka': cfg.get('banka_adi',''),

            })



        elif path == '/api/kredi':

            self.send_json({'records': KREDI_DATA})



        elif path == '/api/likidite':

            # Aylık likidite analizi: nakit açığı + optimal kredi kullanım önerisi

            AYLAR_L = ['Ocak','Şubat','Mart','Nisan','Mayıs','Haziran',

                       'Temmuz','Ağustos','Eylül','Ekim','Kasım','Aralık']

            def get_yil_ay(rec):

                t = rec.get('tarih','')

                yil = t[6:10] if len(t)>=10 else ''

                return yil, rec.get('ay','')



            # Aylık nakit pozisyonu (gelir - gider)

            ay_map = {}

            for rec in DATA:

                yil, ay = get_yil_ay(rec)

                if not yil or not ay: continue

                key = f"{yil}|{ay}"

                if key not in ay_map:

                    ay_map[key] = {'yil':yil,'ay':ay,'gelir':0,'gider':0,'gider_bekl':0}

                ay_map[key]['gider'] += rec.get('tutar',0)

                if rec.get('durum') != 'ÖDENDİ':

                    ay_map[key]['gider_bekl'] += rec.get('tutar',0)

            for rec in GELIR_DATA:

                yil, ay = get_yil_ay(rec)

                if not yil or not ay: continue

                key = f"{yil}|{ay}"

                if key not in ay_map:

                    ay_map[key] = {'yil':yil,'ay':ay,'gelir':0,'gider':0,'gider_bekl':0}

                ay_map[key]['gelir'] += rec.get('tutar',0)



            # Kullanılabilir kredi limitleri

            kredi_kullanilabilir = []

            for kr in KREDI_DATA:

                kalan = float(kr.get('kalan_limit', 0) or 0)

                if kalan > 0:

                    kredi_kullanilabilir.append({

                        'firma':    kr.get('firma',''),

                        'banka':    kr.get('banka',''),

                        'tur':      kr.get('tur',''),

                        'limit':    float(kr.get('limit',0) or 0),

                        'kullanilan': float(kr.get('kullanilan',0) or 0),

                        'kalan':    kalan,

                        'faiz_yillik': float(kr.get('faiz_yillik',0) or 0),

                        'faiz_aylik':  float(kr.get('faiz_aylik',0) or 0),

                        'bitis':    kr.get('bitis',''),

                        'ad':       kr.get('ad',''),

                    })

            # Faize göre sırala (en ucuz önce)

            kredi_kullanilabilir.sort(key=lambda x: x['faiz_yillik'] or 99)



            # Toplam banka bakiyesi

            toplam_banka = sum(float(b.get('bakiye',0) or 0) for b in BANKA_DATA)

            banka_detay = [{

                'banka': b.get('banka',''), 'firma': b.get('firma',''),

                'hesap_turu': b.get('hesap_turu',''), 'bakiye': float(b.get('bakiye',0) or 0),

                'para_birimi': b.get('para_birimi','TL')

            } for b in BANKA_DATA]



            # Her ay için likidite analizi

            all_keys = sorted(ay_map.keys(),

                key=lambda k: (k.split('|')[0],

                               AYLAR_L.index(k.split('|')[1]) if k.split('|')[1] in AYLAR_L else 99))



            likidite = []

            kumulatif_banka = toplam_banka  # Dönem başı nakit

            for key in all_keys:

                o = ay_map[key]

                denge = o['gelir'] - o['gider']

                nakit_ihtiyac = max(0, -denge)  # Sadece negatif dengede ihtiyaç var

                kumulatif_banka += denge  # Bu ay sonunda banka

                # Önerilen kredi kullanımı (en ucuzdan en pahalıya)

                oneriler = []

                kalan_ihtiyac = nakit_ihtiyac

                toplam_faiz_maliyeti = 0

                for kr in kredi_kullanilabilir:

                    if kalan_ihtiyac <= 0: break

                    kullan = min(kalan_ihtiyac, kr['kalan'])

                    if kullan <= 0: continue

                    faiz_ay = (kr['faiz_aylik'] or kr['faiz_yillik']/12 if kr['faiz_yillik'] else 0)

                    faiz_maliyet = kullan * faiz_ay

                    toplam_faiz_maliyeti += faiz_maliyet

                    oneriler.append({

                        'firma': kr['firma'], 'banka': kr['banka'], 'tur': kr['tur'],

                        'kredi_no': kr.get('kredi_no',''),

                        'ad': kr['ad'], 'kullan': round(kullan,2),

                        'faiz_aylik': round(kr['faiz_aylik'],4) if kr['faiz_aylik'] else round(kr['faiz_yillik']/12,4) if kr['faiz_yillik'] else 0,

                        'faiz_yillik': round(kr['faiz_yillik'],4),

                        'tahmini_faiz': round(faiz_maliyet,2),

                        'kalan_limit_sonra': round(kr['kalan']-kullan,2),

                    })

                    kalan_ihtiyac -= kullan



                likidite.append({

                    'key': key, 'yil': o['yil'], 'ay': o['ay'],

                    'gelir': round(o['gelir'],2),

                    'gider': round(o['gider'],2),

                    'denge': round(denge,2),

                    'nakit_ihtiyac': round(nakit_ihtiyac,2),

                    'kumulatif_banka': round(kumulatif_banka,2),

                    'oneriler': oneriler,

                    'karsilanamayan': round(max(0, kalan_ihtiyac),2),

                    'toplam_faiz_maliyeti': round(toplam_faiz_maliyeti,2),

                    'pozitif': denge >= 0,

                })



            self.send_json({

                'likidite': likidite,

                'toplam_banka': round(toplam_banka,2),

                'banka_detay': banka_detay,

                'kredi_kullanilabilir': kredi_kullanilabilir,

                'toplam_kalan_limit': round(sum(k['kalan'] for k in kredi_kullanilabilir),2),

            })



        elif path == '/api/gelir':

            # Gelir listesi — filtre destekli

            params = {k: v[0] for k,v in qs.items()}

            result = []

            for i,g in enumerate(GELIR_DATA):

                yil = g.get('tarih','')[6:10] if len(g.get('tarih',''))>=10 else ''

                if params.get('yil')   and yil != params['yil']:         continue

                if params.get('ay')    and g.get('ay') != params['ay']:  continue

                if params.get('firma') and g.get('firma') != params['firma']: continue

                if params.get('tur')   and g.get('tur') != params['tur']:     continue

                r2 = dict(g); r2['_idx'] = i

                result.append(r2)

            self.send_json({'records': result, 'total': len(GELIR_DATA)})



        elif path == '/api/nakit/ozet':

            AYLAR_L = ['Ocak','Şubat','Mart','Nisan','Mayıs','Haziran',

                       'Temmuz','Ağustos','Eylül','Ekim','Kasım','Aralık']

            def get_yil_ay(rec):

                t = rec.get('tarih','')

                yil = t[6:10] if len(t)>=10 else ''

                ay  = rec.get('ay','')

                return yil, ay



            # Gider: hem toplam hem kalem bazlı grupla

            gider_map = {}

            for rec_idx, rec in enumerate(DATA):

                yil, ay = get_yil_ay(rec)

                if not yil or not ay: continue

                key = f"{yil}|{ay}"

                if key not in gider_map:

                    gider_map[key] = {'yil':yil,'ay':ay,'gider':0,'gider_odendi':0,'gider_bekl':0,'kalemler':{}}

                gider_map[key]['gider'] += rec.get('tutar',0)

                if rec.get('durum')=='ÖDENDİ':

                    gider_map[key]['gider_odendi'] += rec.get('tutar',0)

                else:

                    gider_map[key]['gider_bekl'] += rec.get('tutar',0)

                # Kalem bazlı: tur + firma

                tur = rec.get('tur','Diğer') or 'Diğer'

                firma = rec.get('firma','') or ''

                kkey = f"{tur}||{firma}"

                if kkey not in gider_map[key]['kalemler']:

                    gider_map[key]['kalemler'][kkey] = {

                        'tur': tur, 'firma': firma, 'tutar': 0, 'odendi': 0, 'bekl': 0, 'adet': 0, 'satirlar': []

                    }

                gider_map[key]['kalemler'][kkey]['tutar'] += rec.get('tutar', 0)

                gider_map[key]['kalemler'][kkey]['adet']  += 1

                if rec.get('durum') == 'ÖDENDİ':

                    gider_map[key]['kalemler'][kkey]['odendi'] += rec.get('tutar', 0)

                else:

                    gider_map[key]['kalemler'][kkey]['bekl'] += rec.get('tutar', 0)

                gider_map[key]['kalemler'][kkey]['satirlar'].append({

                    'rec_idx': rec_idx,

                    'tarih': rec.get('tarih',''),

                    'yer':   rec.get('yer',''),

                    'tutar': round(rec.get('tutar',0), 2),

                    'durum': rec.get('durum',''),

                    'aciklama': rec.get('aciklama',''),

                    'firma': rec.get('firma',''),

                })



            # Gelir: hem toplam hem kalem bazlı grupla

            gelir_map = {}

            for rec in GELIR_DATA:

                yil, ay = get_yil_ay(rec)

                if not yil or not ay: continue

                key = f"{yil}|{ay}"

                if key not in gelir_map:

                    gelir_map[key] = {'yil':yil,'ay':ay,'gelir':0,'gelir_tahsil':0,'gelir_bekl':0,'kalemler':{}}

                gelir_map[key]['gelir'] += rec.get('tutar',0)

                if rec.get('durum')=='TAHSİL EDİLDİ':

                    gelir_map[key]['gelir_tahsil'] += rec.get('tutar',0)

                else:

                    gelir_map[key]['gelir_bekl'] += rec.get('tutar',0)

                # Kalem bazlı: tur + firma

                tur = rec.get('tur','Diğer') or 'Diğer'

                firma = rec.get('firma','') or ''

                kkey = f"{tur}||{firma}"

                if kkey not in gelir_map[key]['kalemler']:

                    gelir_map[key]['kalemler'][kkey] = {

                        'tur': tur, 'firma': firma, 'tutar': 0, 'tahsil': 0, 'bekl': 0, 'adet': 0, 'satirlar': []

                    }

                gelir_map[key]['kalemler'][kkey]['tutar'] += rec.get('tutar', 0)

                gelir_map[key]['kalemler'][kkey]['adet']  += 1

                if rec.get('durum') == 'TAHSİL EDİLDİ':

                    gelir_map[key]['kalemler'][kkey]['tahsil'] += rec.get('tutar', 0)

                else:

                    gelir_map[key]['kalemler'][kkey]['bekl'] += rec.get('tutar', 0)

                gelir_map[key]['kalemler'][kkey]['satirlar'].append({

                    'tarih':    rec.get('tarih',''),

                    'aciklama': rec.get('aciklama',''),

                    'tutar':    round(rec.get('tutar',0), 2),

                    'durum':    rec.get('durum',''),

                })



            all_keys = sorted(

                set(gider_map.keys()) | set(gelir_map.keys()),

                key=lambda k: (k.split('|')[0], AYLAR_L.index(k.split('|')[1]) if k.split('|')[1] in AYLAR_L else 99)

            )

            ozet = []

            for key in all_keys:

                yil, ay = key.split('|')

                g  = gider_map.get(key, {'gider':0,'gider_odendi':0,'gider_bekl':0,'kalemler':{}})

                gl = gelir_map.get(key, {'gelir':0,'gelir_tahsil':0,'gelir_bekl':0,'kalemler':{}})

                denge         = gl['gelir'] - g['gider']

                finansman_iht = max(0, g['gider'] - gl['gelir'])

                # Kalemleri round ile temizle

                def clean_kalem(km, include_satirlar=False):

                    out = {}

                    for k, v in km.items():

                        if k == 'satirlar':

                            if include_satirlar:

                                out[k] = sorted(v, key=lambda s: s.get('tarih',''))

                            # Satirlar varsayılan olarak hariç - ayrı endpoint'ten gelecek

                        elif isinstance(v, float):

                            out[k] = round(v, 2)

                        else:

                            out[k] = v

                    return out



                gider_kalemler = sorted(

                    [clean_kalem(km) for km in g['kalemler'].values()],

                    key=lambda x: -x['tutar']

                )

                gelir_kalemler = sorted(

                    [clean_kalem(km) for km in gl['kalemler'].values()],

                    key=lambda x: -x['tutar']

                )

                ozet.append({

                    'key': key, 'yil': yil, 'ay': ay,

                    'gelir': round(gl['gelir'],2),

                    'gelir_tahsil': round(gl['gelir_tahsil'],2),

                    'gelir_bekl': round(gl['gelir_bekl'],2),

                    'gider': round(g['gider'],2),

                    'gider_odendi': round(g['gider_odendi'],2),

                    'gider_bekl': round(g['gider_bekl'],2),

                    'denge': round(denge,2),

                    'finansman_iht': round(finansman_iht,2),

                    'pozitif': denge >= 0,

                    'gider_kalemler': gider_kalemler,

                    'gelir_kalemler': gelir_kalemler,

                })

            self.send_json({'ozet': ozet})



        elif path == '/api/nakit/detail':

            # Belirli bir ay için kalem+satır detayı (lazy load)

            key_req = qs.get('key',[''])[0]

            AYLAR_L2 = ['Ocak','Şubat','Mart','Nisan','Mayıs','Haziran',

                        'Temmuz','Ağustos','Eylül','Ekim','Kasım','Aralık']

            def get_yil_ay2(rec):

                t = rec.get('tarih','')

                return t[6:10] if len(t)>=10 else '', rec.get('ay','')



            gider_k, gelir_k = {}, {}

            for rec in DATA:

                yil2, ay2 = get_yil_ay2(rec)

                k2 = f"{yil2}|{ay2}"

                if k2 != key_req: continue

                tur2 = rec.get('tur','Diğer') or 'Diğer'

                firma2 = rec.get('firma','') or ''

                kk = f"{tur2}||{firma2}"

                if kk not in gider_k:

                    gider_k[kk] = {'tur':tur2,'firma':firma2,'tutar':0,'odendi':0,'bekl':0,'adet':0,'satirlar':[]}

                gider_k[kk]['tutar']  += rec.get('tutar',0)

                gider_k[kk]['adet']   += 1

                if rec.get('durum')=='ÖDENDİ': gider_k[kk]['odendi'] += rec.get('tutar',0)

                else: gider_k[kk]['bekl'] += rec.get('tutar',0)

                gider_k[kk]['satirlar'].append({'tarih':rec.get('tarih',''),'yer':rec.get('yer',''),'tutar':round(rec.get('tutar',0),2),'durum':rec.get('durum',''),'aciklama':rec.get('aciklama','')})



            for rec in GELIR_DATA:

                yil2, ay2 = get_yil_ay2(rec)

                k2 = f"{yil2}|{ay2}"

                if k2 != key_req: continue

                tur2 = rec.get('tur','Diğer') or 'Diğer'

                firma2 = rec.get('firma','') or ''

                kk = f"{tur2}||{firma2}"

                if kk not in gelir_k:

                    gelir_k[kk] = {'tur':tur2,'firma':firma2,'tutar':0,'tahsil':0,'bekl':0,'adet':0,'satirlar':[]}

                gelir_k[kk]['tutar'] += rec.get('tutar',0)

                gelir_k[kk]['adet']  += 1

                if rec.get('durum')=='TAHSİL EDİLDİ': gelir_k[kk]['tahsil'] += rec.get('tutar',0)

                else: gelir_k[kk]['bekl'] += rec.get('tutar',0)

                gelir_k[kk]['satirlar'].append({'tarih':rec.get('tarih',''),'aciklama':rec.get('aciklama',''),'tutar':round(rec.get('tutar',0),2),'durum':rec.get('durum','')})



            def rnd(km):

                out = {}

                for k,v in km.items():

                    if k=='satirlar': out[k]=sorted(v,key=lambda s:s.get('tarih',''))

                    elif isinstance(v,float): out[k]=round(v,2)

                    else: out[k]=v

                return out



            self.send_json({

                'gider_kalemler': sorted([rnd(km) for km in gider_k.values()],key=lambda x:-x['tutar']),

                'gelir_kalemler': sorted([rnd(km) for km in gelir_k.values()],key=lambda x:-x['tutar']),

            })



        elif path == '/api/banka':

            self.send_json({'records': BANKA_DATA})



        elif path == '/api/kredi':

            self.send_json({'records': KREDI_DATA})



        elif path == '/api/finansman/oneri':

            # Belirtilen ay/yıl için finansman ihtiyacını hesapla ve öneri üret

            ay  = qs.get('ay',  [''])[0]

            yil = qs.get('yil', [''])[0]



            def get_yil_ay_f(rec):

                t = rec.get('tarih','')

                return (t[6:10] if len(t)>=10 else ''), rec.get('ay','')



            # O ay gider toplamı

            gider_bekl = sum(

                r.get('tutar',0) for r in DATA

                if r.get('durum') != 'ÖDENDİ'

                and (not ay  or r.get('ay','')          == ay)

                and (not yil or r.get('tarih','')[6:10] == yil)

            )

            # O ay gelir toplamı (tahsil edilmemiş)

            gelir_bekl = sum(

                r.get('tutar',0) for r in GELIR_DATA

                if r.get('durum') != 'TAHSİL EDİLDİ'

                and (not ay  or r.get('ay','')          == ay)

                and (not yil or r.get('tarih','')[6:10] == yil)

            )

            # Banka toplam bakiye

            toplam_banka = sum(b.get('bakiye',0) for b in BANKA_DATA)

            # Net nakit açığı: gider - gelir - banka_bakiye

            net_acik = max(0, gider_bekl - gelir_bekl - toplam_banka)



            # Kredileri maliyete göre sırala (faiz oranı en düşük önce)

            kullanilabilir = []

            for k in KREDI_DATA:

                bos = k.get('limit',0) - k.get('kullanilan',0)

                if bos > 0:

                    kullanilabilir.append({

                        'id':      k.get('id',''),

                        'banka':   k.get('banka',''),

                        'tur':     k.get('tur',''),

                        'firma':   k.get('firma',''),

                        'limit':   k.get('limit',0),

                        'kullanilan': k.get('kullanilan',0),

                        'bos':     round(bos,2),

                        'faiz':    k.get('faiz',0),

                        'aciklama': k.get('aciklama',''),

                    })

            kullanilabilir.sort(key=lambda x: x['faiz'])



            # Öneri oluştur: en ucuz kredilerden başlayarak açığı kapat

            oneriler = []

            kalan_acik = net_acik

            toplam_kullanilacak = 0

            tahmini_maliyet = 0



            for k in kullanilabilir:

                if kalan_acik <= 0:

                    break

                kullan = min(k['bos'], kalan_acik)

                aylik_faiz = round(kullan * k['faiz'] / 100, 2)

                oneriler.append({

                    'banka':   k['banka'],

                    'tur':     k['tur'],

                    'firma':   k['firma'],

                    'bos':     k['bos'],

                    'faiz':    k['faiz'],

                    'kullan':  round(kullan, 2),

                    'aylik_faiz': aylik_faiz,

                    'oncelik': len(oneriler) + 1,

                })

                toplam_kullanilacak += kullan

                tahmini_maliyet     += aylik_faiz

                kalan_acik          -= kullan



            karsilandi = net_acik > 0 and kalan_acik <= 0

            self.send_json({

                'ay': ay, 'yil': yil,

                'gider_bekl': round(gider_bekl, 2),

                'gelir_bekl': round(gelir_bekl, 2),

                'toplam_banka': round(toplam_banka, 2),

                'net_acik': round(net_acik, 2),

                'toplam_kullanilacak': round(toplam_kullanilacak, 2),

                'tahmini_maliyet': round(tahmini_maliyet, 2),

                'kalan_karsilanamiyan': round(max(0, kalan_acik), 2),

                'karsilandi': karsilandi,

                'oneriler': oneriler,

                'kullanilabilir_toplam': round(sum(k['bos'] for k in kullanilabilir), 2),

            })



        elif path == '/api/export/excel':

            mode   = qs.get('mode',['all'])[0]

            params = {k:v[0] for k,v in qs.items() if k!='mode'}

            # yil filtresi apply_filter üzerinden çalışır

            data   = export_excel_data(mode, params)

            fname  = f"nakit_akis_{mode}_{datetime.now().strftime('%Y%m%d')}.xlsx"

            self.send_file(data, fname,

                'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')



        elif path == '/api/export/csv':

            import csv, io as sio

            buf = sio.StringIO()

            w = csv.writer(buf)

            w.writerow(['Firma','Ödeme Türü','Ödeme Yeri','Tarih','Ay','Durum',

                        'Anapara','Faiz','BSMV','Tutar','Açıklama'])

            for d in DATA:

                w.writerow([d.get(k,'') for k in

                    ['firma','tur','yer','tarih','ay','durum','anapara','faiz','bsmv','tutar','aciklama']])

            data = ('\ufeff'+buf.getvalue()).encode('utf-8')

            self.send_file(data, f"nakit_akis_{datetime.now().strftime('%Y%m%d')}.csv",

                           'text/csv; charset=utf-8')



        elif self.path == '/api/whapi/config/save':

            try:

                p = json.loads(body)

                WHAPI_CONFIG.update(p)

                with open(WHAPI_FILE, 'w', encoding='utf-8') as f:

                    json.dump(WHAPI_CONFIG, f, ensure_ascii=False, indent=2)

                self.send_json({'ok': True})

            except Exception as e:

                self.send_json({'ok': False, 'msg': str(e)})



        elif self.path == '/api/whapi/send':

            try:

                import urllib.request as ureq

                p = json.loads(body)

                token = WHAPI_CONFIG.get('token','')

                if not token:

                    self.send_json({'ok': False, 'msg': 'Whapi token girilmemiş'}); return

                to = p.get('to','').strip().lstrip('+')

                chat_id = p.get('chat_id', to + '@s.whatsapp.net')

                chat_ad = p.get('chat_ad', to)

                msg = p.get('message','').strip()

                if not to or not msg:

                    self.send_json({'ok': False, 'msg': 'Numara ve mesaj zorunlu'}); return

                api_url = 'https://gate.whapi.cloud/messages/text'

                payload = json.dumps({'to': to, 'body': msg}).encode('utf-8')

                req = ureq.Request(api_url, data=payload,

                    headers={'Authorization': f'Bearer {token}', 'Content-Type': 'application/json'})

                with ureq.urlopen(req, timeout=15) as resp:

                    result = json.loads(resp.read())

                # Kalıcı kayıt

                import time as _time

                yeni_mesaj = {

                    'id': result.get('sent_message',{}).get('id', str(_time.time())),

                    'from_me': True, 'text': {'body': msg},

                    'timestamp': int(_time.time()), 'type': 'text'

                }

                if chat_id not in WHAPI_MSGS:

                    WHAPI_MSGS[chat_id] = {'ad': chat_ad, 'mesajlar': []}

                WHAPI_MSGS[chat_id]['mesajlar'].append(yeni_mesaj)

                WHAPI_MSGS[chat_id]['son_mesaj'] = msg

                WHAPI_MSGS[chat_id]['son_zaman'] = int(_time.time())

                with open(WHAPI_MSG_FILE, 'w', encoding='utf-8') as f:

                    json.dump(WHAPI_MSGS, f, ensure_ascii=False, indent=2)

                self.send_json({'ok': True, 'result': result})

            except Exception as e:

                self.send_json({'ok': False, 'msg': str(e)})



        elif self.path == '/api/whapi/messages':

            try:

                import urllib.request as ureq, time as _time

                p = json.loads(body)

                token = WHAPI_CONFIG.get('token','')

                if not token:

                    self.send_json({'ok': False, 'msg': 'Whapi token girilmemiş'}); return

                chat_id = p.get('chat_id','')

                chat_ad = p.get('chat_ad', chat_id)

                limit = p.get('limit', 50)

                api_url = f'https://gate.whapi.cloud/messages/list/{chat_id}?count={limit}'

                req = ureq.Request(api_url, headers={'Authorization': f'Bearer {token}'})

                with ureq.urlopen(req, timeout=15) as resp:

                    result = json.loads(resp.read())

                msgs = result.get('messages', [])

                # Kalıcı kayıt — mevcut mesajlarla birleştir

                if chat_id not in WHAPI_MSGS:

                    WHAPI_MSGS[chat_id] = {'ad': chat_ad, 'mesajlar': []}

                mevcut_ids = {m.get('id') for m in WHAPI_MSGS[chat_id]['mesajlar']}

                yeni_sayisi = 0

                for m in msgs:

                    if m.get('id') not in mevcut_ids:

                        WHAPI_MSGS[chat_id]['mesajlar'].append(m)

                        yeni_sayisi += 1

                # Zaman sırasına göre sırala

                WHAPI_MSGS[chat_id]['mesajlar'].sort(key=lambda x: x.get('timestamp',0))

                WHAPI_MSGS[chat_id]['ad'] = chat_ad

                if msgs:

                    son = msgs[-1]

                    WHAPI_MSGS[chat_id]['son_mesaj'] = son.get('text',{}).get('body','') or son.get('caption','')

                    WHAPI_MSGS[chat_id]['son_zaman'] = son.get('timestamp', int(_time.time()))

                with open(WHAPI_MSG_FILE, 'w', encoding='utf-8') as f:

                    json.dump(WHAPI_MSGS, f, ensure_ascii=False, indent=2)

                self.send_json({'ok': True, 'messages': WHAPI_MSGS[chat_id]['mesajlar'], 'yeni': yeni_sayisi})

            except Exception as e:

                self.send_json({'ok': False, 'msg': str(e)})



        elif self.path == '/api/whapi/chats':

            try:

                import urllib.request as ureq

                token = WHAPI_CONFIG.get('token','')

                if not token:

                    self.send_json({'ok': False, 'msg': 'Whapi token girilmemiş'}); return

                api_url = 'https://gate.whapi.cloud/chats?count=50'

                req = ureq.Request(api_url, headers={'Authorization': f'Bearer {token}'})

                with ureq.urlopen(req, timeout=15) as resp:

                    result = json.loads(resp.read())

                chats = result.get('chats', [])

                # Whapi'den gelen chat listesini kalıcı kayıtla zenginleştir

                for c in chats:

                    cid = c.get('id','')

                    if cid in WHAPI_MSGS:

                        c['_kayitli_mesaj_sayisi'] = len(WHAPI_MSGS[cid]['mesajlar'])

                self.send_json({'ok': True, 'chats': chats, 'kayitli': {k: {'ad': v['ad'], 'mesaj_sayisi': len(v['mesajlar']), 'son_zaman': v.get('son_zaman',0), 'son_mesaj': v.get('son_mesaj','')} for k,v in WHAPI_MSGS.items()}})

            except Exception as e:

                self.send_json({'ok': False, 'msg': str(e)})



        elif self.path == '/api/whapi/config':

            self.send_json({'ok': True, 'config': WHAPI_CONFIG})



        elif self.path == '/api/docs/list' or self.path.startswith('/api/docs/list?'):

            from urllib.parse import parse_qs as _pqs2, urlparse as _up2

            qs = _pqs2(_up2(self.path).query)

            q     = qs.get('q',[''])[0].lower()

            kat   = qs.get('kategori',[''])[0]

            docs  = DOCS_META[:]

            if q:

                docs = [d for d in docs if q in d.get('ad','').lower() or q in d.get('not','').lower() or q in d.get('kategori','').lower()]

            if kat:

                docs = [d for d in docs if d.get('kategori','') == kat]

            docs.sort(key=lambda x: x.get('tarih',''), reverse=True)

            kategoriler = sorted(set(d.get('kategori','Genel') for d in DOCS_META))

            self.send_json({'ok': True, 'docs': docs, 'kategoriler': kategoriler, 'toplam': len(DOCS_META)})



        elif self.path.startswith('/api/docs/indir/') or self.path.startswith('/api/docs/onizle/'):

            is_indir = self.path.startswith('/api/docs/indir/')

            doc_id = self.path.split('/')[-1]

            doc = next((d for d in DOCS_META if d.get('id') == doc_id), None)

            if not doc:

                self.send_response(404); self.send_header('Content-Type','text/plain'); self.end_headers()

                self.wfile.write(b'Belge bulunamadi'); return

            dosya_adi = doc.get('dosya_adi','')

            search_dirs = [DOCS_DIR, APP_DIR, os.path.join(APP_DIR,'belgeler')]

            filepath = None

            for sd in search_dirs:

                fp = os.path.join(sd, dosya_adi)

                if os.path.exists(fp): filepath = fp; break

            if not filepath:

                self.send_response(404); self.send_header('Content-Type','text/plain'); self.end_headers()

                self.wfile.write(('Dosya yok: '+dosya_adi).encode()); return

            import mimetypes as _mt

            mime = _mt.guess_type(filepath)[0] or 'application/octet-stream'

            fsize = os.path.getsize(filepath)

            disp = 'attachment' if is_indir else 'inline'

            orijinal = doc.get('orijinal_ad', dosya_adi)

            try: orijinal_enc = orijinal.encode('latin-1').decode('latin-1')

            except: orijinal_enc = orijinal.encode('ascii','replace').decode('ascii')

            self.send_response(200)

            self.send_header('Content-Type', mime)

            self.send_header('Content-Disposition', f'{disp}; filename="{orijinal_enc}"')

            self.send_header('Content-Length', str(fsize))

            self.send_header('Accept-Ranges', 'bytes')

            self.end_headers()

            try:

                with open(filepath, 'rb') as fobj:

                    while True:

                        chunk = fobj.read(65536)

                        if not chunk: break

                        self.wfile.write(chunk)

            except (BrokenPipeError, ConnectionResetError):

                pass



        elif self.path == '/api/whapi/kayitli_sohbetler':

            import time as _t2

            ozet = {}

            for k, v in WHAPI_MSGS.items():

                ozet[k] = {

                    'ad': v.get('ad', k),

                    'mesaj_sayisi': len(v.get('mesajlar', [])),

                    'son_zaman': v.get('son_zaman', 0),

                    'son_mesaj': v.get('son_mesaj', ''),

                    'okunmadi': v.get('okunmadi', 0),

                }

            self.send_json({'ok': True, 'sohbetler': ozet})



        elif self.path.startswith('/api/whapi/kayitli_mesajlar?'):

            from urllib.parse import parse_qs as _pqs, urlparse as _uparse

            qs = _pqs(_uparse(self.path).query)

            chat_id = qs.get('chat_id', [''])[0]

            if chat_id and chat_id in WHAPI_MSGS:

                # Okundu işaretle

                WHAPI_MSGS[chat_id]['okunmadi'] = 0

                with open(WHAPI_MSG_FILE, 'w', encoding='utf-8') as f:

                    json.dump(WHAPI_MSGS, f, ensure_ascii=False, indent=2)

                self.send_json({'ok': True, 'mesajlar': WHAPI_MSGS[chat_id].get('mesajlar', []), 'ad': WHAPI_MSGS[chat_id].get('ad', '')})

            else:

                self.send_json({'ok': True, 'mesajlar': [], 'ad': ''})



        elif self.path == '/api/whapi/kayitli_mesajlar':

            # Sadece kalıcı kayıtlı mesajları döndür (offline çalışır)

            p = json.loads(body) if body else {}

            chat_id = p.get('chat_id','')

            if chat_id:

                self.send_json({'ok': True, 'mesajlar': WHAPI_MSGS.get(chat_id, {}).get('mesajlar', []), 'ad': WHAPI_MSGS.get(chat_id, {}).get('ad','')})

            else:

                # Tüm sohbet özeti

                ozet = {k: {'ad': v['ad'], 'mesaj_sayisi': len(v['mesajlar']), 'son_zaman': v.get('son_zaman',0), 'son_mesaj': v.get('son_mesaj','')} for k,v in WHAPI_MSGS.items()}

                self.send_json({'ok': True, 'sohbetler': ozet})



        elif self.path == '/api/whapi/mesaj_sil':

            try:

                p = json.loads(body)

                chat_id = p.get('chat_id','')

                if chat_id and chat_id in WHAPI_MSGS:

                    del WHAPI_MSGS[chat_id]

                    with open(WHAPI_MSG_FILE, 'w', encoding='utf-8') as f:

                        json.dump(WHAPI_MSGS, f, ensure_ascii=False, indent=2)

                self.send_json({'ok': True})

            except Exception as e:

                self.send_json({'ok': False, 'msg': str(e)})



        elif self.path == '/api/whapi/webhook':

            # Whapi.Cloud'dan gelen gerçek zamanlı mesajlar

            try:

                import time as _wt

                payload = json.loads(body) if body else {}

                messages = payload.get('messages', [])

                degisti = False

                for m in messages:

                    chat_id = m.get('chat_id', '')

                    if not chat_id:

                        continue

                    # Gönderen adı

                    ad = m.get('chat_name') or m.get('from_name') or chat_id

                    metin = ''

                    if m.get('text'):

                        metin = m['text'].get('body', '')

                    elif m.get('caption'):

                        metin = m['caption']

                    elif m.get('type'):

                        metin = '[' + m['type'] + ']'



                    if chat_id not in WHAPI_MSGS:

                        WHAPI_MSGS[chat_id] = {'ad': ad, 'mesajlar': [], 'okunmadi': 0}

                    

                    mevcut_ids = {x.get('id') for x in WHAPI_MSGS[chat_id]['mesajlar']}

                    if m.get('id') not in mevcut_ids:

                        WHAPI_MSGS[chat_id]['mesajlar'].append(m)

                        WHAPI_MSGS[chat_id]['son_mesaj'] = metin

                        WHAPI_MSGS[chat_id]['son_zaman'] = m.get('timestamp', int(_wt.time()))

                        WHAPI_MSGS[chat_id]['ad'] = ad

                        if not m.get('from_me'):

                            WHAPI_MSGS[chat_id]['okunmadi'] = WHAPI_MSGS[chat_id].get('okunmadi', 0) + 1

                        degisti = True



                if degisti:

                    with open(WHAPI_MSG_FILE, 'w', encoding='utf-8') as f:

                        json.dump(WHAPI_MSGS, f, ensure_ascii=False, indent=2)

                self.send_json({'ok': True, 'processed': len(messages)})

            except Exception as e:

                self.send_json({'ok': False, 'msg': str(e)})



        else:

            self.send_response(404); self.end_headers()



    def do_POST(self):

        global DATA,GELIR_DATA,BANKA_DATA,KREDI_DATA,CARI_DATA,SOURCE_FILE,SOURCE_SHEET,AUTO_RELOAD,NOTLAR,API_CONFIG,LOGO_CONFIG,AI_CONFIG,OTEL_DATA,OTEL_CONFIG,HATIRLATMALAR,GMAIL_CONFIG,WHAPI_CONFIG,WHAPI_MSGS,DOCS_META,BANKA_KART_DATA

        length = int(self.headers.get('Content-Length',0))

        # Body'yi önce oku

        body = b''

        remaining = length

        while remaining > 0:

            chunk_size = min(remaining, 65536)

            chunk = self.rfile.read(chunk_size)

            if not chunk: break

            body += chunk; remaining -= len(chunk)



        # ── LOGIN API ─────────────────────────────────────────────────────────────

        if self.path == '/api/login':

            try:

                p = json.loads(body)

                username = p.get('username','').strip()

                password = p.get('password','')

                users = load_users()

                u = users.get(username)

                if u and u['password'] == hash_password(password):

                    token = create_session(username)

                    self.send_response(200)

                    self.send_header('Content-Type','application/json; charset=utf-8')

                    self.send_header('Set-Cookie',f'na_token={token}; Path=/; HttpOnly; SameSite=Lax; Max-Age={SESSION_TTL}')

                    self.end_headers()

                    self.wfile.write(json.dumps({'ok':True,'username':username,'ad':u.get('ad','')},ensure_ascii=False).encode('utf-8'))

                else:

                    self.send_json({'ok':False,'msg':'Kullanici adi veya sifre hatali.'},401)

            except Exception as e:

                self.send_json({'ok':False,'msg':str(e)},500)

            return



        # ── AUTH KONTROL ───────────────────────────────────────────────────────────

        # docs/yukle multipart içinde kendi token kontrolü yapar

        if self.path not in ('/api/docs/yukle',):

            sess = check_auth(self)

            if not sess:

                self.send_json({'ok':False,'msg':'Oturum suresi doldu. Sayfayi yenileyin.'},401); return

        else:

            sess = check_auth(self)  # None olabilir, yukle handler kendi kontrol eder



        # ── KULLANICI YÖNETİMİ ────────────────────────────────────────────────────

        if self.path == '/api/users/add':

            users = load_users()

            if users.get(sess['username'],{}).get('rol') != 'admin':

                self.send_json({'ok':False,'msg':'Sadece admin ekleyebilir.'}); return

            p = json.loads(body)

            uname=p.get('username','').strip(); pw=p.get('password',''); ad=p.get('ad','').strip(); rol=p.get('rol','user')

            if not uname or not pw: self.send_json({'ok':False,'msg':'Zorunlu alan eksik.'}); return

            if uname in users: self.send_json({'ok':False,'msg':'Bu kullanici zaten var.'}); return

            users[uname]={'password':hash_password(pw),'ad':ad or uname,'rol':rol,'created_at':datetime.now().strftime('%d.%m.%Y %H:%M')}

            save_users(users); self.send_json({'ok':True}); return



        if self.path == '/api/users/delete':

            users = load_users()

            if users.get(sess['username'],{}).get('rol') != 'admin':

                self.send_json({'ok':False,'msg':'Sadece admin silebilir.'}); return

            p = json.loads(body); uname=p.get('username','')

            if uname == 'admin': self.send_json({'ok':False,'msg':'Admin silinemez.'}); return

            if uname in users: del users[uname]; save_users(users)

            self.send_json({'ok':True}); return



        if self.path == '/api/users/change_password':

            p = json.loads(body); users=load_users()

            target=p.get('username',sess['username'])

            if target != sess['username'] and users.get(sess['username'],{}).get('rol') != 'admin':

                self.send_json({'ok':False,'msg':'Yetersiz yetki.'}); return

            new_pw=p.get('new_password','')

            if len(new_pw)<6: self.send_json({'ok':False,'msg':'Sifre en az 6 karakter olmali.'}); return

            if target in users: users[target]['password']=hash_password(new_pw); save_users(users)

            self.send_json({'ok':True}); return



        if self.path == '/api/users/list':

            users=load_users()

            if users.get(sess['username'],{}).get('rol') != 'admin':

                self.send_json({'ok':False,'msg':'Sadece admin gorebilir.'}); return

            lst=[{'username':k,'ad':v.get('ad',''),'rol':v.get('rol','user'),'created_at':v.get('created_at','')} for k,v in users.items()]

            self.send_json({'ok':True,'users':lst}); return



        # ── İPOTEK API ────────────────────────────────────────────────────────────

        if self.path == '/api/banka_kart/kaydet':

            try:

                kayit = json.loads(body)

                mevcut = next((i for i,r in enumerate(BANKA_KART_DATA) if str(r.get('id'))==str(kayit.get('id'))), None)

                if mevcut is not None: BANKA_KART_DATA[mevcut]=kayit

                else: BANKA_KART_DATA.append(kayit)

                save_banka_kart(); self.send_json({'ok':True})

            except Exception as e: self.send_json({'ok':False,'msg':str(e)})

            return



        if self.path == '/api/banka_kart/sil':

            try:

                p=json.loads(body)

                BANKA_KART_DATA[:]=[r for r in BANKA_KART_DATA if str(r.get('id'))!=str(p.get('id'))]

                save_banka_kart(); self.send_json({'ok':True})

            except Exception as e: self.send_json({'ok':False,'msg':str(e)})

            return



        if self.path == '/api/banka_kart/aktar_kredi':

            # Bu banka kartındaki limitleri KREDI_DATA'ya aktar

            try:

                p = json.loads(body)

                kart_id = str(p.get('id'))

                kart = next((r for r in BANKA_KART_DATA if str(r.get('id'))==kart_id), None)

                if not kart:

                    self.send_json({'ok':False,'msg':'Banka kartı bulunamadı'}); return

                banka = kart.get('banka','')

                firma = kart.get('firma','ULUSAL')

                # Önce bu bankaya ait eski kayıtları temizle

                KREDI_DATA[:] = [k for k in KREDI_DATA if not (k.get('banka')==banka and k.get('kaynak')=='banka_kart')]

                # Yeni kayıtlar ekle

                eklendi = 0

                limit_turleri = [

                    ('taksitli','TAKSİTLİ KREDİ'),('rotatif','ROTATİF'),

                    ('kmh','KMH'),('kredi_karti','KREDİ KARTI'),

                    ('teminat','TEMİNAT MEKTUBU'),('diger','DİĞER'),

                ]

                for alan, tur_adi in limit_turleri:

                    limit = float(kart.get(f'limit_{alan}',0) or 0)

                    if limit > 0:

                        kullanilan = float(kart.get(f'kullanim_{alan}',0) or 0)

                        KREDI_DATA.append({

                            'id': f'bk_{kart_id}_{alan}_{int(datetime.now().timestamp())}',

                            'firma': firma, 'banka': banka, 'tur': tur_adi,

                            'limit': limit, 'kullanilan': kullanilan,

                            'kalan_limit': round(limit - kullanilan, 2),

                            'faiz_yillik': float(kart.get(f'faiz_{alan}',0) or 0) / 100,

                            'faiz_aylik': float(kart.get(f'faiz_{alan}',0) or 0) / 100 / 12,

                            'aktif': True, 'ad': tur_adi, 'bitis': '',

                            'kaynak': 'banka_kart', 'aciklama': f'{banka} - Banka Kart aktarımı'

                        })

                        eklendi += 1

                save_kredi()

                self.send_json({'ok':True,'eklendi':eklendi,'msg':f'{eklendi} kredi limiti aktarıldı'})

            except Exception as e:

                self.send_json({'ok':False,'msg':str(e)})

            return



        if self.path == '/api/ipotek/kaydet':

            try:

                kayit = json.loads(body)

                mevcut = next((i for i,r in enumerate(IPOTEK_DATA) if str(r.get('id'))==str(kayit.get('id'))), None)

                if mevcut is not None: IPOTEK_DATA[mevcut]=kayit

                else: IPOTEK_DATA.append(kayit)

                save_ipotek(); self.send_json({'ok':True})

            except Exception as e: self.send_json({'ok':False,'msg':str(e)})

            return



        if self.path == '/api/ipotek/sil':

            try:

                p=json.loads(body)

                IPOTEK_DATA[:]=[r for r in IPOTEK_DATA if str(r.get('id'))!=str(p.get('id'))]

                save_ipotek(); self.send_json({'ok':True})

            except Exception as e: self.send_json({'ok':False,'msg':str(e)})

            return



        if self.path == '/api/ipotek/excel_yukle':

            try:

                import cgi, tempfile

                ctype=self.headers.get('Content-Type','')

                environ={'REQUEST_METHOD':'POST','CONTENT_TYPE':ctype,'CONTENT_LENGTH':str(length)}

                form=cgi.FieldStorage(fp=io.BytesIO(body),environ=environ,keep_blank_values=True)

                fileitem=form['file']

                ext=os.path.splitext(fileitem.filename)[1].lower()

                tmp=tempfile.NamedTemporaryFile(suffix=ext,delete=False)

                tmp.write(fileitem.file.read()); tmp.close()

                import pandas as pd

                df=pd.read_excel(tmp.name,header=0,dtype=str).fillna('')

                os.unlink(tmp.name)

                def nc(col): return str(col).upper().strip().replace('İ','I').replace('Ş','S').replace('Ğ','G').replace('Ü','U').replace('Ö','O').replace('Ç','C').replace(' ','').replace('.','')

                col_map={}

                for col in df.columns:

                    c=nc(col)

                    if c in ('SNO','S.NO','NO'): col_map['sno']=col

                    elif 'KOD' in c and 'TASIN' in c: col_map['kod']=col

                    elif 'KOD' in c: col_map.setdefault('kod',col)

                    elif 'SAHIB' in c: col_map['sahip']=col

                    elif c in ('ILI','IL'): col_map['il']=col

                    elif 'ILCE' in c: col_map['ilce']=col

                    elif 'ADRES' in c: col_map['adres']=col

                    elif c=='ADA': col_map['ada']=col

                    elif 'PARSEL' in c: col_map['parsel']=col

                    elif c in ('BB','B.B'): col_map['bb']=col

                    elif c=='KAT': col_map['kat']=col

                    elif c=='BLOK': col_map['blok']=col

                    elif 'NITELIK' in c: col_map['nitelik']=col

                    elif 'IPOTEK' in c and 'TAR' not in c and 'TUTAR' not in c and 'DEGER' not in c: col_map['banka']=col

                    elif 'TUTAR' in c: col_map['tutar']=col

                    elif 'IPOTEK' in c and 'TAR' in c: col_map['ipotek_tarih']=col

                    elif 'EKSPERTIZ' in c and 'TAR' in c: col_map['exp_tarih']=col

                    elif 'EKSPERTIZ' in c and 'DEGER' in c: col_map['exp_deger']=col

                    elif 'TASINMAZ' in c and 'TIP' in c: col_map['tbl_tip']=col

                    elif 'DASK' in c and 'POLICE' in c: col_map['dask_no']=col

                    elif 'DASK' in c and 'VADE' in c: col_map['dask_vade']=col

                    elif 'KONUT' in c and 'POLICE' in c: col_map['konut_no']=col

                    elif 'KONUT' in c and 'VADE' in c: col_map['konut_vade']=col

                kayitlar=[]

                for _,row in df.iterrows():

                    def gv(f): return str(row.get(col_map.get(f,''),'') if col_map.get(f) else '').strip().replace('nan','')

                    def gn(f):

                        v=gv(f).replace('.','').replace(',','.')

                        try: return float(v)

                        except: return 0

                    il=gv('il')

                    if not il: continue

                    kayitlar.append({'id':str(int(datetime.now().timestamp()*1000)+len(kayitlar)),'sno':gv('sno') or str(len(IPOTEK_DATA)+len(kayitlar)+1),'kod':gv('kod'),'sahip':gv('sahip'),'il':il,'ilce':gv('ilce'),'adres':gv('adres'),'ada':gv('ada'),'parsel':gv('parsel'),'bb':gv('bb'),'kat':gv('kat'),'blok':gv('blok'),'nitelik':gv('nitelik'),'banka':gv('banka'),'tutar':gn('tutar'),'ipotek_tarih':gv('ipotek_tarih'),'exp_deger':gn('exp_deger'),'exp_tarih':gv('exp_tarih'),'tbl_tip':gv('tbl_tip'),'dask_no':gv('dask_no'),'dask_vade':gv('dask_vade'),'konut_no':gv('konut_no'),'konut_vade':gv('konut_vade')})

                IPOTEK_DATA.extend(kayitlar)

                save_ipotek()

                self.send_json({'ok':True,'count':len(kayitlar),'msg':f'{len(kayitlar)} tasınmaz yuklendi'})

            except Exception as e:

                traceback.print_exc(); self.send_json({'ok':False,'msg':str(e)})

            return



        # ── JSON İÇE AKTAR ────────────────────────────────────────────────────────

        if self.path == '/api/import/json':

            try:

                p=json.loads(body); tur=p.get('tur','gider'); data_in=p.get('data',[])

                if tur=='gider': DATA[:]=data_in; save_data(); self.send_json({'ok':True,'msg':f'{len(DATA)} gider kaydı yüklendi'})

                elif tur=='gelir': GELIR_DATA[:]=data_in; save_gelir(); self.send_json({'ok':True,'msg':f'{len(GELIR_DATA)} gelir kaydı yüklendi'})

                elif tur=='banka': BANKA_DATA[:]=data_in; save_banka(); self.send_json({'ok':True,'msg':f'{len(BANKA_DATA)} banka kaydı yüklendi'})

                elif tur=='kredi': KREDI_DATA[:]=data_in; save_kredi(); self.send_json({'ok':True,'msg':f'{len(KREDI_DATA)} kredi kaydı yüklendi'})

                elif tur=='cari': CARI_DATA[:]=data_in; save_cari(); self.send_json({'ok':True,'msg':f'{len(CARI_DATA)} cari kaydı yüklendi'})

                elif tur=='notlar': NOTLAR[:]=data_in; save_notlar(); self.send_json({'ok':True,'msg':f'{len(NOTLAR)} not yüklendi'})

                elif tur=='otel':

                    if isinstance(data_in,dict): OTEL_DATA.clear(); OTEL_DATA.update(data_in)

                    else: OTEL_DATA['otel1']=data_in

                    save_otel_data(); self.send_json({'ok':True,'msg':'Otel verisi yüklendi'})

                elif tur=='hatirlatma': HATIRLATMALAR[:]=data_in; save_hatirlatmalar(); self.send_json({'ok':True,'msg':f'{len(HATIRLATMALAR)} hatırlatıcı yüklendi'})

                else: self.send_json({'ok':False,'msg':'Bilinmeyen tur'})

            except Exception as e: self.send_json({'ok':False,'msg':str(e)})

            return



        # ── Kaynak dosya ayarla ───────────────────────────────────────────────

        if self.path == '/api/source/set':

            try:

                payload = json.loads(body)

                fp = payload.get('file','').strip()

                if fp and not os.path.exists(fp):

                    self.send_json({'ok':False,'msg':'Dosya bulunamadi: '+fp}); return

                SOURCE_FILE  = fp or None

                SOURCE_SHEET = payload.get('sheet','').strip() or None

                AUTO_RELOAD  = payload.get('auto', True)

                SOURCE_MTIME = None

                save_config()

                if SOURCE_FILE:

                    ok = reload_from_source(force=True)

                    msg = LAST_SYNC_MSG

                    self.send_json({'ok':ok,'msg':msg,'count':len(DATA)})

                else:

                    self.send_json({'ok':True,'msg':'Kaynak dosya kaldirildi','count':len(DATA)})

            except Exception as e:

                traceback.print_exc()

                self.send_json({'ok':False,'msg':type(e).__name__+': '+str(e)})



        elif self.path == '/api/docs/yukle':

            import uuid as _uuid, mimetypes as _mime

            try:

                ct = self.headers.get('Content-Type','')

                if 'multipart/form-data' not in ct:

                    self.send_json({'ok':False,'msg':'multipart gerekli'}); return

                # Content-Length'e göre tam oku

                content_length = int(self.headers.get('Content-Length', 0))

                full_body = body

                if content_length > len(body):

                    full_body = body + self.rfile.read(content_length - len(body))

                # Boundary ayır

                boundary = None

                for part in ct.split(';'):

                    part = part.strip()

                    if part.startswith('boundary='):

                        boundary = part[9:].strip().strip('"').encode()

                        break

                if not boundary:

                    self.send_json({'ok':False,'msg':'Boundary bulunamadi'}); return

                dosya_adi_orijinal = ''

                dosya_data = b''

                kategori = 'Genel'

                not_metni = ''

                _token_field = ''

                sess_override = None

                # Kendi session kontrolü - cookie veya form token

                _yukle_sess = check_auth(self)

                # Her parçayı işle

                for chunk in full_body.split(b'--' + boundary):

                    if b'Content-Disposition' not in chunk: continue

                    # Header ve body ayrımı

                    sep = chunk.find(b'\r\n\r\n')

                    if sep == -1: continue

                    hdr_raw = chunk[:sep].decode('utf-8','ignore')

                    cdata = chunk[sep+4:]

                    # Sondaki CRLF ve -- temizle

                    if cdata.endswith(b'\r\n'):

                        cdata = cdata[:-2]

                    if cdata.endswith(b'--'):

                        cdata = cdata[:-2]

                    if 'filename=' in hdr_raw:

                        # Dosya adını al

                        fn_part = hdr_raw.split('filename=')[1]

                        fn = fn_part.split('"')[1] if '"' in fn_part else fn_part.split(';')[0].strip()

                        # UTF-8 decode dene

                        try: fn = fn.encode('latin-1').decode('utf-8')

                        except: pass

                        dosya_adi_orijinal = fn

                        dosya_data = cdata

                    elif 'name="kategori"' in hdr_raw:

                        kategori = cdata.decode('utf-8','ignore').strip() or 'Genel'

                    elif 'name="not"' in hdr_raw:

                        not_metni = cdata.decode('utf-8','ignore').strip()

                    elif 'name="_token"' in hdr_raw:

                        _token_field = cdata.decode('utf-8','ignore').strip()

                # Form'daki _token ile session kontrol et

                if _token_field and not _yukle_sess:

                    _yukle_sess = get_session(_token_field)

                # Hâlâ auth yok - 401

                if not _yukle_sess:

                    self.send_json({'ok':False,'msg':'Oturum bulunamadi — sayfayi yenileyip tekrar deneyin'},401); return

                if not dosya_data:

                    self.send_json({'ok':False,'msg':'Dosya verisi bos — boyut: '+str(content_length)}); return

                if not dosya_adi_orijinal:

                    dosya_adi_orijinal = 'belge_' + str(_uuid.uuid4())[:6]

                doc_id = str(_uuid.uuid4())[:8]

                ext = os.path.splitext(dosya_adi_orijinal)[1].lower()

                dosya_adi_disk = doc_id + (ext or '.bin')

                filepath = os.path.join(DOCS_DIR, dosya_adi_disk)

                os.makedirs(DOCS_DIR, exist_ok=True)

                with open(filepath, 'wb') as f:

                    f.write(dosya_data)

                boyut = len(dosya_data)

                mime = _mime.guess_type(dosya_adi_orijinal)[0] or 'application/octet-stream'

                meta = {

                    'id': doc_id,

                    'ad': dosya_adi_orijinal,

                    'orijinal_ad': dosya_adi_orijinal,

                    'dosya_adi': dosya_adi_disk,

                    'kategori': kategori,

                    'not': not_metni,

                    'boyut': boyut,

                    'mime': mime,

                    'tarih': __import__('datetime').datetime.now().strftime('%Y-%m-%d %H:%M'),

                    'ext': ext.lstrip('.')

                }

                # Aynı isimde dosya varsa güncelle, yoksa ekle

                mevcut = next((i for i,d in enumerate(DOCS_META) if d.get('orijinal_ad')==dosya_adi_orijinal and d.get('kategori')==kategori), None)

                if mevcut is not None:

                    # Eski fiziksel dosyayı sil

                    eski = os.path.join(DOCS_DIR, DOCS_META[mevcut].get('dosya_adi',''))

                    if os.path.exists(eski) and eski != filepath:

                        os.remove(eski)

                    DOCS_META[mevcut] = meta

                else:

                    DOCS_META.append(meta)

                with open(DOCS_META_FILE, 'w', encoding='utf-8') as f:

                    json.dump(DOCS_META, f, ensure_ascii=False, indent=2)

                self.send_json({'ok':True,'doc':meta})

            except Exception as e:

                import traceback; traceback.print_exc()

                self.send_json({'ok':False,'msg':str(e)})



        elif self.path == '/api/docs/not_guncelle':

            try:

                p = json.loads(body)

                doc = next((d for d in DOCS_META if d.get('id') == p.get('id')), None)

                if doc:

                    doc['not'] = p.get('not','')

                    doc['kategori'] = p.get('kategori', doc.get('kategori','Genel'))

                    doc['ad'] = p.get('ad', doc.get('ad',''))

                    with open(DOCS_META_FILE, 'w', encoding='utf-8') as f:

                        json.dump(DOCS_META, f, ensure_ascii=False, indent=2)

                    self.send_json({'ok':True})

                else:

                    self.send_json({'ok':False,'msg':'Bulunamadi'})

            except Exception as e:

                self.send_json({'ok':False,'msg':str(e)})



        elif self.path == '/api/docs/temizle_duplicate':

            # Aynı orijinal_ad'a sahip kayıtlardan sadece en yeniyi tut

            goruldu = {}

            temiz = []

            for d in sorted(DOCS_META, key=lambda x: x.get('tarih',''), reverse=True):

                k = d.get('orijinal_ad','')

                if k not in goruldu:

                    goruldu[k] = True

                    temiz.append(d)

                else:

                    # Fiziksel dosyayı sil

                    fp = os.path.join(DOCS_DIR, d.get('dosya_adi',''))

                    if os.path.exists(fp):

                        os.remove(fp)

            DOCS_META[:] = temiz

            with open(DOCS_META_FILE, 'w', encoding='utf-8') as f:

                json.dump(DOCS_META, f, ensure_ascii=False, indent=2)

            self.send_json({'ok': True, 'kalan': len(DOCS_META)})



        elif self.path == '/api/docs/sil':

            try:

                p = json.loads(body)

                doc = next((d for d in DOCS_META if d.get('id') == p.get('id')), None)

                if doc:

                    filepath = os.path.join(DOCS_DIR, doc.get('dosya_adi',''))

                    if os.path.exists(filepath):

                        os.remove(filepath)

                    DOCS_META[:] = [d for d in DOCS_META if d.get('id') != p.get('id')]

                    with open(DOCS_META_FILE, 'w', encoding='utf-8') as f:

                        json.dump(DOCS_META, f, ensure_ascii=False, indent=2)

                self.send_json({'ok':True})

            except Exception as e:

                self.send_json({'ok':False,'msg':str(e)})



        elif self.path == '/api/docs/webhook':

            # Webhook ile gelen belgeler için placeholder

            self.send_json({'ok':True})



        elif self.path == '/api/record/add':

            rec = json.loads(body)

            DATA.append(rec); save_data()

            self.send_json({'ok':True,'count':len(DATA)})



        elif self.path == '/api/record/edit':

            p = json.loads(body); idx = p.get('idx'); rec = p.get('record')

            if 0 <= idx < len(DATA):

                DATA[idx] = rec; save_data()

                self.send_json({'ok':True})

            else:

                self.send_json({'ok':False,'error':'Geçersiz index'},400)



        elif self.path == '/api/record/delete':

            p = json.loads(body); idx = p.get('idx')

            if 0 <= idx < len(DATA):

                DATA.pop(idx); save_data()

                self.send_json({'ok':True})

            else:

                self.send_json({'ok':False},400)



        elif self.path == '/api/record/toggle':

            p = json.loads(body); idx = p.get('idx')

            if 0 <= idx < len(DATA):

                DATA[idx]['durum'] = 'ÖDENDİ' if DATA[idx].get('durum')!='ÖDENDİ' else 'ÖDENMEDİ'

                save_data()

                # Excel kaynak dosyasına da yaz

                excel_ok = write_back_to_excel(idx)

                self.send_json({'ok':True,'durum':DATA[idx]['durum'],'excel':excel_ok})

            else:

                self.send_json({'ok':False},400)



        elif self.path == '/api/cari/save':

            CARI_DATA = json.loads(body)

            save_cari()

            self.send_json({'ok': True, 'count': len(CARI_DATA)})



        elif self.path == '/api/cari/add':

            rec = json.loads(body)

            rec['_id'] = len(CARI_DATA)

            CARI_DATA.append(rec)

            save_cari()

            self.send_json({'ok': True, 'count': len(CARI_DATA)})



        elif self.path == '/api/cari/edit':

            p = json.loads(body); idx = p.get('idx'); rec = p.get('record')

            if 0 <= idx < len(CARI_DATA):

                CARI_DATA[idx] = rec; save_cari()

                self.send_json({'ok': True})

            else:

                self.send_json({'ok': False}, 400)



        elif self.path == '/api/cari/delete':

            p = json.loads(body); idx = p.get('idx')

            if 0 <= idx < len(CARI_DATA):

                CARI_DATA.pop(idx); save_cari()

                self.send_json({'ok': True})

            else:

                self.send_json({'ok': False}, 400)



        elif self.path == '/api/cari/toggle':

            # Durum değiştir: AÇIK → KAPALI

            p = json.loads(body); idx = p.get('idx')

            if 0 <= idx < len(CARI_DATA):

                cur = CARI_DATA[idx].get('durum', 'AÇIK')

                CARI_DATA[idx]['durum'] = 'KAPALI' if cur == 'AÇIK' else 'AÇIK'

                save_cari()

                self.send_json({'ok': True, 'durum': CARI_DATA[idx]['durum']})

            else:

                self.send_json({'ok': False}, 400)



        elif self.path == '/api/db/sync':

            try:

                results = db_sync_all()

                self.send_json({'ok': True, 'results': results})

            except Exception as e:

                self.send_json({'ok': False, 'msg': str(e)})



        elif self.path == '/api/otel/config/save':

            OTEL_CONFIG = json.loads(body)

            save_otel_config()

            self.send_json({'ok': True})



        elif self.path == '/api/otel/rezervasyon/add':

            p       = json.loads(body)

            otel_id = p.get('otel_id','otel1')

            rezerv  = p.get('rezervasyon',{})

            from datetime import datetime as _dt

            rezerv['id']         = int(_dt.now().timestamp()*1000)

            rezerv['created_at'] = _dt.now().strftime('%d.%m.%Y %H:%M')

            if otel_id not in OTEL_DATA: OTEL_DATA[otel_id] = []

            OTEL_DATA[otel_id].append(rezerv)

            save_otel_data()

            self.send_json({'ok': True, 'id': rezerv['id']})



        elif self.path == '/api/otel/rezervasyon/edit':

            p       = json.loads(body)

            otel_id = p.get('otel_id','otel1')

            rev_id  = p.get('id')

            rezerv  = p.get('rezervasyon',{})

            from datetime import datetime as _dt

            rezerv['updated_at'] = _dt.now().strftime('%d.%m.%Y %H:%M')

            lst = OTEL_DATA.get(otel_id, [])

            for i, r in enumerate(lst):

                if r.get('id') == rev_id:

                    rezerv['id'] = rev_id

                    rezerv['created_at'] = r.get('created_at','')

                    lst[i] = rezerv

                    break

            save_otel_data()

            self.send_json({'ok': True})



        elif self.path == '/api/otel/rezervasyon/delete':

            p       = json.loads(body)

            otel_id = p.get('otel_id','otel1')

            rev_id  = p.get('id')

            OTEL_DATA[otel_id] = [r for r in OTEL_DATA.get(otel_id,[]) if r.get('id') != rev_id]

            save_otel_data()

            self.send_json({'ok': True})



        elif self.path == '/api/otel/import_excel':

            p       = json.loads(body)

            otel_id = p.get('otel_id','otel1')

            rows    = p.get('rows', [])

            from datetime import datetime as _dt

            added = 0

            for row in rows:

                row['id']         = int(_dt.now().timestamp()*1000) + added

                row['created_at'] = _dt.now().strftime('%d.%m.%Y %H:%M')

                if otel_id not in OTEL_DATA: OTEL_DATA[otel_id] = []

                OTEL_DATA[otel_id].append(row)

                added += 1

            save_otel_data()

            self.send_json({'ok': True, 'count': added})



        elif self.path == '/api/hatirlatma/add':

            from datetime import datetime as _dt

            h = json.loads(body)

            h['id']         = int(_dt.now().timestamp()*1000)

            h['olusturuldu'] = _dt.now().strftime('%d.%m.%Y %H:%M')

            h['tamamlandi']  = False

            HATIRLATMALAR.append(h)

            HATIRLATMALAR.sort(key=lambda x:(x.get('tarih',''),x.get('saat','')))

            save_hatirlatmalar()

            self.send_json({'ok': True, 'hatirlatma': h})



        elif self.path == '/api/hatirlatma/edit':

            p = json.loads(body)

            hid = p.get('id')

            for i,h in enumerate(HATIRLATMALAR):

                if h.get('id') == hid:

                    p['data']['id'] = hid

                    p['data']['olusturuldu'] = h.get('olusturuldu','')

                    HATIRLATMALAR[i] = p['data']

                    break

            HATIRLATMALAR.sort(key=lambda x:(x.get('tarih',''),x.get('saat','')))

            save_hatirlatmalar()

            self.send_json({'ok': True})



        elif self.path == '/api/hatirlatma/toggle':

            p = json.loads(body); hid = p.get('id')

            for h in HATIRLATMALAR:

                if h.get('id') == hid:

                    h['tamamlandi'] = not h.get('tamamlandi', False)

                    break

            save_hatirlatmalar()

            self.send_json({'ok': True})



        elif self.path == '/api/hatirlatma/delete':

            hid = json.loads(body).get('id')

            HATIRLATMALAR = [h for h in HATIRLATMALAR if h.get('id') != hid]

            save_hatirlatmalar()

            self.send_json({'ok': True})



        elif self.path == '/api/notlar/save':

            NOTLAR = json.loads(body)

            save_notlar()

            self.send_json({'ok': True})



        elif self.path == '/api/notlar/add':

            not_rec = json.loads(body)

            from datetime import datetime as _dt

            not_rec['id']     = int(_dt.now().timestamp() * 1000)

            not_rec['tarih']  = _dt.now().strftime('%d.%m.%Y %H:%M')

            NOTLAR.insert(0, not_rec)

            save_notlar()

            self.send_json({'ok': True, 'not': not_rec})



        elif self.path == '/api/notlar/edit':

            p = json.loads(body)

            not_id = p.get('id')

            from datetime import datetime as _dt

            for i, n in enumerate(NOTLAR):

                if n.get('id') == not_id:

                    NOTLAR[i]['baslik']     = p.get('baslik', n.get('baslik',''))

                    NOTLAR[i]['metin']      = p.get('metin',  n.get('metin',''))

                    NOTLAR[i]['renk']       = p.get('renk',   n.get('renk','default'))

                    NOTLAR[i]['guncellendi']= _dt.now().strftime('%d.%m.%Y %H:%M')

                    break

            save_notlar()

            self.send_json({'ok': True})



        elif self.path == '/api/notlar/delete':

            p = json.loads(body); not_id = p.get('id')

            NOTLAR = [n for n in NOTLAR if n.get('id') != not_id]

            save_notlar()

            self.send_json({'ok': True})



        elif self.path == '/api/ai/config/save':

            AI_CONFIG = json.loads(body)

            save_ai_config()

            self.send_json({'ok': True})



        elif self.path == '/api/ai/chat':

            p = json.loads(body)

            messages = p.get('messages', [])

            model    = p.get('model', None)

            result   = ai_chat(messages, model)

            self.send_json(result)



        elif self.path == '/api/gmail/config/save':

            GMAIL_CONFIG = json.loads(body)

            save_gmail_config()

            self.send_json({'ok': True})



        elif self.path == '/api/gmail/send':

            p = json.loads(body)

            result = gmail_mail_gonder(p.get('kime',''), p.get('konu',''), p.get('body',''))

            self.send_json(result)



        elif self.path == '/api/logo/config/save':

            LOGO_CONFIG = json.loads(body)

            save_logo_config()

            self.send_json({'ok': True})



        elif self.path == '/api/banka_api/save':

            API_CONFIG = json.loads(body)

            save_api_config()

            self.send_json({'ok': True, 'count': len(API_CONFIG)})



        elif self.path == '/api/banka/import_ekstre':

            # Excel/CSV ekstre import - banka formatını otomatik tanı

            import tempfile, shutil

            ct = self.headers.get('Content-Type','')

            boundary = None

            for part in ct.split(';'):

                p = part.strip()

                if p.startswith('boundary='): boundary = p[9:].strip(); break

            if not boundary:

                self.send_json({'ok':False,'msg':'Boundary bulunamadı'}); return



            boundary_bytes = ('--'+boundary).encode()

            parts = body.split(boundary_bytes)

            file_data = None; filename = 'ekstre.xlsx'

            for part in parts:

                if b'filename=' in part:

                    header_end = part.find(b'\r\n\r\n')

                    if header_end != -1:

                        file_data = part[header_end+4:].rstrip(b'\r\n--')

                        hdrs = part[:header_end].decode('utf-8', errors='ignore')

                        for seg in hdrs.split(';'):

                            seg = seg.strip()

                            if seg.startswith('filename='): filename = seg[9:].strip('"')

                        break

            if not file_data:

                self.send_json({'ok':False,'msg':'Dosya verisi alınamadı'}); return



            ext = os.path.splitext(filename)[1].lower()

            tmp_fd, tmp_path = tempfile.mkstemp(suffix=ext or '.xlsx')

            os.close(tmp_fd)

            try:

                with open(tmp_path,'wb') as f: f.write(file_data)

                result = parse_banka_ekstre(tmp_path, ext)

                self.send_json(result)

            except Exception as e:

                traceback.print_exc()

                self.send_json({'ok':False,'msg':str(e)})

            finally:

                if os.path.exists(tmp_path): os.unlink(tmp_path)



        elif self.path == '/api/banka/import_excel':

            # Banka bakiyeleri Excel import

            import tempfile

            ct = self.headers.get('Content-Type','')

            boundary = None

            for part in ct.split(';'):

                p = part.strip()

                if p.startswith('boundary='): boundary = p[9:].strip(); break

            if not boundary:

                self.send_json({'ok':False,'msg':'Boundary bulunamadı'}); return



            boundary_bytes = ('--'+boundary).encode()

            parts = body.split(boundary_bytes)

            file_data = None; filename = 'banka.xlsx'

            for part in parts:

                if b'filename=' in part:

                    header_end = part.find(b'\r\n\r\n')

                    if header_end != -1:

                        file_data = part[header_end+4:].rstrip(b'\r\n--')

                        hdrs = part[:header_end].decode('utf-8', errors='ignore')

                        for seg in hdrs.split(';'):

                            seg = seg.strip()

                            if seg.startswith('filename='): filename = seg[9:].strip('"')

                        break

            if not file_data:

                self.send_json({'ok':False,'msg':'Dosya verisi alınamadı'}); return



            ext = os.path.splitext(filename)[1].lower()

            tmp_fd, tmp_path = tempfile.mkstemp(suffix=ext or '.xlsx')

            os.close(tmp_fd)

            try:

                with open(tmp_path,'wb') as f: f.write(file_data)

                import warnings; warnings.filterwarnings('ignore')



                if ext == '.csv':

                    df = None

                    for enc in ['utf-8-sig','utf-8','cp1254','latin-1']:

                        try:

                            df = pd.read_csv(tmp_path, encoding=enc, dtype=str)

                            break

                        except: pass

                else:

                    # Header satırını bul

                    df_raw = pd.read_excel(tmp_path, header=None, dtype=str)

                    header_row = 0

                    for i in range(min(10, len(df_raw))):

                        row = ' '.join(df_raw.iloc[i].fillna('').astype(str)).upper()

                        if any(k in row for k in ['BANKA','BAKIYE','FİRMA','FIRMA']):

                            header_row = i; break

                    df = pd.read_excel(tmp_path, header=header_row, dtype=str)



                if df is None or df.empty:

                    self.send_json({'ok':False,'msg':'Dosya okunamadı'}); return



                df.columns = [str(c).strip().upper()

                              .replace('İ','I').replace('Ş','S')

                              .replace('Ğ','G').replace('Ü','U')

                              .replace('Ö','O').replace('Ç','C')

                              for c in df.columns]



                # Sütun eşleştir

                def find_col(aliases):

                    for col in df.columns:

                        for a in aliases:

                            if a in col: return col

                    return None



                col_banka  = find_col(['BANKA','BANK'])

                col_firma  = find_col(['FIRMA','GRUP','SIRKET'])

                col_bakiye = find_col(['BAKIYE','BALANCE','TUTAR','AMOUNT'])

                col_hesap  = find_col(['HESAP','TUR','TYPE','ACCOUNT'])

                col_tarih  = find_col(['TARIH','DATE','GUN'])



                if not col_banka or not col_bakiye:

                    self.send_json({'ok':False,

                        'msg':f'Banka veya Bakiye sütunu bulunamadı. Mevcut sütunlar: {list(df.columns)}'})

                    return



                records = []

                for _, row in df.iterrows():

                    banka = str(row.get(col_banka,'')).strip()

                    if not banka or banka == 'nan': continue

                    try:

                        bakiye_str = str(row.get(col_bakiye,0)).replace('.','').replace(',','.').strip()

                        bakiye = float(bakiye_str) if bakiye_str and bakiye_str != 'nan' else 0

                    except: bakiye = 0



                    firma = str(row.get(col_firma,'ULUSAL')).strip() if col_firma else 'ULUSAL'

                    if firma == 'nan': firma = 'ULUSAL'

                    hesap = str(row.get(col_hesap,'Vadesiz TL')).strip() if col_hesap else 'Vadesiz TL'

                    if hesap == 'nan': hesap = 'Vadesiz TL'

                    tarih = ''

                    if col_tarih:

                        tv = str(row.get(col_tarih,'')).strip()

                        if tv and tv != 'nan': tarih = tv[:10]

                    if not tarih:

                        from datetime import datetime

                        tarih = datetime.now().strftime('%Y-%m-%d')



                    records.append({

                        'banka': banka, 'firma': firma,

                        'hesap_turu': hesap, 'bakiye': round(bakiye,2),

                        'tarih': tarih, 'kaynak': 'excel_import'

                    })



                self.send_json({'ok':True,'count':len(records),'records':records})

            except Exception as e:

                traceback.print_exc()

                self.send_json({'ok':False,'msg':str(e)})

            finally:

                if os.path.exists(tmp_path): os.unlink(tmp_path)



        elif self.path == '/api/banka/save':

            BANKA_DATA = json.loads(body)

            save_banka()

            self.send_json({'ok':True,'count':len(BANKA_DATA)})



        elif self.path == '/api/kredi/save':

            KREDI_DATA = json.loads(body)

            save_kredi()

            self.send_json({'ok':True,'count':len(KREDI_DATA)})



        elif self.path == '/api/kredi/import_excel':

            # Kaynak Excel'den ROTATİF-KMH-VİNOV sayfasını oku

            if not SOURCE_FILE or not os.path.exists(SOURCE_FILE):

                self.send_json({'ok':False,'msg':'Kaynak dosya tanımlanmamış'}); return

            try:

                import warnings, shutil, tempfile

                warnings.filterwarnings('ignore')

                ext = os.path.splitext(SOURCE_FILE)[1]

                tmp_fd, tmp_path = tempfile.mkstemp(suffix=ext)

                os.close(tmp_fd)

                shutil.copy2(SOURCE_FILE, tmp_path)

                df = pd.read_excel(tmp_path, sheet_name='ROTATİF-KMH-VİNOV', header=1)

                df.columns = df.columns.astype(str).str.strip()

                os.unlink(tmp_path)

                records = []

                for _, row in df.iterrows():

                    firma = str(row.get('FİRMA','')).strip()

                    banka = str(row.get('BANKA','')).strip()

                    if not firma or firma == 'nan': continue

                    try: limit = float(row.get('LİMİT') or 0)

                    except: limit = 0

                    try: kullanilan = float(row.get('KULLANILAN ANAPARA') or 0)

                    except: kullanilan = 0

                    try: kalan = float(row.get('KALAN LİMİT') or 0)

                    except: kalan = limit - kullanilan

                    try: faiz_y = float(row.get('YILLIK FAİZ ORANI') or 0)

                    except: faiz_y = 0

                    try: faiz_a = float(row.get('AYLIK  FAİZ ORANI') or 0)

                    except: faiz_a = 0

                    tur = str(row.get('KREDİ TÜRÜ','')).strip()

                    if not tur or tur == 'nan': tur = 'DİĞER'

                    bitis = ''

                    try:

                        bv = row.get('BİTİŞ TARİHİ')

                        if pd.notna(bv):

                            bitis = pd.Timestamp(bv).strftime('%d.%m.%Y')

                    except: pass

                    if limit <= 0 and kalan <= 0: continue

                    records.append({

                        'id': f"{firma}_{banka}_{tur}_{len(records)}",

                        'firma': firma, 'banka': banka, 'tur': tur,

                        'ad': f"{banka} - {tur}",

                        'limit': round(limit,2), 'kullanilan': round(kullanilan,2),

                        'kalan_limit': round(kalan,2),

                        'faiz_yillik': round(faiz_y,4), 'faiz_aylik': round(faiz_a,4),

                        'bitis': bitis, 'aktif': True,

                    })

                KREDI_DATA = records

                save_kredi()

                self.send_json({'ok':True,'count':len(records),'records':records})

            except Exception as e:

                traceback.print_exc()

                self.send_json({'ok':False,'msg':str(e)})



        elif self.path == '/api/gelir/add':

            rec = json.loads(body)

            GELIR_DATA.append(rec); save_gelir()

            self.send_json({'ok':True,'count':len(GELIR_DATA)})



        elif self.path == '/api/gelir/edit':

            p = json.loads(body); idx = p.get('idx'); rec = p.get('record')

            if 0 <= idx < len(GELIR_DATA):

                GELIR_DATA[idx] = rec; save_gelir()

                self.send_json({'ok':True})

            else:

                self.send_json({'ok':False},400)



        elif self.path == '/api/gelir/delete':

            p = json.loads(body); idx = p.get('idx')

            if 0 <= idx < len(GELIR_DATA):

                GELIR_DATA.pop(idx); save_gelir()

                self.send_json({'ok':True})

            else:

                self.send_json({'ok':False},400)



        elif self.path == '/api/gelir/toggle':

            p = json.loads(body); idx = p.get('idx')

            if 0 <= idx < len(GELIR_DATA):

                cur = GELIR_DATA[idx].get('durum','')

                GELIR_DATA[idx]['durum'] = 'TAHSİL EDİLDİ' if cur != 'TAHSİL EDİLDİ' else 'BEKLENİYOR'

                save_gelir()

                self.send_json({'ok':True,'durum':GELIR_DATA[idx]['durum']})

            else:

                self.send_json({'ok':False},400)



        elif self.path == '/api/banka/save':

            new_list = json.loads(body)

            BANKA_DATA.clear(); BANKA_DATA.extend(new_list)

            save_banka()

            self.send_json({'ok': True, 'count': len(BANKA_DATA)})



        elif self.path == '/api/kredi/save':

            new_list = json.loads(body)

            KREDI_DATA.clear(); KREDI_DATA.extend(new_list)

            save_kredi()

            self.send_json({'ok': True, 'count': len(KREDI_DATA)})



        elif self.path == '/api/import/confirm':

            p = json.loads(body)

            records = p.get('records',[])

            if p.get('merge'): DATA.extend(records)

            else: DATA.clear(); DATA.extend(records)

            save_data()

            self.send_json({'ok':True,'total':len(DATA)})



        else:

            self.send_response(404); self.end_headers()





# ── HTML FRONTEND ──────────────────────────────────────────────────────────────

def get_html(initial_page=""):

    return r'''<!DOCTYPE html>

<html lang="tr">

<head>

<meta charset="UTF-8">

<meta name="viewport" content="width=device-width,initial-scale=1">

<title>Nakit Akış v25</title>

<style>

html,body{height:100%;margin:0;padding:0;overflow:hidden;display:flex;flex-direction:column;background:var(--bg);color:var(--t1);font-family:var(--sans);font-size:14px}body{display:flex;flex-direction:column;background:var(--bg);color:var(--t1);font-family:var(--sans);font-size:14px}:root{--bg:#0d1117;--s1:#161c26;--s2:#1c2333;--s3:#21293a;--s4:#283042;--b1:#2d3748;--b2:#3d4f6b;--acc:#4f9cf9;--acc-d:#2563eb;--acc2:#818cf8;--g:#34d399;--g-d:#059669;--r:#f87171;--r-d:#dc2626;--am:#fbbf24;--am-d:#d97706;--t1:#e2e8f0;--t2:#94a3b8;--t3:#4a5568;--mono:Consolas,'Courier New',monospace;--sans:'Segoe UI',system-ui,sans-serif;--rad:6px;--radl:10px}

*{box-sizing:border-box;margin:0;padding:0}





.hdr{background:var(--s1);border-bottom:1px solid var(--b1);height:50px;display:flex;align-items:center;padding:0 18px;gap:12px;flex-shrink:0;z-index:50}

.logo{font-family:var(--mono);font-size:13px;font-weight:700;color:var(--acc);letter-spacing:.5px}

.logo em{color:var(--t3);font-style:normal;font-weight:400}

.hdr-sep{width:1px;height:20px;background:var(--b1)}

.hdr-src{font-size:11px;font-family:var(--mono);display:flex;align-items:center;gap:6px}

.src-dot{width:7px;height:7px;border-radius:50%;flex-shrink:0}

.src-dot.ok{background:var(--g)}.src-dot.err{background:var(--r)}.src-dot.none{background:var(--t3)}

.hdr-sp{flex:1}

.lay{display:flex;flex:1;min-height:0;overflow:hidden}

.sbar{width:200px;flex-shrink:0;background:var(--s1);border-right:1px solid var(--b1);display:flex;flex-direction:column;padding:10px 0;overflow-y:auto}

.nav-sec{font-size:9px;color:var(--t3);padding:12px 16px 4px;letter-spacing:.1em;text-transform:uppercase;font-weight:600;font-family:var(--mono)}

.nav{display:flex;align-items:center;gap:9px;padding:7px 16px;cursor:pointer;color:var(--t2);font-size:12.5px;transition:all .12s;border-left:2px solid transparent;user-select:none}

.nav:hover{color:var(--t1);background:var(--s2)}.nav.on{color:var(--acc);background:var(--s2);border-left-color:var(--acc)}

.main{flex:1;overflow-y:auto;padding:20px}

.view{display:none;animation:fadein .15s ease;width:100%}.view.on{display:block;width:100%;min-height:50px}

@keyframes fadein{from{opacity:0;transform:translateY(4px)}to{opacity:1;transform:none}}

.ptit{font-size:18px;font-weight:600;margin-bottom:2px}.psub{font-size:12px;color:var(--t2);margin-bottom:16px}

.mets{display:grid;grid-template-columns:repeat(4,1fr);gap:10px;margin-bottom:16px}

.met{background:var(--s1);border:1px solid var(--b1);border-radius:var(--radl);padding:14px 15px;position:relative;overflow:hidden}

.met-stripe{position:absolute;top:0;left:0;right:0;height:2px}

.mlb{font-size:10px;color:var(--t2);font-family:var(--mono);text-transform:uppercase;letter-spacing:.06em;margin-bottom:8px}

.mvl{font-family:var(--mono);font-size:19px;font-weight:600;line-height:1}

.msb{font-size:10px;color:var(--t3);margin-top:4px;font-family:var(--mono)}

.btn{display:inline-flex;align-items:center;gap:5px;padding:6px 12px;border-radius:var(--rad);font-size:12.5px;cursor:pointer;transition:all .12s;border:1px solid;font-weight:500;font-family:var(--sans);white-space:nowrap}

.bp{background:var(--acc);border-color:var(--acc);color:#fff}.bp:hover{background:var(--acc-d)}

.bo{background:transparent;border-color:var(--b2);color:var(--t2)}.bo:hover{background:var(--s3);color:var(--t1)}

.bg2{background:var(--g-d);border-color:var(--g-d);color:#fff}.bg2:hover{filter:brightness(1.1)}

.bgh{background:transparent;border-color:transparent;color:var(--t3);padding:4px 7px}.bgh:hover{background:var(--s3);color:var(--t1)}

.bam2{background:rgba(245,158,11,.15);border-color:var(--am-d);color:var(--am)}.bam2:hover{background:rgba(245,158,11,.25)}

.bsm{padding:4px 9px;font-size:12px}

.tbr{display:flex;align-items:center;gap:7px;margin-bottom:12px;flex-wrap:wrap}

.si{background:var(--s2);border:1px solid var(--b1);border-radius:var(--rad);padding:6px 10px;color:var(--t1);font-size:12.5px;width:190px;outline:none}

.si:focus{border-color:var(--acc)}.si::placeholder{color:var(--t3)}

.fi{background:var(--s2);border:1px solid var(--b1);border-radius:var(--rad);padding:6px 8px;color:var(--t1);font-size:12.5px;outline:none;cursor:pointer}

.fi:focus{border-color:var(--acc)}

.sp{flex:1}

.tc{background:var(--s1);border:1px solid var(--b1);border-radius:var(--radl);overflow:hidden}

table{width:100%;border-collapse:collapse;font-size:12.5px}

th{background:var(--s2);padding:8px 10px;text-align:left;font-family:var(--mono);font-size:9.5px;color:var(--t3);text-transform:uppercase;letter-spacing:.06em;border-bottom:1px solid var(--b1);cursor:pointer;user-select:none;white-space:nowrap}

th:hover{color:var(--t2)}

td{padding:7px 10px;border-bottom:1px solid var(--b1);vertical-align:middle}

tr:last-child td{border-bottom:none}

tr:hover td{background:rgba(255,255,255,.02)}

.num{text-align:right;font-family:var(--mono)}

.clip{overflow:hidden;text-overflow:ellipsis;white-space:nowrap}

.bdg{display:inline-block;padding:2px 7px;border-radius:99px;font-size:10.5px;font-weight:600;font-family:var(--mono)}

.bdg.cst{cursor:pointer;transition:opacity .12s}.bdg.cst:hover{opacity:.75}

.bgi{background:rgba(52,211,153,.12);color:var(--g);border:1px solid rgba(52,211,153,.25)}

.bri{background:rgba(248,113,113,.12);color:var(--r);border:1px solid rgba(248,113,113,.25)}

.bbi{background:rgba(79,156,249,.12);color:var(--acc);border:1px solid rgba(79,156,249,.25)}

.bpui{background:rgba(129,140,248,.12);color:var(--acc2);border:1px solid rgba(129,140,248,.25)}

.bami{background:rgba(251,191,36,.12);color:var(--am);border:1px solid rgba(251,191,36,.25)}

.bcyi{background:rgba(34,211,238,.12);color:#22d3ee;border:1px solid rgba(34,211,238,.25)}

.bami2{background:rgba(245,158,11,.12);color:var(--am);border:1px solid rgba(245,158,11,.25)}

.pgn{display:flex;align-items:center;gap:5px;padding:9px 13px;border-top:1px solid var(--b1);background:var(--s2)}

.pgi{font-size:10.5px;color:var(--t3);font-family:var(--mono)}.pgsp{flex:1}

.pbn{width:25px;height:25px;border-radius:5px;border:1px solid var(--b1);background:transparent;color:var(--t2);cursor:pointer;font-size:11px;display:flex;align-items:center;justify-content:center}

.pbn:hover{background:var(--s4);color:var(--t1)}.pbn.on{background:var(--acc);border-color:var(--acc);color:#fff}

.ov{display:none;position:fixed;inset:0;background:rgba(0,0,0,.7);z-index:300;align-items:center;justify-content:center;backdrop-filter:blur(4px)}

.ov.on{display:flex}

.mod{background:var(--s1);border:1px solid var(--b2);border-radius:14px;padding:22px;width:500px;max-width:96vw;max-height:92vh;overflow-y:auto;animation:mi .18s ease}

@keyframes mi{from{opacity:0;transform:scale(.96)translateY(8px)}to{opacity:1;transform:none}}

.mtit{font-size:15px;font-weight:600;margin-bottom:18px;display:flex;align-items:center;justify-content:space-between}

.mclose{background:transparent;border:none;color:var(--t3);cursor:pointer;font-size:18px;line-height:1;padding:2px 6px;border-radius:4px}.mclose:hover{color:var(--t1);background:var(--s3)}

.fg{display:grid;grid-template-columns:1fr 1fr;gap:10px}

.ff{grid-column:1/-1}

.fgr{display:flex;flex-direction:column;gap:4px}

.flb{font-size:9.5px;color:var(--t2);font-family:var(--mono);text-transform:uppercase;letter-spacing:.06em}

.fi2,.fse{background:var(--s2);border:1px solid var(--b1);border-radius:var(--rad);padding:7px 9px;color:var(--t1);font-size:12.5px;width:100%;outline:none;font-family:var(--sans);transition:border-color .12s}

.fi2:focus,.fse:focus{border-color:var(--acc)}

.mac{display:flex;gap:7px;justify-content:flex-end;margin-top:18px;padding-top:14px;border-top:1px solid var(--b1)}

.cgrd{display:grid;grid-template-columns:1fr 1fr;gap:12px;margin-bottom:16px}

.cc{background:var(--s1);border:1px solid var(--b1);border-radius:var(--radl);padding:15px}

.cct{font-family:var(--mono);font-size:9.5px;color:var(--t3);text-transform:uppercase;letter-spacing:.06em;margin-bottom:12px}

.cwrap{position:relative;height:200px}

.cload{text-align:center;color:var(--t3);padding:40px 0;font-size:12px}

.mgrd{display:grid;grid-template-columns:repeat(4,1fr);gap:9px;margin-bottom:16px}

.mca{background:var(--s1);border:1px solid var(--b1);border-radius:var(--rad);padding:11px;cursor:pointer;transition:all .12s}

.mca:hover,.mca.sel{border-color:var(--acc);background:var(--s2)}

.mcn{font-family:var(--mono);font-size:9.5px;color:var(--t2);text-transform:uppercase;letter-spacing:.05em;margin-bottom:4px}

.mct{font-family:var(--mono);font-size:13px;font-weight:600}

.mcp{font-size:10px;color:var(--r);margin-top:2px;font-family:var(--mono)}

.egrd{display:grid;grid-template-columns:1fr 1fr;gap:10px}

.ec{background:var(--s1);border:1px solid var(--b1);border-radius:var(--radl);padding:16px}

.eico{font-size:22px;margin-bottom:8px}.etit{font-size:13.5px;font-weight:500;margin-bottom:5px}

.edsc{font-size:11.5px;color:var(--t2);margin-bottom:12px;line-height:1.55}

.ebr{display:flex;gap:6px;flex-wrap:wrap}

.tc2{position:fixed;bottom:18px;right:18px;z-index:999;display:flex;flex-direction:column;gap:6px;pointer-events:none}

.tst{background:var(--s2);border:1px solid var(--b2);border-radius:var(--rad);padding:9px 13px;font-size:12.5px;display:flex;align-items:center;gap:8px;animation:tin .2s ease;box-shadow:0 8px 28px rgba(0,0,0,.5);pointer-events:auto}

@keyframes tin{from{opacity:0;transform:translateX(12px)}to{opacity:1;transform:none}}

.tst.s{border-left:3px solid var(--g)}.tst.e{border-left:3px solid var(--r)}.tst.i{border-left:3px solid var(--acc)}

hr.dv{border:none;border-top:1px solid var(--b1);margin:14px 0}

::-webkit-scrollbar{width:5px;height:5px}

::-webkit-scrollbar-track{background:var(--s1)}

::-webkit-scrollbar-thumb{background:var(--b2);border-radius:3px}

/* AYARLAR */

.src-card{background:var(--s2);border:1px solid var(--b1);border-radius:var(--radl);padding:18px;margin-bottom:14px}

.src-card h3{font-size:13px;font-weight:600;margin-bottom:14px;color:var(--t1)}

.src-status-box{background:var(--s3);border:1px solid var(--b1);border-radius:var(--rad);padding:12px 14px;font-family:var(--mono);font-size:11.5px;margin-bottom:12px}

.src-row{display:flex;align-items:center;justify-content:space-between;margin-bottom:6px;font-size:12px}

.src-row:last-child{margin-bottom:0}

.src-key{color:var(--t2)}.src-val{font-family:var(--mono);font-size:11.5px}

.tgl{display:flex;align-items:center;gap:8px;font-size:12px;color:var(--t2);cursor:pointer;user-select:none}

.tgl input[type=checkbox]{width:14px;height:14px;cursor:pointer;accent-color:var(--acc)}

.info-box{background:rgba(79,156,249,.08);border:1px solid rgba(79,156,249,.2);border-radius:var(--rad);padding:10px 12px;font-size:11.5px;color:var(--t2);line-height:1.6;margin-bottom:12px}



.typing-dot{width:7px;height:7px;border-radius:50%;background:var(--t3);animation:typing-bounce 1.2s infinite ease-in-out}

.typing-dot:nth-child(1){animation-delay:0s}

.typing-dot:nth-child(2){animation-delay:.2s}

.typing-dot:nth-child(3){animation-delay:.4s}

@keyframes typing-bounce{0%,80%,100%{transform:scale(.8);opacity:.5}40%{transform:scale(1.1);opacity:1}}

</style>

<script src="https://cdnjs.cloudflare.com/ajax/libs/Chart.js/4.4.0/chart.umd.min.js" crossorigin="anonymous"></script>

</head>

<body>

<div class="hdr">

  <div class="logo">NAKİT<em>/</em>AKİŞ <em style="font-size:10px">v3</em></div>

  <div class="hdr-sep"></div>

  <div class="hdr-src" id="hdr-src">

    <div class="src-dot none" id="src-dot"></div>

    <span id="src-label">Kaynak dosya tanımlanmadı</span>

  </div>

  <div class="hdr-sp"></div>

  <button class="btn bo bsm" id="sync-btn" onclick="manualSync()" style="display:none">↻ Şimdi Senkronize Et</button>

  <button class="btn bp bsm" onclick="openAdd()">+ Yeni Kayıt</button>

</div>



<div class="lay">

<nav class="sbar">

  <div class="nav-sec">Görünüm</div>

  <div class="nav on" id="nv-db"     onclick="showView('db')">◈ DASHBOARD</div>

  <div class="nav"    id="nv-banka-kart" onclick="showView('banka-kart')">🏦 BANKA KARTLARI</div>

  <div class="nav"    id="nv-tbl"    onclick="showView('tbl')">≡ GİDER TABLOSU</div>

  <div class="nav"    id="nv-gelir"  onclick="showView('gelir')">＋ GELİR TABLOSU</div>

  <div class="nav"    id="nv-nakit"  onclick="showView('nakit')">⇌ NAKİT AKIŞ DENGESİ</div>

  <div class="nav"    id="nv-finans" onclick="showView('finans')">⬡ FİNANSMAN YÖNETİMİ</div>

  <div class="nav"    id="nv-liki"   onclick="showView('liki')">◎ LİKİDİTE & FİNANSMAN</div>

  <div class="nav"    id="nv-cari"   onclick="showView('cari')">⇄ CARİ HESAPLAR</div>

  <div class="nav"    id="nv-banka-api" onclick="showView('banka-api')">🔌 BANKA API & EKSTRE</div>

  <div class="nav"    id="nv-kur"      onclick="showView('kur')">💱 DÖVİZ KURLARI</div>

  <div class="nav"    id="nv-logo"     onclick="showView('logo')">🧾 FATURALAR (LOGO)</div>

  <div class="nav"    id="nv-satis"    onclick="showView('satis')">📈 SATIŞ ANALİZİ</div>

  <div class="nav"    id="nv-stok"     onclick="showView('stok')">📦 STOK LİSTESİ (LOGO)</div>

  <div class="nav"    id="nv-cari-logo" onclick="showView('cari-logo')">👥 CARİ LİSTESİ (LOGO)</div>

  <div class="nav"    id="nv-gmail"    onclick="showView('gmail')">📧 GMAİL</div>

  <div class="nav"    id="nv-docs"   onclick="showView('docs')">📁 DÖKÜMAN YÖNETİMİ</div>

  <div class="nav"    id="nv-notlar"   onclick="showView('notlar')">📝 NOTLAR</div>

  <div class="nav"    id="nv-hat"     onclick="showView('hat')">🔔 HATIRLATICI</div>

  <div class="nav"    id="nv-ai"       onclick="showView('ai')">🤖 AI ASISTAN</div>

  <div class="nav"    id="nv-wa"      onclick="showView('wa')">💬 WHATSAPP</div>

  <div class="nav"    id="nv-otel"    onclick="showView('otel')">🏨 OTEL YÖNETİMİ</div>

  <div class="nav"    id="nv-mo"     onclick="showView('mo')">◷ AYLIK ÖZET</div>

  <div class="nav"    id="nv-ipotek"  onclick="showView('ipotek')">🏠 İPOTEK TAKİBİ</div>

  <div class="nav-sec">Hesap</div>

  <div class="nav"    id="nv-kullanicilar" onclick="showView('kullanicilar')">👤 KULLANICILAR</div>

  <div class="nav"    id="nv-sifre"  onclick="showView('sifre')">🔑 ŞİFRE DEĞİŞTİR</div>

  <div class="nav"    onclick="cikisYap()" style="color:#f87171">🚪 ÇIKIŞ YAP</div>

  <div style="padding:8px 14px;font-size:11px;color:#9ca3af" id="nav-user-label">...</div>

  <div class="nav-sec">Veri</div>

  <div class="nav"    id="nv-src"    onclick="showView('src')">⚙ KAYNAK DOSYA</div>

  <div class="nav"    id="nv-exp"    onclick="showView('exp')">⬇ DIŞA AKTAR</div>

  <div class="nav"    id="nv-import" onclick="showView('import')">⬆ VERİ İÇE AKTAR</div>

  <div class="nav"    id="nv-veritabani" onclick="showView('db')">🗄 VERİTABANI</div>

</nav>



<div class="main">



<!-- DASHBOARD -->

<div class="view on" id="vw-db">

  <div class="ptit">Dashboard</div>

  <div class="psub" id="db-sub">Bağlanıyor...</div>

  <div class="mets">

    <div class="met"><div class="met-stripe" style="background:var(--acc)"></div><div class="mlb">Toplam Ödeme</div><div class="mvl" style="color:var(--acc)" id="st-tot">₺ 0</div><div class="msb" id="st-tc">0 kayıt</div></div>

    <div class="met"><div class="met-stripe" style="background:var(--g)"></div><div class="mlb">Ödendi</div><div class="mvl" style="color:var(--g)" id="st-paid">₺ 0</div><div class="msb" id="st-pc">0 kayıt</div></div>

    <div class="met"><div class="met-stripe" style="background:var(--r)"></div><div class="mlb">Bekleyen</div><div class="mvl" style="color:var(--r)" id="st-pend">₺ 0</div><div class="msb" id="st-bc">0 kayıt</div></div>

    <div class="met"><div class="met-stripe" style="background:var(--am)"></div><div class="mlb">Yaklaşan (30 gün)</div><div class="mvl" style="color:var(--am)" id="st-soon">₺ 0</div><div class="msb" id="st-sc">0 kayıt</div></div>

  </div>

  <div class="cgrd">

    <div class="cc"><div class="cct">Aylık Toplam Ödeme</div><div class="cwrap"><div id="ch-mo"></div></div></div>

    <div class="cc"><div class="cct">Tür Bazlı Dağılım</div><div class="cwrap"><div id="ch-pie"></div></div></div>

    <div class="cc" style="grid-column:1/-1"><div class="cct">Ödendi / Bekleyen — Aylık Karşılaştırma</div><div class="cwrap" style="height:180px"><div id="ch-st" style="height:100%"></div></div></div>

  </div>

</div>



<!-- TABLE -->

<div class="view" id="vw-tbl">

  <div class="ptit">Detay Tablosu</div>

  <div class="psub">Durum sütununa <strong style="color:var(--acc)">çift tıklayarak</strong> ÖDENDİ/ÖDENMEDİ değiştirebilirsiniz — değişiklik Excel kaynak dosyasına da yazılır</div>

  <div class="tbr" style="flex-wrap:wrap;gap:6px">

    <input class="si" type="text" id="fs" placeholder="🔍 Ara..." oninput="doFilter()" style="width:150px">

    <select class="fi" id="fay" onchange="doFilter()"><option value="">Tüm Aylar</option></select>

    <select class="fi" id="ffr" onchange="doFilter()"><option value="">Tüm Firmalar</option></select>

    <select class="fi" id="ftr" onchange="doFilter()"><option value="">Tüm Türler</option></select>

    <select class="fi" id="fyer" onchange="doFilter()"><option value="">Tüm Ödeme Yerleri</option></select>

    <select class="fi" id="fyil" onchange="doFilter()"><option value="">Tüm Yıllar</option></select>

    <select class="fi" id="fdu" onchange="doFilter()"><option value="">Tüm Durumlar</option><option>ÖDENDİ</option><option>ÖDENMEDİ</option></select>

    <div style="display:flex;align-items:center;gap:4px">

      <input class="si" type="date" id="ftarih-bas" style="width:130px" title="Tarih başlangıç" onkeydown="if(event.key==='Enter')doFilter()">

      <span style="color:var(--t3);font-size:11px">—</span>

      <input class="si" type="date" id="ftarih-bit" style="width:130px" title="Tarih bitiş" onkeydown="if(event.key==='Enter')doFilter()">

      <button class="btn bp bsm" onclick="doFilter()" style="padding:4px 10px;font-size:11px;font-weight:600">🔍 Filtrele</button>

      <button class="btn bgh bsm" onclick="giderTarihSifirla()" title="Tarihi temizle" style="padding:4px 7px;font-size:11px">✕</button>

      <button class="btn bgh bsm" onclick="hizliTarih('gider','bu-ay')" style="font-size:10px;padding:3px 6px">Bu Ay</button>

      <button class="btn bgh bsm" onclick="hizliTarih('gider','bu-yil')" style="font-size:10px;padding:3px 6px">Bu Yıl</button>

      <button class="btn bgh bsm" onclick="hizliTarih('gider','son-3')" style="font-size:10px;padding:3px 6px">Son 3 Ay</button>

    </div>

    <div class="sp"></div>

    <button class="btn bo bsm" onclick="doExpExcel('filtered')">⬇ Aktar</button>

    <button class="btn bp bsm" onclick="openAdd()">+ Ekle</button>

  </div>

  <div class="tc">

    <table>

      <thead><tr>

        <th onclick="srt('firma')"   style="width:80px">Firma</th>

        <th onclick="srt('tur')"     style="width:145px">Tür</th>

        <th onclick="srt('yer')"     style="width:150px">Ödeme Yeri</th>

        <th onclick="srt('kredi_no')" style="width:120px">Kredi/Çek No</th>

        <th onclick="srt('tarih')"   style="width:92px">Tarih</th>

        <th onclick="srt('ay')"      style="width:68px">Ay</th>

        <th onclick="srt('durum')"   style="width:98px">Durum</th>

        <th onclick="srt('anapara')" class="num" style="width:112px">Anapara ₺</th>

        <th onclick="srt('faiz')"    class="num" style="width:98px">Faiz ₺</th>

        <th onclick="srt('tutar')"   class="num" style="width:118px">Tutar ₺</th>

        <th style="width:65px"></th>

      </tr></thead>

      <tbody id="tbody"></tbody>

    </table>

    <div class="pgn" id="pgn"></div>

  </div>

</div>



<!-- MONTHLY -->

<div class="view" id="vw-mo">

  <div class="ptit">Aylık Özet</div>

  <div class="psub">Aya tıklayarak detayını görün · Filtreler seçili aya da uygulanır</div>



  <!-- Üst filtreler -->

  <div class="tbr" style="margin-bottom:10px">

    <select class="fi" id="mo-yil" onchange="renderMonthly()"><option value="">Tüm Yıllar</option></select>

    <select class="fi" id="mo-firma" onchange="renderMonthly()"><option value="">Tüm Firmalar</option></select>

    <select class="fi" id="mo-tur" onchange="renderMonthly()"><option value="">Tüm Türler</option></select>

    <select class="fi" id="mo-durum" onchange="renderMonthly()">

      <option value="">Tüm Durumlar</option>

      <option value="ÖDENDİ">Ödendi</option>

      <option value="ÖDENMEDİ">Ödenmedi</option>

    </select>

    <div class="sp"></div>

    <button class="btn bo bsm" onclick="resetMoFilters()">↺ Sıfırla</button>

    <button class="btn bo bsm" onclick="doExpExcelMo()">⬇ Excel</button>

  </div>



  <!-- Özet bar -->

  <div id="mo-summary-bar" style="display:none;background:var(--s2);border:1px solid var(--b1);border-radius:var(--rad);padding:9px 14px;margin-bottom:12px;display:flex;gap:20px;font-family:var(--mono);font-size:11.5px;flex-wrap:wrap"></div>



  <!-- Ay kartları -->

  <div class="mgrd" id="mgrd"></div>



  <!-- Seçili ay detayı -->

  <div id="mdet"></div>

</div>



<div class="view" id="vw-docs">

  <div class="ptit">📁 DÖKÜMAN YÖNETİMİ</div>

  <div class="psub">Tüm belgelerinizi tek yerde saklayın — PDF, resim, Word, Excel ve daha fazlası</div>



  <!-- Üst araç çubuğu -->

  <div style="display:flex;gap:8px;margin-bottom:12px;align-items:center;flex-wrap:wrap">

    <input class="si" id="docs-ara" placeholder="🔍 Belge ara..." style="flex:1;min-width:180px" oninput="docsFiltrele()">

    <select class="fse" id="docs-kat-filtre" onchange="docsFiltrele()" style="min-width:150px;font-size:12px">

      <option value="">Tüm Kategoriler</option>

    </select>

    <label class="btn bp bsm" style="cursor:pointer">

      ⬆ Belge Yükle

      <input type="file" id="docs-file-input" multiple style="display:none" onchange="docsYukle(this.files)">

    </label>

    <button class="btn bgh bsm" onclick="docsYeniKlasorModal()">📂 Kategori Ekle</button>

  </div>



  <!-- İstatistik kartları -->

  <div id="docs-stats" style="display:flex;gap:10px;margin-bottom:12px;flex-wrap:wrap"></div>



  <!-- Sürükle bırak + yükleme durumu yan yana -->

  <div style="display:grid;grid-template-columns:1fr 1fr;gap:10px;margin-bottom:14px">

    <div id="docs-drop-zone"

      style="border:2px dashed var(--b2);border-radius:var(--radl);padding:18px;text-align:center;color:var(--t3);font-size:13px;cursor:pointer;transition:all .2s"

      ondragover="event.preventDefault();this.style.borderColor='var(--acc)';this.style.color='var(--acc)';this.style.background='rgba(79,156,249,.05)'"

      ondragleave="this.style.borderColor='var(--b2)';this.style.color='var(--t3)';this.style.background=''"

      ondrop="event.preventDefault();this.style.borderColor='var(--b2)';this.style.color='var(--t3)';this.style.background='';docsYukle(event.dataTransfer.files)"

      onclick="document.getElementById('docs-file-input').click()">

      <div style="font-size:28px;margin-bottom:6px">📂</div>

      <div>Dosyaları buraya sürükleyin<br>veya tıklayarak seçin</div>

      <div style="font-size:10px;margin-top:4px;color:var(--t3)">Tüm dosya türleri desteklenir</div>

    </div>

    <div id="docs-yukle-durum" style="background:var(--s2);border:1px solid var(--b1);border-radius:var(--radl);padding:16px;display:flex;flex-direction:column;justify-content:center">

      <div style="font-size:12px;font-weight:600;color:var(--t2);margin-bottom:8px">📤 Yükleme Durumu</div>

      <div id="docs-progress-text" style="font-size:12px;color:var(--t3)">Bekliyor...</div>

      <div style="height:6px;background:var(--b1);border-radius:3px;margin-top:8px;overflow:hidden">

        <div id="docs-progress-bar" style="height:100%;background:var(--acc);border-radius:3px;width:0%;transition:width .3s"></div>

      </div>

      <div id="docs-yukle-log" style="margin-top:8px;font-size:11px;color:var(--t3);max-height:80px;overflow-y:auto"></div>

    </div>

  </div>



  <!-- Belge grid -->

  <div id="docs-grid" style="display:grid;grid-template-columns:repeat(auto-fill,minmax(200px,1fr));gap:10px"></div>

  <div id="docs-bos" style="display:none;text-align:center;padding:50px;color:var(--t3)">

    <div style="font-size:48px;margin-bottom:12px">📭</div>

    <div style="font-size:14px;font-weight:600">Henüz belge yok</div>

    <div style="font-size:12px;margin-top:6px">Yukarıdan dosya yükleyin veya sürükleyip bırakın</div>

  </div>

</div>



<!-- DOKÜMAN ÖNİZLEME MODAL -->

<div id="docs-onizle-mod" style="display:none;position:fixed;inset:0;background:rgba(0,0,0,.88);z-index:9999">

  <div style="display:flex;align-items:center;gap:8px;padding:12px 16px;background:var(--s1);border-bottom:1px solid var(--b1);flex-shrink:0">

    <span id="docs-mod-baslik" style="font-size:14px;font-weight:600;color:var(--t1);flex:1;overflow:hidden;text-overflow:ellipsis;white-space:nowrap"></span>

    <button class="btn bgh bsm" id="docs-mod-indir">⬇ İndir</button>

    <button class="btn bgh bsm" id="docs-mod-wa" style="color:#25d366;border-color:#25d366">💬 WA</button>

    <button class="btn bgh bsm" id="docs-mod-mail" style="color:var(--acc)">📧 Mail</button>

    <button class="btn bgh bsm" onclick="docsOnizleKapat()" style="color:var(--r)">✕ Kapat</button>

  </div>

  <div id="docs-mod-icerik" style="flex:1;overflow:auto;padding:20px;display:flex;justify-content:center;align-items:flex-start"></div>

</div>



<!-- DOKÜMAN DÜZENLE MODAL -->

<div id="docs-edit-mod" class="mod-bg" style="display:none">

  <div class="mod-box" style="max-width:460px">

    <div class="mod-hdr">

      <span style="font-size:14px;font-weight:600">✏ Belge Düzenle</span>

      <button class="btn bgh" onclick="docsEditKapat()">✕</button>

    </div>

    <div class="fg" style="padding:16px;gap:10px">

      <div class="fgr ff">

        <label class="flb">Belge Adı</label>

        <input class="fi2" id="docs-edit-ad" placeholder="Belge adı...">

      </div>

      <div class="fgr ff">

        <label class="flb">Kategori</label>

        <div style="display:flex;gap:6px">

          <select class="fi" id="docs-edit-kat" style="flex:1">

            <option>Genel</option><option>Fatura</option><option>Sözleşme</option>

            <option>Ekstre</option><option>Rapor</option><option>Diğer</option>

          </select>

          <input class="fi2" id="docs-edit-kat-yeni" placeholder="veya yeni..." style="flex:1">

        </div>

      </div>

      <div class="fgr ff">

        <label class="flb">Not</label>

        <textarea class="fi2" id="docs-edit-not" rows="3" placeholder="Bu belge hakkında not..." style="resize:vertical"></textarea>

      </div>

    </div>

    <div style="padding:0 16px 16px;display:flex;gap:8px;justify-content:flex-end">

      <button class="btn bgh" onclick="docsEditKapat()">İptal</button>

      <button class="btn bp" onclick="docsEditKaydet()">✓ Kaydet</button>

    </div>

  </div>

</div>



<!-- KAYNAK DOSYA AYARLARI -->

<div class="view" id="vw-src">

  <div class="ptit">Kaynak Dosya Ayarları</div>

  <div class="psub">Uygulama bu Excel dosyasını otomatik olarak okuyacak ve değişiklikleri yansıtacak</div>



  <div class="info-box">

    💡 <strong>Nasıl çalışır?</strong><br>

    Bir Excel dosyası tanımladığınızda, uygulama her açılışta ve 30 saniyede bir o dosyayı kontrol eder.

    Dosya değişmişse veriler otomatik güncellenir. Durum değişiklikleriniz (Ödendi/Ödenmedi)

    JSON kayıt dosyasında tutulur, kaynak dosyanız değiştirilmez.

  </div>



  <div class="src-card">

    <h3>Mevcut Durum</h3>

    <div class="src-status-box">

      <div class="src-row"><span class="src-key">Dosya</span><span class="src-val" id="ss-file" style="color:var(--t1)">—</span></div>

      <div class="src-row"><span class="src-key">Sayfa</span><span class="src-val" id="ss-sheet">Otomatik</span></div>

      <div class="src-row"><span class="src-key">Son Senkronizasyon</span><span class="src-val" id="ss-msg" style="color:var(--t2)">—</span></div>

      <div class="src-row"><span class="src-key">Otomatik Yenileme</span><span class="src-val" id="ss-auto">—</span></div>

    </div>

    <button class="btn bg2 bsm" onclick="manualSync()" id="ss-sync-btn">↻ Şimdi Oku</button>

  </div>



  <div class="src-card">

    <h3>Dosya Tanımla</h3>

    <div class="fg">

      <div class="fgr ff">

        <label class="flb">Excel Dosya Yolu</label>

        <div style="display:flex;gap:7px">

          <input class="fi2" type="text" id="src-path"

            placeholder="Dosya seç butonuna tıklayın..." style="flex:1">

          <button class="btn bam2" onclick="browseFile()" id="browse-btn"

            title="Dosya seçici aç" style="flex-shrink:0;white-space:nowrap">

            📂 Dosya Seç

          </button>

        </div>

      </div>

      <div class="fgr">

        <label class="flb">Sayfa Adı (boş = otomatik)</label>

        <input class="fi2" type="text" id="src-sheet"

          placeholder="Boş bırakın — sistem otomatik bulur">

      </div>

      <div class="fgr" style="justify-content:flex-end">

        <label class="flb">&nbsp;</label>

        <label class="tgl">

          <input type="checkbox" id="src-auto" checked>

          30 sn'de bir otomatik yenile

        </label>

      </div>

    </div>

    <div style="display:flex;gap:8px;margin-top:14px;flex-wrap:wrap">

      <button class="btn bp" onclick="saveSource()">✓ Kaydet ve Yükle</button>

      <button class="btn bo" onclick="clearSource()">✕ Kaynağı Kaldır</button>

    </div>

    <div id="src-err" style="font-size:12px;color:var(--r);margin-top:8px;display:none"></div>

  </div>



  <div class="src-card" style="background:rgba(245,158,11,.05);border-color:rgba(245,158,11,.2)">

    <h3 style="color:var(--am)">📁 Dosya Yolunu Nasıl Bulursunuz?</h3>

    <div style="font-size:12px;color:var(--t2);line-height:1.8">

      <b style="color:var(--t1)">Windows:</b><br>

      Excel dosyasına sağ tıklayın → <em>Özellikler</em> → "Konum" alanındaki yol + dosya adı<br>

      Veya: Dosya Gezgini'nde dosyayı seçin → Adres çubuğuna tıklayın → yolu kopyalayın<br><br>

      <b style="color:var(--t1)">Örnek yol:</b>

      <code style="background:var(--s3);padding:2px 6px;border-radius:4px;font-family:var(--mono);font-size:11px">

        C:\Users\MUH2015\Documents\BANKA TABLOLARI\nakit_akis.xlsx

      </code>

    </div>

  </div>

</div>



<!-- GELİR TABLOSU -->

<div class="view" id="vw-gelir">

  <div class="ptit">Gelir Tablosu</div>

  <div class="psub">Gelir tahsilatlarını buradan yönetin · Durum badge'ine çift tıklayarak tahsilat durumunu güncelleyin</div>

  <div class="tbr" style="flex-wrap:wrap;gap:6px">

    <input class="si" type="text" id="gfs" placeholder="🔍 Ara..." oninput="doGelirFilter()" style="width:150px">

    <select class="fi" id="gfyil" onchange="doGelirFilter()"><option value="">Tüm Yıllar</option></select>

    <select class="fi" id="gfay"  onchange="doGelirFilter()"><option value="">Tüm Aylar</option></select>

    <select class="fi" id="gffr"  onchange="doGelirFilter()"><option value="">Tüm Firmalar</option></select>

    <select class="fi" id="gfdu"  onchange="doGelirFilter()">

      <option value="">Tüm Durumlar</option>

      <option>TAHSİL EDİLDİ</option>

      <option>BEKLENİYOR</option>

    </select>

    <div style="display:flex;align-items:center;gap:4px">

      <input class="si" type="date" id="gftarih-bas" style="width:130px" title="Tarih başlangıç" onkeydown="if(event.key==='Enter')doGelirFilter()">

      <span style="color:var(--t3);font-size:11px">—</span>

      <input class="si" type="date" id="gftarih-bit" style="width:130px" title="Tarih bitiş" onkeydown="if(event.key==='Enter')doGelirFilter()">

      <button class="btn bp bsm" onclick="doGelirFilter()" style="padding:4px 10px;font-size:11px;font-weight:600">🔍 Filtrele</button>

      <button class="btn bgh bsm" onclick="gelirTarihSifirla()" title="Tarihi temizle" style="padding:4px 7px;font-size:11px">✕</button>

      <button class="btn bgh bsm" onclick="hizliTarih('gelir','bu-ay')" style="font-size:10px;padding:3px 6px">Bu Ay</button>

      <button class="btn bgh bsm" onclick="hizliTarih('gelir','bu-yil')" style="font-size:10px;padding:3px 6px">Bu Yıl</button>

      <button class="btn bgh bsm" onclick="hizliTarih('gelir','son-3')" style="font-size:10px;padding:3px 6px">Son 3 Ay</button>

    </div>

    <div class="sp"></div>

    <button class="btn bp bsm" onclick="openGelirAdd()">+ Gelir Ekle</button>

  </div>

  <div class="sbar2" id="gelir-stats-bar"></div>

  <div class="tc">

    <table>

      <thead><tr>

        <th style="width:80px">Firma</th>

        <th style="width:150px">Gelir Türü</th>

        <th style="width:150px">Açıklama</th>

        <th style="width:95px">Tarih</th>

        <th style="width:68px">Ay</th>

        <th style="width:105px">Durum</th>

        <th class="num" style="width:130px">Tutar ₺</th>

        <th style="width:65px"></th>

      </tr></thead>

      <tbody id="gelir-tbody"></tbody>

    </table>

    <div class="pgn" id="gelir-pgn"></div>

  </div>

</div>



<!-- NAKİT AKIŞ DENGESİ -->

<div class="view" id="vw-nakit">

  <div class="ptit">Nakit Akış Dengesi</div>

  <div class="psub">Gelir ve gider karşılaştırması · Finansman ihtiyacı analizi</div>

  <div class="tbr">

    <select class="fi" id="nk-yil" onchange="renderNakit()"><option value="">Tüm Yıllar</option></select>

    <div class="sp"></div>

    <button class="btn bo bsm" onclick="expNakitExcel()">⬇ Excel</button>

  </div>



  <!-- Genel özet kartlar -->

  <div id="nakit-top-mets" style="display:grid;grid-template-columns:repeat(4,1fr);gap:10px;margin-bottom:16px"></div>



  <!-- Aylık tablo -->

  <div class="tc" id="nakit-table-wrap">

    <table id="nakit-main-table">

      <thead>

        <tr>

          <th style="width:32px"></th>

          <th style="width:55px">Yıl</th>

          <th style="width:70px">Ay</th>

          <th class="num" style="width:125px">Gelir (Plan) ₺</th>

          <th class="num" style="width:125px">Tahsilat ₺</th>

          <th class="num" style="width:125px">Gider (Plan) ₺</th>

          <th class="num" style="width:125px">Ödenen ₺</th>

          <th class="num" style="width:125px">Net Denge ₺</th>

          <th class="num" style="width:125px">Finansman İht. ₺</th>

          <th style="width:80px">Durum</th>

        </tr>

      </thead>

      <tbody id="nakit-tbody"></tbody>

      <tfoot id="nakit-tfoot"></tfoot>

    </table>

  </div>



  <!-- Bar grafik placeholder -->

  <div style="margin-top:16px;background:var(--s1);border:1px solid var(--b1);border-radius:var(--radl);padding:16px">

    <div style="font-family:var(--mono);font-size:9.5px;color:var(--t3);text-transform:uppercase;letter-spacing:.06em;margin-bottom:14px">Aylık Gelir / Gider / Net Denge</div>

    <div id="nakit-chart-wrap" style="overflow-x:auto"></div>

  </div>

</div>



<!-- LİKİDİTE & FİNANSMAN -->

<div class="view" id="vw-liki">

  <div class="ptit">Likidite &amp; Finansman Modeli</div>

  <div class="psub">Banka bakiyesi + kullanılabilir limitler · Nakit açığı için optimal kredi önerisi</div>



  <!-- Özet kartlar -->

  <div id="liki-mets" style="display:grid;grid-template-columns:repeat(4,1fr);gap:10px;margin-bottom:16px"></div>



  <!-- Sekmeler -->

  <div style="display:flex;gap:0;border-bottom:1px solid var(--b1);margin-bottom:16px">

    <button class="btn bgh" id="ltab-analiz" onclick="likiTab('analiz')" style="border-radius:0;border-bottom:2px solid var(--acc);color:var(--acc)">◎ Aylık Analiz</button>

    <button class="btn bgh" id="ltab-banka" onclick="likiTab('banka')" style="border-radius:0">🏦 Banka Bakiyeleri</button>

    <button class="btn bgh" id="ltab-kredi" onclick="likiTab('kredi')" style="border-radius:0">💳 Kredi Limitleri</button>

  </div>



  <!-- ANALİZ SEKMESİ -->

  <div id="ltab-analiz-view">

    <div class="tbr">

      <select class="fi" id="liki-yil" onchange="renderLikidite()"><option value="">Tüm Yıllar</option></select>

      <div class="sp"></div>

      <button class="btn bo bsm" onclick="expLikiCSV()">⬇ CSV</button>

    </div>

    <div class="tc">

      <table>

        <thead><tr>

          <th style="width:30px"></th>

          <th style="width:55px">Yıl</th>

          <th style="width:70px">Ay</th>

          <th class="num" style="width:120px">Gelir ₺</th>

          <th class="num" style="width:120px">Gider ₺</th>

          <th class="num" style="width:120px">Net Denge ₺</th>

          <th class="num" style="width:130px">Nakit İhtiyaç ₺</th>

          <th class="num" style="width:130px">Kümülatif Banka ₺</th>

          <th class="num" style="width:130px">Tahmini Faiz ₺</th>

          <th style="width:90px">Durum</th>

        </tr></thead>

        <tbody id="liki-tbody"></tbody>

        <tfoot id="liki-tfoot"></tfoot>

      </table>

    </div>

  </div>



  <!-- BANKA SEKMESİ -->

  <div id="ltab-banka-view" style="display:none">

    <div class="tbr">

      <div class="sp"></div>

      <button class="btn bp bsm" onclick="addBankaSatir()">+ Hesap Ekle</button>

      <button class="btn bo bsm" onclick="saveBanka()">💾 Kaydet</button>

    </div>

    <div class="tc">

      <table>

        <thead><tr>

          <th style="width:80px">Firma</th>

          <th style="width:150px">Banka</th>

          <th style="width:150px">Hesap Türü</th>

          <th class="num" style="width:150px">Bakiye ₺</th>

          <th style="width:120px">Güncelleme Tarihi</th>

          <th style="width:60px">Sil</th>

        </tr></thead>

        <tbody id="banka-tbody"></tbody>

      </table>

    </div>

    <div id="banka-toplam" style="padding:10px 14px;background:var(--s2);border:1px solid var(--b1);border-radius:var(--rad);margin-top:10px;font-family:var(--mono);font-size:13px"></div>

  </div>



  <!-- KREDİ LİMİTLERİ SEKMESİ -->

  <div id="ltab-kredi-view" style="display:none">

    <div class="tbr">

      <div class="sp"></div>

      <button class="btn bam2 bsm" id="kredi-import-btn" onclick="importKrediExcel()">⬆ Excelden Yükle</button>

      <button class="btn bsm" style="background:#6366f1;color:#fff;border:none" onclick="krediSablonIndir()">📋 Şablon İndir</button>

      <button class="btn bp bsm" onclick="addKrediSatir()">+ Limit Ekle</button>

      <button class="btn bo bsm" onclick="saveKredi()">💾 Kaydet</button>

    </div>

    <div class="tc">

      <table>

        <thead><tr>

          <th style="width:70px">Firma</th>

          <th style="width:130px">Banka</th>

          <th style="width:110px">Kredi Türü</th>

          <th style="width:120px">Kredi/K.Kart No</th>

          <th class="num" style="width:120px">Limit ₺</th>

          <th class="num" style="width:120px">Kullanılan ₺</th>

          <th class="num" style="width:120px">Kalan Limit ₺</th>

          <th class="num" style="width:90px">Yıllık Faiz %</th>

          <th style="width:90px">Bitiş</th>

          <th style="width:55px">Aktif</th>

          <th style="width:40px">Sil</th>

        </tr></thead>

        <tbody id="kredi-tbody"></tbody>

      </table>

    </div>

    <div id="kredi-toplam" style="padding:10px 14px;background:var(--s2);border:1px solid var(--b1);border-radius:var(--rad);margin-top:10px;font-family:var(--mono);font-size:13px"></div>

  </div>

</div>



<!-- FİNANSMAN YÖNETİMİ -->

<div class="view" id="vw-finans">

  <div class="ptit">Finansman Yönetimi</div>

  <div class="psub">Banka bakiyeleri, kredi limitleri ve aylık finansman öneri motoru</div>



  <!-- İki panel: Banka + Kredi -->

  <div style="display:grid;grid-template-columns:1fr 1fr;gap:12px;margin-bottom:16px">



    <!-- BANKA BAKİYELERİ -->

    <div class="src-card" style="margin:0">

      <div style="display:flex;align-items:center;justify-content:space-between;margin-bottom:12px;flex-wrap:wrap;gap:6px">

        <h3 style="font-size:13px;font-weight:600">Banka Bakiyeleri</h3>

        <div style="display:flex;gap:6px">

          <button class="btn bam2 bsm" onclick="showBankaExcelImport()">⬆ Excel'den Yükle</button>

          <button class="btn bp bsm" onclick="addBankaRow()">+ Ekle</button>

          <button class="btn bo bsm" onclick="saveBankaFinans()">💾 Kaydet</button>

        </div>

      </div>

      <!-- Excel import alanı -->

      <div id="banka-excel-import" style="display:none;margin-bottom:12px;background:var(--s3);border:1px solid var(--b2);border-radius:var(--rad);padding:12px">

        <div style="font-size:11.5px;color:var(--t2);margin-bottom:8px">

          Excel dosyanızda şu sütunlar olmalı: <strong>Banka Adı, Firma, Bakiye, Hesap Türü</strong> (sıra önemli değil)

        </div>

        <div style="display:flex;gap:8px;align-items:center;flex-wrap:wrap">

          <input type="file" id="banka-excel-file" accept=".xlsx,.xls,.csv" style="font-size:12px;color:var(--t2);background:var(--s2);border:1px solid var(--b1);border-radius:var(--rad);padding:5px 8px;flex:1">

          <button class="btn bp bsm" onclick="importBankaExcel()">Yükle</button>

          <button class="btn bo bsm" onclick="document.getElementById('banka-excel-import').style.display='none'">İptal</button>

        </div>

        <div id="banka-import-result" style="margin-top:8px;font-size:12px;display:none"></div>

        <div style="margin-top:8px;font-size:11px;color:var(--t3)">

          💡 Şablon: <a onclick="downloadBankaTemplate()" style="color:var(--acc);cursor:pointer">Örnek Excel şablonunu indir</a>

        </div>

      </div>

      <div id="banka-list"></div>

      <div style="border-top:1px solid var(--b1);margin-top:10px;padding-top:10px;display:flex;justify-content:space-between;font-family:var(--mono);font-size:11.5px">

        <span style="color:var(--t2)">Toplam kullanılabilir nakit</span>

        <span style="color:var(--g);font-weight:600" id="banka-toplam">₺ 0</span>

      </div>

    </div>



    <!-- KREDİ LİMİTLERİ -->

    <div class="src-card" style="margin:0">

      <div style="display:flex;align-items:center;justify-content:space-between;margin-bottom:12px;flex-wrap:wrap;gap:6px">

        <h3 style="font-size:13px;font-weight:600">Kredi Limitleri</h3>

        <div style="display:flex;gap:6px">

          <button class="btn bam2 bsm" id="kredi-finans-import-btn" onclick="importKrediExcelFinans()">⬆ Kaynak Excel'den Yükle</button>

          <button class="btn bp bsm" onclick="addKrediRow()">+ Ekle</button>

          <button class="btn bo bsm" onclick="saveKrediFinans()">💾 Kaydet</button>

        </div>

      </div>

      <div id="kredi-list"></div>

      <div style="border-top:1px solid var(--b1);margin-top:10px;padding-top:10px;display:flex;justify-content:space-between;font-family:var(--mono);font-size:11.5px">

        <span style="color:var(--t2)">Toplam kullanılabilir kredi</span>

        <span style="color:var(--acc);font-weight:600" id="kredi-bos-toplam">₺ 0</span>

      </div>

    </div>

  </div>



  <!-- FİNANSMAN ÖNERİ MOTORU -->

  <div class="src-card" style="margin:0;background:rgba(79,156,249,.04);border-color:rgba(79,156,249,.2)">

    <div style="display:flex;align-items:center;justify-content:space-between;margin-bottom:14px;flex-wrap:wrap;gap:8px">

      <h3 style="font-size:13px;font-weight:600;color:var(--acc)">⬡ Finansman Öneri Motoru</h3>

      <div style="display:flex;gap:8px;align-items:center;flex-wrap:wrap">

        <select class="fi" id="fn-yil" style="font-size:12px"><option value="">Tüm dönem</option></select>

        <select class="fi" id="fn-ay"  style="font-size:12px"><option value="">Tüm aylar</option></select>

        <button class="btn bp bsm" onclick="hesaplaOneri()">Analiz Et</button>

      </div>

    </div>



    <!-- Özet metrikler -->

    <div id="fn-mets" style="display:grid;grid-template-columns:repeat(4,1fr);gap:9px;margin-bottom:14px"></div>



    <!-- Öneri kartları -->

    <div id="fn-oneriler"></div>



    <!-- Kullanılabilir kredi tablosu -->

    <div id="fn-kredi-tablo" style="margin-top:12px"></div>

  </div>

</div>



<!-- CARİ HESAPLAR -->

<div class="view" id="vw-cari">

  <div class="ptit">Cari Hesaplar</div>

  <div class="psub">Alacaklı ve borçlu olduğunuz firma hesap dökümleri</div>



  <!-- Özet metrikler -->

  <div id="cari-mets" style="display:grid;grid-template-columns:repeat(4,1fr);gap:10px;margin-bottom:14px"></div>



  <!-- Sekmeler -->

  <div style="display:flex;gap:0;border-bottom:1px solid var(--b1);margin-bottom:14px">

    <button class="btn bgh" id="ctab-ozet"  onclick="cariTab('ozet')"  style="border-radius:0;border-bottom:2px solid var(--acc);color:var(--acc)">📊 Firma Özeti</button>

    <button class="btn bgh" id="ctab-detay" onclick="cariTab('detay')" style="border-radius:0">≡ Detay Hareketler</button>

  </div>



  <!-- ÖZET SEKMESİ -->

  <div id="ctab-ozet-view">

    <div class="tbr">

      <input class="si" type="text" id="cari-ozet-q" placeholder="🔍 Firma ara..." oninput="renderCariOzet()">

      <select class="fi" id="cari-ozet-tur" onchange="renderCariOzet()">

        <option value="">Tümü</option>

        <option value="ALACAK">Sadece Alacaklılar</option>

        <option value="BORC">Sadece Borçlular</option>

      </select>

      <div class="sp"></div>

      <button class="btn bo bsm" onclick="expCariExcel()">⬇ Excel</button>

      <button class="btn bp bsm" onclick="openCariAdd()">+ Hareket Ekle</button>

    </div>

    <div class="tc">

      <table>

        <thead><tr>

          <th style="width:180px">Cari Firma</th>

          <th style="width:80px">Grup Firma</th>

          <th class="num" style="width:140px">Alacak ₺</th>

          <th class="num" style="width:140px">Borç ₺</th>

          <th class="num" style="width:140px">Net Bakiye ₺</th>

          <th style="width:80px">Durum</th>

          <th style="width:70px">Hareket</th>

          <th style="width:60px"></th>

        </tr></thead>

        <tbody id="cari-ozet-tbody"></tbody>

        <tfoot id="cari-ozet-tfoot"></tfoot>

      </table>

    </div>

  </div>



  <!-- DETAY SEKMESİ -->

  <div id="ctab-detay-view" style="display:none">

    <div class="tbr">

      <input class="si" type="text" id="cari-q" placeholder="🔍 Ara..." oninput="renderCariDetay()">

      <select class="fi" id="cari-firma-f" onchange="renderCariDetay()"><option value="">Tüm Firmalar</option></select>

      <select class="fi" id="cari-tur-f" onchange="renderCariDetay()">

        <option value="">Tüm Türler</option>

        <option value="ALACAK">Alacak</option>

        <option value="BORÇ">Borç</option>

      </select>

      <select class="fi" id="cari-durum-f" onchange="renderCariDetay()">

        <option value="">Tüm Durumlar</option>

        <option value="AÇIK">Açık</option>

        <option value="KAPALI">Kapalı</option>

        <option value="VADEDE">Vadede</option>

      </select>

      <div class="sp"></div>

      <button class="btn bp bsm" onclick="openCariAdd()">+ Ekle</button>

    </div>

    <div class="tc">

      <table>

        <thead><tr>

          <th style="width:160px">Cari Firma</th>

          <th style="width:75px">Grup</th>

          <th style="width:65px">Tür</th>

          <th style="width:80px">Belge No</th>

          <th style="width:90px">Belge Tarihi</th>

          <th style="width:90px">Vade Tarihi</th>

          <th class="num" style="width:130px">Tutar ₺</th>

          <th style="width:90px">Durum</th>

          <th style="width:130px">Açıklama</th>

          <th style="width:65px"></th>

        </tr></thead>

        <tbody id="cari-detay-tbody"></tbody>

      </table>

      <div class="pgn" id="cari-pgn"></div>

    </div>

  </div>

</div>



<!-- CARİ MODAL -->

<div class="ov" id="cari-mod">

  <div class="mod">

    <div class="mtit"><span id="cari-mtit">Yeni Cari Hareket</span><button class="mclose" onclick="closeCariMod()">✕</button></div>

    <div class="fg">

      <div class="fgr ff"><label class="flb">Cari Firma Adı</label>

        <input class="fi2" type="text" id="cm-firma" placeholder="Firma ünvanı" list="cari-firma-list">

        <datalist id="cari-firma-list"></datalist>

      </div>

      <div class="fgr"><label class="flb">Grup Firma</label>

        <select class="fse" id="cm-grup"><option>ULUSAL</option><option>ABC</option><option>BRG</option><option>BRK</option><option>-</option></select>

      </div>

      <div class="fgr"><label class="flb">Hareket Türü</label>

        <select class="fse" id="cm-tur">

          <option value="ALACAK">ALACAK (Bize borçlu)</option>

          <option value="BORÇ">BORÇ (Biz borçluyuz)</option>

        </select>

      </div>

      <div class="fgr"><label class="flb">Belge No / Fatura No</label>

        <input class="fi2" type="text" id="cm-belge" placeholder="FTR-2024-001"></div>

      <div class="fgr"><label class="flb">Belge / İşlem Tarihi</label>

        <input class="fi2" type="date" id="cm-tarih"></div>

      <div class="fgr"><label class="flb">Vade Tarihi</label>

        <input class="fi2" type="date" id="cm-vade"></div>

      <div class="fgr"><label class="flb">Tutar (₺)</label>

        <input class="fi2" type="number" id="cm-tutar" placeholder="0"></div>

      <div class="fgr"><label class="flb">Durum</label>

        <select class="fse" id="cm-durum">

          <option value="AÇIK">AÇIK</option>

          <option value="VADEDE">VADEDE</option>

          <option value="KAPALI">KAPALI</option>

        </select>

      </div>

      <div class="fgr ff"><label class="flb">Açıklama</label>

        <input class="fi2" type="text" id="cm-aciklama" placeholder="Mal/hizmet türü, notlar..."></div>

    </div>

    <div class="mac">

      <button class="btn bo" onclick="closeCariMod()">İptal</button>

      <button class="btn bp" id="cm-save" onclick="saveCari()">Kaydet</button>

    </div>

  </div>

</div>



<!-- BANKA API & EKSTRE -->

<div class="view" id="vw-banka-api">

  <div class="ptit">Banka API & Ekstre İmport</div>

  <div class="psub">API entegrasyonu hazır bekliyor · Şimdilik ekstre dosyası yükleyerek bakiye güncelleyin</div>



  <!-- Sekmeler -->

  <div style="display:flex;gap:0;border-bottom:1px solid var(--b1);margin-bottom:16px">

    <button class="btn bgh" id="batab-ekstre" onclick="baTab('ekstre')" style="border-radius:0;border-bottom:2px solid var(--acc);color:var(--acc)">📄 Ekstre Import</button>

    <button class="btn bgh" id="batab-api"    onclick="baTab('api')"    style="border-radius:0">🔌 API Ayarları</button>

  </div>



  <!-- EKSTRE SEKMESİ -->

  <div id="batab-ekstre-view">

    <div class="info-box" style="margin-bottom:14px">

      📌 <strong>Banka ekstrenizi Excel veya CSV olarak indirip buraya yükleyin.</strong><br>

      Vakıfbank, Garanti BBVA, İş Bankası, YapıKredi, Akbank, Halkbank, Ziraat Bankası, TEB, QNB Finansbank ve Denizbank formatları otomatik tanınır.

      Son bakiye <strong>Banka Bakiyeleri</strong> tablosuna otomatik eklenir.

    </div>



    <!-- Yükleme alanı -->

    <div id="ekstre-drop-zone" style="border:2px dashed var(--b2);border-radius:var(--radl);padding:32px;text-align:center;cursor:pointer;transition:all .2s;margin-bottom:16px"

      onclick="document.getElementById('ekstre-file-input').click()"

      ondragover="event.preventDefault();this.style.borderColor='var(--acc)'"

      ondragleave="this.style.borderColor='var(--b2)'"

      ondrop="event.preventDefault();this.style.borderColor='var(--b2)';handleEkstreFile(event.dataTransfer.files[0])">

      <div style="font-size:28px;margin-bottom:8px">📂</div>

      <div style="font-size:13px;color:var(--t2)">Ekstre dosyasını buraya sürükleyin veya tıklayın</div>

      <div style="font-size:11px;color:var(--t3);margin-top:4px">Excel (.xlsx, .xls) veya CSV (.csv) · Maks 10MB</div>

      <input type="file" id="ekstre-file-input" style="display:none" accept=".xlsx,.xls,.csv,.txt" onchange="handleEkstreFile(this.files[0])">

    </div>



    <!-- Sonuç önizleme -->

    <div id="ekstre-result" style="display:none">

      <div id="ekstre-info" style="background:var(--s2);border:1px solid var(--b1);border-radius:var(--rad);padding:12px 16px;margin-bottom:12px;display:flex;gap:20px;flex-wrap:wrap;font-family:var(--mono);font-size:12px"></div>

      <div class="tc" style="max-height:320px;overflow-y:auto">

        <table>

          <thead><tr>

            <th style="width:90px">Tarih</th>

            <th>Açıklama</th>

            <th class="num" style="width:120px">Borç ₺</th>

            <th class="num" style="width:120px">Alacak ₺</th>

            <th class="num" style="width:120px">Bakiye ₺</th>

          </tr></thead>

          <tbody id="ekstre-tbody"></tbody>

        </table>

      </div>

      <div style="display:flex;gap:8px;margin-top:12px;flex-wrap:wrap">

        <button class="btn bp" onclick="saveEkstreBakiye()">💾 Bakiyeyi Kaydet</button>

        <button class="btn bo" onclick="resetEkstre()">✕ İptal</button>

      </div>

    </div>



    <!-- Geçmiş yüklemeler -->

    <div style="margin-top:20px">

      <div style="font-family:var(--mono);font-size:10px;color:var(--t3);text-transform:uppercase;letter-spacing:.06em;margin-bottom:8px">Son Yüklenen Ekstre Bakiyeleri</div>

      <div id="ekstre-gecmis" style="font-size:12px;color:var(--t3)">—</div>

    </div>

  </div>



  <!-- API AYARLARI SEKMESİ -->

  <div id="batab-api-view" style="display:none">

    <div class="info-box" style="margin-bottom:16px;border-color:rgba(251,191,36,.3);background:rgba(251,191,36,.05)">

      🔌 <strong>Banka API Ayarları</strong> — API anahtarlarınızı bankadan aldıktan sonra buraya girin.<br>

      Şu an kayıt aşamasında, aktif ettiğinizde gerçek zamanlı bakiye çekmeye başlar.

    </div>



    <div id="api-list" style="display:grid;gap:10px;margin-bottom:16px"></div>



    <button class="btn bp bsm" onclick="addApiEntry()">+ Yeni Banka API Ekle</button>



    <div id="api-edit-area" style="display:none;margin-top:16px;background:var(--s2);border:1px solid var(--b1);border-radius:var(--radl);padding:18px">

      <div style="font-weight:600;font-size:13px;margin-bottom:14px" id="api-edit-title">Yeni API Bağlantısı</div>

      <div class="fg">

        <div class="fgr"><label class="flb">Banka Adı</label>

          <select class="fse" id="api-banka">

            <option>VAKIFBANK</option><option>GARANTİ BBVA</option><option>İŞ BANKASI</option>

            <option>YAPİ KREDİ</option><option>AKBANK</option><option>HALKBANK</option>

            <option>ZİRAAT BANKASI</option><option>TEB</option><option>QNB FİNANSBANK</option>

            <option>DENİZBANK</option><option>DİĞER</option>

          </select></div>

        <div class="fgr"><label class="flb">Hesap / Tanım</label>

          <input class="fi2" type="text" id="api-tanim" placeholder="Ana hesap, Maaş hesabı..."></div>

        <div class="fgr ff"><label class="flb">API Base URL</label>

          <input class="fi2" type="text" id="api-url" placeholder="https://api.vakifbank.com.tr/v1"></div>

        <div class="fgr"><label class="flb">API Key / Client ID</label>

          <input class="fi2" type="text" id="api-key" placeholder="API anahtarını girin"></div>

        <div class="fgr"><label class="flb">API Secret / Client Secret</label>

          <input class="fi2" type="password" id="api-secret" placeholder="Gizli anahtar"></div>

        <div class="fgr"><label class="flb">Hesap No / IBAN</label>

          <input class="fi2" type="text" id="api-hesap" placeholder="TR00 0000 0000 0000 0000 00"></div>

        <div class="fgr"><label class="flb">Firma</label>

          <select class="fse" id="api-firma"><option>ULUSAL</option><option>ABC</option><option>BRG</option><option>BRK</option></select></div>

        <div class="fgr" style="align-items:flex-end">

          <label class="tgl" style="margin-bottom:8px">

            <input type="checkbox" id="api-aktif"> Aktif (API hazır olduğunda işaretle)

          </label></div>

        <div class="fgr ff"><label class="flb">Notlar / Başvuru Durumu</label>

          <input class="fi2" type="text" id="api-not" placeholder="Başvuru yapıldı, bekleniyor..."></div>

      </div>

      <div style="display:flex;gap:8px;margin-top:14px;flex-wrap:wrap">

        <button class="btn bp" onclick="saveApiEntry()">✓ Kaydet</button>

        <button class="btn bo" onclick="document.getElementById('api-edit-area').style.display='none'">İptal</button>

        <button class="btn bo" id="api-test-btn" onclick="testApiConn()" style="margin-left:auto">⚡ Bağlantı Test Et</button>

      </div>

      <div id="api-test-result" style="margin-top:8px;font-size:12px;display:none"></div>

    </div>

  </div>

</div>



<!-- DÖVİZ KURLARI -->

<div class="view" id="vw-kur">

  <div class="ptit">Döviz Kurları</div>

  <div class="psub">TCMB (Türkiye Cumhuriyet Merkez Bankası) resmi döviz kurları</div>



  <!-- Tarih seçici ve özet -->

  <div style="display:flex;align-items:center;gap:10px;margin-bottom:16px;flex-wrap:wrap">

    <div style="display:flex;align-items:center;gap:8px;background:var(--s2);border:1px solid var(--b1);border-radius:var(--rad);padding:8px 12px">

      <label style="font-size:12px;color:var(--t2)">Tarih:</label>

      <input type="date" id="kur-tarih" class="fi2" style="width:150px"

        value="" onchange="">

    </div>

    <button class="btn bp" onclick="kurCek()">📡 Kurları Getir</button>

    <button class="btn bo bsm" onclick="kurBugun()">Bugün</button>

    <div class="sp"></div>

    <button class="btn bo bsm" onclick="kurExport()">⬇ CSV İndir</button>

  </div>



  <!-- Sonuç bilgi barı -->

  <div id="kur-info" style="display:none;background:var(--s2);border:1px solid var(--b1);border-radius:var(--rad);padding:10px 14px;margin-bottom:14px;font-size:12px;font-family:var(--mono);display:flex;gap:16px;align-items:center;flex-wrap:wrap"></div>



  <!-- Ana kurlar özet kartları -->

  <div id="kur-ozet-kartlar" style="display:grid;grid-template-columns:repeat(5,1fr);gap:10px;margin-bottom:16px"></div>



  <!-- Tüm kurlar tablosu -->

  <div class="tc" id="kur-tablo-wrap" style="display:none">

    <table>

      <thead><tr>

        <th style="width:60px">Kod</th>

        <th style="width:40px">Birim</th>

        <th>Para Birimi</th>

        <th class="num" style="width:130px">Döviz Alış ₺</th>

        <th class="num" style="width:130px">Döviz Satış ₺</th>

        <th class="num" style="width:130px">Efektif Alış ₺</th>

        <th class="num" style="width:130px">Efektif Satış ₺</th>

      </tr></thead>

      <tbody id="kur-tbody"></tbody>

    </table>

  </div>



  <!-- Geçmiş kurlar -->

  <div style="margin-top:20px" id="kur-gecmis-wrap" style="display:none">

    <div style="font-family:var(--mono);font-size:10px;color:var(--t3);text-transform:uppercase;letter-spacing:.06em;margin-bottom:8px">Sorgulanan Tarihler</div>

    <div id="kur-gecmis" style="display:flex;gap:6px;flex-wrap:wrap"></div>

  </div>

</div>



<!-- LOGO FATURALAR -->

<div class="view" id="vw-logo">

  <div class="ptit">Faturalar — Logo Tiger 3</div>

  <div class="psub">Satış ve alış faturalarınız doğrudan Logo veritabanından</div>



  <!-- Bağlantı durumu barı -->

  <div id="logo-conn-bar" style="background:var(--s2);border:1px solid var(--b1);border-radius:var(--rad);padding:10px 14px;margin-bottom:14px;display:flex;align-items:center;gap:12px;flex-wrap:wrap">

    <div class="src-dot none" id="logo-dot"></div>

    <span id="logo-conn-msg" style="font-size:12px;color:var(--t2)">Logo bağlantısı test ediliyor...</span>

    <div class="sp"></div>

    <button class="btn bo bsm" onclick="showLogoAyarlar()">⚙ Bağlantı Ayarları</button>

  </div>



  <!-- Bağlantı ayarları (gizli) -->

  <div id="logo-ayarlar" style="display:none;background:var(--s2);border:1px solid var(--b2);border-radius:var(--radl);padding:18px;margin-bottom:14px">

    <div style="font-weight:600;font-size:13px;margin-bottom:14px">Logo Tiger 3 — SQL Server Bağlantı Ayarları</div>

    <div class="fg">

      <div class="fgr ff"><label class="flb">SQL Server Adı / IP</label>

        <input class="fi2" type="text" id="logo-server" placeholder="SUNUCU\LOGO veya 192.168.1.10"></div>

      <div class="fgr"><label class="flb">Port</label>

        <input class="fi2" type="number" id="logo-port" value="1433"></div>

      <div class="fgr ff"><label class="flb">Veritabanı Adı</label>

        <input class="fi2" type="text" id="logo-db" placeholder="LGS_SIRKET"></div>

      <div class="fgr"><label class="flb">SQL Kullanıcı</label>

        <input class="fi2" type="text" id="logo-user" placeholder="logo_user"></div>

      <div class="fgr"><label class="flb">Şifre</label>

        <input class="fi2" type="password" id="logo-pass" placeholder="••••••••"></div>

    </div>

    <div style="margin-top:12px">

      <div style="font-size:12px;font-weight:600;color:var(--acc);margin-bottom:8px">🏢 Firma Listesi

        <span style="font-size:10px;color:var(--t3);font-weight:400;margin-left:6px">Logo'da kaç firma varsa hepsini ekleyin (001, 002...)</span>

      </div>

      <div id="logo-firma-listesi" style="margin-bottom:8px;display:flex;flex-direction:column;gap:6px"></div>

      <button class="btn bo bsm" onclick="logoFirmaEkle()">+ Firma Ekle</button>

    </div>

    <div class="info-box" style="margin-top:12px;margin-bottom:12px">

      🔒 Şifre yalnızca bu bilgisayarda yerel olarak saklanır. Dışarıya gönderilmez.

      IT'den <strong>salt okunur SQL kullanıcısı</strong> isteyin — uygulama sadece SELECT yapar.

    </div>

    <div style="display:flex;gap:8px;flex-wrap:wrap">

      <button class="btn bp" onclick="logoAyarKaydet()">✓ Kaydet</button>

      <button class="btn bg2 bsm" onclick="logoTestEt()" id="logo-test-btn">⚡ Bağlantı Test Et</button>

      <button class="btn bo" onclick="document.getElementById('logo-ayarlar').style.display='none'">İptal</button>

    </div>

    <div id="logo-test-result" style="margin-top:10px;font-size:12px;display:none"></div>

  </div>



  <!-- Fatura filtreleri -->

  <!-- Fatura filtre alanı -->

  <div style="background:var(--s2);border:1px solid var(--b1);border-radius:var(--radl);padding:12px;margin-bottom:10px" id="logo-filter-bar">

    <div style="display:flex;gap:8px;flex-wrap:wrap;margin-bottom:8px">

      <!-- Tip seçici -->

      <div style="display:flex;gap:0;border:1px solid var(--b1);border-radius:var(--rad);overflow:hidden">

        <button class="btn" id="logo-tab-satis" onclick="logoTab('satis')"

          style="border-radius:0;background:var(--acc);color:#fff;border:none;padding:5px 12px;font-size:12px">📤 Satış</button>

        <button class="btn bgh" id="logo-tab-alis" onclick="logoTab('alis')"

          style="border-radius:0;border:none;padding:5px 12px;font-size:12px">📥 Alış</button>

        <button class="btn bgh" id="logo-tab-tumu" onclick="logoTab('tumu')"

          style="border-radius:0;border:none;padding:5px 12px;font-size:12px">📋 Tümü</button>

      </div>

      <input class="si" type="date" id="logo-bas" style="width:135px" title="Başlangıç tarihi">

      <input class="si" type="date" id="logo-bit" style="width:135px" title="Bitiş tarihi">

      <input class="si" type="text" id="logo-cari-f" placeholder="Cari adı / kodu..." style="width:170px">

      <input class="si" type="number" id="logo-min-tutar" placeholder="Min ₺" style="width:90px">

      <input class="si" type="number" id="logo-max-tutar" placeholder="Max ₺" style="width:90px">

      <button class="btn bp bsm" onclick="logoFaturaYukle()">🔍 Getir</button>

      <button class="btn bgh bsm" onclick="hizliTarih('logo','bu-ay')" style="font-size:10px;padding:3px 6px">Bu Ay</button>

      <button class="btn bgh bsm" onclick="hizliTarih('logo','bu-yil')" style="font-size:10px;padding:3px 6px">Bu Yıl</button>

      <button class="btn bgh bsm" onclick="hizliTarih('logo','son-3')" style="font-size:10px;padding:3px 6px">Son 3 Ay</button>

      <div class="sp"></div>

      <button class="btn bo bsm" onclick="logoExport()">⬇ CSV</button>

    </div>

    <!-- Firma ve yıl çoklu seçim -->

    <div style="display:flex;gap:12px;flex-wrap:wrap;align-items:flex-start">

      <div>

        <div style="font-size:10px;color:var(--t3);margin-bottom:4px;text-transform:uppercase;letter-spacing:.05em">Firma Seçimi</div>

        <div id="logo-firma-secim" style="display:flex;gap:6px;flex-wrap:wrap;min-height:26px">

          <span style="color:var(--t3);font-size:11px">Bağlantı kurulunca firmalar görünür</span>

        </div>

      </div>

      <div>

        <div style="font-size:10px;color:var(--t3);margin-bottom:4px;text-transform:uppercase;letter-spacing:.05em">Yıl Seçimi</div>

        <div id="logo-yil-secim" style="display:flex;gap:6px;flex-wrap:wrap;min-height:26px">

          <span style="color:var(--t3);font-size:11px">Veriler yüklendikçe görünür</span>

        </div>

      </div>

    </div>

  </div>



  <!-- Özet metrikler -->

  <div id="logo-mets" style="display:grid;grid-template-columns:repeat(4,1fr);gap:10px;margin-bottom:14px"></div>



  <!-- Fatura tablosu -->

  <div class="tc" id="logo-tablo">

    <table>

      <thead><tr>

        <th style="width:30px"></th>

        <th style="width:120px">Fatura No</th>

        <th style="width:85px">Tarih</th>

        <th style="width:40px">Yıl</th>

        <th style="width:80px">Firma</th>

        <th>Cari Unvan</th>

        <th style="width:65px">Cari Kod</th>

        <th class="num" style="width:120px">Net Toplam ₺</th>

        <th class="num" style="width:90px">KDV ₺</th>

        <th class="num" style="width:120px">Brüt Toplam ₺</th>

        <th style="width:65px">Tür</th>

      </tr></thead>

      <tbody id="logo-tbody"></tbody>

      <tfoot id="logo-tfoot"></tfoot>

    </table>

    <div class="pgn" id="logo-pgn"></div>

  </div>

</div>



<!-- STOK LİSTESİ -->

<div class="view" id="vw-stok">

  <div class="ptit">Stok Listesi — Logo Tiger 3</div>

  <div class="psub">Logo veritabanından canlı stok kartları — firma ve yıl bazlı</div>



  <!-- Bağlantı durumu barı -->

  <div id="stok-conn-bar" style="background:var(--s2);border:1px solid var(--b1);border-radius:var(--rad);padding:10px 14px;margin-bottom:12px;display:flex;align-items:center;gap:12px;flex-wrap:wrap">

    <div class="src-dot none" id="stok-conn-dot"></div>

    <span id="stok-conn-msg" style="font-size:12px;color:var(--t2)">Logo bağlantısı kontrol ediliyor...</span>

    <div class="sp"></div>

    <button class="btn bo bsm" onclick="showStokAyarlar()">⚙ Bağlantı Ayarları</button>

  </div>



  <!-- Bağlantı ayarları (stok için - fatura ile paylaşılan) -->

  <div id="stok-ayarlar" style="display:none;background:var(--s2);border:1px solid var(--b2);border-radius:var(--radl);padding:16px;margin-bottom:12px">

    <div style="font-size:13px;font-weight:600;margin-bottom:12px">Logo Tiger 3 — SQL Server Bağlantı Ayarları</div>

    <div class="fg">

      <div class="fgr ff"><label class="flb">SQL Server Adı / IP</label>

        <input class="fi2" type="text" id="stok-logo-server" placeholder="SUNUCU\LOGO veya 192.168.1.10"></div>

      <div class="fgr"><label class="flb">Port</label>

        <input class="fi2" type="number" id="stok-logo-port" value="1433"></div>

      <div class="fgr ff"><label class="flb">Veritabanı Adı</label>

        <input class="fi2" type="text" id="stok-logo-db" placeholder="LGS_SIRKET"></div>

      <div class="fgr"><label class="flb">SQL Kullanıcı</label>

        <input class="fi2" type="text" id="stok-logo-user" placeholder="logo_user"></div>

      <div class="fgr"><label class="flb">Şifre</label>

        <input class="fi2" type="password" id="stok-logo-pass" placeholder="••••••••"></div>

    </div>

    <div style="margin-top:10px">

      <div style="font-size:12px;font-weight:600;color:var(--acc);margin-bottom:6px">🏢 Firma Listesi</div>

      <div id="stok-logo-firma-listesi" style="margin-bottom:8px;display:flex;flex-direction:column;gap:6px"></div>

      <button class="btn bo bsm" onclick="stokLogoFirmaEkle()">+ Firma Ekle</button>

    </div>

    <div class="info-box" style="margin-top:10px">

      🔒 Faturalar sayfasındaki bağlantı ile aynı ayarlar kullanılır. Kaydettiğinizde tüm Logo sayfaları güncellenir.

    </div>

    <div style="display:flex;gap:8px;margin-top:12px">

      <button class="btn bp" onclick="stokLogoAyarKaydet()">✓ Kaydet</button>

      <button class="btn bg2 bsm" onclick="stokLogoTestEt()">⚡ Test Et</button>

      <button class="btn bo" onclick="document.getElementById('stok-ayarlar').style.display='none'">İptal</button>

    </div>

    <div id="stok-test-result" style="margin-top:8px;font-size:12px;display:none"></div>

  </div>

  <div style="background:var(--s2);border:1px solid var(--b1);border-radius:var(--radl);padding:12px;margin-bottom:10px">

    <div style="display:flex;gap:8px;flex-wrap:wrap;margin-bottom:8px">

      <input class="si" type="text" id="stok-q" placeholder="🔍 Stok kodu veya adı..." style="width:200px" onkeydown="if(event.key==='Enter')stokYukle()">

      <select class="fi" id="stok-grup-f" style="width:120px"><option value="">Tüm Gruplar</option></select>

      <div style="display:flex;align-items:center;gap:4px">

        <input class="si" type="date" id="stok-bas" style="width:130px" title="Hareket başlangıç tarihi">

        <span style="color:var(--t3);font-size:11px">—</span>

        <input class="si" type="date" id="stok-bit" style="width:130px" title="Hareket bitiş tarihi">

      </div>

      <label style="display:flex;align-items:center;gap:4px;font-size:12px;color:var(--t2);cursor:pointer">

        <input type="checkbox" id="stok-aktif-hareket"> Sadece hareketli

      </label>

      <button class="btn bp bsm" onclick="stokYukle()">🔍 Ara</button>

      <button class="btn bgh bsm" onclick="hizliTarih('stok','bu-yil')" style="font-size:10px;padding:3px 6px">Bu Yıl</button>

      <button class="btn bgh bsm" onclick="hizliTarih('stok','son-3')" style="font-size:10px;padding:3px 6px">Son 3 Ay</button>

      <div class="sp"></div>

      <span id="stok-count" style="font-size:12px;color:var(--t3);align-self:center"></span>

      <button class="btn bo bsm" onclick="stokExport()">⬇ CSV</button>

    </div>

    <div style="display:flex;gap:16px;flex-wrap:wrap">

      <div>

        <div style="font-size:10px;color:var(--t3);margin-bottom:4px">Firma:</div>

        <div id="stok-firma-secim" style="display:flex;gap:6px;flex-wrap:wrap"></div>

      </div>

      <div>

        <div style="font-size:10px;color:var(--t3);margin-bottom:4px">Yıl:</div>

        <div id="stok-yil-secim" style="display:flex;gap:6px;flex-wrap:wrap"></div>

      </div>

    </div>

  </div>

  <div class="tc">

    <table>

      <thead><tr>

        <th style="width:120px">Stok Kodu</th>

        <th>Stok Adı</th>

        <th style="width:75px">Firma</th>

        <th style="width:90px">Grup</th>

        <th style="width:55px">Birim</th>

        <th class="num" style="width:85px">Stok Bakiye</th>

        <th class="num" style="width:90px">Dönem Sat. Adet</th>

        <th class="num" style="width:110px">Dönem Sat. ₺</th>

        <th class="num" style="width:90px">Son Alış ₺</th>

        <th class="num" style="width:90px">Son Satış ₺</th>

      </tr></thead>

      <tbody id="stok-tbody"></tbody>

    </table>

    <div class="pgn" id="stok-pgn"></div>

  </div>

</div>



<!-- CARİ LİSTESİ (LOGO) -->

<div class="view" id="vw-cari-logo">

  <div class="ptit">Cari Hesap Listesi — Logo Tiger 3</div>

  <div class="psub">Logo veritabanından canlı cari kartlar — firma, yıl ve bakiye analiziyle</div>



  <!-- Bağlantı durumu barı -->

  <div id="cari-logo-conn-bar" style="background:var(--s2);border:1px solid var(--b1);border-radius:var(--rad);padding:10px 14px;margin-bottom:12px;display:flex;align-items:center;gap:12px;flex-wrap:wrap">

    <div class="src-dot none" id="cari-logo-conn-dot"></div>

    <span id="cari-logo-conn-msg" style="font-size:12px;color:var(--t2)">Logo bağlantısı kontrol ediliyor...</span>

    <div class="sp"></div>

    <button class="btn bo bsm" onclick="showCariLogoAyarlar()">⚙ Bağlantı Ayarları</button>

  </div>



  <!-- Bağlantı ayarları (cari için) -->

  <div id="cari-logo-ayarlar" style="display:none;background:var(--s2);border:1px solid var(--b2);border-radius:var(--radl);padding:16px;margin-bottom:12px">

    <div style="font-size:13px;font-weight:600;margin-bottom:12px">Logo Tiger 3 — SQL Server Bağlantı Ayarları</div>

    <div class="fg">

      <div class="fgr ff"><label class="flb">SQL Server Adı / IP</label>

        <input class="fi2" type="text" id="cari-logo-server" placeholder="SUNUCU\LOGO veya 192.168.1.10"></div>

      <div class="fgr"><label class="flb">Port</label>

        <input class="fi2" type="number" id="cari-logo-port" value="1433"></div>

      <div class="fgr ff"><label class="flb">Veritabanı Adı</label>

        <input class="fi2" type="text" id="cari-logo-db" placeholder="LGS_SIRKET"></div>

      <div class="fgr"><label class="flb">SQL Kullanıcı</label>

        <input class="fi2" type="text" id="cari-logo-user" placeholder="logo_user"></div>

      <div class="fgr"><label class="flb">Şifre</label>

        <input class="fi2" type="password" id="cari-logo-pass" placeholder="••••••••"></div>

    </div>

    <div style="margin-top:10px">

      <div style="font-size:12px;font-weight:600;color:var(--acc);margin-bottom:6px">🏢 Firma Listesi</div>

      <div id="cari-logo-firma-listesi" style="margin-bottom:8px;display:flex;flex-direction:column;gap:6px"></div>

      <button class="btn bo bsm" onclick="cariLogoFirmaEkle()">+ Firma Ekle</button>

    </div>

    <div class="info-box" style="margin-top:10px">

      🔒 Faturalar sayfasındaki bağlantı ile aynı ayarlar kullanılır. Kaydettiğinizde tüm Logo sayfaları güncellenir.

    </div>

    <div style="display:flex;gap:8px;margin-top:12px">

      <button class="btn bp" onclick="cariLogoAyarKaydet()">✓ Kaydet</button>

      <button class="btn bg2 bsm" onclick="cariLogoTestEt()">⚡ Test Et</button>

      <button class="btn bo" onclick="document.getElementById('cari-logo-ayarlar').style.display='none'">İptal</button>

    </div>

    <div id="cari-logo-test-result" style="margin-top:8px;font-size:12px;display:none"></div>

  </div>

  <div style="background:var(--s2);border:1px solid var(--b1);border-radius:var(--radl);padding:12px;margin-bottom:10px">

    <div style="display:flex;gap:8px;flex-wrap:wrap;margin-bottom:8px">

      <input class="si" type="text" id="cari-logo-q" placeholder="🔍 Cari kodu, adı veya vergi no..." style="width:220px" onkeydown="if(event.key==='Enter')cariLogoYukle()">

      <select class="fi" id="cari-logo-tip">

        <option value="">Tümü</option>

        <option value="musteri">Müşteriler</option>

        <option value="tedarikci">Tedarikçiler</option>

      </select>

      <div style="display:flex;align-items:center;gap:4px">

        <input class="si" type="date" id="cari-bas" style="width:130px" title="İşlem başlangıç tarihi">

        <span style="color:var(--t3);font-size:11px">—</span>

        <input class="si" type="date" id="cari-bit" style="width:130px" title="İşlem bitiş tarihi">

      </div>

      <select class="fi" id="cari-sirala">

        <option value="kod">Koda Göre</option>

        <option value="ciro_desc">Ciroya Göre ↓</option>

        <option value="bakiye_desc">Bakiyeye Göre ↓</option>

      </select>

      <button class="btn bp bsm" onclick="cariLogoYukle()">🔍 Ara</button>

      <button class="btn bgh bsm" onclick="hizliTarih('cari','bu-yil')" style="font-size:10px;padding:3px 6px">Bu Yıl</button>

      <button class="btn bgh bsm" onclick="hizliTarih('cari','son-3')" style="font-size:10px;padding:3px 6px">Son 3 Ay</button>

      <div class="sp"></div>

      <span id="cari-logo-count" style="font-size:12px;color:var(--t3);align-self:center"></span>

      <button class="btn bo bsm" onclick="cariLogoExport()">⬇ CSV</button>

    </div>

    <div style="display:flex;gap:16px;flex-wrap:wrap">

      <div>

        <div style="font-size:10px;color:var(--t3);margin-bottom:4px">Firma:</div>

        <div id="cari-firma-secim" style="display:flex;gap:6px;flex-wrap:wrap"></div>

      </div>

      <div>

        <div style="font-size:10px;color:var(--t3);margin-bottom:4px">Yıl (dönem ciro için):</div>

        <div id="cari-yil-secim" style="display:flex;gap:6px;flex-wrap:wrap"></div>

      </div>

    </div>

  </div>

  <div class="tc">

    <table>

      <thead><tr>

        <th style="width:105px">Cari Kodu</th>

        <th>Cari Adı</th>

        <th style="width:70px">Tür</th>

        <th style="width:70px">Firma</th>

        <th class="num" style="width:120px">Dönem Satış ₺</th>

        <th class="num" style="width:120px">Dönem Alış ₺</th>

        <th class="num" style="width:110px">Cari Bakiye ₺</th>

        <th style="width:90px">Vergi No</th>

        <th style="width:110px">Şehir</th>

        <th style="width:110px">Telefon</th>

      </tr></thead>

      <tbody id="cari-logo-tbody"></tbody>

    </table>

    <div class="pgn" id="cari-logo-pgn"></div>

  </div>

</div>



<!-- GMAIL -->

<div class="view" id="vw-gmail">

  <div class="ptit">Gmail</div>

  <div class="psub">sahinerikin@gmail.com — Google hesabı entegrasyonu</div>



  <!-- Bağlantı durumu -->

  <div id="gmail-conn-bar" style="background:var(--s2);border:1px solid var(--b1);border-radius:var(--rad);padding:10px 14px;margin-bottom:14px;display:flex;align-items:center;gap:12px;flex-wrap:wrap">

    <div class="src-dot none" id="gmail-dot"></div>

    <span id="gmail-conn-msg" style="font-size:12px;color:var(--t2)">Gmail durumu kontrol ediliyor...</span>

    <div class="sp"></div>

    <button class="btn bo bsm" onclick="showGmailAyarlar()">⚙ API Ayarları</button>

  </div>



  <!-- API Ayarları -->

  <div id="gmail-ayarlar" style="display:none;background:var(--s2);border:1px solid var(--b2);border-radius:var(--radl);padding:18px;margin-bottom:14px">

    <div style="font-weight:600;font-size:13px;margin-bottom:10px">Gmail API Ayarları — Google Cloud Console</div>

    <div class="info-box" style="margin-bottom:14px">

      <strong>Kurulum adımları:</strong><br>

      1. <a href="https://console.cloud.google.com" target="_blank" style="color:var(--acc)">console.cloud.google.com</a> → Yeni proje oluştur<br>

      2. <strong>API &amp; Services → Enable APIs</strong> → "Gmail API" etkinleştir<br>

      3. <strong>Credentials → Create Credentials → OAuth 2.0 Client ID</strong><br>

      4. Application type: <strong>Web application</strong><br>

      5. Authorized redirect URI: <code style="background:var(--s3);padding:2px 6px;border-radius:4px">http://localhost:5678/api/gmail/oauth/callback</code><br>

      6. Client ID ve Client Secret'ı kopyalayıp aşağıya yapıştırın

    </div>

    <div class="fg">

      <div class="fgr ff"><label class="flb">Client ID</label>

        <input class="fi2" type="text" id="gmail-client-id" placeholder="....apps.googleusercontent.com"></div>

      <div class="fgr ff"><label class="flb">Client Secret</label>

        <input class="fi2" type="password" id="gmail-client-secret" placeholder="GOCSPX-..."></div>

      <div class="fgr"><label class="flb">Gmail Adresi</label>

        <input class="fi2" type="text" id="gmail-email" value="sahinerikin@gmail.com"></div>

    </div>

    <div style="display:flex;gap:8px;margin-top:14px;flex-wrap:wrap">

      <button class="btn bp" onclick="gmailAyarKaydet()">✓ Kaydet</button>

      <button class="btn bg2 bsm" onclick="gmailOAuthBaslat()" id="gmail-oauth-btn">🔐 Google ile Bağlan</button>

      <button class="btn bo" onclick="document.getElementById('gmail-ayarlar').style.display='none'">İptal</button>

    </div>

    <div id="gmail-ayar-result" style="margin-top:10px;font-size:12px;display:none"></div>

  </div>



  <!-- Mail listesi ve yaz butonu -->

  <div class="tbr" id="gmail-filter-bar">

    <input class="si" type="text" id="gmail-q" placeholder="🔍 Konu, gönderen..." style="width:220px" onkeydown="if(event.key==='Enter')gmailYukle()">

    <button class="btn bp bsm" onclick="gmailYukle()">🔍 Getir</button>

    <div class="sp"></div>

    <button class="btn bp" onclick="openMailYaz()">✏ Yeni Mail Yaz</button>

  </div>



  <!-- Mail listesi -->

  <div class="tc" id="gmail-liste-wrap">

    <table>

      <thead><tr>

        <th style="width:20px"></th>

        <th style="width:220px">Kimden</th>

        <th>Konu</th>

        <th style="width:170px">Tarih</th>

      </tr></thead>

      <tbody id="gmail-tbody"></tbody>

    </table>

  </div>



  <!-- Mail detay -->

  <div id="gmail-detay" style="display:none;background:var(--s2);border:1px solid var(--b1);border-radius:var(--radl);padding:18px;margin-top:14px">

    <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:12px;flex-wrap:wrap;gap:8px">

      <div>

        <div id="gd-konu" style="font-weight:600;font-size:14px"></div>

        <div id="gd-kimden" style="font-size:12px;color:var(--t3);margin-top:3px"></div>

        <div id="gd-tarih" style="font-size:11px;color:var(--t3)"></div>

      </div>

      <div style="display:flex;gap:8px">

        <button class="btn bp bsm" id="gd-cevapla-btn" onclick="openMailCevapla()">↩ Cevapla</button>

        <button class="btn bo bsm" onclick="document.getElementById('gmail-detay').style.display='none'">✕ Kapat</button>

      </div>

    </div>

    <div id="gd-body" style="font-size:13px;color:var(--t2);white-space:pre-wrap;max-height:400px;overflow-y:auto;border-top:1px solid var(--b1);padding-top:12px;font-family:var(--mono)"></div>

  </div>

</div>



<!-- MAIL YAZ MODAL -->

<div class="ov" id="mail-mod">

  <div class="mod" style="width:600px;max-width:95vw">

    <div class="mtit"><span id="mail-mtit">Yeni Mail</span><button class="mclose" onclick="closeMailMod()">✕</button></div>

    <div class="fg">

      <div class="fgr ff"><label class="flb">Kime</label>

        <input class="fi2" type="email" id="mail-kime" placeholder="ornek@gmail.com"></div>

      <div class="fgr ff"><label class="flb">Konu</label>

        <input class="fi2" type="text" id="mail-konu" placeholder="Mail konusu"></div>

      <div class="fgr ff"><label class="flb">Mesaj</label>

        <textarea class="fi2" id="mail-body" rows="10" style="resize:vertical;min-height:180px;font-family:var(--mono);font-size:13px"

          placeholder="Mesajınızı buraya yazın..."></textarea></div>

    </div>

    <div class="mac">

      <button class="btn bo" onclick="closeMailMod()">İptal</button>

      <button class="btn bp" onclick="mailGonder()">📤 Gönder</button>

    </div>

  </div>

</div>



<!-- NOTLAR -->

<div class="view" id="vw-notlar">

  <div class="ptit">Notlar</div>

  <div class="psub">Anlık notlar — tarih damgalı, aranabilir</div>



  <!-- Not ekle -->

  <div style="background:var(--s2);border:1px solid var(--b1);border-radius:var(--radl);padding:16px;margin-bottom:16px">

    <div style="display:flex;gap:10px;align-items:flex-start">

      <div style="flex:1">

        <input class="fi2" type="text" id="not-baslik" placeholder="Başlık (isteğe bağlı)"

          style="margin-bottom:8px;width:100%">

        <textarea class="fi2" id="not-metin" rows="4"

          style="resize:vertical;min-height:90px;width:100%;font-size:13px"

          placeholder="Notunuzu buraya yazın... (Ctrl+Enter ile kaydet)"

          onkeydown="if(event.ctrlKey&&event.key==='Enter')notEkle()"></textarea>

      </div>

      <div style="display:flex;flex-direction:column;gap:6px">

        <select class="fse" id="not-renk" style="width:100px">

          <option value="default">⬜ Normal</option>

          <option value="mavi">🟦 Mavi</option>

          <option value="yesil">🟩 Yeşil</option>

          <option value="sari">🟨 Sarı</option>

          <option value="kirmizi">🟥 Kırmızı</option>

        </select>

        <button class="btn bp" onclick="notEkle()" style="height:36px">+ Ekle</button>

      </div>

    </div>

  </div>



  <!-- Arama + filtre -->

  <div class="tbr" style="margin-bottom:12px">

    <input class="si" type="text" id="not-q" placeholder="🔍 Not ara..." oninput="renderNotlar()">

    <select class="fi" id="not-renk-f" onchange="renderNotlar()">

      <option value="">Tüm Renkler</option>

      <option value="default">Normal</option>

      <option value="mavi">Mavi</option>

      <option value="yesil">Yeşil</option>

      <option value="sari">Sarı</option>

      <option value="kirmizi">Kırmızı</option>

    </select>

    <div class="sp"></div>

    <span id="not-count" style="font-size:12px;color:var(--t3)"></span>

  </div>



  <!-- Not listesi -->

  <div id="not-listesi" style="display:grid;grid-template-columns:repeat(auto-fill,minmax(300px,1fr));gap:12px"></div>

</div>



<!-- NOT DÜZENLEME MODALI -->

<div class="ov" id="not-edit-mod">

  <div class="mod" style="width:560px;max-width:95vw">

    <div class="mtit">

      <span>Notu Düzenle</span>

      <button class="mclose" onclick="closeNotEditMod()">✕</button>

    </div>

    <div class="fg">

      <div class="fgr ff">

        <label class="flb">Başlık</label>

        <input class="fi2" type="text" id="not-edit-baslik" placeholder="Başlık (isteğe bağlı)">

      </div>

      <div class="fgr ff">

        <label class="flb">Not Metni</label>

        <textarea class="fi2" id="not-edit-metin" rows="8"

          style="resize:vertical;min-height:160px;font-size:13px"

          placeholder="Not metni..."

          onkeydown="if(event.ctrlKey&&event.key==='Enter')notEditKaydet()"></textarea>

      </div>

      <div class="fgr">

        <label class="flb">Renk</label>

        <select class="fse" id="not-edit-renk">

          <option value="default">⬜ Normal</option>

          <option value="mavi">🟦 Mavi</option>

          <option value="yesil">🟩 Yeşil</option>

          <option value="sari">🟨 Sarı</option>

          <option value="kirmizi">🟥 Kırmızı</option>

        </select>

      </div>

      <div id="not-edit-tarih" style="font-size:11px;color:var(--t3);font-family:var(--mono);padding:4px 0"></div>

    </div>

    <div class="mac">

      <button class="btn bo" onclick="closeNotEditMod()">İptal</button>

      <button class="btn bp" onclick="notEditKaydet()">✓ Kaydet <span style="opacity:.6;font-size:10px">(Ctrl+Enter)</span></button>

    </div>

  </div>

</div>



<!-- AI ASİSTAN -->

<div class="view" id="vw-ai">

  <div class="ptit">AI Asistan</div>

  <div class="psub">ChatGPT ile finansal danışman — nakit akış verilerinizi anlık olarak biliyor</div>



  <!-- API Ayarları -->

  <div id="ai-ayar-bar" style="display:none;background:var(--s2);border:1px solid var(--b2);border-radius:var(--radl);padding:16px;margin-bottom:14px">

    <div style="font-weight:600;font-size:13px;margin-bottom:12px">OpenAI API Ayarları</div>

    <div class="info-box" style="margin-bottom:12px">

      📌 API key almak için: <a href="https://platform.openai.com/api-keys" target="_blank" style="color:var(--acc)">platform.openai.com/api-keys</a>

      → <strong>Create new secret key</strong> → kopyalayıp aşağıya yapıştırın.<br>

      ChatGPT Plus üyeliği API erişimi <strong>sağlamaz</strong> — platform.openai.com'da ayrıca kredi yüklemeniz gerekir (çok ucuz: ~$0.01/soru).

    </div>

    <div class="fg">

      <div class="fgr ff"><label class="flb">API Key</label>

        <input class="fi2" type="password" id="ai-api-key" placeholder="sk-..."></div>

      <div class="fgr"><label class="flb">Model</label>

        <select class="fse" id="ai-model">

          <option value="gpt-4o-mini">GPT-4o Mini (Hızlı, ucuz — önerilen)</option>

          <option value="gpt-4o">GPT-4o (Güçlü, pahalı)</option>

          <option value="gpt-4-turbo">GPT-4 Turbo</option>

          <option value="gpt-3.5-turbo">GPT-3.5 Turbo (En ucuz)</option>

        </select></div>

    </div>

    <div style="display:flex;gap:8px;margin-top:12px">

      <button class="btn bp" onclick="aiAyarKaydet()">✓ Kaydet</button>

      <button class="btn bo" onclick="document.getElementById('ai-ayar-bar').style.display='none'">İptal</button>

    </div>

    <div id="ai-ayar-result" style="margin-top:8px;font-size:12px;display:none"></div>

  </div>



  <!-- Chat alanı -->

  <div style="display:flex;flex-direction:column;height:calc(100vh - 280px);min-height:400px">



    <!-- Hızlı sorular -->

    <div id="ai-hizli" style="display:flex;gap:6px;flex-wrap:wrap;margin-bottom:12px">

      <button class="btn bgh bsm" onclick="aiHizliSor('Bu ayki nakit durumumuzu özetler misin?')">📊 Aylık özet</button>

      <button class="btn bgh bsm" onclick="aiHizliSor('Ödenmemiş faturalar için ne yapmalıyım?')">⚠ Ödenmemiş</button>

      <button class="btn bgh bsm" onclick="aiHizliSor('Likidite riskimiz var mı? Değerlendirme yap.')">💧 Likidite riski</button>

      <button class="btn bgh bsm" onclick="aiHizliSor('Kredi limitlerimi verimli kullanıyor muyum?')">💳 Kredi analizi</button>

      <button class="btn bgh bsm" onclick="aiHizliSor('Nakit akışımı iyileştirmek için ne önerirsin?')">💡 Öneri</button>

      <div class="sp"></div>

      <button class="btn bo bsm" onclick="document.getElementById('ai-ayar-bar').style.display=document.getElementById('ai-ayar-bar').style.display===''?'none':''">⚙ API Ayarları</button>

      <button class="btn bgh bsm" onclick="aiTemizle()">🗑 Temizle</button>

    </div>



    <!-- Mesaj listesi -->

    <div id="ai-msgs" style="flex:1;overflow-y:auto;background:var(--s2);border:1px solid var(--b1);border-radius:var(--radl);padding:16px;display:flex;flex-direction:column;gap:12px;margin-bottom:12px">

      <div id="ai-welcome" style="text-align:center;padding:30px 20px">

        <div style="font-size:32px;margin-bottom:12px">🤖</div>

        <div style="font-size:14px;font-weight:600;margin-bottom:8px">Nakit Akış AI Asistanı</div>

        <div style="font-size:12px;color:var(--t3);max-width:400px;margin:0 auto">

          Finansal verilerinizi anlık olarak biliyor. Nakit akış, ödeme planı, likidite analizi

          ve finansman kararları konusunda size yardımcı olabilirim.

        </div>

        <div id="ai-conn-status" style="margin-top:16px;font-size:12px"></div>

      </div>

    </div>



    <!-- Mesaj yazma alanı -->

    <div style="display:flex;gap:8px;align-items:flex-end">

      <textarea id="ai-input" class="fi2"

        style="flex:1;resize:none;min-height:44px;max-height:120px;font-size:13px;padding:10px 12px;font-family:var(--ff)"

        placeholder="Sorunuzu yazın... (Enter ile gönder, Shift+Enter yeni satır)"

        onkeydown="if(event.key==='Enter'&&!event.shiftKey){event.preventDefault();aiGonder();}

                   this.style.height='auto';this.style.height=Math.min(this.scrollHeight,120)+'px';"></textarea>

      <button class="btn bp" id="ai-send-btn" onclick="aiGonder()"

        style="height:44px;padding:0 18px;white-space:nowrap">

        ➤ Gönder

      </button>

    </div>

  </div>

</div>



<!-- VERİTABANI YÖNETİMİ -->

<div class="view" id="vw-db">

  <div class="ptit">Veritabanı Yönetimi</div>

  <div class="psub">SQLite veritabanı — tüm veriler <code style="background:var(--s3);padding:2px 6px;border-radius:4px">nakit_akis.db</code> dosyasına kaydedilir</div>



  <!-- Durum kartları -->

  <div id="db-status-grid" style="display:grid;grid-template-columns:repeat(3,1fr);gap:10px;margin-bottom:16px"></div>



  <!-- Kontroller -->

  <div style="display:flex;gap:8px;margin-bottom:16px;flex-wrap:wrap">

    <button class="btn bp" onclick="dbSync()">🔄 Manuel Senkronize Et</button>

    <button class="btn bo bsm" onclick="dbDurumYukle()">↻ Yenile</button>

    <div class="sp"></div>

    <span id="db-sync-msg" style="font-size:12px;color:var(--t3);align-self:center"></span>

  </div>



  <!-- Tablo detayları -->

  <div style="display:grid;grid-template-columns:1fr 1fr;gap:10px;margin-bottom:20px" id="db-tablo-detay"></div>



  <!-- SQL Sorgu terminali -->

  <div style="background:var(--s2);border:1px solid var(--b1);border-radius:var(--radl);padding:16px">

    <div style="font-weight:600;font-size:13px;margin-bottom:10px">🔍 SQL Sorgu Terminali</div>

    <div style="font-size:11px;color:var(--t3);margin-bottom:10px">

      Sadece SELECT sorguları. Tablo adları: <code style="background:var(--s3);padding:1px 5px;border-radius:3px">giderler</code>

      <code style="background:var(--s3);padding:1px 5px;border-radius:3px">gelirler</code>

      <code style="background:var(--s3);padding:1px 5px;border-radius:3px">banka_bakiyeleri</code>

      <code style="background:var(--s3);padding:1px 5px;border-radius:3px">kredi_limitleri</code>

      <code style="background:var(--s3);padding:1px 5px;border-radius:3px">cari_hareketler</code>

      <code style="background:var(--s3);padding:1px 5px;border-radius:3px">notlar</code>

    </div>

    <!-- Hazır sorgular -->

    <div style="display:flex;gap:6px;flex-wrap:wrap;margin-bottom:10px">

      <button class="btn bgh bsm" onclick="dbHazirSorgu('SELECT firma, SUM(tutar) as toplam, COUNT(*) as adet FROM giderler GROUP BY firma ORDER BY toplam DESC')">Firma bazlı gider</button>

      <button class="btn bgh bsm" onclick="dbHazirSorgu('SELECT ay, SUM(tutar) as toplam FROM giderler WHERE durum != &quot;ÖDENDİ&quot; GROUP BY ay ORDER BY tarih')">Aylık bekleyen</button>

      <button class="btn bgh bsm" onclick="dbHazirSorgu('SELECT banka, firma, bakiye FROM banka_bakiyeleri ORDER BY bakiye DESC')">Banka bakiyeleri</button>

      <button class="btn bgh bsm" onclick="dbHazirSorgu('SELECT cari_firma, SUM(CASE WHEN tur=&quot;ALACAK&quot; THEN tutar ELSE -tutar END) as net FROM cari_hareketler GROUP BY cari_firma ORDER BY net DESC')">Cari net bakiye</button>

      <button class="btn bgh bsm" onclick="dbHazirSorgu('SELECT tablo, islem, kayit_sayisi, created_at FROM sync_log ORDER BY id DESC LIMIT 20')">Sync geçmişi</button>

    </div>

    <div style="display:flex;gap:8px;margin-bottom:10px">

      <textarea id="db-sql-input" class="fi2"

        style="flex:1;font-family:var(--mono);font-size:12px;min-height:60px;resize:vertical"

        placeholder="SELECT * FROM giderler WHERE durum = 'ÖDENMEDİ' LIMIT 20"

        onkeydown="if(event.ctrlKey&&event.key==='Enter')dbSorguCalistir()"></textarea>

      <button class="btn bp" onclick="dbSorguCalistir()" style="height:60px;align-self:flex-start">▶ Çalıştır<br><span style="font-size:10px;opacity:.7">Ctrl+Enter</span></button>

    </div>

    <!-- Sonuçlar -->

    <div id="db-sorgu-result" style="display:none">

      <div id="db-sorgu-info" style="font-size:11px;color:var(--t3);margin-bottom:8px"></div>

      <div class="tc" style="max-height:350px;overflow:auto">

        <table id="db-sorgu-table"></table>

      </div>

      <button class="btn bo bsm" style="margin-top:8px" onclick="dbSorguExport()">⬇ CSV İndir</button>

    </div>

    <div id="db-sorgu-hata" style="display:none;color:var(--r);font-size:12px;font-family:var(--mono);padding:8px;background:rgba(239,68,68,.08);border-radius:var(--rad)"></div>

  </div>

</div>



<!-- OTEL YÖNETİMİ -->

<div class="view" id="vw-otel">

  <div class="ptit">Otel Yönetimi</div>

  <div class="psub">Rezervasyonlar, doluluk takvimi ve gelir analizi</div>



  <!-- Otel seçici -->

  <div style="display:flex;gap:0;margin-bottom:14px;border:1px solid var(--b1);border-radius:var(--rad);overflow:hidden;width:fit-content">

    <button class="btn" id="otel-tab-1" onclick="otelTab('otel1')"

      style="border-radius:0;background:var(--acc);color:#fff;border:none;padding:7px 20px;font-size:13px">🏨 Otel 1</button>

    <button class="btn bgh" id="otel-tab-2" onclick="otelTab('otel2')"

      style="border-radius:0;border:none;padding:7px 20px;font-size:13px">🏨 Otel 2</button>

    <button class="btn bgh" onclick="showOtelAyarlar()"

      style="border-radius:0;border:none;padding:7px 14px;font-size:12px;border-left:1px solid var(--b1)">⚙</button>

  </div>



  <!-- Otel ayarları -->

  <div id="otel-ayarlar" style="display:none;background:var(--s2);border:1px solid var(--b2);border-radius:var(--radl);padding:16px;margin-bottom:14px">

    <div style="font-weight:600;font-size:13px;margin-bottom:12px">Otel Ayarları</div>

    <div style="display:grid;grid-template-columns:1fr 1fr;gap:16px">

      <div>

        <div style="font-size:12px;font-weight:600;margin-bottom:8px;color:var(--acc)">🏨 Otel 1</div>

        <div class="fg">

          <div class="fgr ff"><label class="flb">Otel Adı</label><input class="fi2" type="text" id="otel1-ad" placeholder="Otel adı"></div>

          <div class="fgr"><label class="flb">Oda Sayısı</label><input class="fi2" type="number" id="otel1-oda" placeholder="50"></div>

          <div class="fgr ff"><label class="flb">Konum</label><input class="fi2" type="text" id="otel1-konum" placeholder="Şehir / Bölge"></div>

          <div class="fgr ff"><label class="flb">PMS Yazılımı</label><input class="fi2" type="text" id="otel1-pms" placeholder="Opera, Protel, Excel..."></div>

        </div>

      </div>

      <div>

        <div style="font-size:12px;font-weight:600;margin-bottom:8px;color:var(--acc)">🏨 Otel 2</div>

        <div class="fg">

          <div class="fgr ff"><label class="flb">Otel Adı</label><input class="fi2" type="text" id="otel2-ad" placeholder="Otel adı"></div>

          <div class="fgr"><label class="flb">Oda Sayısı</label><input class="fi2" type="number" id="otel2-oda" placeholder="50"></div>

          <div class="fgr ff"><label class="flb">Konum</label><input class="fi2" type="text" id="otel2-konum" placeholder="Şehir / Bölge"></div>

          <div class="fgr ff"><label class="flb">PMS Yazılımı</label><input class="fi2" type="text" id="otel2-pms" placeholder="Opera, Protel, Excel..."></div>

        </div>

      </div>

    </div>

    <div class="info-box" style="margin-top:12px">

      📌 PMS yazılımının adını ve SQL bağlantı bilgilerini öğrenince otomatik senkronizasyon aktif edilecek.

    </div>

    <div style="display:flex;gap:8px;margin-top:12px">

      <button class="btn bp" onclick="otelAyarKaydet()">✓ Kaydet</button>

      <button class="btn bo" onclick="document.getElementById('otel-ayarlar').style.display='none'">İptal</button>

    </div>

  </div>



  <!-- Sekmeler -->

  <div style="display:flex;gap:0;border-bottom:1px solid var(--b1);margin-bottom:14px">

    <button class="btn bgh" id="ostab-ozet"    onclick="otelSekme('ozet')"    style="border-radius:0;border-bottom:2px solid var(--acc);color:var(--acc)">📊 Özet</button>

    <button class="btn bgh" id="ostab-takvim"  onclick="otelSekme('takvim')"  style="border-radius:0">📅 Doluluk Takvimi</button>

    <button class="btn bgh" id="ostab-odalar"  onclick="otelSekme('odalar')"  style="border-radius:0">🚪 Oda Bazlı</button>

    <button class="btn bgh" id="ostab-rezerv"  onclick="otelSekme('rezerv')"  style="border-radius:0">📋 Rezervasyonlar</button>

    <button class="btn bgh" id="ostab-gelir"   onclick="otelSekme('gelir')"   style="border-radius:0">💰 Gelir Analizi</button>

  </div>



  <!-- ÖZET SEKMESİ -->

  <div id="ostab-ozet-view">

    <div id="otel-mets" style="display:grid;grid-template-columns:repeat(4,1fr);gap:10px;margin-bottom:16px"></div>

    <div style="display:grid;grid-template-columns:1fr 1fr;gap:14px">

      <div id="otel-bugun-checkin-card" class="src-card" style="margin:0"></div>

      <div id="otel-bugun-checkout-card" class="src-card" style="margin:0"></div>

    </div>

    <div id="otel-kanal-chart" style="margin-top:14px"></div>

  </div>



  <!-- TAKVIM SEKMESİ -->

  <div id="ostab-takvim-view" style="display:none">

    <div id="otel-takvim-grid" style="overflow-x:auto"></div>

  </div>



  <!-- REZERVASYON SEKMESİ -->

  <div id="ostab-rezerv-view" style="display:none">

    <div class="tbr">

      <input class="si" type="text" id="otel-q" placeholder="🔍 Misafir adı, tel, rezervasyon no..." oninput="otelRezervYukle()">

      <select class="fi" id="otel-durum-f" onchange="otelRezervYukle()">

        <option value="">Tüm Durumlar</option>

        <option value="Rezerve">Rezerve</option>

        <option value="Check-in">Check-in</option>

        <option value="Check-out">Check-out</option>

        <option value="Konaklamakta">Konaklamakta</option>

        <option value="İptal">İptal</option>

      </select>

      <select class="fi" id="otel-kanal-f" onchange="otelRezervYukle()"><option value="">Tüm Kanallar</option></select>

      <input class="si" type="date" id="otel-bas-f" onchange="otelRezervYukle()" style="width:140px">

      <input class="si" type="date" id="otel-bit-f" onchange="otelRezervYukle()" style="width:140px">

      <div class="sp"></div>

      <button class="btn bam2 bsm" onclick="showOtelExcelImport()">⬆ Excel Import</button>

      <button class="btn bp bsm" onclick="openOtelRezervAdd()">+ Rezervasyon Ekle</button>

    </div>

    <!-- Excel import alanı -->

    <div id="otel-excel-import" style="display:none;background:var(--s2);border:1px solid var(--b2);border-radius:var(--rad);padding:12px;margin-bottom:10px">

      <div style="font-size:12px;color:var(--t2);margin-bottom:8px">

        Excel sütunları: <strong>Rezervasyon No, Misafir Adı, Giriş Tarihi, Çıkış Tarihi, Oda No, Kanal, Tutar, Durum, Telefon, Açıklama</strong>

      </div>

      <div style="display:flex;gap:8px;align-items:center">

        <input type="file" id="otel-excel-file" accept=".xlsx,.xls,.csv" style="font-size:12px;flex:1">

        <button class="btn bp bsm" onclick="otelExcelImport()">Yükle</button>

        <button class="btn bo bsm" onclick="document.getElementById('otel-excel-import').style.display='none'">İptal</button>

      </div>

      <div id="otel-import-result" style="margin-top:6px;font-size:12px;display:none"></div>

    </div>

    <div class="tc">

      <table>

        <thead><tr>

          <th style="width:110px">Rezerv. No</th>

          <th style="width:160px">Misafir</th>

          <th style="width:90px">Giriş</th>

          <th style="width:90px">Çıkış</th>

          <th style="width:50px">Gece</th>

          <th style="width:60px">Oda</th>

          <th style="width:80px">Kanal</th>

          <th class="num" style="width:110px">Tutar ₺</th>

          <th style="width:90px">Durum</th>

          <th style="width:80px"></th>

        </tr></thead>

        <tbody id="otel-tbody"></tbody>

        <tfoot id="otel-tfoot"></tfoot>

      </table>

      <div class="pgn" id="otel-pgn"></div>

    </div>

  </div>



  <!-- ODA BAZLI SEKMESİ -->

  <div id="ostab-odalar-view" style="display:none">

    <div class="tbr">

      <input class="si" type="text" id="oda-no-f" placeholder="🔍 Oda no ara..." oninput="otelOdalarRender()">

      <select class="fi" id="oda-durum-f" onchange="otelOdalarRender()">

        <option value="">Tüm Odalar</option>

        <option value="dolu">Şu an Dolu</option>

        <option value="bos">Şu an Boş</option>

      </select>

      <div class="sp"></div>

      <button class="btn bam2 bsm" onclick="otelOdaEkle()">+ Oda Ekle</button>

    </div>

    <div id="oda-grid" style="display:grid;grid-template-columns:repeat(auto-fill,minmax(220px,1fr));gap:10px;margin-top:10px"></div>

  </div>



  <!-- GELİR ANALİZİ SEKMESİ -->

  <div id="ostab-gelir-view" style="display:none">

    <div id="otel-gelir-content"></div>

  </div>

</div>



<!-- REZERVASYON MODAL -->

<div class="ov" id="otel-rezerv-mod">

  <div class="mod" style="width:600px;max-width:95vw">

    <div class="mtit"><span id="otel-mod-title">Yeni Rezervasyon</span><button class="mclose" onclick="closeOtelMod()">✕</button></div>

    <div class="fg">

      <div class="fgr"><label class="flb">Rezervasyon No</label>

        <input class="fi2" type="text" id="orm-no" placeholder="RES-2024-001"></div>

      <div class="fgr"><label class="flb">Misafir Adı Soyadı</label>

        <input class="fi2" type="text" id="orm-misafir" placeholder="Ad Soyad"></div>

      <div class="fgr"><label class="flb">Giriş Tarihi</label>

        <input class="fi2" type="date" id="orm-giris"></div>

      <div class="fgr"><label class="flb">Çıkış Tarihi</label>

        <input class="fi2" type="date" id="orm-cikis"></div>

      <div class="fgr"><label class="flb">Oda No / Tipi</label>

        <input class="fi2" type="text" id="orm-oda" placeholder="101 veya Standart"></div>

      <div class="fgr"><label class="flb">Kanal</label>

        <select class="fse" id="orm-kanal">

          <option>Direkt</option><option>Booking.com</option><option>Airbnb</option>

          <option>Expedia</option><option>Tur Operatörü</option><option>Acente</option>

          <option>Telefon</option><option>Walk-in</option><option>Diğer</option>

        </select></div>

      <div class="fgr"><label class="flb">Toplam Tutar ₺</label>

        <input class="fi2" type="number" id="orm-tutar" placeholder="0"></div>

      <div class="fgr"><label class="flb">Durum</label>

        <select class="fse" id="orm-durum">

          <option>Rezerve</option><option>Check-in</option><option>Konaklamakta</option>

          <option>Check-out</option><option>İptal</option><option>No-show</option>

        </select></div>

      <div class="fgr"><label class="flb">Telefon</label>

        <input class="fi2" type="tel" id="orm-tel" placeholder="+90 555 000 00 00"></div>

      <div class="fgr ff"><label class="flb">Açıklama / Not</label>

        <input class="fi2" type="text" id="orm-aciklama" placeholder="Özel istek, notlar..."></div>

    </div>

    <div class="mac">

      <button class="btn bo" onclick="closeOtelMod()">İptal</button>

      <button class="btn bp" onclick="otelRezervKaydet()">✓ Kaydet</button>

    </div>

  </div>

</div>



<!-- WHATSAPP -->

<div class="view" id="vw-wa">

  <div class="ptit">WhatsApp</div>

  <div class="psub">Hızlı mesaj şablonları ve kişi yönetimi</div>



  <!-- Sekmeler -->

  <div style="display:flex;gap:0;border-bottom:1px solid var(--b1);margin-bottom:14px">

    <button class="btn bgh" id="watab-gonder"  onclick="waTab('gonder')"  style="border-radius:0;border-bottom:2px solid var(--acc);color:var(--acc)">📤 Mesaj Gönder</button>

    <button class="btn bgh" id="watab-sablon"  onclick="waTab('sablon')"  style="border-radius:0">📝 Şablonlar</button>

    <button class="btn bgh" id="watab-liste"   onclick="waTab('liste')"   style="border-radius:0">👥 Kişi Listesi</button>

    <button class="btn bgh" id="watab-gelen"   onclick="waTab('gelen')"   style="border-radius:0">📥 Gelen Mesajlar</button>

    <button class="btn bgh" id="watab-baglanti" onclick="waTab('baglanti')" style="border-radius:0">⚙ Bağlantı</button>

  </div>



  <!-- MESAJ GÖNDER SEKMESİ -->

  <div id="watab-gonder-view">

    <div style="display:grid;grid-template-columns:1fr 1fr;gap:16px">

      <div>

        <div style="font-size:12px;font-weight:600;margin-bottom:10px;color:var(--t2)">📱 Mesaj Oluştur</div>

        <div class="fg">

          <div class="fgr ff"><label class="flb">Telefon Numarası</label>

            <div style="display:flex;gap:6px">

              <input class="fi2" type="tel" id="wa-tel" placeholder="+90 555 000 00 00" style="flex:1">

              <button class="btn bgh bsm" onclick="waKisiSec()" title="Kişi listesinden seç">👥</button>

            </div>

          </div>

          <div class="fgr ff"><label class="flb">Mesaj</label>

            <textarea class="fi2" id="wa-mesaj" rows="6" style="resize:vertical;min-height:120px;font-size:13px"

              placeholder="Mesajınızı buraya yazın..."></textarea>

          </div>

          <div class="fgr ff"><label class="flb">Şablon Kullan</label>

            <select class="fse" id="wa-sablon-sec" onchange="waSablonUygula()">

              <option value="">— Şablon seç —</option>

            </select>

          </div>

        </div>

        <div style="display:flex;gap:8px;margin-top:4px">

          <button class="btn bp" onclick="waGonder()" style="flex:1">

            <span style="font-size:16px">💬</span> WhatsApp'ta Aç

          </button>

        </div>

        <div class="info-box" style="margin-top:10px;font-size:11px">

          💡 Butona tıklayınca WhatsApp Web veya telefonunuzdaki uygulama açılır, mesaj hazır gelir, siz gönderirsiniz.

        </div>

      </div>

      <div>

        <div style="font-size:12px;font-weight:600;margin-bottom:10px;color:var(--t2)">⚡ Hızlı Şablonlar</div>

        <div id="wa-hizli-sablon" style="display:flex;flex-direction:column;gap:6px"></div>

      </div>

    </div>

  </div>



  <!-- ŞABLONLAR SEKMESİ -->

  <div id="watab-sablon-view" style="display:none">

    <div class="tbr"><div class="sp"></div>

      <button class="btn bp bsm" onclick="openWaSablonEkle()">+ Şablon Ekle</button>

    </div>

    <div id="wa-sablon-listesi" style="display:grid;grid-template-columns:repeat(auto-fill,minmax(280px,1fr));gap:10px"></div>

  </div>



  <!-- KİŞİ LİSTESİ SEKMESİ -->

  <div id="watab-liste-view" style="display:none">

    <div class="tbr">

      <input class="si" type="text" id="wa-kisi-q" placeholder="🔍 İsim veya numara ara..." oninput="waKisiListeRender()">

      <div class="sp"></div>

      <button class="btn bp bsm" onclick="openWaKisiEkle()">+ Kişi Ekle</button>

    </div>

    <div id="wa-kisi-listesi" style="display:grid;grid-template-columns:repeat(auto-fill,minmax(240px,1fr));gap:8px;margin-top:10px"></div>



  <!-- GELEN MESAJLAR SEKMESİ -->

  </div>

  <div id="watab-gelen-view" style="display:none">

    <div style="display:grid;grid-template-columns:300px 1fr;gap:0;height:calc(100vh - 185px);border:1px solid var(--b1);border-radius:var(--rad);overflow:hidden">

      <!-- Sol: Kalıcı sohbet listesi -->

      <div style="background:var(--s1);border-right:1px solid var(--b1);display:flex;flex-direction:column">

        <div style="padding:10px 12px;border-bottom:1px solid var(--b1);display:flex;gap:6px;align-items:center;background:var(--s2)">

          <span style="font-size:13px;font-weight:700;color:var(--t1)">💬 Mesajlar</span>

          <div class="sp"></div>

          <span id="wa-toplam-okunmadi" style="display:none;background:var(--r);color:#fff;border-radius:10px;font-size:10px;padding:1px 7px;font-weight:700"></span>

          <button class="btn bgh bsm" onclick="waSohbetListeYukle(true)" title="Yenile" style="font-size:14px">🔄</button>

        </div>

        <input class="si" id="wa-sohbet-ara" placeholder="🔍 Sohbet ara..." oninput="waSohbetFiltrele()" style="margin:8px;width:calc(100% - 16px);font-size:12px">

        <div id="wa-sohbet-listesi" style="flex:1;overflow-y:auto">

          <div style="font-size:12px;color:var(--t3);padding:20px;text-align:center">Yükleniyor...</div>

        </div>

      </div>

      <!-- Sağ: Mesaj ekranı -->

      <div style="background:var(--s2);display:flex;flex-direction:column">

        <div id="wa-mesaj-baslik" style="padding:11px 16px;border-bottom:1px solid var(--b1);display:flex;align-items:center;gap:10px;background:var(--s1)">

          <span style="font-size:22px">💬</span>

          <div>

            <div style="font-size:13px;font-weight:700;color:var(--t1)" id="wa-baslik-ad">← Sol taraftan sohbet seçin</div>

            <div style="font-size:10px;color:var(--t3)" id="wa-baslik-alt">Kalıcı mesaj geçmişi</div>

          </div>

          <div class="sp"></div>

          <button class="btn bgh bsm" id="wa-yenile-btn" onclick="waMesajlarYenile()" style="display:none" title="Mesajları güncelle">🔄</button>

          <button class="btn bgh bsm" id="wa-sil-btn" onclick="waSohbetiSil()" style="display:none;color:var(--r)" title="Bu sohbeti sil">🗑</button>

        </div>

        <div id="wa-mesaj-listesi" style="flex:1;overflow-y:auto;padding:16px;display:flex;flex-direction:column;gap:6px;background:var(--bg)">

          <div style="font-size:12px;color:var(--t3);text-align:center;margin-top:40px">Sol taraftan bir sohbet seçin</div>

        </div>

        <div style="padding:10px 12px;border-top:1px solid var(--b1);display:flex;gap:8px;background:var(--s1)">

          <textarea id="wa-yanit-mesaj" rows="2" class="fi2" style="flex:1;resize:none;font-size:13px;background:var(--s2)" placeholder="Mesaj yaz... (Enter = gönder, Shift+Enter = yeni satır)"

            onkeydown="if(event.key==='Enter'&&!event.shiftKey){event.preventDefault();waYanitGonder();}"></textarea>

          <button class="btn bp" onclick="waYanitGonder()" style="align-self:flex-end;padding:10px 18px;font-size:14px">➤</button>

        </div>

      </div>

    </div>

    <!-- Webhook URL bilgisi -->

    <div id="wa-webhook-bilgi" style="margin-top:8px;font-size:11px;color:var(--t3);font-family:var(--mono)"></div>

  </div>



  <!-- BAĞLANTI SEKMESİ -->

  <div id="watab-baglanti-view" style="display:none">

    <div style="max-width:560px">

      <div class="info-box" style="margin-bottom:16px;font-size:12px;line-height:1.7">

        <strong>📱 Whapi.Cloud ile Bağlantı Kurma</strong><br>

        1. <a href="https://whapi.cloud" target="_blank" style="color:var(--acc)">whapi.cloud</a> adresine git ve ücretsiz kayıt ol<br>

        2. Hesabına giriş yap → Dashboard → <strong>New Channel</strong> oluştur<br>

        3. QR kodu WhatsApp'tan tara (bağlı cihazlar gibi)<br>

        4. Channel sayfasında <strong>Token</strong>'ı kopyala ve aşağıya yapıştır<br>

        5. Kaydet — artık gelen/giden mesajları buradan yönetebilirsin

      </div>

      <div class="fg">

        <div class="fgr ff"><label class="flb">Whapi API Token</label>

          <input class="fi2" id="whapi-token" type="password" placeholder="Bearer token buraya..." style="font-family:var(--mono);font-size:12px">

        </div>

        <div class="fgr ff"><label class="flb">Bağlantı Durumu</label>

          <div id="whapi-durum" style="font-size:12px;padding:8px;background:var(--s3);border-radius:var(--rad);border:1px solid var(--b1);color:var(--t3)">

            Token girilmedi

          </div>

        </div>

      </div>

      <div style="display:flex;gap:8px;margin-top:8px">

        <button class="btn bp" onclick="whapiKaydet()">💾 Token Kaydet</button>

        <button class="btn bgh" onclick="whapiTest()">🔌 Bağlantıyı Test Et</button>

      </div>

    </div>

  </div>

  </div>

</div>



<!-- WA ŞABLON MODAL -->

<div class="ov" id="wa-sablon-mod">

  <div class="mod" style="width:520px">

    <div class="mtit"><span id="wa-sablon-mod-title">Yeni Şablon</span><button class="mclose" onclick="closeWaSablonMod()">✕</button></div>

    <div class="fg">

      <div class="fgr ff"><label class="flb">Şablon Adı</label>

        <input class="fi2" type="text" id="wasm-ad" placeholder="Rezervasyon Onayı"></div>

      <div class="fgr"><label class="flb">Kategori</label>

        <select class="fse" id="wasm-kategori">

          <option>Otel</option><option>Ödeme</option><option>Cari</option>

          <option>Genel</option><option>Hatırlatma</option>

        </select></div>

      <div class="fgr ff"><label class="flb">Mesaj İçeriği</label>

        <textarea class="fi2" id="wasm-metin" rows="6" style="resize:vertical;min-height:120px;font-size:13px"

          placeholder="Sayın {MISAFIR_ADI},&#10;Rezervasyonunuz onaylanmıştır.&#10;Giriş: {GIRIS}  Çıkış: {CIKIS}"></textarea>

      </div>

      <div style="font-size:11px;color:var(--t3);padding:4px 0">

        Değişkenler: {AD} {TARIH} {SAAT} {KONU} {TUTAR} {TELEFON}

      </div>

    </div>

    <div class="mac">

      <button class="btn bo" onclick="closeWaSablonMod()">İptal</button>

      <button class="btn bp" onclick="waSablonKaydet()">✓ Kaydet</button>

    </div>

  </div>

</div>



<!-- WA KİŞİ MODAL -->

<div class="ov" id="wa-kisi-mod">

  <div class="mod" style="width:440px">

    <div class="mtit"><span>Kişi Ekle / Düzenle</span><button class="mclose" onclick="closeWaKisiMod()">✕</button></div>

    <div class="fg">

      <div class="fgr ff"><label class="flb">Ad Soyad</label>

        <input class="fi2" type="text" id="wakm-ad" placeholder="Ad Soyad"></div>

      <div class="fgr ff"><label class="flb">Telefon (+90 ile)</label>

        <input class="fi2" type="tel" id="wakm-tel" placeholder="+90555000000"></div>

      <div class="fgr"><label class="flb">Grup</label>

        <select class="fse" id="wakm-grup">

          <option>Misafir</option><option>Cari</option><option>Tedarikçi</option>

          <option>Personel</option><option>Diğer</option>

        </select></div>

      <div class="fgr ff"><label class="flb">Not</label>

        <input class="fi2" type="text" id="wakm-not" placeholder="Notlar..."></div>

    </div>

    <div class="mac">

      <button class="btn bo" onclick="closeWaKisiMod()">İptal</button>

      <button class="btn bp" onclick="waKisiKaydet()">✓ Kaydet</button>

    </div>

  </div>

</div>



<!-- HATIRLATICI -->

<div class="view" id="vw-hat">

  <div class="ptit">Hatırlatıcı</div>

  <div class="psub">Görevler, randevular ve önemli tarihler</div>



  <!-- Ekle formu -->

  <div style="background:var(--s2);border:1px solid var(--b1);border-radius:var(--radl);padding:16px;margin-bottom:16px">

    <div style="display:grid;grid-template-columns:1fr 140px 100px auto auto;gap:8px;align-items:end;flex-wrap:wrap" id="hat-form-grid">

      <div>

        <label class="flb">Hatırlatma</label>

        <input class="fi2" type="text" id="hat-baslik" placeholder="Ne hatırlatayım?" onkeydown="if(event.key==='Enter')hatEkle()">

      </div>

      <div>

        <label class="flb">Tarih</label>

        <input class="fi2" type="date" id="hat-tarih">

      </div>

      <div>

        <label class="flb">Saat</label>

        <input class="fi2" type="time" id="hat-saat">

      </div>

      <div>

        <label class="flb">Öncelik</label>

        <select class="fse" id="hat-oncelik">

          <option value="normal">Normal</option>

          <option value="yuksek">🔴 Yüksek</option>

          <option value="dusuk">⬇ Düşük</option>

        </select>

      </div>

      <div style="padding-top:20px">

        <button class="btn bp" onclick="hatEkle()" style="height:36px;padding:0 16px">+ Ekle</button>

      </div>

    </div>

    <div style="margin-top:8px">

      <input class="fi2" type="text" id="hat-not" placeholder="Not / açıklama (isteğe bağlı)" style="width:100%">

    </div>

  </div>



  <!-- Filtreler ve özet -->

  <div class="tbr" style="margin-bottom:12px">

    <div style="display:flex;gap:0;border:1px solid var(--b1);border-radius:var(--rad);overflow:hidden">

      <button class="btn" id="hf-tumu"     onclick="hatFiltre('tumu')"      style="border-radius:0;background:var(--acc);color:#fff;border:none;padding:5px 12px;font-size:12px">Tümü</button>

      <button class="btn bgh" id="hf-bugun"    onclick="hatFiltre('bugun')"    style="border-radius:0;border:none;padding:5px 12px;font-size:12px">Bugün</button>

      <button class="btn bgh" id="hf-bekleyen" onclick="hatFiltre('bekleyen')" style="border-radius:0;border:none;padding:5px 12px;font-size:12px">Bekleyen</button>

      <button class="btn bgh" id="hf-tamamlandi" onclick="hatFiltre('tamamlandi')" style="border-radius:0;border:none;padding:5px 12px;font-size:12px">Tamamlanan</button>

    </div>

    <input class="si" type="text" id="hat-q" placeholder="🔍 Ara..." oninput="renderHatirlatma()" style="width:180px">

    <div class="sp"></div>

    <div id="hat-ozet" style="font-size:12px;color:var(--t3)"></div>

  </div>



  <!-- Liste -->

  <div id="hat-listesi"></div>

</div>



<!-- HATIRLATICI DÜZENLE MODAL -->

<div class="ov" id="hat-edit-mod">

  <div class="mod" style="width:500px">

    <div class="mtit"><span>Hatırlatmayı Düzenle</span><button class="mclose" onclick="closeHatMod()">✕</button></div>

    <div class="fg">

      <div class="fgr ff"><label class="flb">Başlık</label>

        <input class="fi2" type="text" id="hem-baslik"></div>

      <div class="fgr"><label class="flb">Tarih</label>

        <input class="fi2" type="date" id="hem-tarih"></div>

      <div class="fgr"><label class="flb">Saat</label>

        <input class="fi2" type="time" id="hem-saat"></div>

      <div class="fgr"><label class="flb">Öncelik</label>

        <select class="fse" id="hem-oncelik">

          <option value="normal">Normal</option>

          <option value="yuksek">🔴 Yüksek</option>

          <option value="dusuk">⬇ Düşük</option>

        </select></div>

      <div class="fgr ff"><label class="flb">Not</label>

        <input class="fi2" type="text" id="hem-not"></div>

    </div>

    <div class="mac">

      <button class="btn bo" onclick="closeHatMod()">İptal</button>

      <button class="btn bp" onclick="hatEditKaydet()">✓ Kaydet</button>

    </div>

  </div>

</div>



<!-- SATIŞ ANALİZİ -->

<div class="view" id="vw-satis">

  <div class="ptit">Satış Analizi</div>

  <div class="psub">Logo veritabanından ürün bazlı satış, alış-satış farkı ve stok analizleri</div>



  <!-- Filtreler -->

  <div style="background:var(--s2);border:1px solid var(--b1);border-radius:var(--radl);padding:12px;margin-bottom:14px">

    <div style="display:flex;gap:8px;flex-wrap:wrap;margin-bottom:8px">

      <input class="si" type="date" id="sa-bas" style="width:135px" title="Başlangıç">

      <input class="si" type="date" id="sa-bit" style="width:135px" title="Bitiş">

      <input class="si" type="text" id="sa-stok-q" placeholder="🔍 Ürün adı / kodu..." style="width:200px">

      <button class="btn bp" onclick="satisAnalizYukle()">📊 Analiz Et</button>

      <button class="btn bgh bsm" onclick="hizliTarih('sa','bu-ay')" style="font-size:10px;padding:3px 6px">Bu Ay</button>

      <button class="btn bgh bsm" onclick="hizliTarih('sa','bu-yil')" style="font-size:10px;padding:3px 6px">Bu Yıl</button>

      <button class="btn bgh bsm" onclick="hizliTarih('sa','son-3')" style="font-size:10px;padding:3px 6px">Son 3 Ay</button>

      <div class="sp"></div>

      <button class="btn bo bsm" onclick="satisExport()">⬇ CSV</button>

    </div>

    <div style="display:flex;gap:16px;flex-wrap:wrap">

      <div>

        <div style="font-size:10px;color:var(--t3);margin-bottom:4px">Firma:</div>

        <div id="sa-firma-secim" style="display:flex;gap:6px;flex-wrap:wrap"></div>

      </div>

      <div>

        <div style="font-size:10px;color:var(--t3);margin-bottom:4px">Yıl:</div>

        <div id="sa-yil-secim" style="display:flex;gap:6px;flex-wrap:wrap"></div>

      </div>

    </div>

  </div>



  <!-- Yükleniyor -->

  <div id="sa-loading" style="display:none;text-align:center;padding:40px;color:var(--t3)">

    ⏳ Logo veritabanı sorgulanıyor...

  </div>



  <!-- Sekmeler -->

  <div style="display:flex;gap:0;border-bottom:1px solid var(--b1);margin-bottom:14px" id="sa-tabs">

    <button class="btn bgh" id="satab-ozet"    onclick="saTab('ozet')"    style="border-radius:0;border-bottom:2px solid var(--acc);color:var(--acc)">📊 Genel Özet</button>

    <button class="btn bgh" id="satab-urun"    onclick="saTab('urun')"    style="border-radius:0">📦 Ürün Bazlı</button>

    <button class="btn bgh" id="satab-donem"   onclick="saTab('donem')"   style="border-radius:0">📅 Dönem Analizi</button>

    <button class="btn bgh" id="satab-marj"    onclick="saTab('marj')"    style="border-radius:0">💹 Alış-Satış Marjı</button>

    <button class="btn bgh" id="satab-stok"    onclick="saTab('stok')"    style="border-radius:0">⚠ Stok Durumu</button>

  </div>



  <!-- GENEL ÖZET -->

  <div id="satab-ozet-view">

    <div id="sa-mets" style="display:grid;grid-template-columns:repeat(4,1fr);gap:10px;margin-bottom:16px"></div>

    <div style="display:grid;grid-template-columns:1fr 1fr;gap:14px">

      <div id="sa-aylik-grafik" class="src-card" style="margin:0"></div>

      <div id="sa-firma-grafik" class="src-card" style="margin:0"></div>

    </div>

  </div>



  <!-- ÜRÜN BAZLI -->

  <div id="satab-urun-view" style="display:none">

    <div class="tbr" style="margin-bottom:10px">

      <input class="si" type="text" id="sa-urun-q" placeholder="🔍 Ürün filtrele..." oninput="saUrunRender()">

      <select class="fi" id="sa-urun-sirala" onchange="saUrunRender()">

        <option value="tutar">Tutara Göre ↓</option>

        <option value="adet">Adete Göre ↓</option>

        <option value="fatura">Fatura Sayısına Göre ↓</option>

      </select>

      <div class="sp"></div>

      <span id="sa-urun-count" style="font-size:12px;color:var(--t3)"></span>

    </div>

    <div class="tc">

      <table>

        <thead><tr>

          <th style="width:30px">#</th>

          <th style="width:120px">Stok Kodu</th>

          <th>Ürün Adı</th>

          <th style="width:80px">Grup</th>

          <th style="width:70px">Firma</th>

          <th class="num" style="width:90px">Toplam Adet</th>

          <th class="num" style="width:90px">Fatura Sayısı</th>

          <th class="num" style="width:130px">Toplam Tutar ₺</th>

          <th class="num" style="width:110px">Ort. Birim Fiyat ₺</th>

          <th class="num" style="width:80px">Pay %</th>

        </tr></thead>

        <tbody id="sa-urun-tbody"></tbody>

      </table>

      <div class="pgn" id="sa-urun-pgn"></div>

    </div>

  </div>



  <!-- DÖNEM ANALİZİ -->

  <div id="satab-donem-view" style="display:none">

    <div class="tbr" style="margin-bottom:10px">

      <select class="fi" id="sa-donem-tip" onchange="saDonemRender()">

        <option value="yillik">Yıllık</option>

        <option value="aylik" selected>Aylık</option>

        <option value="haftalik">Haftalık</option>

      </select>

      <select class="fi" id="sa-donem-yil-f" onchange="saDonemRender()">

        <option value="">Tüm Yıllar</option>

      </select>

    </div>

    <div id="sa-donem-grafik" style="margin-bottom:16px"></div>

    <div class="tc">

      <table>

        <thead><tr>

          <th style="width:80px">Dönem</th>

          <th class="num" style="width:110px">Toplam Adet</th>

          <th class="num" style="width:130px">Toplam Tutar ₺</th>

          <th class="num" style="width:130px">Ort. Fatura ₺</th>

          <th class="num" style="width:90px">Fatura Sayısı</th>

          <th class="num" style="width:90px">Ürün Çeşidi</th>

          <th style="width:120px">Değişim</th>

        </tr></thead>

        <tbody id="sa-donem-tbody"></tbody>

      </table>

    </div>

  </div>



  <!-- ALIŞ-SATIŞ MARJI -->

  <div id="satab-marj-view" style="display:none">

    <div class="info-box" style="margin-bottom:12px">

      💹 Alış ve satış fiyatları Logo'daki <strong>son alış/satış fiyatı</strong> alanlarından çekilmektedir.

      Fark = Satış Fiyatı − Alış Fiyatı · Marj = (Fark / Satış Fiyatı) × 100

    </div>

    <div class="tbr" style="margin-bottom:10px">

      <input class="si" type="text" id="sa-marj-q" placeholder="🔍 Ürün filtrele..." oninput="saMarjRender()">

      <select class="fi" id="sa-marj-sirala" onchange="saMarjRender()">

        <option value="marj_desc">Marj ↓ (Yüksekten)</option>

        <option value="marj_asc">Marj ↑ (Düşükten)</option>

        <option value="tutar">Tutara Göre</option>

        <option value="negatif">Sadece Negatif Marj</option>

      </select>

    </div>

    <div class="tc">

      <table>

        <thead><tr>

          <th style="width:120px">Stok Kodu</th>

          <th>Ürün Adı</th>

          <th style="width:70px">Firma</th>

          <th class="num" style="width:110px">Son Alış ₺</th>

          <th class="num" style="width:110px">Son Satış ₺</th>

          <th class="num" style="width:110px">Fark ₺</th>

          <th class="num" style="width:90px">Marj %</th>

          <th class="num" style="width:110px">Toplam Kâr Est. ₺</th>

        </tr></thead>

        <tbody id="sa-marj-tbody"></tbody>

      </table>

      <div class="pgn" id="sa-marj-pgn"></div>

    </div>

  </div>



  <!-- STOK DURUMU -->

  <div id="satab-stok-view" style="display:none">

    <div style="display:flex;gap:8px;margin-bottom:12px;flex-wrap:wrap">

      <button class="btn bgh" id="ss-tumu" onclick="saStokFiltre('tumu')" style="font-size:12px">Tümü</button>

      <button class="btn bgh" id="ss-negatif" onclick="saStokFiltre('negatif')" style="font-size:12px;color:var(--r)">⚠ Negatif Stok</button>

      <button class="btn bgh" id="ss-sifir" onclick="saStokFiltre('sifir')" style="font-size:12px;color:var(--am)">Sıfır Stok</button>

      <button class="btn bgh" id="ss-kritik" onclick="saStokFiltre('kritik')" style="font-size:12px;color:var(--am)">Kritik Seviye (&lt;10)</button>

      <div class="sp"></div>

      <span id="ss-count" style="font-size:12px;color:var(--t3);align-self:center"></span>

    </div>

    <div class="tc">

      <table>

        <thead><tr>

          <th style="width:120px">Stok Kodu</th>

          <th>Ürün Adı</th>

          <th style="width:80px">Grup</th>

          <th style="width:70px">Firma</th>

          <th class="num" style="width:110px">Stok Bakiye</th>

          <th class="num" style="width:110px">Son Alış ₺</th>

          <th class="num" style="width:110px">Son Satış ₺</th>

          <th style="width:80px">Durum</th>

        </tr></thead>

        <tbody id="sa-stok-tbody"></tbody>

      </table>

      <div class="pgn" id="sa-stok-pgn"></div>

    </div>

  </div>

</div>



<!-- EXPORT -->

<div class="view" id="vw-exp">

  <div class="ptit">Dışa Aktar</div>

  <div class="psub">Verilerinizi farklı formatlarda indirin</div>

  <div class="egrd">

    <div class="ec"><div class="eico">📊</div><div class="etit">Excel (.xlsx)</div>

      <div class="edsc">Renk kodlamalı, otomatik filtre ve kolon genişlikleri ile profesyonel format.</div>

      <div class="ebr">

        <button class="btn bp bsm" onclick="doExpExcel('all')">Tüm Veriler</button>

        <button class="btn bo bsm" onclick="doExpExcel('filtered')">Filtrelenmiş</button>

        <button class="btn bo bsm" onclick="doExpExcel('summary')">Aylık Özet</button>

      </div>

    </div>

    <div class="ec"><div class="eico">📄</div><div class="etit">CSV</div>

      <div class="edsc">Sade metin formatı. Diğer muhasebe yazılımlarına aktarım için uygundur.</div>

      <button class="btn bo bsm" onclick="doExpCSV()">CSV İndir</button>

    </div>

    <div class="ec" style="grid-column:1/-1"><div class="eico">🖨️</div><div class="etit">Yazdır / PDF</div>

      <div class="edsc">Tarayıcının yazdır penceresinden "PDF olarak kaydet" seçeneğini kullanın.</div>

      <button class="btn bo bsm" onclick="window.print()">Yazdır (Ctrl+P)</button>

    </div>

  </div>

</div>



<!-- BANKA KARTLARI VIEW -->

<div class="view" id="vw-banka-kart">

  <div class="ptit">🏦 Banka Kartları</div>

  <div class="psub">Bankalar ile ilişki yönetimi — limitler, iletişim ve ipotek takibi</div>



  <!-- Özet Kartlar -->

  <div class="mets" style="margin-bottom:16px" id="bk-ozet-mets">

    <div class="met"><div class="met-stripe" style="background:#1e3a5f"></div><div class="mlb">Toplam Banka</div><div class="mvl" id="bk-toplam-banka" style="color:#1e3a5f">0</div><div class="msb">banka</div></div>

    <div class="met"><div class="met-stripe" style="background:#0ea5e9"></div><div class="mlb">Toplam Limit</div><div class="mvl" id="bk-toplam-limit" style="color:#0ea5e9">₺0</div><div class="msb">toplam</div></div>

    <div class="met"><div class="met-stripe" style="background:#f59e0b"></div><div class="mlb">Kullanılan</div><div class="mvl" id="bk-toplam-kullanim" style="color:#f59e0b">₺0</div><div class="msb">kullanım</div></div>

    <div class="met"><div class="met-stripe" style="background:#10b981"></div><div class="mlb">Kullanılabilir</div><div class="mvl" id="bk-toplam-kalan" style="color:#10b981">₺0</div><div class="msb">kalan limit</div></div>

  </div>



  <!-- Araç Çubuğu -->

  <div class="tbr" style="margin-bottom:14px">

    <input class="si" type="text" id="bk-ara" placeholder="🔍 Banka ara..." oninput="bankaKartFiltrele()" style="width:200px">

    <button class="btn bp bsm" onclick="bankaKartEkleAc()">+ Yeni Banka Ekle</button>

  </div>



  <!-- Banka Kartları Listesi -->

  <div id="bk-liste" style="display:grid;gap:16px"></div>

</div>





<!-- İPOTEK TAKİBİ VIEW -->

<div class="view" id="vw-ipotek">

  <div class="ptit">🏠 İPOTEK TAKİBİ</div>

  <div class="psub">Bankalara verilen ipotekli taşınmazların takibi</div>

  <div class="mets" style="margin-bottom:16px">

    <div class="met"><div class="met-stripe" style="background:#6366f1"></div><div class="mlb">Toplam Taşınmaz</div><div class="mvl" id="ipt-toplam-sayi" style="color:#6366f1">0</div><div class="msb">kayıt</div></div>

    <div class="met"><div class="met-stripe" style="background:#0ea5e9"></div><div class="mlb">Toplam Ekspertiz</div><div class="mvl" id="ipt-toplam-exp" style="color:#0ea5e9">₺0</div><div class="msb">toplam değer</div></div>

    <div class="met"><div class="met-stripe" style="background:#f59e0b"></div><div class="mlb">Toplam İpotek</div><div class="mvl" id="ipt-toplam-limit" style="color:#f59e0b">₺0</div><div class="msb">toplam tutar</div></div>

    <div class="met"><div class="met-stripe" style="background:#10b981"></div><div class="mlb">Durum</div><div class="mvl" id="ipt-limit-fark" style="color:#10b981">₺0</div><div class="msb" id="ipt-limit-fark-lbl">ekspertiz - ipotek</div></div>

  </div>

  <div class="tbr" style="margin-bottom:12px;flex-wrap:wrap;gap:8px">

    <input class="si" type="text" id="ipt-ara" placeholder="🔍 Ara..." oninput="ipotekFiltrele()" style="width:180px">

    <select class="fi" id="ipt-banka-f" onchange="ipotekFiltrele()" style="width:150px"><option value="">Tüm Bankalar</option></select>

    <select class="fi" id="ipt-il-f" onchange="ipotekFiltrele()" style="width:120px"><option value="">Tüm İller</option></select>

    <select class="fi" id="ipt-sahip-f" onchange="ipotekFiltrele()" style="width:150px"><option value="">Tüm Sahipler</option></select>

    <button class="btn bp bsm" onclick="ipotekEkleAc()">+ Yeni Ekle</button>

    <button class="btn bo bsm" onclick="ipotekExcel()">⬇ Excel İndir</button>

    <button class="btn bsm" style="background:#6366f1;color:#fff;border:none" onclick="ipotekSablonIndir()">📋 Şablon</button>

    <label class="btn bsm" style="background:#0ea5e9;color:#fff;border:none;cursor:pointer">⬆ Excel Yükle<input type="file" id="ipt-excel-yukle" accept=".xlsx,.xls" style="display:none" onchange="ipotekExcelYukle(this)"></label>

  </div>

  <div style="overflow-x:auto;border-radius:10px;border:1px solid var(--b2)">

  <table style="width:100%;border-collapse:collapse;font-size:11px;min-width:1800px">

    <thead><tr style="background:#1e3a5f;color:#fff">

      <th style="padding:8px 6px;text-align:center">S.No</th>

      <th style="padding:8px 6px">Kod</th>

      <th style="padding:8px 6px">Sahibi</th>

      <th style="padding:8px 6px">İli</th>

      <th style="padding:8px 6px">İlçe</th>

      <th style="padding:8px 6px">Adres</th>

      <th style="padding:8px 6px;text-align:center">Ada</th>

      <th style="padding:8px 6px;text-align:center">Parsel</th>

      <th style="padding:8px 6px;text-align:center">B.B</th>

      <th style="padding:8px 6px;text-align:center">Kat</th>

      <th style="padding:8px 6px;text-align:center">Blok</th>

      <th style="padding:8px 6px">Nitelik</th>

      <th style="padding:8px 6px">İpotek Bankası</th>

      <th style="padding:8px 6px;text-align:right">Tutar (₺)</th>

      <th style="padding:8px 6px;text-align:center">İpotek Tarihi</th>

      <th style="padding:8px 6px;text-align:right">Ekspertiz (₺)</th>

      <th style="padding:8px 6px;text-align:center">Exp. Tarihi</th>

      <th style="padding:8px 6px">Taşınmaz Tipi</th>

      <th style="padding:8px 6px">DASK Poliçe</th>

      <th style="padding:8px 6px;text-align:center">DASK Vade</th>

      <th style="padding:8px 6px">Konut Poliçe</th>

      <th style="padding:8px 6px;text-align:center">Konut Vade</th>

      <th style="padding:8px 6px;text-align:center">İşlem</th>

    </tr></thead>

    <tbody id="ipt-tbody">

      <tr><td colspan="23" style="text-align:center;padding:30px;color:var(--t3)">Kayıt yok — Şablon indirip yükleyin veya + Yeni Ekle</td></tr>

    </tbody>

  </table>

  </div>

</div>



<!-- İPOTEK MODAL -->

<div class="ov" id="ipt-mod" style="display:none">

  <div class="mod" style="max-width:760px;width:96%;max-height:90vh;overflow-y:auto">

    <div class="mtit"><span id="ipt-mod-baslik">Yeni Taşınmaz</span><button class="mclose" onclick="ipotekKapat()">✕</button></div>

    <div style="display:grid;grid-template-columns:1fr 1fr 1fr;gap:10px;padding:16px">

      <div><label class="flb">S.No</label><input class="fi2" id="ipt-sno" placeholder="1"></div>

      <div><label class="flb">Kod</label><input class="fi2" id="ipt-kod" placeholder="IPT-001"></div>

      <div><label class="flb">Sahibi</label><input class="fi2" id="ipt-sahip" placeholder="Ad Soyad"></div>

      <div><label class="flb">İli *</label><input class="fi2" id="ipt-il" placeholder="İstanbul"></div>

      <div><label class="flb">İlçe</label><input class="fi2" id="ipt-ilce" placeholder="Şişli"></div>

      <div><label class="flb">Adres</label><input class="fi2" id="ipt-adres" placeholder="Mahalle, Cadde..."></div>

      <div><label class="flb">Ada</label><input class="fi2" id="ipt-ada" placeholder="123"></div>

      <div><label class="flb">Parsel</label><input class="fi2" id="ipt-parsel" placeholder="42"></div>

      <div><label class="flb">B.B</label><input class="fi2" id="ipt-bb" placeholder="5"></div>

      <div><label class="flb">Kat</label><input class="fi2" id="ipt-kat" placeholder="3"></div>

      <div><label class="flb">Blok</label><input class="fi2" id="ipt-blok" placeholder="A"></div>

      <div><label class="flb">Nitelik</label><input class="fi2" id="ipt-nitelik" placeholder="Mesken..."></div>

      <div><label class="flb">İpotek Bankası</label><input class="fi2" id="ipt-banka" placeholder="Akbank"></div>

      <div><label class="flb">İpotek Tutarı (₺)</label><input class="fi2" type="number" id="ipt-tutar" placeholder="0" oninput="ipotekOtoCap()"></div>

      <div><label class="flb">İpotek Tarihi</label><input class="fi2" type="date" id="ipt-ipotek-tarih"></div>

      <div><label class="flb">Ekspertiz Değeri (₺)</label><input class="fi2" type="number" id="ipt-exp-deger" placeholder="0" oninput="ipotekOtoCap()"></div>

      <div><label class="flb">Ekspertiz Tarihi</label><input class="fi2" type="date" id="ipt-exp-tarih"></div>

      <div><label class="flb">Taşınmaz Tipi</label><input class="fi2" id="ipt-tasınmaz-tip" placeholder="Kat Mülkiyeti..."></div>

      <div><label class="flb">DASK Poliçe No</label><input class="fi2" id="ipt-dask-no"></div>

      <div><label class="flb">DASK Vade</label><input class="fi2" id="ipt-dask-vade" placeholder="01.01.2025"></div>

      <div></div>

      <div><label class="flb">Konut Poliçe No</label><input class="fi2" id="ipt-konut-no"></div>

      <div><label class="flb">Konut Vade</label><input class="fi2" id="ipt-konut-vade" placeholder="01.01.2025"></div>

      <div style="grid-column:1/-1;background:#f0fdf4;border-radius:8px;padding:10px;display:flex;gap:20px;align-items:center">

        <span style="font-size:12px;color:#6b7280">Fark:</span>

        <strong id="ipt-fark-preview" style="color:#10b981">₺0</strong>

        <strong id="ipt-durum-preview" style="font-size:12px">—</strong>

      </div>

    </div>

    <div style="padding:0 16px 16px;display:flex;gap:10px">

      <button class="btn bp" onclick="ipotekKaydet()" style="flex:1">💾 Kaydet</button>

      <button class="btn bo" onclick="ipotekKapat()">İptal</button>

    </div>

  </div>

</div>



<!-- KULLANICILAR VIEW -->

<div class="view" id="vw-kullanicilar">

  <div class="ptit">👤 Kullanıcı Yönetimi</div>

  <div class="psub">Sisteme erişebilecek kullanıcıları yönetin (sadece admin)</div>

  <div style="max-width:600px">

    <div style="background:#f9fafb;border-radius:12px;padding:20px;margin-bottom:20px;border:1px solid #e5e7eb">

      <div style="font-weight:600;margin-bottom:14px;font-size:14px">Yeni Kullanıcı Ekle</div>

      <div style="display:grid;grid-template-columns:1fr 1fr;gap:10px;margin-bottom:10px">

        <div><label style="font-size:12px;color:#6b7280">Kullanıcı Adı *</label><input id="yeni-uname" type="text" placeholder="kullanici" style="width:100%;padding:8px 10px;border:1.5px solid #d1d5db;border-radius:7px;font-size:13px;margin-top:4px"></div>

        <div><label style="font-size:12px;color:#6b7280">Ad Soyad</label><input id="yeni-ad" type="text" placeholder="Ali Yilmaz" style="width:100%;padding:8px 10px;border:1.5px solid #d1d5db;border-radius:7px;font-size:13px;margin-top:4px"></div>

        <div><label style="font-size:12px;color:#6b7280">Şifre *</label><input id="yeni-pw" type="password" placeholder="min. 6 karakter" style="width:100%;padding:8px 10px;border:1.5px solid #d1d5db;border-radius:7px;font-size:13px;margin-top:4px"></div>

        <div><label style="font-size:12px;color:#6b7280">Rol</label><select id="yeni-rol" style="width:100%;padding:8px 10px;border:1.5px solid #d1d5db;border-radius:7px;font-size:13px;margin-top:4px"><option value="user">Kullanıcı</option><option value="admin">Admin</option></select></div>

      </div>

      <button onclick="kullaniciEkle()" style="background:#1e3a5f;color:#fff;border:none;padding:9px 20px;border-radius:8px;cursor:pointer;font-size:13px;font-weight:600">+ Kullanıcı Ekle</button>

    </div>

    <div style="font-weight:600;margin-bottom:10px;font-size:14px">Mevcut Kullanıcılar</div>

    <div id="kullanici-liste"><p style="color:#9ca3af">Yükleniyor...</p></div>

  </div>

</div>



<!-- ŞİFRE DEĞİŞTİR VIEW -->

<div class="view" id="vw-sifre">

  <div class="ptit">🔑 ŞİFRE DEĞİŞTİR</div>

  <div class="psub">Kendi hesabınızın şifresini değiştirin</div>

  <div style="max-width:360px;background:#f9fafb;border-radius:12px;padding:24px;border:1px solid #e5e7eb">

    <div style="margin-bottom:14px"><label style="font-size:13px;font-weight:600">Yeni Şifre</label><input id="yeni-sifre" type="password" placeholder="min. 6 karakter" style="width:100%;padding:10px 12px;border:1.5px solid #d1d5db;border-radius:8px;font-size:14px;margin-top:6px"></div>

    <div style="margin-bottom:18px"><label style="font-size:13px;font-weight:600">Tekrar</label><input id="yeni-sifre2" type="password" placeholder="tekrar girin" style="width:100%;padding:10px 12px;border:1.5px solid #d1d5db;border-radius:8px;font-size:14px;margin-top:6px"></div>

    <button onclick="sifreDegistir()" style="width:100%;background:#1e3a5f;color:#fff;border:none;padding:11px;border-radius:8px;cursor:pointer;font-size:14px;font-weight:600">Şifreyi Kaydet</button>

  </div>

</div>



<!-- VERİ İÇE AKTAR VIEW -->

<div class="view" id="vw-import">

  <div class="ptit">⬆ VERİ İÇE AKTAR</div>

  <div class="psub">JSON veya Excel dosyalarınızı yükleyerek mevcut verilerinizi aktarın</div>

  <div style="max-width:640px;display:grid;gap:20px">

    <div style="background:#f9fafb;border:1px solid #e5e7eb;border-radius:12px;padding:22px">

      <div style="font-weight:700;font-size:15px;margin-bottom:6px;color:#1e3a5f">📂 Tüm Verileri Yükle (Toplu)</div>

      <div style="font-size:13px;color:#6b7280;margin-bottom:14px">Tüm JSON dosyalarınızı tek seferde seçip yükleyin.</div>

      <input type="file" id="imp-toplu-files" accept=".json" multiple style="width:100%;padding:8px;border:1.5px solid #d1d5db;border-radius:7px;font-size:13px;background:#fff;margin-bottom:12px">

      <div style="display:flex;gap:10px;align-items:center">

        <button onclick="topluYukle()" style="background:#1d4ed8;color:#fff;border:none;padding:9px 20px;border-radius:8px;cursor:pointer;font-size:13px;font-weight:600">⬆ Tümünü Yükle</button>

        <span id="imp-toplu-msg" style="font-size:13px;color:#6b7280"></span>

      </div>

      <div id="imp-toplu-detail" style="margin-top:10px"></div>

    </div>

    <div style="background:#f9fafb;border:1px solid #e5e7eb;border-radius:12px;padding:22px">

      <div style="font-weight:700;font-size:15px;margin-bottom:6px;color:#1e3a5f">📂 Tek JSON Yükle</div>

      <div style="display:grid;grid-template-columns:1fr 1fr;gap:10px;margin-bottom:12px">

        <select id="imp-json-tur" style="padding:8px 10px;border:1.5px solid #d1d5db;border-radius:7px;font-size:13px">

          <option value="gider">Gider (nakit_akis_data.json)</option>

          <option value="gelir">Gelir (nakit_akis_gelir.json)</option>

          <option value="banka">Banka (nakit_akis_banka.json)</option>

          <option value="kredi">Kredi (nakit_akis_kredi.json)</option>

          <option value="cari">Cari (nakit_akis_cari.json)</option>

          <option value="notlar">Notlar (nakit_akis_notlar.json)</option>

          <option value="otel">Otel (nakit_akis_otel.json)</option>

          <option value="hatirlatma">Hatırlatıcı</option>

        </select>

        <input type="file" id="imp-json-file" accept=".json" style="padding:6px;border:1.5px solid #d1d5db;border-radius:7px;font-size:13px;background:#fff">

      </div>

      <div style="display:flex;gap:10px;align-items:center">

        <button onclick="jsonYukle()" style="background:#1e3a5f;color:#fff;border:none;padding:9px 20px;border-radius:8px;cursor:pointer;font-size:13px;font-weight:600">⬆ JSON Yükle</button>

        <span id="imp-json-msg" style="font-size:13px;color:#6b7280"></span>

      </div>

    </div>

  </div>

</div>



</div>

</div>





<!-- BANKA KARTI MODAL -->

<div class="ov" id="bk-mod" style="display:none">

  <div class="mod" style="max-width:860px;width:96%;max-height:92vh;overflow-y:auto">

    <div class="mtit"><span id="bk-mod-baslik">Yeni Banka Kartı</span><button class="mclose" onclick="bankaKartKapat()">✕</button></div>

    <div style="padding:16px;display:grid;gap:14px">



      <!-- BANKA BİLGİLERİ -->

      <div style="background:var(--s2);border-radius:10px;padding:14px;border:1px solid var(--b2)">

        <div style="font-size:12px;font-weight:700;color:var(--acc);margin-bottom:12px;letter-spacing:.05em">🏦 BANKA BİLGİLERİ</div>

        <div style="display:grid;grid-template-columns:1fr 1fr 1fr;gap:10px">

          <div><label class="flb">Banka Adı *</label>

            <input class="fi2" type="text" id="bk-banka" list="bk-banka-list" placeholder="Banka adı yazın veya seçin..." autocomplete="off">
              <datalist id="bk-banka-list">
                <option>VAKIFBANK</option><option>ZİRAAT BANKASI</option><option>HALKBANK</option>
                <option>GARANTİ BBVA</option><option>İŞ BANKASI</option><option>YAPI KREDİ</option>
                <option>AKBANK</option><option>TEB</option><option>QNB FİNANSBANK</option>
                <option>DENİZBANK</option><option>ING BANK</option><option>HSBC</option>
                <option>BURGAN BANK</option><option>DİĞER</option>
              </datalist>

          </div>

          <div><label class="flb">Çalışılan Şube</label><input class="fi2" id="bk-sube" placeholder="Şişli Şubesi"></div>

          <div><label class="flb">Firma</label>

            <select class="fi2" id="bk-firma">

              <option>ULUSAL</option><option>ABC</option><option>BRG</option><option>BRK</option>

            </select>

          </div>

        </div>

      </div>



      <!-- HESAP NUMARALARI -->

      <div style="background:var(--s2);border-radius:10px;padding:14px;border:1px solid var(--b2)">

        <div style="font-size:12px;font-weight:700;color:var(--acc);margin-bottom:10px;letter-spacing:.05em">💳 HESAP NUMARALARI</div>

        <div id="bk-hesaplar" style="display:grid;gap:8px"></div>

        <button class="btn bgh bsm" onclick="bankaKartHesapEkle()" style="margin-top:8px;font-size:12px">+ Hesap Ekle</button>

      </div>



      <!-- İLETİŞİM -->

      <div style="display:grid;grid-template-columns:1fr 1fr;gap:12px">

        <!-- Şube Müdürü -->

        <div style="background:var(--s2);border-radius:10px;padding:14px;border:1px solid var(--b2)">

          <div style="font-size:12px;font-weight:700;color:var(--am);margin-bottom:10px">👤 ŞUBE MÜDÜRÜ</div>

          <div style="display:grid;gap:8px">

            <div><label class="flb">Ad Soyad</label><input class="fi2" id="bk-mudur-ad" placeholder="Ali Yılmaz"></div>

            <div><label class="flb">Telefon</label><input class="fi2" id="bk-mudur-tel" placeholder="0212 000 00 00"></div>

            <div><label class="flb">E-posta</label><input class="fi2" id="bk-mudur-mail" placeholder="ali@banka.com.tr"></div>

          </div>

        </div>

        <!-- İlgili Personel -->

        <div style="background:var(--s2);border-radius:10px;padding:14px;border:1px solid var(--b2)">

          <div style="font-size:12px;font-weight:700;color:var(--g);margin-bottom:10px">👤 İLGİLİ PERSONEL</div>

          <div style="display:grid;gap:8px">

            <div><label class="flb">Ad Soyad</label><input class="fi2" id="bk-personel-ad" placeholder="Ayşe Kaya"></div>

            <div><label class="flb">Telefon</label><input class="fi2" id="bk-personel-tel" placeholder="0212 000 00 01"></div>

            <div><label class="flb">E-posta</label><input class="fi2" id="bk-personel-mail" placeholder="ayse@banka.com.tr"></div>

          </div>

        </div>

      </div>



      <!-- KREDİ LİMİTLERİ -->

      <div style="background:var(--s2);border-radius:10px;padding:14px;border:1px solid var(--b2)">

        <div style="font-size:12px;font-weight:700;color:var(--acc);margin-bottom:12px;letter-spacing:.05em">💰 KREDİ LİMİTLERİ</div>

        <div style="overflow-x:auto">

        <table style="width:100%;border-collapse:collapse;font-size:12px">

          <thead><tr style="background:var(--s3)">

            <th style="padding:7px 10px;text-align:left;color:var(--t2)">Kredi Türü</th>

            <th style="padding:7px 10px;text-align:right;color:var(--t2)">Toplam Limit ₺</th>

            <th style="padding:7px 10px;text-align:right;color:var(--t2)">Kullanım ₺</th>

            <th style="padding:7px 10px;text-align:right;color:var(--t2)">Kalan Limit ₺</th>

            <th style="padding:7px 10px;text-align:right;color:var(--t2)">Faiz % (Yıllık)</th>

          </tr></thead>

          <tbody>

            <tr><td style="padding:6px 10px;font-weight:600">Taksitli Kredi</td>

              <td><input class="fi2" type="number" id="bk-limit-taksitli" placeholder="0" style="text-align:right;font-family:var(--mono);font-size:11px" oninput="bkHesapla()"></td>

              <td><input class="fi2" type="number" id="bk-kullanim-taksitli" placeholder="0" style="text-align:right;font-family:var(--mono);font-size:11px" oninput="bkHesapla()"></td>

              <td style="text-align:right;font-family:var(--mono);font-size:12px;color:var(--g);padding:6px 10px;font-weight:600" id="bk-kalan-taksitli">₺0</td>

              <td><input class="fi2" type="number" id="bk-faiz-taksitli" placeholder="0.0" step="0.1" style="text-align:right;font-family:var(--mono);font-size:11px"></td>

            </tr>

            <tr style="background:var(--s3)"><td style="padding:6px 10px;font-weight:600">Rotatif Kredi</td>

              <td><input class="fi2" type="number" id="bk-limit-rotatif" placeholder="0" style="text-align:right;font-family:var(--mono);font-size:11px" oninput="bkHesapla()"></td>

              <td><input class="fi2" type="number" id="bk-kullanim-rotatif" placeholder="0" style="text-align:right;font-family:var(--mono);font-size:11px" oninput="bkHesapla()"></td>

              <td style="text-align:right;font-family:var(--mono);font-size:12px;color:var(--g);padding:6px 10px;font-weight:600" id="bk-kalan-rotatif">₺0</td>

              <td><input class="fi2" type="number" id="bk-faiz-rotatif" placeholder="0.0" step="0.1" style="text-align:right;font-family:var(--mono);font-size:11px"></td>

            </tr>

            <tr><td style="padding:6px 10px;font-weight:600">KMH</td>

              <td><input class="fi2" type="number" id="bk-limit-kmh" placeholder="0" style="text-align:right;font-family:var(--mono);font-size:11px" oninput="bkHesapla()"></td>

              <td><input class="fi2" type="number" id="bk-kullanim-kmh" placeholder="0" style="text-align:right;font-family:var(--mono);font-size:11px" oninput="bkHesapla()"></td>

              <td style="text-align:right;font-family:var(--mono);font-size:12px;color:var(--g);padding:6px 10px;font-weight:600" id="bk-kalan-kmh">₺0</td>

              <td><input class="fi2" type="number" id="bk-faiz-kmh" placeholder="0.0" step="0.1" style="text-align:right;font-family:var(--mono);font-size:11px"></td>

            </tr>

            <tr style="background:var(--s3)"><td style="padding:6px 10px;font-weight:600">Kredi Kartı</td>

              <td><input class="fi2" type="number" id="bk-limit-kredi_karti" placeholder="0" style="text-align:right;font-family:var(--mono);font-size:11px" oninput="bkHesapla()"></td>

              <td><input class="fi2" type="number" id="bk-kullanim-kredi_karti" placeholder="0" style="text-align:right;font-family:var(--mono);font-size:11px" oninput="bkHesapla()"></td>

              <td style="text-align:right;font-family:var(--mono);font-size:12px;color:var(--g);padding:6px 10px;font-weight:600" id="bk-kalan-kredi_karti">₺0</td>

              <td><input class="fi2" type="number" id="bk-faiz-kredi_karti" placeholder="0.0" step="0.1" style="text-align:right;font-family:var(--mono);font-size:11px"></td>

            </tr>

            <tr><td style="padding:6px 10px;font-weight:600">Teminat Mektubu</td>

              <td><input class="fi2" type="number" id="bk-limit-teminat" placeholder="0" style="text-align:right;font-family:var(--mono);font-size:11px" oninput="bkHesapla()"></td>

              <td><input class="fi2" type="number" id="bk-kullanim-teminat" placeholder="0" style="text-align:right;font-family:var(--mono);font-size:11px" oninput="bkHesapla()"></td>

              <td style="text-align:right;font-family:var(--mono);font-size:12px;color:var(--g);padding:6px 10px;font-weight:600" id="bk-kalan-teminat">₺0</td>

              <td><input class="fi2" type="number" id="bk-faiz-teminat" placeholder="0.0" step="0.1" style="text-align:right;font-family:var(--mono);font-size:11px"></td>

            </tr>

            <tr style="background:var(--s3)"><td style="padding:6px 10px;font-weight:600">Diğer</td>

              <td><input class="fi2" type="number" id="bk-limit-diger" placeholder="0" style="text-align:right;font-family:var(--mono);font-size:11px" oninput="bkHesapla()"></td>

              <td><input class="fi2" type="number" id="bk-kullanim-diger" placeholder="0" style="text-align:right;font-family:var(--mono);font-size:11px" oninput="bkHesapla()"></td>

              <td style="text-align:right;font-family:var(--mono);font-size:12px;color:var(--g);padding:6px 10px;font-weight:600" id="bk-kalan-diger">₺0</td>

              <td><input class="fi2" type="number" id="bk-faiz-diger" placeholder="0.0" step="0.1" style="text-align:right;font-family:var(--mono);font-size:11px"></td>

            </tr>

            <tr style="border-top:2px solid var(--acc);background:rgba(79,156,249,.06)">

              <td style="padding:8px 10px;font-weight:700;color:var(--acc)">TOPLAM</td>

              <td style="text-align:right;font-family:var(--mono);font-weight:700;color:var(--acc);padding:8px 10px" id="bk-toplam-limit-modal">₺0</td>

              <td style="text-align:right;font-family:var(--mono);font-weight:700;color:var(--am);padding:8px 10px" id="bk-toplam-kullanim-modal">₺0</td>

              <td style="text-align:right;font-family:var(--mono);font-weight:700;color:var(--g);padding:8px 10px" id="bk-toplam-kalan-modal">₺0</td>

              <td></td>

            </tr>

          </tbody>

        </table>

        </div>

      </div>



      <!-- NOTLAR -->

      <div><label class="flb">Notlar</label>

        <textarea class="fi2" id="bk-notlar" rows="2" placeholder="Özel notlar, hatırlatmalar..." style="width:100%;resize:vertical"></textarea>

      </div>

    </div>



    <div style="padding:0 16px 16px;display:flex;gap:10px">

      <button class="btn bp" onclick="bankaKartKaydet()" style="flex:1">💾 Kaydet</button>

      <button class="btn bo" onclick="bankaKartKapat()">İptal</button>

    </div>

  </div>

</div>

<!-- MODAL: Ekle/Düzenle -->

<div class="ov" id="mod">

  <div class="mod">

    <div class="mtit"><span id="mtit">Yeni Ödeme Kaydı</span><button class="mclose" onclick="closeMod()">✕</button></div>

    <div class="fg">

      <div class="fgr"><label class="flb">Grup Firması</label>

        <select class="fse" id="mfr"><option>ULUSAL</option><option>ABC</option><option>BRG</option><option>BRK</option></select></div>

      <div class="fgr"><label class="flb">Ödeme Türü</label>

        <select class="fse" id="mtu"><option>TAKSİTLİ KREDİ ÖDEMESİ</option><option>ÇEK ÖDEMESİ</option><option>MAAŞ</option><option>CARİ ÖDEME</option><option>DİĞER</option></select></div>

      <div class="fgr ff"><label class="flb">Ödeme Yapılacak Yer / Banka</label>

        <input class="fi2" type="text" id="mye" placeholder="YAPIKREDİ, VAKIFBANK..."></div>

      <div class="fgr"><label class="flb">Kredi No / Çek No</label>

        <input class="fi2" type="text" id="mkn" placeholder="Kredi veya Çek numarası"></div>

      <div class="fgr"><label class="flb">Vade / Ödeme Tarihi</label>

        <input class="fi2" type="date" id="mta"></div>

      <div class="fgr"><label class="flb">Durum</label>

        <select class="fse" id="mdu"><option>ÖDENMEDİ</option><option>ÖDENDİ</option></select></div>

      <div class="fgr"><label class="flb">Anapara (₺)</label>

        <input class="fi2" type="number" id="man" placeholder="0" oninput="autoCalc()"></div>

      <div class="fgr"><label class="flb">Faiz (₺)</label>

        <input class="fi2" type="number" id="mfa" placeholder="0" oninput="autoCalc()"></div>

      <div class="fgr"><label class="flb">BSMV (₺)</label>

        <input class="fi2" type="number" id="mbs" placeholder="0" oninput="autoCalc()"></div>

      <div class="fgr"><label class="flb">Ödeme Tutarı (₺)</label>

        <input class="fi2" type="number" id="mto" placeholder="Otomatik hesaplanır"></div>

      <div class="fgr ff"><label class="flb">Açıklama</label>

        <input class="fi2" type="text" id="mac" placeholder="İsteğe bağlı"></div>

    </div>

    <div class="mac">

      <button class="btn bo" onclick="closeMod()">İptal</button>

      <button class="btn bp" id="msb" onclick="saveRec()">Kaydet</button>

    </div>

  </div>

</div>



<!-- GELİR MODAL -->

<div class="ov" id="gelir-mod">

  <div class="mod">

    <div class="mtit"><span id="gelir-mtit">Yeni Gelir Kaydı</span><button class="mclose" onclick="closeGelirMod()">✕</button></div>

    <div class="fg">

      <div class="fgr"><label class="flb">Firma</label>

        <select class="fse" id="gm-firma"><option>ULUSAL</option><option>ABC</option><option>BRG</option><option>BRK</option></select></div>

      <div class="fgr"><label class="flb">Gelir Türü</label>

        <select class="fse" id="gm-tur">

          <option>FLO TAHSİLATI</option><option>MÜŞTERİ TAHSİLATI</option>

          <option>OTEL GELİRİ</option><option>KİRA GELİRİ</option><option>DİĞER GELİR</option>

        </select></div>

      <div class="fgr ff"><label class="flb">Açıklama / Müşteri</label>

        <input class="fi2" type="text" id="gm-aciklama" placeholder="Müşteri adı, fatura no..."></div>

      <div class="fgr"><label class="flb">Tarih</label>

        <input class="fi2" type="date" id="gm-tarih"></div>

      <div class="fgr"><label class="flb">Durum</label>

        <select class="fse" id="gm-durum"><option>BEKLENİYOR</option><option>TAHSİL EDİLDİ</option></select></div>

      <div class="fgr ff"><label class="flb">Tutar (₺)</label>

        <input class="fi2" type="number" id="gm-tutar" placeholder="0"></div>

    </div>

    <div class="mac">

      <button class="btn bo" onclick="closeGelirMod()">İptal</button>

      <button class="btn bp" id="gm-save" onclick="saveGelir()">Kaydet</button>

    </div>

  </div>

</div>



<div class="tc2" id="toasts"></div>



<script>

const AYLAR=['Ocak','Şubat','Mart','Nisan','Mayıs','Haziran','Temmuz','Ağustos','Eylül','Ekim','Kasım','Aralık'];

const FC={ULUSAL:'bbi',ABC:'bpui',BRG:'bami',BRK:'bgi'};

let allData=[],filtered=[],sCol='tarih',sDir=1,pg=1,PPG=25,editIdx=null;



function tarihToISO(t){

  // DD.MM.YYYY → YYYY-MM-DD  veya  zaten YYYY-MM-DD

  if(!t) return '';

  if(t.includes('.')){ const [d,m,y]=t.split('.'); return `${y}-${m.padStart(2,'0')}-${d.padStart(2,'0')}`; }

  return t;

}

let charts={mo:null,pie:null,st:null};

// Gelir state

let gelirData=[],gelirFiltered=[],gelirPg=1,gelirEditIdx=null;



async function api(url,opts={}){
  opts.credentials='include';
  // Token'ı her zaman header olarak gönder
  const cookieStr = document.cookie.split(';').map(c=>c.trim()).find(c=>c.startsWith('na_token='))||'';
  const tok = cookieStr ? cookieStr.substring('na_token='.length) : '';
  if(tok){
    if(!opts.headers) opts.headers = {};
    else if(typeof opts.headers === 'string') opts.headers = {'Content-Type': opts.headers};
    opts.headers['X-Token'] = tok;
  }
  let resp;
  try { resp = await fetch(url, opts); } 
  catch(e) { console.error('API fetch error:', e); throw e; }
  if(resp.status===401 && url !== '/api/login'){
    toast('Oturum süresi doldu, yeniden giriş yapılıyor...','e');
    setTimeout(()=>{ window.location.href='/login'; }, 2000);
  }

  return resp;

}



// ── STATS & CHARTS ────────────────────────────────────────────────────────────

async function loadStats(){

  const r=await api('/api/stats'); const d=await r.json();

  const fm=n=>{n=n||0;if(n>=1e6)return'₺'+(n/1e6).toFixed(2)+' M';if(n>=1e3)return'₺'+(n/1e3).toFixed(0)+'K';return'₺'+Math.round(n).toLocaleString('tr-TR');};

  document.getElementById('st-tot').textContent=fm(d.total);

  document.getElementById('st-tc').textContent=d.total_count+' kayıt';

  document.getElementById('st-paid').textContent=fm(d.odendi);

  document.getElementById('st-pc').textContent=d.odendi_count+' kayıt';

  document.getElementById('st-pend').textContent=fm(d.bekleyen);

  document.getElementById('st-bc').textContent=d.bekleyen_count+' kayıt';

  document.getElementById('st-soon').textContent=fm(d.soon);

  document.getElementById('st-sc').textContent=d.soon_count+' kayıt · 30 gün';

  document.getElementById('db-sub').textContent=d.total_count

    ?`${d.total_count} kayıt yüklü · ${new Date().toLocaleString('tr-TR')}`

    :'Veri yok — Kaynak dosya tanımlayın veya kayıt ekleyin';

  populateFilters(d.firmas,d.turler,d.aylar,d.yillar);

}





let _chartsLoaded = false;



function fmtM(n){ n=n||0; if(n>=1e6) return (n/1e6).toFixed(1)+'M ₺'; if(n>=1e3) return Math.round(n/1000)+'K ₺'; return Math.round(n)+' ₺'; }



function svgBar(elId, labels, datasets, legend){

  const el=document.getElementById(elId); if(!el) return;

  const W=el.clientWidth||460, H=160;

  const PL=44, PR=8, PT=12, PB=38;

  const cW=W-PL-PR, cH=H-PT-PB;

  if(!labels.length){ el.innerHTML='<div style="color:var(--t3);font-size:11px;text-align:center;padding:50px 0">Veri yok</div>'; return; }

  const allV=datasets.flatMap(d=>d.values);

  const maxV=Math.max(...allV,1);

  const gCount=labels.length, dsCount=datasets.length;

  const barW=Math.max(3, Math.floor(cW/(gCount*(dsCount+0.5)+0.5)));

  const gW=barW*dsCount+2;

  let s=`<svg width="${W}" height="${H}" xmlns="http://www.w3.org/2000/svg" style="display:block;overflow:visible">`;

  // Grid

  [0,.25,.5,.75,1].forEach(p=>{

    const y=PT+cH*(1-p);

    s+=`<line x1="${PL}" y1="${y.toFixed(1)}" x2="${W-PR}" y2="${y.toFixed(1)}" stroke="#1e2535" stroke-width="1"/>`;

    s+=`<text x="${PL-3}" y="${(y+4).toFixed(1)}" text-anchor="end" font-size="8" fill="#3d4f6b">${p?fmtM(maxV*p):''}</text>`;

  });

  // Bars

  labels.forEach((lbl,i)=>{

    const gx=PL+i*(gW+Math.max(4,barW));

    datasets.forEach((ds,di)=>{

      const v=ds.values[i]||0;

      const bH=Math.max(1,Math.round(v/maxV*cH));

      s+=`<rect x="${gx+di*barW}" y="${PT+cH-bH}" width="${barW-1}" height="${bH}" fill="${ds.color}" rx="1" opacity=".85"/>`;

    });

    const lx=gx+gW/2;

    const parts=lbl.split('|');

    s+=`<text x="${lx.toFixed(1)}" y="${H-PB+13}" text-anchor="middle" font-size="8" fill="#4a5568">${parts[0].slice(0,3)}</text>`;

    if(parts[1]) s+=`<text x="${lx.toFixed(1)}" y="${H-PB+22}" text-anchor="middle" font-size="7" fill="#3d4f6b">${parts[1]}</text>`;

  });

  // Legend

  if(legend){ let lx=PL; legend.forEach(l=>{ s+=`<rect x="${lx}" y="${H-10}" width="8" height="6" fill="${l.c}" rx="1"/><text x="${lx+11}" y="${H-4}" font-size="8" fill="#94a3b8">${l.n}</text>`; lx+=l.n.length*5.5+18; }); }

  s+='</svg>';

  el.innerHTML=s;

}



function svgPie(elId, labels, values, colors){

  const el=document.getElementById(elId); if(!el) return;

  const W=el.clientWidth||460, H=155;

  const cx=75, cy=H/2, r=62;

  const total=values.reduce((a,b)=>a+b,0);

  if(!total){ el.innerHTML='<div style="color:var(--t3);font-size:11px;text-align:center;padding:50px 0">Veri yok</div>'; return; }

  let s=`<svg width="${W}" height="${H}" xmlns="http://www.w3.org/2000/svg" style="display:block">`;

  let a=-Math.PI/2;

  values.forEach((v,i)=>{

    const sl=(v/total)*Math.PI*2;

    const x1=cx+r*Math.cos(a), y1=cy+r*Math.sin(a);

    const x2=cx+r*Math.cos(a+sl), y2=cy+r*Math.sin(a+sl);

    s+=`<path d="M${cx},${cy} L${x1.toFixed(1)},${y1.toFixed(1)} A${r},${r} 0 ${sl>Math.PI?1:0},1 ${x2.toFixed(1)},${y2.toFixed(1)} Z" fill="${colors[i%colors.length]}" stroke="#161b27" stroke-width="1.5"/>`;

    a+=sl;

  });

  labels.slice(0,7).forEach((l,i)=>{

    const ly=10+i*21; if(ly+10>H) return;

    const pct=Math.round(values[i]/total*100);

    const short=l.length>17?l.slice(0,15)+'…':l;

    s+=`<rect x="${cx*2+12}" y="${ly}" width="8" height="8" fill="${colors[i%colors.length]}" rx="1"/>`;

    s+=`<text x="${cx*2+24}" y="${ly+7}" font-size="8" fill="#94a3b8">${short} ${pct}%</text>`;

  });

  s+='</svg>';

  el.innerHTML=s;

}



function loadCharts(){

  if(!allData||allData.length===0){

    ['ch-mo','ch-pie','ch-st'].forEach(id=>{

      const el=document.getElementById(id);

      if(el) el.innerHTML='<div style="color:var(--t3);font-size:11px;padding:50px 0;text-align:center">Veri yok — Kaynak Dosya tanımlayın</div>';

    });

    _chartsLoaded=false; return;

  }

  const AY=['Ocak','Şubat','Mart','Nisan','Mayıs','Haziran','Temmuz','Ağustos','Eylül','Ekim','Kasım','Aralık'];

  const bm={}, bt={};

  allData.forEach(d=>{

    const yil=d.tarih?d.tarih.slice(6,10):'????', ay=d.ay||'';

    if(!ay) return;

    const k=`${ay}|${yil}`;

    if(!bm[k]) bm[k]={tot:0,paid:0,pend:0,yil,ay,idx:AY.indexOf(ay)};

    bm[k].tot+=d.tutar||0;

    if(d.durum==='ÖDENDİ') bm[k].paid+=d.tutar||0; else bm[k].pend+=d.tutar||0;

    bt[d.tur]=(bt[d.tur]||0)+(d.tutar||0);

  });

  const ms=Object.entries(bm).sort(([,a],[,b])=>a.yil!==b.yil?a.yil.localeCompare(b.yil):a.idx-b.idx);

  const lbls=ms.map(([k])=>k);

  svgBar('ch-mo', lbls, [{values:ms.map(([,v])=>Math.round(v.tot)), color:'#4f9cf9'}], null);

  const tk=Object.keys(bt).sort((a,b)=>bt[b]-bt[a]).slice(0,7);

  svgPie('ch-pie', tk, tk.map(k=>bt[k]), ['#3b82f6','#6366f1','#10b981','#f59e0b','#ef4444','#8b5cf6','#06b6d4']);

  svgBar('ch-st', lbls, [{values:ms.map(([,v])=>Math.round(v.paid)),color:'#10b981'},{values:ms.map(([,v])=>Math.round(v.pend)),color:'#ef4444'}], [{n:'Ödendi',c:'#10b981'},{n:'Bekleyen',c:'#ef4444'}]);

  _chartsLoaded=true;

}



// ── SOURCE STATUS ─────────────────────────────────────────────────────────────

async function loadSourceStatus(){

  const r=await api('/api/source/status'); const d=await r.json();

  const dot=document.getElementById('src-dot');

  const lbl=document.getElementById('src-label');

  const btn=document.getElementById('sync-btn');

  if(!d.file){

    dot.className='src-dot none'; lbl.textContent='Kaynak dosya tanımlanmadı'; btn.style.display='none';

  } else if(!d.exists){

    dot.className='src-dot err'; lbl.textContent='Dosya bulunamadı!'; btn.style.display='';

  } else {

    dot.className='src-dot '+(d.ok?'ok':'err');

    const fname=d.file.split('\\').pop().split('/').pop();

    lbl.textContent=fname+(d.msg?' — '+d.msg.split(' — ').pop():'');

    btn.style.display='';

  }

  // Ayarlar sayfası

  document.getElementById('ss-file').textContent=d.file||'Tanımlanmadı';

  document.getElementById('ss-sheet').textContent=d.sheet||'Otomatik';

  const msgEl = document.getElementById('ss-msg');

  const msgLines = (d.msg||'—').split('\n');

  msgEl.innerHTML = msgLines.map(l=>`<div>${l}</div>`).join('');

  msgEl.style.color=d.ok?'var(--g)':d.file?'var(--r)':'var(--t3)';

  document.getElementById('ss-auto').textContent=d.auto?'Açık (30 sn)':'Kapalı';

  if(d.file) document.getElementById('src-path').value=d.file;

  // src-sheet alanı kullanıcı tarafından doldurulur, otomatik set edilmez

  document.getElementById('src-auto').checked=d.auto;

}



async function browseFile(){

  const btn = document.getElementById('browse-btn');

  btn.textContent = '⏳ Açılıyor...';

  btn.disabled = true;

  try {

    const r = await api('/api/source/browse');

    const d = await r.json();

    if(d.ok && d.path){

      document.getElementById('src-path').value = d.path;

      document.getElementById('src-path').value = d.path;

      document.getElementById('src-sheet').value = '';  // Sayfa adını sıfırla, otomatik bulacak

      toast('Dosya seçildi: '+d.path.split('\\\\').pop().split('/').pop()+' · Sayfa adı boş bırakıldı (otomatik)','s');

    } else if(d.cancelled){

      // kullanıcı iptal etti, sessiz geç

    } else {

      toast('Dosya seçici açılamadı: '+(d.error||'Bilinmeyen hata'),'e');

    }

  } catch(e){ toast('Hata: '+e.message,'e'); }

  btn.textContent = '📂 Dosya Seç';

  btn.disabled = false;

}



async function saveSource(){

  const fp=document.getElementById('src-path').value.trim();

  const sh=document.getElementById('src-sheet').value.trim();

  const au=document.getElementById('src-auto').checked;

  const errEl=document.getElementById('src-err');

  errEl.style.display='none';

  if(!fp){ errEl.textContent='Lütfen dosya yolunu girin'; errEl.style.display='block'; return; }

  const btn=document.querySelector('#vw-src .btn.bp');

  btn.textContent='⏳ Yükleniyor...'; btn.disabled=true;

  const r=await api('/api/source/set',{method:'POST',headers:{'Content-Type':'application/json'},

    body:JSON.stringify({file:fp,sheet:sh,auto:au})});

  const d=await r.json();

  btn.textContent='✓ Kaydet ve Yükle'; btn.disabled=false;

  if(d.ok){

    toast(`✓ ${d.count} kayıt yüklendi!`,'s');

    loadSourceStatus(); reload();

  } else {

    const errMsg = d.msg || d.error || 'Bilinmeyen hata';

    const lines = errMsg.split('\n');

    errEl.innerHTML = lines.map(l=>'<div>'+l+'</div>').join('');

    if(errMsg.includes('Kayıt bulunamadı')){

      errEl.innerHTML += '<div style="margin-top:6px;color:var(--am)">👉 Sayfa Adı kutusunu <b>tamamen boş bırakıp</b> tekrar deneyin.</div>';

      document.getElementById('src-sheet').value='';

    }

    errEl.style.display='block';

    toast(lines[0]||'Hata','e');

  }

}



async function clearSource(){

  if(!confirm('Kaynak dosya tanımını kaldırmak istediğinizden emin misiniz? Mevcut veriler silinmez.'))return;

  await api('/api/source/set',{method:'POST',headers:{'Content-Type':'application/json'},

    body:JSON.stringify({file:'',auto:true})});

  document.getElementById('src-path').value='';

  document.getElementById('src-sheet').value='';

  loadSourceStatus();

  toast('Kaynak dosya kaldırıldı','i');

}



async function manualSync(){

  const btn=document.getElementById('sync-btn');

  const sb=document.getElementById('ss-sync-btn');

  if(btn) btn.textContent='⏳';

  const r=await api('/api/source/reload'); const d=await r.json();

  if(btn) btn.textContent='↻ Şimdi Senkronize Et';

  if(d.ok){

    toast(`✓ ${d.count} kayıt güncellendi`,'s');

    reload(); loadSourceStatus();

  } else {

    toast('Hata: '+d.msg,'e');

    loadSourceStatus();

  }

}



// Otomatik durum yenileme (15 sn'de bir sadece header'ı güncelle)

setInterval(()=>{ loadSourceStatus(); },15000);



// ── TABLE ─────────────────────────────────────────────────────────────────────

function populateFilters(firmas,turler,aylar,yillar){

  const fill=(id,vals,def)=>{

    const s=document.getElementById(id); const cur=s.value;

    s.innerHTML=`<option value="">${def}</option>`;

    vals.forEach(v=>{const o=document.createElement('option');o.value=v;o.text=v;s.appendChild(o);});

    s.value=cur;

  };

  fill('fyil', yillar||[], 'Tüm Yıllar');

  fill('fay',  aylar,      'Tüm Aylar');

  fill('ffr',  firmas,     'Tüm Firmalar');

  fill('ftr',  turler,     'Tüm Türler');

  // Ödeme yerleri - allData'dan doldur

  const yerSel = document.getElementById('fyer');

  if(yerSel && allData?.length){

    const yerler = [...new Set(allData.map(d=>d.yer||'').filter(Boolean))].sort((a,b)=>a.localeCompare(b,'tr'));

    const curYer = yerSel.value;

    yerSel.innerHTML = '<option value="">Tüm Ödeme Yerleri</option>' + yerler.map(y=>`<option${y===curYer?' selected':''}>${y}</option>`).join('');

  }

}



async function doFilter(){

  const q=(document.getElementById('fs').value||'').toLowerCase();

  const yil=document.getElementById('fyil').value;

  const ay=document.getElementById('fay').value;

  const fr=document.getElementById('ffr').value;

  const tr=document.getElementById('ftr').value;

  const du=document.getElementById('fdu').value;

  const tBas = document.getElementById('ftarih-bas')?.value||'';

  const tBit = document.getElementById('ftarih-bit')?.value||'';

  const params=new URLSearchParams();

  if(q)params.set('q',q);if(yil)params.set('yil',yil);if(ay)params.set('ay',ay);

  const yer=document.getElementById('fyer')?.value||'';

  if(fr)params.set('firma',fr);if(tr)params.set('tur',tr);if(du)params.set('durum',du);

  if(yer)params.set('yer',yer);

  if(tBas)params.set('tarih_bas',tBas);

  if(tBit)params.set('tarih_bit',tBit);

  const r=await api('/api/data?'+params); const d=await r.json();

  filtered=d.records;

  // _idx backend'den geliyor, allData'yı da güncelle ki referanslar tutarlı olsun

  if(!params.toString()) allData=d.records;

  doSort(); pg=1; renderTbl();

}



function srt(col){if(sCol===col)sDir*=-1;else{sCol=col;sDir=1;}doSort();renderTbl();}

function doSort(){

  filtered.sort((a,b)=>{

    let va=a[sCol],vb=b[sCol];

    if(sCol==='tarih'&&va&&vb){

      const pa=va.split('.'),pb=vb.split('.');

      return(new Date(`${pa[2]}-${pa[1]}-${pa[0]}`)-new Date(`${pb[2]}-${pb[1]}-${pb[0]}`))*sDir;

    }

    if(typeof va==='number')return(va-vb)*sDir;

    return String(va||'').localeCompare(String(vb||''),'tr')*sDir;

  });

}



function fmtN(n){return Math.round(n||0).toLocaleString('tr-TR');}



function renderTbl(){

  const start=(pg-1)*PPG,end=start+PPG;

  document.getElementById('tbody').innerHTML=filtered.slice(start,end).map(d=>{

    const idx = (d._idx !== undefined && d._idx >= 0) ? d._idx : allData.indexOf(d);

    const fc=FC[d.firma]||'bbi';

    return`<tr>

      <td><span class="bdg ${fc}">${d.firma}</span></td>

      <td class="clip" style="max-width:145px" title="${d.tur}"><span style="font-size:12px">${d.tur}</span></td>

      <td class="clip" style="max-width:150px" title="${d.yer||''}">${d.yer||'—'}</td>

      <td style="font-family:var(--mono);font-size:11px;color:var(--acc)">${d.kredi_no||'—'}</td>

      <td style="font-family:var(--mono);font-size:11.5px">${d.tarih||'—'}</td>

      <td style="font-size:12px">${d.ay||'—'}<span style="font-family:var(--mono);font-size:9px;color:var(--t3);display:block;line-height:1">${d.tarih?d.tarih.slice(6,10):''}</span></td>

      <td><span class="bdg ${d.durum==='ÖDENDİ'?'bgi':'bri'}" 

        ondblclick="toggleSt(${idx})"

        title="Çift tıkla: ÖDENDİ / ÖDENMEDİ değiştir"

        style="cursor:pointer;user-select:none">${d.durum||'—'}</span></td>

      <td class="num">${d.anapara?fmtN(d.anapara):'—'}</td>

      <td class="num">${d.faiz?fmtN(d.faiz):'—'}</td>

      <td class="num" style="font-weight:600">${fmtN(d.tutar)}</td>

      <td style="white-space:nowrap">

        <button class="btn bgh" onclick="openEdit(${idx})" title="Düzenle">✏</button>

        <button class="btn bgh" style="color:var(--r)" onclick="delRec(${idx})" title="Sil">✕</button>

      </td>

    </tr>`;

  }).join('');

  renderPgn();

}



function renderPgn(){

  const tot=Math.ceil(filtered.length/PPG)||1;

  let h=`<span class="pgi">${filtered.length} kayıt</span><div class="pgsp"></div>`;

  if(pg>1)h+=`<button class="pbn" onclick="goPg(${pg-1})">‹</button>`;

  for(let i=Math.max(1,pg-2);i<=Math.min(tot,pg+2);i++)

    h+=`<button class="pbn${i===pg?' on':''}" onclick="goPg(${i})">${i}</button>`;

  if(pg<tot)h+=`<button class="pbn" onclick="goPg(${pg+1})">›</button>`;

  document.getElementById('pgn').innerHTML=h;

}

function goPg(p){pg=p;renderTbl();}



async function toggleSt(idx){

  const r=await api('/api/record/toggle',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({idx})});

  const d=await r.json();

  if(d.ok){

    // allData ve filtered içindeki objeleri güncelle

    if(allData[idx]) allData[idx].durum=d.durum;

    filtered.forEach(r=>{ if(r._idx===idx) r.durum=d.durum; });

    renderTbl(); loadStats();

    const excelMsg = d.excel===true ? ' · Excel ✓' : (d.excel===false ? ' · Excel yazılamadı' : '');

    toast((d.durum==='ÖDENDİ'?'✓ Ödendi':'⚠ Ödenmedi')+excelMsg, d.durum==='ÖDENDİ'?'s':'i');

  } else {

    toast('Hata: durum değiştirilemedi (idx='+idx+')','e');

  }

}



async function delRec(idx){

  if(!confirm('Bu kaydı silmek istediğinizden emin misiniz?'))return;

  const r=await api('/api/record/delete',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({idx})});

  if((await r.json()).ok){toast('Kayıt silindi','i');reload();}

}



// ── MODAL ─────────────────────────────────────────────────────────────────────

function openAdd(){

  editIdx=null;

  document.getElementById('mtit').textContent='Yeni Ödeme Kaydı';

  document.getElementById('msb').textContent='Kaydet';

  ['mye','mkn','mta','man','mfa','mbs','mto','mac'].forEach(id=>document.getElementById(id).value='');

  document.getElementById('mfr').value='ULUSAL';

  document.getElementById('mtu').selectedIndex=0;

  document.getElementById('mdu').value='ÖDENMEDİ';

  document.getElementById('mod').classList.add('on');

}

function openEdit(idx){

  editIdx=idx; const d=allData[idx];

  document.getElementById('mtit').textContent='Kaydı Düzenle';

  document.getElementById('msb').textContent='Güncelle';

  document.getElementById('mfr').value=d.firma||'ULUSAL';

  document.getElementById('mtu').value=d.tur||'';

  document.getElementById('mye').value=d.yer||'';

  document.getElementById('mkn').value=d.kredi_no||'';

  if(d.tarih&&d.tarih.length===10){const p=d.tarih.split('.');document.getElementById('mta').value=`${p[2]}-${p[1]}-${p[0]}`;}

  else document.getElementById('mta').value='';

  document.getElementById('mdu').value=d.durum||'ÖDENMEDİ';

  document.getElementById('man').value=d.anapara||'';

  document.getElementById('mfa').value=d.faiz||'';

  document.getElementById('mbs').value=d.bsmv||'';

  document.getElementById('mto').value=d.tutar||'';

  document.getElementById('mac').value=d.aciklama||'';

  document.getElementById('mod').classList.add('on');

}

function closeMod(){document.getElementById('mod').classList.remove('on');}

function autoCalc(){

  const a=parseFloat(document.getElementById('man').value)||0;

  const f=parseFloat(document.getElementById('mfa').value)||0;

  const b=parseFloat(document.getElementById('mbs').value)||0;

  if(a||f||b)document.getElementById('mto').value=(a+f+b).toFixed(2);

}

async function saveRec(){

  const tv=document.getElementById('mta').value;

  let ts='',ay='';

  if(tv){const d=new Date(tv);ts=d.toLocaleDateString('tr-TR',{day:'2-digit',month:'2-digit',year:'numeric'});ay=AYLAR[d.getMonth()];}

  const an=parseFloat(document.getElementById('man').value)||0;

  const fa=parseFloat(document.getElementById('mfa').value)||0;

  const bs=parseFloat(document.getElementById('mbs').value)||0;

  let to=parseFloat(document.getElementById('mto').value)||0;

  if(!to)to=an+fa+bs;

  const rec={firma:document.getElementById('mfr').value,tur:document.getElementById('mtu').value,

    yer:document.getElementById('mye').value,kredi_no:document.getElementById('mkn').value,tarih:ts,ay,durum:document.getElementById('mdu').value,

    anapara:an,faiz:fa,bsmv:bs,tutar:to,aciklama:document.getElementById('mac').value};

  const url=editIdx!==null?'/api/record/edit':'/api/record/add';

  const body=editIdx!==null?JSON.stringify({idx:editIdx,record:rec}):JSON.stringify(rec);

  const r=await api(url,{method:'POST',headers:{'Content-Type':'application/json'},body});

  if((await r.json()).ok){

    toast(editIdx!==null?'Kayıt güncellendi':'Yeni kayıt eklendi','s');

    closeMod();reload();

  }

}



// ── MONTHLY ───────────────────────────────────────────────────────────────────

// ── MONTHLY ───────────────────────────────────────────────────────────────────

const AYLAR_S=['Ocak','Şubat','Mart','Nisan','Mayıs','Haziran','Temmuz','Ağustos','Eylül','Ekim','Kasım','Aralık'];

let moActiveKey = null; // Açık olan ay kartı



function getMoFilters(){

  return {

    yil:   document.getElementById('mo-yil')?.value   || '',

    firma: document.getElementById('mo-firma')?.value  || '',

    tur:   document.getElementById('mo-tur')?.value    || '',

    durum: document.getElementById('mo-durum')?.value  || '',

  };

}



function resetMoFilters(){

  ['mo-yil','mo-firma','mo-tur','mo-durum'].forEach(id=>{

    const el=document.getElementById(id); if(el) el.value='';

  });

  moActiveKey=null;

  renderMonthly();

}



async function renderMonthly(){

  const r=await api('/api/data'); const d=await r.json();

  const f=getMoFilters();



  // Filtre seçeneklerini doldur (ilk çalıştırmada)

  const fillSel=(id,vals,def)=>{

    const s=document.getElementById(id); if(!s) return;

    const cur=s.value;

    s.innerHTML=`<option value="">${def}</option>`;

    vals.forEach(v=>{const o=document.createElement('option');o.value=v;o.text=v;s.appendChild(o);});

    s.value=cur;

  };

  const allRecs=d.records;

  fillSel('mo-yil',[...new Set(allRecs.map(r=>r.tarih?.slice(6,10)).filter(Boolean))].sort(),'Tüm Yıllar');

  fillSel('mo-firma',[...new Set(allRecs.map(r=>r.firma).filter(Boolean))].sort(),'Tüm Firmalar');

  fillSel('mo-tur',[...new Set(allRecs.map(r=>r.tur).filter(Boolean))].sort(),'Tüm Türler');



  // Filtre uygula

  const recs=allRecs.filter(rec=>{

    const recYil=rec.tarih?.slice(6,10)||'';

    if(f.yil   && recYil!==f.yil)           return false;

    if(f.firma && rec.firma!==f.firma)       return false;

    if(f.tur   && rec.tur!==f.tur)           return false;

    if(f.durum && rec.durum!==f.durum)       return false;

    return true;

  });



  // Yıl+Ay gruplama

  const bm={};

  recs.forEach(rec=>{

    if(!rec.ay||!rec.tarih) return;

    const yil=rec.tarih.slice(6,10)||'????';

    const key=yil+'|'+rec.ay;

    if(!bm[key]) bm[key]={yil,ay:rec.ay,tot:0,paid:0,pend:0,count:0};

    bm[key].tot   += rec.tutar;

    bm[key].count += 1;

    if(rec.durum==='ÖDENDİ') bm[key].paid+=rec.tutar;

    else                      bm[key].pend+=rec.tutar;

  });



  const keys=Object.keys(bm).sort((a,b)=>{

    const [ya,ma]=a.split('|'); const [yb,mb]=b.split('|');

    return ya!==yb?ya.localeCompare(yb):AYLAR_S.indexOf(ma)-AYLAR_S.indexOf(mb);

  });



  const fm=n=>{n=n||0;if(n>=1e6)return(n/1e6).toFixed(2)+' M ₺';return Math.round(n).toLocaleString('tr-TR')+' ₺';};



  // Özet bar

  const totAll=recs.reduce((s,r)=>s+r.tutar,0);

  const paidAll=recs.filter(r=>r.durum==='ÖDENDİ').reduce((s,r)=>s+r.tutar,0);

  const pendAll=totAll-paidAll;

  const filterActive=f.yil||f.firma||f.tur||f.durum;

  const sb=document.getElementById('mo-summary-bar');

  if(sb){

    sb.style.display='flex';

    sb.innerHTML=`

      <span style="color:var(--t3)">Gösterilen:</span>

      <span style="color:var(--acc);font-weight:600">${fm(totAll)}</span>

      <span style="color:var(--t3)">·</span>

      <span style="color:var(--g)">Ödendi: ${fm(paidAll)}</span>

      <span style="color:var(--t3)">·</span>

      <span style="color:var(--r)">Bekleyen: ${fm(pendAll)}</span>

      <span style="color:var(--t3)">·</span>

      <span style="color:var(--t2)">${recs.length} kayıt · ${keys.length} dönem</span>

      ${filterActive?'<span style="color:var(--am);margin-left:4px">⚑ Filtre aktif</span>':''}

    `;

  }



  // Kartlar

  let html=''; let prevYil='';

  for(const key of keys){

    const g=bm[key];

    const isActive=key===moActiveKey;

    if(g.yil!==prevYil){

      const yilTot=keys.filter(k=>k.startsWith(g.yil+'|')).reduce((s,k)=>s+bm[k].tot,0);

      html+=`<div style="grid-column:1/-1;display:flex;align-items:baseline;gap:10px;padding:8px 0 4px;border-bottom:1px solid var(--b2);margin-bottom:2px">

        <span style="font-family:var(--mono);font-size:12px;color:var(--acc);font-weight:700">${g.yil}</span>

        <span style="font-family:var(--mono);font-size:10px;color:var(--t3)">Toplam: ${fm(yilTot)}</span>

      </div>`;

      prevYil=g.yil;

    }

    const pct=g.tot>0?Math.round(g.paid/g.tot*100):0;

    html+=`<div class="mca${isActive?' sel':''}" onclick="showMoDet('${g.ay}','${g.yil}','${key}')">

      <div class="mcn">${g.ay} <span style="color:var(--t3);font-weight:400">${g.count} kayıt</span></div>

      <div class="mct">${fm(g.tot)}</div>

      ${g.pend>0

        ? `<div class="mcp">Bekl: ${fm(g.pend)}</div>

           <div style="margin-top:5px;height:3px;background:var(--b1);border-radius:2px;overflow:hidden">

             <div style="height:100%;width:${pct}%;background:var(--g);border-radius:2px"></div>

           </div>`

        : '<div style="font-size:10px;color:var(--g);margin-top:3px;font-family:var(--mono)">✓ Tümü ödendi</div>'

      }

    </div>`;

  }



  if(!keys.length){

    html='<div style="grid-column:1/-1;text-align:center;padding:30px;color:var(--t3);font-size:13px">Filtrelerle eşleşen dönem bulunamadı</div>';

  }



  document.getElementById('mgrd').innerHTML=html;



  // Aktif kartı yeniden aç

  if(moActiveKey && bm[moActiveKey]){

    const [yil,ay]=moActiveKey.split('|');

    showMoDet(ay,yil,moActiveKey);

  } else {

    document.getElementById('mdet').innerHTML='';

  }

}



async function showMoDet(ay, yil, key){

  moActiveKey=key||null;

  // Kartları güncelle

  document.querySelectorAll('.mca').forEach(el=>el.classList.remove('sel'));

  if(key){

    const cards=document.querySelectorAll('.mca');

    // onclick içinde key geçiyor, data attribute'e bakalım

    cards.forEach(el=>{

      if(el.getAttribute('onclick')?.includes(`'${key}'`)) el.classList.add('sel');

    });

  }



  const f=getMoFilters();

  const params=new URLSearchParams();

  params.set('ay',ay);

  params.set('yil',yil);

  if(f.firma) params.set('firma',f.firma);

  if(f.tur)   params.set('tur',f.tur);

  if(f.durum) params.set('durum',f.durum);



  const r=await api('/api/data?'+params); const d=await r.json();

  let rows=d.records;



  const fm=n=>Math.round(n||0).toLocaleString('tr-TR');

  const fmM=n=>{n=n||0;if(n>=1e6)return(n/1e6).toFixed(2)+' M ₺';return fm(n)+' ₺';};



  // Hesaplar

  const tot=rows.reduce((s,d)=>s+d.tutar,0);

  const od=rows.filter(d=>d.durum==='ÖDENDİ').reduce((s,d)=>s+d.tutar,0);

  const pend=tot-od;

  const pct=tot>0?Math.round(od/tot*100):0;



  // Firma bazlı dağılım

  const byFirma={}, byTur={}, byDurum={'ÖDENDİ':0,'ÖDENMEDİ':0};

  rows.forEach(d=>{

    byFirma[d.firma]=(byFirma[d.firma]||0)+d.tutar;

    byTur[d.tur]=(byTur[d.tur]||0)+d.tutar;

    byDurum[d.durum]=(byDurum[d.durum]||0)+d.tutar;

  });



  // Tablo sıralama: tarih artan

  rows=[...rows].sort((a,b)=>{

    const pa=a.tarih?.split('.')||[], pb=b.tarih?.split('.')||[];

    if(pa.length===3&&pb.length===3){

      return new Date(`${pa[2]}-${pa[1]}-${pa[0]}`)-new Date(`${pb[2]}-${pb[1]}-${pb[0]}`);

    }

    return 0;

  });



  // Alt filtre seçenekleri (bu ayın içinde)

  const turleri=[...new Set(rows.map(r=>r.tur))];

  const firmalari=[...new Set(rows.map(r=>r.firma))];



  document.getElementById('mdet').innerHTML=`

    <hr class="dv">

    <div style="display:flex;align-items:center;justify-content:space-between;margin-bottom:12px;flex-wrap:wrap;gap:8px">

      <div style="font-weight:600;font-size:14px">${yil} — ${ay}

        <span style="font-size:11px;color:var(--t3);font-weight:400;margin-left:6px">${rows.length} kayıt</span>

      </div>

      <div style="display:flex;gap:7px;flex-wrap:wrap">

        <select class="fi" id="det-firma" onchange="filterDetTable()" style="font-size:11.5px">

          <option value="">Tüm Firmalar</option>

          ${firmalari.map(f=>`<option value="${f}">${f}</option>`).join('')}

        </select>

        <select class="fi" id="det-tur" onchange="filterDetTable()" style="font-size:11.5px">

          <option value="">Tüm Türler</option>

          ${turleri.map(t=>`<option value="${t}">${t.length>30?t.slice(0,30)+'…':t}</option>`).join('')}

        </select>

        <select class="fi" id="det-durum" onchange="filterDetTable()" style="font-size:11.5px">

          <option value="">Tüm Durumlar</option>

          <option value="ÖDENDİ">Ödendi</option>

          <option value="ÖDENMEDİ">Ödenmedi</option>

        </select>

        <button class="btn bo bsm" onclick="doExpExcelMoDetail('${ay}','${yil}')">⬇ Excel</button>

      </div>

    </div>



    <!-- Özet metrikler -->

    <div style="display:flex;gap:9px;margin-bottom:12px">

      <div class="met" style="flex:1;padding:11px"><div class="met-stripe" style="background:var(--g)"></div>

        <div class="mlb">Ödendi</div><div class="mvl" style="color:var(--g);font-size:15px">₺${fm(od)}</div>

        <div class="msb">${rows.filter(r=>r.durum==='ÖDENDİ').length} kayıt</div></div>

      <div class="met" style="flex:1;padding:11px"><div class="met-stripe" style="background:var(--r)"></div>

        <div class="mlb">Bekleyen</div><div class="mvl" style="color:var(--r);font-size:15px">₺${fm(pend)}</div>

        <div class="msb">${rows.filter(r=>r.durum!=='ÖDENDİ').length} kayıt</div></div>

      <div class="met" style="flex:1;padding:11px"><div class="met-stripe" style="background:var(--acc)"></div>

        <div class="mlb">Toplam</div><div class="mvl" style="color:var(--acc);font-size:15px">₺${fm(tot)}</div>

        <div class="msb">${pct}% ödendi</div></div>

    </div>



    <!-- Progress bar -->

    <div style="height:4px;background:var(--b1);border-radius:2px;margin-bottom:14px;overflow:hidden">

      <div style="height:100%;width:${pct}%;background:linear-gradient(90deg,var(--g-d),var(--g));border-radius:2px;transition:width .4s"></div>

    </div>



    <!-- Firma ve tür dağılımı -->

    <div style="display:grid;grid-template-columns:1fr 1fr;gap:10px;margin-bottom:12px">

      <div style="background:var(--s2);border:1px solid var(--b1);border-radius:var(--rad);padding:11px">

        <div style="font-family:var(--mono);font-size:9.5px;color:var(--t3);text-transform:uppercase;letter-spacing:.06em;margin-bottom:8px">Firma Dağılımı</div>

        ${Object.entries(byFirma).sort((a,b)=>b[1]-a[1]).map(([k,v])=>{

          const pctF=tot>0?Math.round(v/tot*100):0;

          const fc=({ULUSAL:'var(--acc)',ABC:'var(--acc2)',BRG:'var(--am)',BRK:'var(--g)'})[k]||'var(--t2)';

          return `<div style="display:flex;align-items:center;gap:7px;margin-bottom:6px">

            <span style="font-size:11px;font-weight:600;color:${fc};min-width:55px;font-family:var(--mono)">${k}</span>

            <div style="flex:1;height:3px;background:var(--b1);border-radius:2px;overflow:hidden">

              <div style="height:100%;width:${pctF}%;background:${fc};border-radius:2px"></div>

            </div>

            <span style="font-size:10.5px;font-family:var(--mono);color:var(--t2);min-width:70px;text-align:right">${fmM(v)}</span>

            <span style="font-size:10px;color:var(--t3);min-width:32px;text-align:right">${pctF}%</span>

          </div>`;

        }).join('')}

      </div>

      <div style="background:var(--s2);border:1px solid var(--b1);border-radius:var(--rad);padding:11px">

        <div style="font-family:var(--mono);font-size:9.5px;color:var(--t3);text-transform:uppercase;letter-spacing:.06em;margin-bottom:8px">Ödeme Türü Dağılımı</div>

        ${Object.entries(byTur).sort((a,b)=>b[1]-a[1]).map(([k,v])=>{

          const pctT=tot>0?Math.round(v/tot*100):0;

          return `<div style="display:flex;align-items:center;gap:7px;margin-bottom:6px">

            <span style="font-size:10px;color:var(--t2);min-width:90px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap" title="${k}">${k.length>14?k.slice(0,14)+'…':k}</span>

            <div style="flex:1;height:3px;background:var(--b1);border-radius:2px;overflow:hidden">

              <div style="height:100%;width:${pctT}%;background:var(--acc);border-radius:2px"></div>

            </div>

            <span style="font-size:10.5px;font-family:var(--mono);color:var(--t2);min-width:70px;text-align:right">${fmM(v)}</span>

            <span style="font-size:10px;color:var(--t3);min-width:32px;text-align:right">${pctT}%</span>

          </div>`;

        }).join('')}

      </div>

    </div>



    <!-- Detay tablosu -->

    <div class="tc" id="det-table-wrap">

      <table id="det-table">

        <thead><tr>

          <th>Firma</th><th>Tür</th><th>Ödeme Yeri</th><th>Tarih</th>

          <th>Durum</th><th class="num">Anapara ₺</th><th class="num">Faiz ₺</th><th class="num">Tutar ₺</th>

        </tr></thead>

        <tbody id="det-tbody">

          ${rows.map(d=>`<tr data-firma="${d.firma}" data-tur="${d.tur}" data-durum="${d.durum}">

            <td><span class="bdg ${FC[d.firma]||'bbi'}">${d.firma}</span></td>

            <td style="font-size:11.5px" class="clip" style="max-width:140px" title="${d.tur}">${d.tur}</td>

            <td style="font-size:11.5px" class="clip" title="${d.yer||''}">${d.yer||'—'}</td>

            <td style="font-family:var(--mono);font-size:11px">${d.tarih||'—'}</td>

            <td><span class="bdg ${d.durum==='ÖDENDİ'?'bgi':'bri'}">${d.durum}</span></td>

            <td class="num">${d.anapara?fmtN(d.anapara):'—'}</td>

            <td class="num">${d.faiz?fmtN(d.faiz):'—'}</td>

            <td class="num" style="font-weight:600">${fmtN(d.tutar)}</td>

          </tr>`).join('')}

        </tbody>

      </table>

      <div id="det-footer" style="padding:8px 13px;background:var(--s2);border-top:1px solid var(--b1);font-size:11.5px;font-family:var(--mono);color:var(--t3)">

        ${rows.length} kayıt · Toplam: <span style="color:var(--acc)">${fmtN(tot)} ₺</span>

      </div>

    </div>

  `;



  // Veriyi saklayalım ki filterDetTable kullanabilsin

  window._moDetRows = rows;

}



function filterDetTable(){

  const fr=document.getElementById('det-firma')?.value||'';

  const tr=document.getElementById('det-tur')?.value||'';

  const du=document.getElementById('det-durum')?.value||'';

  const tbody=document.getElementById('det-tbody');

  if(!tbody||!window._moDetRows) return;



  const filtered=window._moDetRows.filter(d=>

    (!fr||d.firma===fr)&&(!tr||d.tur===tr)&&(!du||d.durum===du)

  );

  const fmtN=n=>Math.round(n||0).toLocaleString('tr-TR');

  const tot=filtered.reduce((s,d)=>s+d.tutar,0);



  tbody.innerHTML=filtered.map(d=>`<tr>

    <td><span class="bdg ${FC[d.firma]||'bbi'}">${d.firma}</span></td>

    <td style="font-size:11.5px" class="clip" title="${d.tur}">${d.tur}</td>

    <td style="font-size:11.5px" class="clip" title="${d.yer||''}">${d.yer||'—'}</td>

    <td style="font-family:var(--mono);font-size:11px">${d.tarih||'—'}</td>

    <td><span class="bdg ${d.durum==='ÖDENDİ'?'bgi':'bri'}">${d.durum}</span></td>

    <td class="num">${d.anapara?fmtN(d.anapara):'—'}</td>

    <td class="num">${d.faiz?fmtN(d.faiz):'—'}</td>

    <td class="num" style="font-weight:600">${fmtN(d.tutar)}</td>

  </tr>`).join('');



  const footer=document.getElementById('det-footer');

  if(footer) footer.innerHTML=`${filtered.length} kayıt · Toplam: <span style="color:var(--acc)">${fmtN(tot)} ₺</span>`;

}



function doExpExcelMo(){

  const f=getMoFilters();

  const params=new URLSearchParams({mode:'filtered'});

  if(f.yil)   params.set('yil',f.yil);

  if(f.firma) params.set('firma',f.firma);

  if(f.tur)   params.set('tur',f.tur);

  if(f.durum) params.set('durum',f.durum);

  const a=document.createElement('a');a.href='/api/export/excel?'+params;a.download='aylik_ozet.xlsx';a.click();

  toast('Excel hazırlanıyor...','i');

}



function doExpExcelMoDetail(ay,yil){

  const params=new URLSearchParams({mode:'filtered'});

  params.set('ay',ay); params.set('yil',yil);

  const f=getMoFilters();

  if(f.firma) params.set('firma',f.firma);

  if(f.tur)   params.set('tur',f.tur);

  if(f.durum) params.set('durum',f.durum);

  const a=document.createElement('a');a.href='/api/export/excel?'+params;a.download=`${yil}_${ay}.xlsx`;a.click();

  toast('Excel hazırlanıyor...','i');

}



// ── EXPORT ────────────────────────────────────────────────────────────────────

function doExpExcel(mode){

  const params=new URLSearchParams({mode});

  const ay=document.getElementById('fay')?.value;

  const fr=document.getElementById('ffr')?.value;

  const du=document.getElementById('fdu')?.value;

  if(mode==='filtered'){if(ay)params.set('ay',ay);if(fr)params.set('firma',fr);if(du)params.set('durum',du);}

  const a=document.createElement('a');a.href='/api/export/excel?'+params;a.download='nakit_akis.xlsx';a.click();

  toast('Excel hazırlanıyor...','i');

}

function doExpCSV(){

  const a=document.createElement('a');a.href='/api/export/csv';a.download='nakit_akis.csv';a.click();

  toast('CSV hazırlanıyor...','i');

}



// ── TOAST ─────────────────────────────────────────────────────────────────────

function toast(msg,type='i'){

  const c=document.getElementById('toasts');

  const t=document.createElement('div'); t.className='tst '+type;

  t.innerHTML=`<span>${{s:'✓',e:'✕',i:'ℹ'}[type]||'ℹ'}</span><span>${msg}</span>`;

  c.appendChild(t);

  setTimeout(()=>{t.style.cssText='opacity:0;transform:translateX(10px);transition:all .25s';setTimeout(()=>t.remove(),260);},3200);

}



// ── NAV ───────────────────────────────────────────────────────────────────────

function showView(v){

  document.querySelectorAll('.view').forEach(el=>el.classList.remove('on'));

  document.querySelectorAll('.nav').forEach(el=>el.classList.remove('on'));

  const vwEl=document.getElementById('vw-'+v); if(vwEl) vwEl.classList.add('on');

  const n=document.getElementById('nv-'+v);if(n)n.classList.add('on');

  const mainEl=document.querySelector('.main'); if(mainEl){mainEl.scrollTop=0;}

  if(v==='ipotek'){renderIpotek();}

  if(v==='banka-kart'){renderBankaKart();}

  if(v==='kullanicilar'){setTimeout(loadKullanicilar,100);}

  if(v==='db'){loadStats(); _chartsLoaded=false; setTimeout(()=>{ loadCharts(); },50);}

  if(v==='tbl'){doFilter();}

  if(v==='gelir'){loadGelir();}

  if(v==='nakit'){renderNakit();}

  if(v==='finans'){renderFinans();}

  if(v==='liki'){renderLikidite();}

  if(v==='cari'){renderCari();}

  if(v==='banka-api'){renderBankaApi();}

  if(v==='kur'){renderKur();}

  if(v==='logo'){renderLogo();}

  if(v==='satis'){renderSatisAnaliz();}

  if(v==='stok'){renderStok();}

  if(v==='cari-logo'){renderCariLogo();}

  if(v==='gmail'){renderGmail();}

  if(v==='notlar'){renderNotlar();}

  if(v==='hat'){renderHatirlatma();}

  if(v==='ai'){renderAI();}

  if(v==='otel'){renderOtel();}

  if(v==='wa'){renderWA();}

  if(v==='db'){renderDB();}

  if(v==='mo'){renderMonthly();}

  if(v==='docs'){renderDocs();}

  if(v==='src'){loadSourceStatus();}

}



async function reload(){

  try {

    const [dr, sr] = await Promise.all([api('/api/data'), api('/api/stats')]);

    const d=await dr.json(); const s=await sr.json();

    allData=d.records; filtered=[...allData]; doSort(); renderTbl();

    const fm2=n=>{n=n||0;return'₺'+n.toLocaleString('tr-TR',{minimumFractionDigits:2,maximumFractionDigits:2});};

    document.getElementById('st-tot').textContent=fm2(s.total);

    document.getElementById('st-tc').textContent=s.total_count+' kayıt';

    document.getElementById('st-paid').textContent=fm2(s.odendi);

    document.getElementById('st-pc').textContent=s.odendi_count+' kayıt';

    document.getElementById('st-pend').textContent=fm2(s.bekleyen);

    document.getElementById('st-bc').textContent=s.bekleyen_count+' kayıt';

    document.getElementById('st-soon').textContent=fm2(s.soon);

    document.getElementById('st-sc').textContent=s.soon_count+' kayıt · 30 gün';

    document.getElementById('db-sub').textContent=s.total_count?`${s.total_count} kayıt yüklü · ${new Date().toLocaleString('tr-TR')}`:'Veri yok';

    populateFilters(s.firmas,s.turler,s.aylar,s.yillar);

    _chartsLoaded=false;

    if(document.getElementById('vw-db').classList.contains('on'))

      setTimeout(()=>{ loadCharts(); }, 80);

  } catch(e){ console.error('reload hatası:', e); }

}



// ── GELİR TABLOSU ─────────────────────────────────────────────────────────────

async function loadGelir(){

  const params=new URLSearchParams();

  const q=(document.getElementById('gfs')?.value||'').toLowerCase();

  const yil=document.getElementById('gfyil')?.value||'';

  const ay=document.getElementById('gfay')?.value||'';

  const fr=document.getElementById('gffr')?.value||'';

  const du=document.getElementById('gfdu')?.value||'';

  if(q) params.set('q',q);

  if(yil) params.set('yil',yil);

  if(ay) params.set('ay',ay);

  if(fr) params.set('firma',fr);

  if(du) params.set('durum',du);

  const r=await api('/api/gelir?'+params); const d=await r.json();

  gelirFiltered=d.records;

  // Fill filters

  const allR=await api('/api/gelir'); const allD=await allR.json();

  gelirData=allD.records;

  const fillSel=(id,vals,def)=>{

    const s=document.getElementById(id); if(!s) return; const cur=s.value;

    s.innerHTML=`<option value="">${def}</option>`;

    vals.forEach(v=>{const o=document.createElement('option');o.value=v;o.text=v;s.appendChild(o);});

    s.value=cur;

  };

  fillSel('gfyil',[...new Set(gelirData.map(r=>r.tarih?.slice(6,10)).filter(Boolean))].sort(),'Tüm Yıllar');

  fillSel('gfay',[...new Set(gelirData.map(r=>r.ay).filter(Boolean))].sort((a,b)=>AYLAR.indexOf(a)-AYLAR.indexOf(b)),'Tüm Aylar');

  fillSel('gffr',[...new Set(gelirData.map(r=>r.firma).filter(Boolean))].sort(),'Tüm Firmalar');

  renderGelirTbl();

}



function doGelirFilter(){ gelirPg=1; loadGelir(); }

function hizliTarih(sayfa, aralik){

  const now = new Date();

  let bas, bit;

  if(aralik==='bu-ay'){

    bas = new Date(now.getFullYear(), now.getMonth(), 1).toISOString().slice(0,10);

    bit = new Date(now.getFullYear(), now.getMonth()+1, 0).toISOString().slice(0,10);

  } else if(aralik==='bu-yil'){

    bas = `${now.getFullYear()}-01-01`;

    bit = `${now.getFullYear()}-12-31`;

  } else if(aralik==='son-3'){

    const d3 = new Date(now); d3.setMonth(d3.getMonth()-3);

    bas = d3.toISOString().slice(0,10);

    bit = now.toISOString().slice(0,10);

  }

  if(sayfa==='gider'){

    document.getElementById('ftarih-bas').value=bas;

    document.getElementById('ftarih-bit').value=bit;

    doFilter();

  } else if(sayfa==='gelir'){

    document.getElementById('gftarih-bas').value=bas;

    document.getElementById('gftarih-bit').value=bit;

    doGelirFilter();

  } else if(sayfa==='logo'){

    document.getElementById('logo-bas').value=bas;

    document.getElementById('logo-bit').value=bit;

    logoFaturaYukle();

  } else if(sayfa==='sa'){

    document.getElementById('sa-bas').value=bas;

    document.getElementById('sa-bit').value=bit;

    satisAnalizYukle();

  } else if(sayfa==='stok'){

    document.getElementById('stok-bas').value=bas;

    document.getElementById('stok-bit').value=bit;

    stokYukle();

  } else if(sayfa==='cari'){

    document.getElementById('cari-bas').value=bas;

    document.getElementById('cari-bit').value=bit;

    cariLogoYukle();

  }

}



function gelirTarihSifirla(){ document.getElementById('gftarih-bas').value=''; document.getElementById('gftarih-bit').value=''; doGelirFilter(); }

function giderTarihSifirla(){ document.getElementById('ftarih-bas').value=''; document.getElementById('ftarih-bit').value=''; doFilter(); }



function renderGelirTbl(){

  const PPG2=25;

  const start=(gelirPg-1)*PPG2, end=start+PPG2;

  const rows=gelirFiltered.slice(start,end);

  const tot=gelirFiltered.reduce((s,r)=>s+r.tutar,0);

  const tahsil=gelirFiltered.filter(r=>r.durum==='TAHSİL EDİLDİ').reduce((s,r)=>s+r.tutar,0);

  const bekl=tot-tahsil;

  // stats bar

  const sb=document.getElementById('gelir-stats-bar');

  if(sb) sb.innerHTML=`<span>${gelirFiltered.length} kayıt</span>&nbsp;|&nbsp;<strong>Toplam: ${fmtN(tot)} ₺</strong>&nbsp;|&nbsp;<span style="color:var(--g)">Tahsil: ${fmtN(tahsil)} ₺</span>&nbsp;|&nbsp;<span style="color:var(--am)">Bekleyen: ${fmtN(bekl)} ₺</span>`;

  document.getElementById('gelir-tbody').innerHTML=rows.map(r=>{

    const idx=r._idx;

    const duCls=r.durum==='TAHSİL EDİLDİ'?'bgi':'bami2';

    const fc=FC[r.firma]||'bbi';

    return`<tr>

      <td><span class="bdg ${fc}">${r.firma}</span></td>

      <td style="font-size:12px">${r.tur}</td>

      <td class="clip" style="max-width:150px;font-size:12px" title="${r.aciklama||''}">${r.aciklama||'—'}</td>

      <td style="font-family:var(--mono);font-size:11.5px">${r.tarih||'—'}</td>

      <td style="font-size:12px">${r.ay||'—'}</td>

      <td><span class="bdg ${duCls}" style="cursor:pointer" ondblclick="toggleGelir(${idx})" title="Çift tıkla: durum değiştir">${r.durum}</span></td>

      <td class="num" style="font-weight:600;color:var(--g)">${fmtN(r.tutar)}</td>

      <td style="white-space:nowrap">

        <button class="btn bgh" onclick="openGelirEdit(${idx})" title="Düzenle">✏</button>

        <button class="btn bgh" style="color:var(--r)" onclick="delGelir(${idx})" title="Sil">✕</button>

      </td>

    </tr>`;

  }).join('');

  // pagination

  const tot2=Math.ceil(gelirFiltered.length/PPG2)||1;

  let ph=`<span class="pgi">${gelirFiltered.length} kayıt</span><div class="pgsp"></div>`;

  if(gelirPg>1) ph+=`<button class="pbn" onclick="gelirPg--;renderGelirTbl()">‹</button>`;

  for(let i=Math.max(1,gelirPg-2);i<=Math.min(tot2,gelirPg+2);i++)

    ph+=`<button class="pbn${i===gelirPg?' on':''}" onclick="gelirPg=${i};renderGelirTbl()">${i}</button>`;

  if(gelirPg<tot2) ph+=`<button class="pbn" onclick="gelirPg++;renderGelirTbl()">›</button>`;

  document.getElementById('gelir-pgn').innerHTML=ph;

}



// GELIR MODAL

function openGelirAdd(){

  gelirEditIdx=null;

  document.getElementById('gelir-mtit').textContent='Yeni Gelir Kaydı';

  document.getElementById('gm-save').textContent='Kaydet';

  document.getElementById('gm-firma').value='ULUSAL';

  document.getElementById('gm-tur').selectedIndex=0;

  document.getElementById('gm-aciklama').value='';

  document.getElementById('gm-tarih').value='';

  document.getElementById('gm-durum').value='BEKLENİYOR';

  document.getElementById('gm-tutar').value='';

  document.getElementById('gelir-mod').classList.add('on');

}

function openGelirEdit(idx){

  gelirEditIdx=idx;

  const r=gelirData.find(x=>x._idx===idx)||gelirData[idx];

  if(!r) return;

  document.getElementById('gelir-mtit').textContent='Gelir Kaydını Düzenle';

  document.getElementById('gm-save').textContent='Güncelle';

  document.getElementById('gm-firma').value=r.firma||'ULUSAL';

  document.getElementById('gm-tur').value=r.tur||'';

  document.getElementById('gm-aciklama').value=r.aciklama||'';

  if(r.tarih&&r.tarih.length===10){

    const p=r.tarih.split('.');

    document.getElementById('gm-tarih').value=`${p[2]}-${p[1]}-${p[0]}`;

  } else document.getElementById('gm-tarih').value='';

  document.getElementById('gm-durum').value=r.durum||'BEKLENİYOR';

  document.getElementById('gm-tutar').value=r.tutar||'';

  document.getElementById('gelir-mod').classList.add('on');

}

function closeGelirMod(){ document.getElementById('gelir-mod').classList.remove('on'); }



async function saveGelir(){

  const tv=document.getElementById('gm-tarih').value;

  let ts='',ay='';

  if(tv){ const d=new Date(tv); ts=d.toLocaleDateString('tr-TR',{day:'2-digit',month:'2-digit',year:'numeric'}); ay=AYLAR[d.getMonth()]; }

  const rec={

    firma:document.getElementById('gm-firma').value,

    tur:document.getElementById('gm-tur').value,

    aciklama:document.getElementById('gm-aciklama').value,

    tarih:ts, ay,

    durum:document.getElementById('gm-durum').value,

    tutar:parseFloat(document.getElementById('gm-tutar').value)||0,

  };

  const url=gelirEditIdx!==null?'/api/gelir/edit':'/api/gelir/add';

  const body=gelirEditIdx!==null?JSON.stringify({idx:gelirEditIdx,record:rec}):JSON.stringify(rec);

  const r=await api(url,{method:'POST',headers:{'Content-Type':'application/json'},body});

  if((await r.json()).ok){

    toast(gelirEditIdx!==null?'Gelir güncellendi':'Gelir eklendi','s');

    closeGelirMod(); loadGelir();

  }

}



async function delGelir(idx){

  if(!confirm('Bu gelir kaydını silmek istediğinizden emin misiniz?')) return;

  const r=await api('/api/gelir/delete',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({idx})});

  if((await r.json()).ok){ toast('Gelir silindi','i'); loadGelir(); }

}



async function toggleGelir(idx){

  const r=await api('/api/gelir/toggle',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({idx})});

  const d=await r.json();

  if(d.ok){ toast(d.durum==='TAHSİL EDİLDİ'?'✓ Tahsil edildi':'⚠ Bekleniyora alındı',d.durum==='TAHSİL EDİLDİ'?'s':'i'); loadGelir(); }

}



// ── NAKİT AKIŞ DENGESİ ────────────────────────────────────────────────────────

async function renderNakit(){

  const yilF=document.getElementById('nk-yil')?.value||'';

  const r=await api('/api/nakit/ozet'); const d=await r.json();

  let ozet=d.ozet||[];



  // Yıl filter seçenekleri

  const yillar=[...new Set(ozet.map(o=>o.yil))].sort();

  const nkYilSel=document.getElementById('nk-yil');

  if(nkYilSel){

    const cur=nkYilSel.value;

    nkYilSel.innerHTML='<option value="">Tüm Yıllar</option>';

    yillar.forEach(y=>{ const o=document.createElement('option');o.value=y;o.text=y;nkYilSel.appendChild(o); });

    nkYilSel.value=cur;

  }

  if(yilF) ozet=ozet.filter(o=>o.yil===yilF);



  // Toplam metrikler

  const totGelir=ozet.reduce((s,o)=>s+o.gelir,0);

  const totTahsil=ozet.reduce((s,o)=>s+o.gelir_tahsil,0);

  const totGider=ozet.reduce((s,o)=>s+o.gider,0);

  const totOdenen=ozet.reduce((s,o)=>s+o.gider_odendi,0);

  const totDenge=totGelir-totGider;

  const totFin=ozet.reduce((s,o)=>s+o.finansman_iht,0);



  document.getElementById('nakit-top-mets').innerHTML=`

    <div class="met"><div class="met-stripe" style="background:var(--g)"></div>

      <div class="mlb">Toplam Gelir (Plan)</div>

      <div class="mvl" style="color:var(--g);font-size:17px">₺${fmtN(totGelir)}</div>

      <div class="msb">Tahsil: ₺${fmtN(totTahsil)}</div></div>

    <div class="met"><div class="met-stripe" style="background:var(--r)"></div>

      <div class="mlb">Toplam Gider (Plan)</div>

      <div class="mvl" style="color:var(--r);font-size:17px">₺${fmtN(totGider)}</div>

      <div class="msb">Ödenen: ₺${fmtN(totOdenen)}</div></div>

    <div class="met"><div class="met-stripe" style="background:${totDenge>=0?'var(--g)':'var(--r)'}"></div>

      <div class="mlb">Net Denge</div>

      <div class="mvl" style="color:${totDenge>=0?'var(--g)':'var(--r)'};font-size:17px">${totDenge>=0?'+':''}₺${fmtN(totDenge)}</div>

      <div class="msb">${totDenge>=0?'Nakit fazlası':'Finansman gerekli'}</div></div>

    <div class="met"><div class="met-stripe" style="background:var(--am)"></div>

      <div class="mlb">Toplam Finansman İht.</div>

      <div class="mvl" style="color:var(--am);font-size:17px">₺${fmtN(totFin)}</div>

      <div class="msb">${ozet.filter(o=>!o.pozitif).length} ay negatif</div></div>

  `;



  // Tablo - kalem detayı desteği

  let prevYil='';

  window._nakitOzet = ozet; // detay açma için sakla

  const rows = [];

  ozet.forEach((o,idx)=>{

    if(o.yil!==prevYil){

      rows.push(`<tr><td colspan="10" style="background:var(--s3);font-family:var(--mono);font-size:10px;color:var(--acc);font-weight:700;padding:6px 10px;border-bottom:1px solid var(--b2)">${o.yil}</td></tr>`);

      prevYil=o.yil;

    }

    const dCls=o.denge>=0?'color:var(--g)':'color:var(--r)';

    const fCls=o.finansman_iht>0?'color:var(--am);font-weight:600':'color:var(--t3)';

    const durum=o.denge>=0?`<span class="bdg bgi">Pozitif</span>`:`<span class="bdg bri">Negatif</span>`;

    const hasDetail=(o.gider_kalemler?.length||0)+(o.gelir_kalemler?.length||0)>0;

    rows.push(`<tr id="nkrow-${idx}" style="${!o.pozitif?'background:rgba(248,113,113,.04)':''};cursor:${hasDetail?'pointer':'default'}" onclick="${hasDetail?`toggleNakitDet(${idx})`:''}">

      <td style="text-align:center;color:var(--t3);font-size:11px" id="nkexp-${idx}">${hasDetail?'▶':''}</td>

      <td style="font-family:var(--mono);font-size:11px;color:var(--t3)">${o.yil}</td>

      <td style="font-weight:500">${o.ay}</td>

      <td class="num" style="color:var(--g)">${o.gelir?fmtN(o.gelir):'—'}</td>

      <td class="num" style="color:var(--g-d)">${o.gelir_tahsil?fmtN(o.gelir_tahsil):'—'}</td>

      <td class="num" style="color:var(--r)">${o.gider?fmtN(o.gider):'—'}</td>

      <td class="num" style="color:var(--r-d)">${o.gider_odendi?fmtN(o.gider_odendi):'—'}</td>

      <td class="num" style="${dCls}">${o.denge>=0?'+':''}${fmtN(o.denge)}</td>

      <td class="num" style="${fCls}">${o.finansman_iht>0?fmtN(o.finansman_iht):'—'}</td>

      <td>${durum}</td>

    </tr>`);

    // Detay satırı (başlangıçta gizli)

    rows.push(`<tr id="nkdet-${idx}" style="display:none"><td colspan="10" style="padding:0;background:var(--s2);border-bottom:1px solid var(--b1)">

      <div id="nkdet-inner-${idx}"></div>

    </td></tr>`);

  });

  document.getElementById('nakit-tbody').innerHTML=rows.join('');



  // Footer toplam satırı

  document.getElementById('nakit-tfoot').innerHTML=`

    <tr style="border-top:2px solid var(--b2);font-weight:600">

      <td></td>

      <td colspan="2" style="padding:8px 10px;font-family:var(--mono);font-size:11px;color:var(--t2)">TOPLAM</td>

      <td class="num" style="color:var(--g)">${fmtN(totGelir)}</td>

      <td class="num" style="color:var(--g-d)">${fmtN(totTahsil)}</td>

      <td class="num" style="color:var(--r)">${fmtN(totGider)}</td>

      <td class="num" style="color:var(--r-d)">${fmtN(totOdenen)}</td>

      <td class="num" style="color:${totDenge>=0?'var(--g)':'var(--r)'}">${totDenge>=0?'+':''}${fmtN(totDenge)}</td>

      <td class="num" style="color:var(--am)">${totFin>0?fmtN(totFin):'—'}</td>

      <td></td>

    </tr>`;



  // Mini bar chart (SVG)

  renderNakitChart(ozet);

}



function toggleInnerDet(kId){

  const row=document.getElementById(kId);

  const ic=document.getElementById(kId+'-ic');

  if(!row) return;

  const open=row.style.display!=='none';

  row.style.display=open?'none':'';

  if(ic) ic.textContent=open?'▶':'▼';

}



async function toggleNakitDet(idx){

  const detRow=document.getElementById('nkdet-'+idx);

  const expIcon=document.getElementById('nkexp-'+idx);

  if(!detRow) return;

  const isOpen=detRow.style.display!=='none';

  if(isOpen){detRow.style.display='none';if(expIcon)expIcon.textContent='▶';return;}

  detRow.style.display='';

  if(expIcon) expIcon.textContent='▼';

  const o=window._nakitOzet?.[idx];

  if(!o) return;

  const inner=document.getElementById('nkdet-inner-'+idx);

  if(!inner) return;

  const fmtR=n=>Math.round(n||0).toLocaleString('tr-TR');

  const FC2={ULUSAL:'var(--acc)',ABC:'var(--acc2)',BRG:'var(--am)',BRK:'var(--g)'};



  // Gelir tablosu

  let gHtml='';

  if(o.gelir_kalemler&&o.gelir_kalemler.length>0){

    const gRows=o.gelir_kalemler.map((k,ki)=>{

      const kId=`gnk${idx}k${ki}`;

      const satirRows=(k.satirlar||[]).map(s=>`<tr style="background:rgba(16,185,129,.04);border-bottom:1px solid rgba(16,185,129,.1)">

        <td colspan="2" style="padding:3px 8px 3px 28px;font-size:11px;color:var(--t2)">

          <span style="font-family:var(--mono);color:var(--g-d)">${s.tarih||'—'}</span>

          ${s.aciklama?`<span style="color:var(--t3);margin-left:6px;font-size:10px">${s.aciklama.slice(0,40)}</span>`:''}

        </td>

        <td style="padding:3px 8px;text-align:right;font-family:var(--mono);font-size:11px"></td>

        <td style="padding:3px 8px;text-align:right;font-family:var(--mono);font-size:11px;color:var(--g)">${fmtR(s.tutar)}</td>

        <td style="padding:3px 8px;text-align:right;font-family:var(--mono);font-size:11px;color:var(--g-d)">${s.durum==='TAHSİL EDİLDİ'?fmtR(s.tutar):'—'}</td>

        <td style="padding:3px 8px;text-align:right;font-family:var(--mono);font-size:11px;color:var(--am)">${s.durum!=='TAHSİL EDİLDİ'?fmtR(s.tutar):'—'}</td>

      </tr>`).join('');

      return `<tr style="border-bottom:1px solid var(--b1);cursor:${k.satirlar?.length?'pointer':'default'}"

        onclick="${k.satirlar?.length?`toggleInnerDet('${kId}')`:''}">

        <td style="padding:5px 8px;font-size:12px">

          ${k.satirlar?.length?`<span id="${kId}-ic" style="color:var(--t3);font-size:10px;margin-right:4px">▶</span>`:''}${k.tur}

        </td>

        <td style="padding:5px 8px"><span style="font-size:11px;font-weight:600;color:${FC2[k.firma]||'var(--t2)'}">${k.firma}</span></td>

        <td style="padding:5px 8px;text-align:right;font-family:var(--mono);font-size:11px;color:var(--t2)">${k.adet}</td>

        <td style="padding:5px 8px;text-align:right;font-family:var(--mono);font-size:12px;color:var(--g);font-weight:600">${fmtR(k.tutar)}</td>

        <td style="padding:5px 8px;text-align:right;font-family:var(--mono);font-size:11px;color:var(--g-d)">${k.tahsil?fmtR(k.tahsil):'—'}</td>

        <td style="padding:5px 8px;text-align:right;font-family:var(--mono);font-size:11px;color:var(--am)">${k.bekl?fmtR(k.bekl):'—'}</td>

      </tr>

      <tr id="${kId}"><td colspan="6" style="padding:0;background:var(--s3)">${satirRows}</td></tr>`;

    }).join('');

    gHtml=`<div style="padding:12px 12px 4px">

      <div style="font-family:var(--mono);font-size:9px;color:var(--g);text-transform:uppercase;letter-spacing:.08em;margin-bottom:6px">● GELİR KALEMLERİ</div>

      <table style="width:100%;border-collapse:collapse;font-size:12px">

        <thead><tr style="background:rgba(16,185,129,.08)">

          <th style="padding:4px 8px;text-align:left;font-size:10px;color:var(--g-d);font-family:var(--mono);border-bottom:1px solid rgba(16,185,129,.2)">Tür</th>

          <th style="padding:4px 8px;font-size:10px;color:var(--g-d);font-family:var(--mono);border-bottom:1px solid rgba(16,185,129,.2)">Firma</th>

          <th style="padding:4px 8px;text-align:right;font-size:10px;color:var(--g-d);font-family:var(--mono);border-bottom:1px solid rgba(16,185,129,.2)">Adet</th>

          <th style="padding:4px 8px;text-align:right;font-size:10px;color:var(--g-d);font-family:var(--mono);border-bottom:1px solid rgba(16,185,129,.2)">Toplam ₺</th>

          <th style="padding:4px 8px;text-align:right;font-size:10px;color:var(--g-d);font-family:var(--mono);border-bottom:1px solid rgba(16,185,129,.2)">Tahsil ₺</th>

          <th style="padding:4px 8px;text-align:right;font-size:10px;color:var(--am);font-family:var(--mono);border-bottom:1px solid rgba(16,185,129,.2)">Bekl. ₺</th>

        </tr></thead>

        <tbody>${gRows}</tbody>

      </table></div>`;

  } else {

    gHtml=`<div style="padding:12px;font-size:11px;color:var(--t3);font-style:italic">Gelir kaydı yok — Gelir Tablosu'ndan ekleyin</div>`;

  }



  // Gider tablosu - türe göre grupla

  let dHtml='';

  if(o.gider_kalemler&&o.gider_kalemler.length>0){

    const byTur={};

    o.gider_kalemler.forEach(k=>{

      if(!byTur[k.tur]) byTur[k.tur]={tur:k.tur,firmalar:[],tutar:0,odendi:0,bekl:0,adet:0};

      byTur[k.tur].firmalar.push(k);

      byTur[k.tur].tutar+=k.tutar; byTur[k.tur].odendi+=k.odendi||0;

      byTur[k.tur].bekl+=k.bekl||0; byTur[k.tur].adet+=k.adet||0;

    });

    const turList=Object.values(byTur).sort((a,b)=>b.tutar-a.tutar);

    const dRows=turList.map((t,ti)=>{

      const pct=o.gider>0?Math.round(t.tutar/o.gider*100):0;

      const fStr=t.firmalar.map(f=>`<span style="color:${FC2[f.firma]||'var(--t2)'}; font-size:11px">${f.firma}</span>`).join(' ');

      const kId=`dnk${idx}t${ti}`;

      // Tüm satırları birleştir (firma bazında ayırt et)

      const allSatirlar = t.firmalar.flatMap(f=>(f.satirlar||[]).map(s=>({...s, firma: f.firma})))

        .sort((a,b)=>(a.tarih||'').localeCompare(b.tarih||''));

      const satirRows = allSatirlar.map((s,si)=>{

        const isDone = s.durum==='ÖDENDİ';

        const rowId = 'nksr-'+idx+'-'+ti+'-'+si;

        return `<tr id="${rowId}" style="background:${isDone?'rgba(52,211,153,.04)':'rgba(248,113,113,.04)'};border-bottom:1px solid var(--b1)">

          <td style="padding:4px 8px 4px 28px;font-size:11px">

            <span style="font-family:var(--mono);color:${isDone?'var(--g)':'var(--am)'};">${s.tarih||'—'}</span>

          </td>

          <td style="padding:4px 8px;font-size:11px;color:var(--t2);max-width:200px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap">

            ${s.yer||s.aciklama||'—'}

          </td>

          <td style="padding:4px 8px"><span style="font-size:10px;color:${FC2[s.firma]||'var(--t2)'};">${s.firma||'—'}</span></td>

          <td style="padding:4px 8px;text-align:right;font-family:var(--mono);font-size:12px;color:var(--r);font-weight:600">${fmtR(s.tutar)}</td>

          <td style="padding:4px 8px;text-align:center">

            <button onclick="nakitToggleDurum(${s.rec_idx},'${rowId}')"

              style="font-size:10px;padding:2px 10px;border-radius:4px;border:none;cursor:pointer;font-weight:600;

                     background:${isDone?'rgba(52,211,153,.2)':'rgba(251,191,36,.2)'};

                     color:${isDone?'var(--g)':'var(--am)'};">

              ${isDone?'✓ ÖDENDİ':'⏳ BEKLİYOR'}

            </button>

          </td>

        </tr>`;

      }).join('');

      return `<tr style="border-bottom:1px solid var(--b1);cursor:${allSatirlar.length?'pointer':'default'}"

        onclick="${allSatirlar.length?`toggleInnerDet('${kId}')`:''}">

        <td style="padding:5px 8px;font-size:12px;font-weight:500">

          ${allSatirlar.length?`<span id="${kId}-ic" style="color:var(--t3);font-size:10px;margin-right:4px">▼</span>`:''}${t.tur}

        </td>

        <td style="padding:5px 8px">${fStr}</td>

        <td style="padding:5px 8px;text-align:right;font-family:var(--mono);font-size:11px;color:var(--t2)">${t.adet}</td>

        <td style="padding:5px 8px;text-align:right;font-family:var(--mono);font-size:12px;color:var(--r);font-weight:600">${fmtR(t.tutar)}</td>

        <td style="padding:5px 8px;text-align:right;font-family:var(--mono);font-size:11px;color:var(--r-d)">${t.odendi?fmtR(t.odendi):'—'}</td>

        <td style="padding:5px 8px;text-align:right;font-family:var(--mono);font-size:11px;color:var(--am)">${t.bekl?fmtR(t.bekl):'—'}</td>

        <td style="padding:5px 8px;text-align:right;font-family:var(--mono);font-size:10px;color:var(--t3)">${pct}%

          <div style="margin-top:2px;height:2px;background:var(--b1);border-radius:1px;overflow:hidden">

            <div style="height:100%;width:${pct}%;background:var(--r);border-radius:1px"></div>

          </div>

        </td>

      </tr>

      <tr id="${kId}"><td colspan="7" style="padding:0;background:var(--s3)">${satirRows}</td></tr>`;

    }).join('');

    dHtml=`<div style="padding:4px 12px 12px">

      <div style="font-family:var(--mono);font-size:9px;color:var(--r);text-transform:uppercase;letter-spacing:.08em;margin-bottom:6px">● GİDER KALEMLERİ</div>

      <table style="width:100%;border-collapse:collapse;font-size:12px">

        <thead><tr style="background:rgba(248,113,113,.08)">

          <th style="padding:4px 8px;text-align:left;font-size:10px;color:var(--r-d);font-family:var(--mono);border-bottom:1px solid rgba(248,113,113,.2)">Ödeme Türü</th>

          <th style="padding:4px 8px;font-size:10px;color:var(--r-d);font-family:var(--mono);border-bottom:1px solid rgba(248,113,113,.2)">Firma(lar)</th>

          <th style="padding:4px 8px;text-align:right;font-size:10px;color:var(--r-d);font-family:var(--mono);border-bottom:1px solid rgba(248,113,113,.2)">Adet</th>

          <th style="padding:4px 8px;text-align:right;font-size:10px;color:var(--r-d);font-family:var(--mono);border-bottom:1px solid rgba(248,113,113,.2)">Toplam ₺</th>

          <th style="padding:4px 8px;text-align:right;font-size:10px;color:var(--r-d);font-family:var(--mono);border-bottom:1px solid rgba(248,113,113,.2)">Ödenen ₺</th>

          <th style="padding:4px 8px;text-align:right;font-size:10px;color:var(--am);font-family:var(--mono);border-bottom:1px solid rgba(248,113,113,.2)">Bekl. ₺</th>

          <th style="padding:4px 8px;text-align:right;font-size:10px;color:var(--r-d);font-family:var(--mono);border-bottom:1px solid rgba(248,113,113,.2)">Pay</th>

        </tr></thead>

        <tbody>${dRows}</tbody>

      </table></div>`;

  }



  inner.innerHTML=`<div style="border-top:2px solid var(--b2)">${dHtml}${gHtml}</div>`;

  // Tüm alt detayları otomatik aç (dnk ve gnk ile başlayan tr'ler)

  setTimeout(()=>{

    ['dnk','gnk'].forEach(prefix=>{

      inner.querySelectorAll('tr').forEach(tr=>{

        if(tr.id && tr.id.startsWith(prefix) && !tr.id.endsWith('-ic')){

          tr.style.display = '';

          const ic = document.getElementById(tr.id + '-ic');

          if(ic) ic.textContent = '▼';

        }

      });

    });

  }, 30);

}



async function nakitToggleDurum(recIdx, rowId){

  try {

    const r = await api('/api/record/toggle',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({idx:recIdx})});

    const d = await r.json();

    if(!d.ok){ toast('Güncelleme hatası','e'); return; }

    const yeniDurum = d.durum;

    const isDone = yeniDurum === 'ÖDENDİ';

    // Satır rengini ve butonu güncelle

    const row = document.getElementById(rowId);

    if(row){

      row.style.background = isDone ? 'rgba(52,211,153,.04)' : 'rgba(248,113,113,.04)';

      const btn = row.querySelector('button');

      if(btn){

        btn.textContent = isDone ? '✓ ÖDENDİ' : '⏳ BEKLİYOR';

        btn.style.background = isDone ? 'rgba(52,211,153,.2)' : 'rgba(251,191,36,.2)';

        btn.style.color = isDone ? 'var(--g)' : 'var(--am)';

      }

      const tarihEl = row.querySelector('span[style*="font-family"]');

      if(tarihEl) tarihEl.style.color = isDone ? 'var(--g)' : 'var(--am)';

    }

    toast(isDone ? '✓ Ödendi olarak işaretlendi' : '⏳ Bekliyor olarak işaretlendi', isDone?'s':'i');

    if(d.excel) toast('Excel güncellendi ✓','s');

    // Nakit özet tablosunu yenile

    setTimeout(()=>renderNakit(), 800);

  } catch(e){

    toast('Hata: '+e.message,'e');

  }

}



function renderNakitChart(ozet){

  const wrap=document.getElementById('nakit-chart-wrap');

  if(!ozet.length){ wrap.innerHTML='<div style="color:var(--t3);font-size:12px;padding:20px 0">Veri yok</div>'; return; }

  const maxVal=Math.max(...ozet.map(o=>Math.max(o.gelir,o.gider)),1);

  const W=Math.max(ozet.length*70,400), H=180, PAD=30, BAR_W=22, GAP=70;

  let svg=`<svg width="${W}" height="${H+60}" xmlns="http://www.w3.org/2000/svg" style="font-family:Consolas,monospace">`;

  // Y eksen çizgisi

  svg+=`<line x1="${PAD}" y1="10" x2="${PAD}" y2="${H+PAD}" stroke="#2d3748" stroke-width="1"/>`;

  svg+=`<line x1="${PAD}" y1="${H+PAD}" x2="${W}" y2="${H+PAD}" stroke="#2d3748" stroke-width="1"/>`;

  ozet.forEach((o,i)=>{

    const x=PAD+i*GAP+8;

    const gelirH=Math.round((o.gelir/maxVal)*(H-20));

    const giderH=Math.round((o.gider/maxVal)*(H-20));

    const dengeH=Math.abs(Math.round((o.denge/maxVal)*(H-20)));

    // Gelir bar

    svg+=`<rect x="${x}" y="${H+PAD-gelirH}" width="${BAR_W}" height="${gelirH}" fill="#059669" opacity=".8" rx="2"/>`;

    // Gider bar

    svg+=`<rect x="${x+BAR_W+2}" y="${H+PAD-giderH}" width="${BAR_W}" height="${giderH}" fill="#dc2626" opacity=".7" rx="2"/>`;

    // Denge çizgisi noktası

    const dengeY=o.denge>=0?H+PAD-gelirH-6:H+PAD-giderH-6;

    svg+=`<circle cx="${x+BAR_W}" cy="${dengeY}" r="4" fill="${o.denge>=0?'#34d399':'#f87171'}"/>`;

    // X etiketi

    svg+=`<text x="${x+BAR_W}" y="${H+PAD+16}" text-anchor="middle" font-size="9" fill="#4a5568">${o.ay.slice(0,3)}</text>`;

    svg+=`<text x="${x+BAR_W}" y="${H+PAD+27}" text-anchor="middle" font-size="9" fill="#3d4f6b">${o.yil}</text>`;

    // Finansman ihtiyacı işareti

    if(o.finansman_iht>0){

      svg+=`<text x="${x+BAR_W}" y="${dengeY-8}" text-anchor="middle" font-size="9" fill="#fbbf24">!</text>`;

    }

  });

  // Legend

  svg+=`<rect x="${PAD}" y="${H+PAD+36}" width="12" height="8" fill="#059669" rx="1"/>

  <text x="${PAD+16}" y="${H+PAD+44}" font-size="9" fill="#94a3b8">Gelir</text>

  <rect x="${PAD+55}" y="${H+PAD+36}" width="12" height="8" fill="#dc2626" rx="1"/>

  <text x="${PAD+71}" y="${H+PAD+44}" font-size="9" fill="#94a3b8">Gider</text>

  <circle cx="${PAD+122}" cy="${H+PAD+40}" r="4" fill="#34d399"/>

  <text x="${PAD+130}" y="${H+PAD+44}" font-size="9" fill="#94a3b8">Denge</text>

  <text x="${PAD+185}" y="${H+PAD+44}" font-size="9" fill="#fbbf24">! = Finansman gerekli</text>`;

  svg+='</svg>';

  wrap.innerHTML=svg;

}



async function expNakitExcel(){

  // Nakit dengesi tablosunu Excel'e aktar

  const r=await api('/api/nakit/ozet'); const d=await r.json();

  const ozet=d.ozet||[];

  if(!ozet.length){ toast('Veri yok','e'); return; }

  // xlsx kütüphanesi yok client'ta, CSV olarak ver

  const hdrs=['Yıl','Ay','Gelir (Plan)','Tahsilat','Gider (Plan)','Ödenen','Net Denge','Finansman İhtiyacı','Durum'];

  const rows=[hdrs,...ozet.map(o=>[o.yil,o.ay,o.gelir,o.gelir_tahsil,o.gider,o.gider_odendi,o.denge,o.finansman_iht,o.pozitif?'Pozitif':'Negatif'])];

  const csv=rows.map(r=>r.map(v=>'"'+String(v).replace(/"/g,'""')+'"').join(',')).join(String.fromCharCode(10));

  const a=document.createElement('a');

  a.href='data:text/csv;charset=utf-8,\uFEFF'+encodeURIComponent(csv);

  a.download='nakit_akis_dengesi.csv'; a.click();

  toast('Nakit dengesi CSV indirildi','s');

}





// ── LİKİDİTE & FİNANSMAN ────────────────────────────────────────────────────



// Sekme geçişi

function likiTab(tab){

  ['analiz','banka','kredi'].forEach(t=>{

    document.getElementById('ltab-'+t+'-view').style.display = t===tab?'':'none';

    const btn = document.getElementById('ltab-'+t);

    if(btn){ btn.style.borderBottom = t===tab?'2px solid var(--acc)':'none'; btn.style.color = t===tab?'var(--acc)':''; }

  });

  if(tab==='banka') renderBankaTbl();

  if(tab==='kredi') renderKrediTbl();

}



// ── ANALİZ ──────────────────────────────────────────────────────────────────

async function renderLikidite(){

  const yilF=document.getElementById('liki-yil')?.value||'';

  const r=await api('/api/likidite'); const d=await r.json();

  let liki=d.likidite||[];



  // Yıl filtresi

  const yillar=[...new Set(liki.map(o=>o.yil))].sort();

  const sel=document.getElementById('liki-yil');

  if(sel){ const cur=sel.value; sel.innerHTML='<option value="">Tüm Yıllar</option>'; yillar.forEach(y=>{const o=document.createElement('option');o.value=y;o.text=y;sel.appendChild(o);}); sel.value=cur; }

  if(yilF) liki=liki.filter(o=>o.yil===yilF);



  // Üst metrikler

  const totNakit=d.toplam_banka||0;

  const totKalan=d.toplam_kalan_limit||0;

  const totIhtiyac=liki.reduce((s,o)=>s+o.nakit_ihtiyac,0);

  const totFaiz=liki.reduce((s,o)=>s+o.toplam_faiz_maliyeti,0);

  document.getElementById('liki-mets').innerHTML=`

    <div class="met"><div class="met-stripe" style="background:var(--acc)"></div>

      <div class="mlb">Mevcut Banka Bakiyesi</div>

      <div class="mvl" style="color:var(--acc);font-size:17px">₺${fmtN(totNakit)}</div>

      <div class="msb">${(d.banka_detay||[]).length} hesap</div></div>

    <div class="met"><div class="met-stripe" style="background:var(--g)"></div>

      <div class="mlb">Kullanılabilir Limit</div>

      <div class="mvl" style="color:var(--g);font-size:17px">₺${fmtN(totKalan)}</div>

      <div class="msb">${(d.kredi_kullanilabilir||[]).length} kredi ürünü</div></div>

    <div class="met"><div class="met-stripe" style="background:${totIhtiyac>0?'var(--r)':'var(--g)'}"></div>

      <div class="mlb">Dönem Toplam Nakit İht.</div>

      <div class="mvl" style="color:${totIhtiyac>0?'var(--r)':'var(--g)'};font-size:17px">₺${fmtN(totIhtiyac)}</div>

      <div class="msb">${liki.filter(o=>!o.pozitif).length} ay açık</div></div>

    <div class="met"><div class="met-stripe" style="background:var(--am)"></div>

      <div class="mlb">Tahmini Faiz Maliyeti</div>

      <div class="mvl" style="color:var(--am);font-size:17px">₺${fmtN(totFaiz)}</div>

      <div class="msb">Optimal kredi kullanımıyla</div></div>

  `;



  window._likiData = liki;

  let prevYil='';

  const rows=[];

  liki.forEach((o,idx)=>{

    if(o.yil!==prevYil){

      rows.push(`<tr><td colspan="10" style="background:var(--s3);font-family:var(--mono);font-size:10px;color:var(--acc);font-weight:700;padding:6px 10px;border-bottom:1px solid var(--b2)">${o.yil}</td></tr>`);

      prevYil=o.yil;

    }

    const dCls=o.denge>=0?'color:var(--g)':'color:var(--r)';

    const iCls=o.nakit_ihtiyac>0?'color:var(--r);font-weight:600':'color:var(--t3)';

    const kCls=o.kumulatif_banka>=0?'color:var(--g)':'color:var(--r)';

    const hasOner=o.oneriler&&o.oneriler.length>0;

    let durum='';

    if(o.pozitif) durum='<span class="bdg bgi">Nakit Fazlası</span>';

    else if(o.karsilanamayan>0) durum='<span class="bdg bri">Yetersiz Limit</span>';

    else durum='<span class="bdg bami">Kredi Önerisi Var</span>';



    rows.push(`<tr style="${!o.pozitif?'background:rgba(248,113,113,.03)':''};cursor:${hasOner?'pointer':'default'}" onclick="${hasOner?`toggleLikiDet(${idx})`:''}">

      <td style="text-align:center;font-size:10px;color:var(--t3)" id="liki-exp-${idx}">${hasOner?'▶':''}</td>

      <td style="font-family:var(--mono);font-size:11px;color:var(--t3)">${o.yil}</td>

      <td style="font-weight:500">${o.ay}</td>

      <td class="num" style="color:var(--g)">${o.gelir?fmtN(o.gelir):'—'}</td>

      <td class="num" style="color:var(--r)">${fmtN(o.gider)}</td>

      <td class="num" style="${dCls}">${o.denge>=0?'+':''}${fmtN(o.denge)}</td>

      <td class="num" style="${iCls}">${o.nakit_ihtiyac>0?fmtN(o.nakit_ihtiyac):'—'}</td>

      <td class="num" style="${kCls}">${fmtN(o.kumulatif_banka)}</td>

      <td class="num" style="color:var(--am)">${o.toplam_faiz_maliyeti>0?fmtN(o.toplam_faiz_maliyeti):'—'}</td>

      <td>${durum}</td>

    </tr>

    <tr id="liki-det-${idx}" style="display:none"><td colspan="10" style="padding:0;background:var(--s2)">

      <div id="liki-det-inner-${idx}"></div>

    </td></tr>`);

  });

  document.getElementById('liki-tbody').innerHTML=rows.join('');



  // Footer

  const totG=liki.reduce((s,o)=>s+o.gelir,0);

  const totD=liki.reduce((s,o)=>s+o.gider,0);

  const totDn=totG-totD;

  document.getElementById('liki-tfoot').innerHTML=`

    <tr style="border-top:2px solid var(--b2);font-weight:600">

      <td colspan="3" style="padding:8px 10px;font-family:var(--mono);font-size:11px;color:var(--t2)">TOPLAM</td>

      <td class="num" style="color:var(--g)">${fmtN(totG)}</td>

      <td class="num" style="color:var(--r)">${fmtN(totD)}</td>

      <td class="num" style="color:${totDn>=0?'var(--g)':'var(--r)'}">${totDn>=0?'+':''}${fmtN(totDn)}</td>

      <td class="num" style="color:var(--r)">${fmtN(totIhtiyac)}</td>

      <td class="num"></td>

      <td class="num" style="color:var(--am)">${fmtN(totFaiz)}</td>

      <td></td>

    </tr>`;

}



function toggleLikiDet(idx){

  const row=document.getElementById('liki-det-'+idx);

  const ic=document.getElementById('liki-exp-'+idx);

  if(!row) return;

  const open=row.style.display!=='none';

  if(open){row.style.display='none';if(ic)ic.textContent='▶';return;}

  row.style.display=''; if(ic)ic.textContent='▼';

  const o=window._likiData?.[idx];

  if(!o) return;

  const inner=document.getElementById('liki-det-inner-'+idx);

  if(!inner||!o.oneriler||!o.oneriler.length) return;

  // İçerik zaten varsa tekrar render etme

  if(inner.innerHTML && inner.innerHTML.length > 50) return;

  const FC2={ULUSAL:'var(--acc)',ABC:'var(--acc2)',BRG:'var(--am)',BRK:'var(--g)'};

  const turRenk={'ROTATİF':'var(--g)','VİNOV':'var(--acc2)','KMH':'var(--acc)','KREDİ KARTI':'var(--am)'};



  inner.innerHTML=`

    <div style="padding:14px 16px">

      <div style="font-family:var(--mono);font-size:10px;color:var(--acc);text-transform:uppercase;letter-spacing:.08em;margin-bottom:10px">

        ● OPTİMAL KREDİ KULLANIM ÖNERİSİ — ${o.ay} ${o.yil}

        <span style="color:var(--r);margin-left:10px">Nakit İhtiyaç: ₺${fmtN(o.nakit_ihtiyac)}</span>

        ${o.karsilanamayan>0?`<span style="color:var(--r);margin-left:10px">⚠ Karşılanamayan: ₺${fmtN(o.karsilanamayan)}</span>`:''}

      </div>

      <table style="width:100%;border-collapse:collapse;font-size:12px">

        <thead><tr style="background:rgba(79,156,249,.1)">

          <th style="padding:5px 9px;text-align:left;font-size:10px;color:var(--acc);font-family:var(--mono);border-bottom:1px solid rgba(79,156,249,.2)">#</th>

          <th style="padding:5px 9px;text-align:left;font-size:10px;color:var(--acc);font-family:var(--mono);border-bottom:1px solid rgba(79,156,249,.2)">Firma</th>

          <th style="padding:5px 9px;text-align:left;font-size:10px;color:var(--acc);font-family:var(--mono);border-bottom:1px solid rgba(79,156,249,.2)">Banka</th>

          <th style="padding:5px 9px;text-align:left;font-size:10px;color:var(--acc);font-family:var(--mono);border-bottom:1px solid rgba(79,156,249,.2)">Kredi Türü</th>

          <th style="padding:5px 9px;text-align:left;font-size:10px;color:var(--acc);font-family:var(--mono);border-bottom:1px solid rgba(79,156,249,.2)">Kredi/Kart No</th>

          <th style="padding:5px 9px;text-align:right;font-size:10px;color:var(--acc);font-family:var(--mono);border-bottom:1px solid rgba(79,156,249,.2)">Kullanılacak ₺</th>

          <th style="padding:5px 9px;text-align:right;font-size:10px;color:var(--acc);font-family:var(--mono);border-bottom:1px solid rgba(79,156,249,.2)">Yıllık Faiz %</th>

          <th style="padding:5px 9px;text-align:right;font-size:10px;color:var(--am);font-family:var(--mono);border-bottom:1px solid rgba(79,156,249,.2)">Tahmini Faiz ₺</th>

          <th style="padding:5px 9px;text-align:right;font-size:10px;color:var(--acc);font-family:var(--mono);border-bottom:1px solid rgba(79,156,249,.2)">Kalan Limit ₺</th>

        </tr></thead>

        <tbody>

          ${o.oneriler.map((k,ki)=>`<tr style="border-bottom:1px solid var(--b1);background:${ki%2?'rgba(255,255,255,.01)':''}">

            <td style="padding:5px 9px;font-family:var(--mono);font-size:10px;color:var(--t3)">${ki+1}</td>

            <td style="padding:5px 9px"><span style="font-size:11px;font-weight:600;color:${FC2[k.firma]||'var(--t2)'}">${k.firma}</span></td>

            <td style="padding:5px 9px;font-size:12px">${k.banka}</td>

            <td style="padding:5px 9px"><span style="font-size:11px;font-weight:500;color:${turRenk[k.tur]||'var(--t2)'}">${k.tur}</span></td>

            <td style="padding:5px 9px;font-family:var(--mono);font-size:11px;color:var(--t2)">${k.kredi_no||'—'}</td>

            <td style="padding:5px 9px;text-align:right;font-family:var(--mono);font-size:12px;color:var(--acc);font-weight:600">${fmtN(k.kullan)}</td>

            <td style="padding:5px 9px;text-align:right;font-family:var(--mono);font-size:11px;color:var(--t2)">${k.faiz_yillik?(k.faiz_yillik*100).toFixed(1)+'%':'—'}</td>

            <td style="padding:5px 9px;text-align:right;font-family:var(--mono);font-size:11px;color:var(--am)">${k.tahmini_faiz?fmtN(k.tahmini_faiz):'—'}</td>

            <td style="padding:5px 9px;text-align:right;font-family:var(--mono);font-size:11px;color:var(--t3)">${fmtN(k.kalan_limit_sonra)}</td>

          </tr>`).join('')}

          <tr style="border-top:1px solid var(--b2);background:rgba(79,156,249,.05)">

            <td colspan="5" style="padding:5px 9px;font-family:var(--mono);font-size:10px;color:var(--t2)">TOPLAM KREDİ KULLANIMI</td>

            <td style="padding:5px 9px;text-align:right;font-family:var(--mono);font-size:12px;color:var(--acc);font-weight:700">₺${fmtN(o.oneriler.reduce((s,k)=>s+k.kullan,0))}</td>

            <td></td>

            <td style="padding:5px 9px;text-align:right;font-family:var(--mono);font-size:11px;color:var(--am)">₺${fmtN(o.toplam_faiz_maliyeti)}</td>

            <td></td>

          </tr>

        </tbody>

      </table>

      ${o.karsilanamayan>0?`

      <div style="margin-top:10px;padding:10px 12px;background:rgba(239,68,68,.1);border:1px solid rgba(239,68,68,.3);border-radius:var(--rad);font-size:12px;color:var(--r)">

        ⚠ <strong>₺${fmtN(o.karsilanamayan)}</strong> nakit ihtiyacı mevcut limitlerle karşılanamıyor. 

        Ek finansman kaynağı veya gider ertelemesi gerekebilir.

      </div>`:''}

    </div>`;

}



// ── BANKA BAKİYELERİ ─────────────────────────────────────────────────────────

let bankaList=[];

async function renderBankaTbl(){

  const r=await api('/api/banka'); const d=await r.json();

  bankaList=[...(d.records||[])];

  drawBankaTbl();

}

function drawBankaTbl(){

  const tbody=document.getElementById('banka-tbody');

  if(!bankaList.length){

    tbody.innerHTML='<tr><td colspan="6" style="padding:20px;text-align:center;color:var(--t3);font-size:12px">Henüz banka hesabı eklenmedi</td></tr>';

    document.getElementById('banka-toplam').textContent='';

    return;

  }

  tbody.innerHTML=bankaList.map((b,i)=>`<tr>

    <td><select class="fi" style="font-size:12px" onchange="bankaList[${i}].firma=this.value">

      ${['ULUSAL','ABC','BRG','BRK'].map(f=>`<option ${b.firma===f?'selected':''}>${f}</option>`).join('')}

    </select></td>

    <td><input class="fi2" type="text" list="banka-datalist" value="${b.banka||''}" style="font-size:12px" oninput="bankaList[${i}].banka=this.value" placeholder="Banka adı seçin veya yazın"></td>

    <td><input class="fi2" type="text" value="${b.hesap_turu||''}" style="font-size:12px" oninput="bankaList[${i}].hesap_turu=this.value" placeholder="Vadesiz, TL vb."></td>

    <td><input class="fi2" type="number" value="${b.bakiye||0}" style="font-size:12px;text-align:right;font-family:var(--mono)" oninput="bankaList[${i}].bakiye=parseFloat(this.value)||0;updateBankaToplam()"></td>

    <td><input class="fi2" type="date" value="${b.tarih||''}" style="font-size:12px" oninput="bankaList[${i}].tarih=this.value"></td>

    <td><button class="btn bgh" title="Bu bankayı sil" style="color:var(--r);font-size:14px;font-weight:700;border:1px solid var(--r);border-radius:var(--rad);padding:2px 8px" onclick="if(confirm('${(b.banka||'Bu banka')} silinsin mi?')){bankaList.splice(${i},1);drawBankaTbl();}">🗑</button></td>

  </tr>`).join('');

  updateBankaToplam();

}

function updateBankaToplam(){

  const tot=bankaList.reduce((s,b)=>s+(parseFloat(b.bakiye)||0),0);

  document.getElementById('banka-toplam').innerHTML=`Toplam Banka Bakiyesi: <strong style="color:var(--acc)">₺${fmtN(tot)}</strong>  ·  ${bankaList.length} hesap`;

}

function addBankaSatir(){

  bankaList.push({firma:'ULUSAL',banka:'',hesap_turu:'Vadesiz TL',bakiye:0,tarih:new Date().toISOString().slice(0,10)});

  drawBankaTbl();

}

async function saveBanka(){

  const r=await api('/api/banka/save',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(bankaList)});

  const d=await r.json();

  if(d.ok){toast(`✓ ${d.count} banka hesabı kaydedildi`,'s');renderLikidite();}

  else toast('Kayıt hatası','e');

}



// ── KREDİ LİMİTLERİ ──────────────────────────────────────────────────────────

let krediList=[];

async function renderKrediTbl(){

  const r=await api('/api/kredi'); const d=await r.json();

  krediList=[...(d.records||[])];

  drawKrediTbl();

}

function drawKrediTbl(){

  const tbody=document.getElementById('kredi-tbody');

  if(!krediList.length){

    tbody.innerHTML='<tr><td colspan="10" style="padding:20px;text-align:center;color:var(--t3);font-size:12px">Kredi limiti eklenmedi. Excelden Yukle butonunu kullanin</td></tr>';

    document.getElementById('kredi-toplam').textContent='';

    return;

  }

  tbody.innerHTML=krediList.map((k,i)=>`<tr style="${!k.aktif?'opacity:.5':''}">

    <td><select class="fi" style="font-size:11px" onchange="krediList[${i}].firma=this.value">

      ${['ULUSAL','ABC','BRG','BRK'].map(f=>`<option ${k.firma===f?'selected':''}>${f}</option>`).join('')}

    </select></td>

    <td><input class="fi2" type="text" value="${k.banka||''}" style="font-size:11px" oninput="krediList[${i}].banka=this.value"></td>

    <td><select class="fi" style="font-size:11px" onchange="krediList[${i}].tur=this.value">

      ${['ROTATİF','KMH','VİNOV','KREDİ KARTI','TAKSİTLİ','DİĞER'].map(t=>`<option ${k.tur===t?'selected':''}>${t}</option>`).join('')}

    </select></td>

    <td><input class="fi2" type="text" value="${k.kredi_no||''}" style="font-size:11px;font-family:var(--mono)" placeholder="Kredi/Kart No" oninput="krediList[${i}].kredi_no=this.value"></td>

    <td><input class="fi2" type="number" value="${k.limit||0}" style="font-size:11px;text-align:right;font-family:var(--mono)" oninput="krediList[${i}].limit=parseFloat(this.value)||0;krediList[${i}].kalan_limit=krediList[${i}].limit-krediList[${i}].kullanilan;drawKrediTbl()"></td>

    <td><input class="fi2" type="number" value="${k.kullanilan||0}" style="font-size:11px;text-align:right;font-family:var(--mono)" oninput="krediList[${i}].kullanilan=parseFloat(this.value)||0;krediList[${i}].kalan_limit=krediList[${i}].limit-krediList[${i}].kullanilan;drawKrediTbl()"></td>

    <td style="text-align:right;font-family:var(--mono);font-size:11px;padding:7px 10px;font-weight:600;color:${(k.kalan_limit||0)>0?'var(--g)':'var(--t3)'}">₺${fmtN(k.kalan_limit||0)}</td>

    <td><input class="fi2" type="number" value="${k.faiz_yillik?Math.round(k.faiz_yillik*1000)/10:0}" step="0.1" style="font-size:11px;text-align:right;font-family:var(--mono)" oninput="krediList[${i}].faiz_yillik=parseFloat(this.value)/100||0;krediList[${i}].faiz_aylik=krediList[${i}].faiz_yillik/12" placeholder="%"></td>

    <td style="font-size:11px;font-family:var(--mono);color:var(--t2);padding:7px 10px">${k.bitis||'—'}</td>

    <td style="text-align:center"><input type="checkbox" ${k.aktif!==false?'checked':''} onchange="krediList[${i}].aktif=this.checked" style="cursor:pointer;accent-color:var(--acc)"></td>

    <td><button class="btn bgh" style="color:var(--r);font-size:11px" onclick="krediList.splice(${i},1);drawKrediTbl()">✕</button></td>

  </tr>`).join('');

  updateKrediToplam();

}

function updateKrediToplam(){

  const aktif=krediList.filter(k=>k.aktif!==false);

  const totLimit=aktif.reduce((s,k)=>s+(k.limit||0),0);

  const totKullanilan=aktif.reduce((s,k)=>s+(k.kullanilan||0),0);

  const totKalan=aktif.reduce((s,k)=>s+(k.kalan_limit||0),0);

  document.getElementById('kredi-toplam').innerHTML=

    `Aktif Limitler: <strong style="color:var(--t1)">₺${fmtN(totLimit)}</strong>  ·  `+

    `Kullanılan: <strong style="color:var(--r)">₺${fmtN(totKullanilan)}</strong>  ·  `+

    `Kullanılabilir: <strong style="color:var(--g)">₺${fmtN(totKalan)}</strong>  ·  `+

    `${aktif.length} ürün`;

}

function addKrediSatir(){

  krediList.push({firma:'ULUSAL',banka:'',tur:'ROTATİF',limit:0,kullanilan:0,kalan_limit:0,faiz_yillik:0,faiz_aylik:0,bitis:'',aktif:true});

  drawKrediTbl();

}

async function saveKredi(){

  // faiz_aylik güncel değil ise yenile

  krediList.forEach(k=>{ if(!k.faiz_aylik&&k.faiz_yillik) k.faiz_aylik=k.faiz_yillik/12; });

  const r=await api('/api/kredi/save',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(krediList)});

  const d=await r.json();

  if(d.ok){toast(`✓ ${d.count} kredi limiti kaydedildi`,'s');renderLikidite();}

  else toast('Kayıt hatası','e');

}

function krediSablonIndir(){

  const a=document.createElement('a');

  a.href='/api/kredi/sablon';

  document.body.appendChild(a);

  a.click();

  document.body.removeChild(a);

  toast('Şablon indiriliyor...','i');

}



async function importKrediExcel(){

  const btn=document.getElementById('kredi-import-btn');

  btn.textContent='⏳ Yükleniyor...'; btn.disabled=true;

  const r=await api('/api/kredi/import_excel',{method:'POST',headers:{'Content-Type':'application/json'},body:'{}'});

  const d=await r.json();

  btn.textContent='⬆ Excelden Yükle'; btn.disabled=false;

  if(d.ok){

    krediList=d.records||[];

    drawKrediTbl();

    toast(`✓ ${d.count} kredi limiti yüklendi (Excel)`,'s');

    await saveKredi();

    renderLikidite();

  } else {

    toast('Hata: '+(d.msg||'Excel okunamadı'),'e');

  }

}



async function expLikiCSV(){

  const liki=window._likiData||[];

  if(!liki.length){toast('Veri yok','e');return;}

  const hdrs=['Yıl','Ay','Gelir','Gider','Net Denge','Nakit İhtiyaç','Kümülatif Banka','Tahmini Faiz','Durum'];

  const csv=[hdrs,...liki.map(o=>[o.yil,o.ay,o.gelir,o.gider,o.denge,o.nakit_ihtiyac,o.kumulatif_banka,o.toplam_faiz_maliyeti,o.pozitif?'Pozitif':o.karsilanamayan>0?'Yetersiz':'Kredi Önerisi'])].map(r=>r.map(v=>'"'+String(v||0).replace(/"/g,'""')+'"').join(',')).join(String.fromCharCode(10));

  const a=document.createElement('a');a.href='data:text/csv;charset=utf-8,\uFEFF'+encodeURIComponent(csv);a.download='likidite_analiz.csv';a.click();

  toast('CSV indirildi','s');

}





// ── FİNANSMAN YÖNETİMİ ────────────────────────────────────────────────────────

let bankaData=[], krediData=[];



async function renderFinans(){

  await loadBanka();

  await loadKredi();

  // Yıl/ay seçeneklerini doldur

  const r=await api('/api/stats'); const d=await r.json();

  const fnYil=document.getElementById('fn-yil');

  const fnAy=document.getElementById('fn-ay');

  if(fnYil&&d.yillar){

    const cur=fnYil.value;

    fnYil.innerHTML='<option value="">Tüm dönem</option>';

    d.yillar.forEach(y=>{const o=document.createElement('option');o.value=y;o.text=y;fnYil.appendChild(o);});

    fnYil.value=cur;

  }

  if(fnAy){

    const cur=fnAy.value;

    fnAy.innerHTML='<option value="">Tüm aylar</option>';

    AYLAR.forEach(m=>{const o=document.createElement('option');o.value=m;o.text=m;fnAy.appendChild(o);});

    fnAy.value=cur;

  }

}



// ── BANKA ──

async function loadBanka(){

  const r=await api('/api/banka'); const d=await r.json();

  bankaData=d.records||[];

  renderBankaList();

}



function renderBankaList(){

  const list=document.getElementById('banka-list');

  if(!list) return;

  if(!bankaData.length){

    list.innerHTML='<div style="font-size:12px;color:var(--t3);padding:8px 0">Henüz banka eklenmedi</div>';

    document.getElementById('banka-toplam').textContent='₺ 0';

    return;

  }

  list.innerHTML=bankaData.map((b,i)=>`

    <div style="display:grid;grid-template-columns:1fr 1fr 1fr auto;gap:6px;margin-bottom:6px;align-items:center">

      <input class="fi2" list="banka-datalist" style="font-size:11.5px;padding:5px 7px" value="${b.banka||''}" placeholder="Banka adı seçin veya yazın"

        oninput="bankaData[${i}].banka=this.value;saveBanka()">

      <input class="fi2" style="font-size:11.5px;padding:5px 7px" value="${b.firma||''}" placeholder="Firma"

        oninput="bankaData[${i}].firma=this.value;saveBanka()">

      <input class="fi2" style="font-size:11.5px;padding:5px 7px;text-align:right;font-family:var(--mono)" type="number"

        value="${b.bakiye||0}" placeholder="Bakiye ₺"

        oninput="bankaData[${i}].bakiye=parseFloat(this.value)||0;updateBankaToplam();saveBanka()">

      <button class="btn bgh" title="Bu bankayı sil" style="color:var(--r);padding:4px 10px;border:1px solid var(--r);border-radius:var(--rad);font-size:14px;font-weight:700" onclick="if(confirm('${(b.banka||'Bu banka')} silinsin mi?')){bankaData.splice(${i},1);renderBankaList();saveBanka();}">🗑</button>

    </div>`).join('');

  updateBankaToplam();

}



function updateBankaToplam(){

  const top=bankaData.reduce((s,b)=>s+(b.bakiye||0),0);

  const el=document.getElementById('banka-toplam');

  if(el) el.textContent='₺ '+Math.round(top).toLocaleString('tr-TR');

}



function addBankaRow(){

  bankaData.push({banka:'',firma:'ULUSAL',bakiye:0});

  renderBankaList();

}



async function saveBanka(){

  await api('/api/banka/save',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(bankaData)});

}



// ── KREDİ ──

async function loadKredi(){

  const r=await api('/api/kredi'); const d=await r.json();

  krediData=d.records||[];

  renderKrediList();

}



// ── BANKA SEÇENEKLERİ YÖNETİMİ ──

const BANKA_SEC_STORAGE_KEY = 'nakit_banka_secenekleri';

let bankaSecenekleri = [];



function loadBankaSec(){

  try {

    const saved = localStorage.getItem(BANKA_SEC_STORAGE_KEY);

    if(saved){

      bankaSecenekleri = JSON.parse(saved);

    } else {

      bankaSecenekleri = [

        'Ziraat Bankası','Vakıfbank','Halkbank','Garanti BBVA',

        'İş Bankası','YapıKredi','Akbank','TEB','QNB Finansbank','Denizbank'

      ];

    }

  } catch(e){

    bankaSecenekleri = ['Ziraat Bankası','Vakıfbank','Halkbank'];

  }

}



function saveBankaSec(){

  try{ localStorage.setItem(BANKA_SEC_STORAGE_KEY, JSON.stringify(bankaSecenekleri)); }catch(e){}

  renderBankaSecList();

  renderKrediList();

  toast('✓ Banka seçenekleri kaydedildi','s');

}



function openBankaSecModal(){

  loadBankaSec();

  renderBankaSecList();

  const m=document.getElementById('banka-sec-mod');

  if(m) m.style.display='flex';

}



function closeBankaSecModal(){

  const m=document.getElementById('banka-sec-mod');

  if(m) m.style.display='none';

}



function renderBankaSecList(){

  const cont=document.getElementById('banka-sec-list');

  if(!cont) return;

  if(!bankaSecenekleri.length){

    cont.innerHTML='<div style="font-size:12px;color:var(--t3);padding:8px 0;text-align:center">Henüz banka eklenmedi</div>';

    return;

  }

  cont.innerHTML=bankaSecenekleri.map((b,i)=>`

    <div style="display:flex;align-items:center;gap:8px;padding:6px 10px;background:var(--s3);border:1px solid var(--b1);border-radius:var(--rad);margin-bottom:5px">

      <span style="flex:1;font-size:12px;color:var(--t1)">${b}</span>

      <button class="btn bgh" style="color:var(--r);padding:3px 9px;font-size:11px;border:1px solid rgba(239,68,68,.3);border-radius:var(--rad)" onclick="deleteBankaSec(${i})">✕ Sil</button>

    </div>`).join('');

}



function addBankaSec(){

  const inp=document.getElementById('banka-sec-input');

  const val=(inp.value||'').trim();

  if(!val){ toast('Banka adı boş olamaz','e'); return; }

  if(bankaSecenekleri.includes(val)){ toast('Bu banka zaten listede','e'); return; }

  bankaSecenekleri.push(val);

  inp.value='';

  renderBankaSecList();

}



function deleteBankaSec(i){

  if(!confirm(bankaSecenekleri[i]+' listeden çıkarılsın mı?')) return;

  bankaSecenekleri.splice(i,1);

  renderBankaSecList();

}



// Başlangıçta yükle

loadBankaSec();



function renderKrediList(){

  const list=document.getElementById('kredi-list');

  if(!list) return;

  if(!krediData.length){

    list.innerHTML='<div style="font-size:12px;color:var(--t3);padding:8px 0">Henüz kredi eklenmedi</div>';

    document.getElementById('kredi-bos-toplam').textContent='₺ 0';

    return;

  }

  list.innerHTML=krediData.map((k,i)=>`

    <div style="background:var(--s3);border:1px solid var(--b1);border-radius:var(--rad);padding:9px;margin-bottom:8px">

      <div style="display:grid;grid-template-columns:1.5fr 1fr 1fr auto;gap:5px;margin-bottom:6px;align-items:center">

        <select class="fi" style="font-size:11px" onchange="krediData[${i}].banka=this.value;saveKredi()">

          <option value="">-- Banka Seçin --</option>

          ${bankaSecenekleri.map(b=>`<option${k.banka===b?' selected':''}>${b}</option>`).join('')}

          ${k.banka && !bankaSecenekleri.includes(k.banka) ? `<option selected value="${k.banka}">${k.banka} (mevcut)</option>` : ''}

        </select>

        <select class="fi" style="font-size:11px" onchange="krediData[${i}].firma=this.value;saveKredi()">

          ${['ULUSAL','ABC','BRG','BRK'].map(f=>`<option${k.firma===f?' selected':''}>${f}</option>`).join('')}

        </select>

        <select class="fi" style="font-size:11px" onchange="krediData[${i}].tur=this.value;saveKredi()">

          ${['Rotatif','Spot Kredi','KMH','Prefinansman','Diğer'].map(t=>`<option${k.tur===t?' selected':''}>${t}</option>`).join('')}

        </select>

        <button class="btn bgh" title="Bu krediyi sil" style="color:var(--r);padding:4px 10px;border:1px solid var(--r);border-radius:var(--rad);font-size:14px;font-weight:700;line-height:1" onclick="if(confirm('${(k.banka||'Bu kredi')} silinsin mi?')){krediData.splice(${i},1);renderKrediList();saveKredi();}">🗑</button>

      </div>

      <div style="display:grid;grid-template-columns:1fr 1fr 1fr 1fr;gap:5px;align-items:center">

        <div>

          <div style="font-size:9.5px;color:var(--t3);font-family:var(--mono);margin-bottom:2px">Limit ₺</div>

          <input class="fi2" type="number" style="font-size:11.5px;padding:4px 7px;font-family:var(--mono)" value="${k.limit||0}"

            oninput="krediData[${i}].limit=parseFloat(this.value)||0;updateKrediToplam();saveKredi()">

        </div>

        <div>

          <div style="font-size:9.5px;color:var(--t3);font-family:var(--mono);margin-bottom:2px">Kullanılan ₺</div>

          <input class="fi2" type="number" style="font-size:11.5px;padding:4px 7px;font-family:var(--mono)" value="${k.kullanilan||0}"

            oninput="krediData[${i}].kullanilan=parseFloat(this.value)||0;updateKrediToplam();saveKredi()">

        </div>

        <div>

          <div style="font-size:9.5px;color:var(--t3);font-family:var(--mono);margin-bottom:2px">Boş ₺</div>

          <div style="font-size:12px;font-family:var(--mono);color:var(--acc);padding:4px 7px;background:var(--s2);border:1px solid var(--b1);border-radius:var(--rad)">

            ${Math.round((k.limit||0)-(k.kullanilan||0)).toLocaleString('tr-TR')}

          </div>

        </div>

        <div>

          <div style="font-size:9.5px;color:var(--t3);font-family:var(--mono);margin-bottom:2px">Faiz % / ay</div>

          <input class="fi2" type="number" step="0.01" style="font-size:11.5px;padding:4px 7px;font-family:var(--mono)" value="${k.faiz||0}"

            oninput="krediData[${i}].faiz=parseFloat(this.value)||0;saveKredi()">

        </div>

      </div>

      <div style="margin-top:5px">

        <div style="height:4px;background:var(--b1);border-radius:2px;overflow:hidden">

          <div style="height:100%;width:${k.limit?Math.min(100,Math.round((k.kullanilan||0)/k.limit*100)):0}%;background:${(k.kullanilan||0)/(k.limit||1)>0.8?'var(--r)':'var(--acc)'};border-radius:2px"></div>

        </div>

        <div style="font-size:10px;color:var(--t3);font-family:var(--mono);margin-top:2px">

          ${k.limit?Math.round((k.kullanilan||0)/k.limit*100):0}% kullanıldı

        </div>

      </div>

    </div>`).join('');

  updateKrediToplam();

}



function updateKrediToplam(){

  const bos=krediData.reduce((s,k)=>s+Math.max(0,(k.limit||0)-(k.kullanilan||0)),0);

  const el=document.getElementById('kredi-bos-toplam');

  if(el) el.textContent='₺ '+Math.round(bos).toLocaleString('tr-TR');

}



function addKrediRow(){

  krediData.push({banka:'',firma:'ULUSAL',tur:'Rotatif',limit:0,kullanilan:0,faiz:0});

  renderKrediList();

}



async function saveKredi(){

  await api('/api/kredi/save',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(krediData)});

}



// ── FİNANSMAN ÖNERİ MOTORU ──

async function hesaplaOneri(){

  const yil=document.getElementById('fn-yil')?.value||'';

  const ay=document.getElementById('fn-ay')?.value||'';

  const btn=document.querySelector('#vw-finans .btn.bp:last-child');

  if(btn){btn.textContent='⏳';btn.disabled=true;}

  const params=new URLSearchParams();

  if(ay) params.set('ay',ay);

  if(yil) params.set('yil',yil);

  const r=await api('/api/finansman/oneri?'+params);

  const d=await r.json();

  if(btn){btn.textContent='Analiz Et';btn.disabled=false;}



  const fmtN=n=>Math.round(n||0).toLocaleString('tr-TR');



  // Özet metrikler

  document.getElementById('fn-mets').innerHTML=`

    <div class="met"><div class="met-stripe" style="background:var(--r)"></div>

      <div class="mlb">Bekleyen Gider</div>

      <div class="mvl" style="color:var(--r);font-size:15px">₺${fmtN(d.gider_bekl)}</div></div>

    <div class="met"><div class="met-stripe" style="background:var(--g)"></div>

      <div class="mlb">Beklenen Gelir</div>

      <div class="mvl" style="color:var(--g);font-size:15px">₺${fmtN(d.gelir_bekl)}</div></div>

    <div class="met"><div class="met-stripe" style="background:var(--acc)"></div>

      <div class="mlb">Banka Bakiyesi</div>

      <div class="mvl" style="color:var(--acc);font-size:15px">₺${fmtN(d.toplam_banka)}</div></div>

    <div class="met"><div class="met-stripe" style="background:${d.net_acik>0?'var(--am)':'var(--g)'}"></div>

      <div class="mlb">Net Finansman İht.</div>

      <div class="mvl" style="color:${d.net_acik>0?'var(--am)':'var(--g)'};font-size:15px">

        ${d.net_acik>0?'₺'+fmtN(d.net_acik):'Gerek yok'}

      </div></div>`;



  const onDiv=document.getElementById('fn-oneriler');



  if(d.net_acik<=0){

    onDiv.innerHTML=`<div style="background:rgba(52,211,153,.1);border:1px solid rgba(52,211,153,.25);border-radius:var(--rad);padding:14px 16px;font-size:13px;color:var(--g)">

      ✓ Bu dönem için finansman ihtiyacı yok. Mevcut bakiye ve gelirler giderleri karşılıyor.

    </div>`;

    document.getElementById('fn-kredi-tablo').innerHTML='';

    return;

  }



  // Öneri kartları

  let onHtml='';

  if(d.oneriler&&d.oneriler.length>0){

    onHtml=`<div style="font-size:12px;font-weight:600;color:var(--t1);margin-bottom:8px">

      Önerilen Kullanım Sırası — Toplam: ₺${fmtN(d.toplam_kullanilacak)} · Tahmini Aylık Maliyet: ₺${fmtN(d.tahmini_maliyet)}

    </div>

    <div style="display:grid;grid-template-columns:repeat(auto-fill,minmax(220px,1fr));gap:8px;margin-bottom:10px">

      ${d.oneriler.map(o=>`

        <div style="background:var(--s2);border:1px solid var(--b1);border-radius:var(--rad);padding:11px;border-left:3px solid var(--acc)">

          <div style="font-family:var(--mono);font-size:9px;color:var(--acc);margin-bottom:4px">ÖNCELİK ${o.oncelik}</div>

          <div style="font-weight:600;font-size:13px;margin-bottom:2px">${o.banka}</div>

          <div style="font-size:11px;color:var(--t2);margin-bottom:6px">${o.tur} · <span style="color:var(--acc2)">${o.firma}</span></div>

          <div style="display:flex;justify-content:space-between;font-family:var(--mono);font-size:11.5px">

            <span style="color:var(--t2)">Kullan</span>

            <span style="color:var(--acc);font-weight:600">₺${fmtN(o.kullan)}</span>

          </div>

          <div style="display:flex;justify-content:space-between;font-family:var(--mono);font-size:11.5px;margin-top:2px">

            <span style="color:var(--t2)">Faiz % / ay</span>

            <span style="color:var(--t1)">${o.faiz}%</span>

          </div>

          <div style="display:flex;justify-content:space-between;font-family:var(--mono);font-size:11.5px;margin-top:2px">

            <span style="color:var(--t2)">Aylık maliyet</span>

            <span style="color:var(--am)">₺${fmtN(o.aylik_faiz)}</span>

          </div>

          <div style="margin-top:6px;height:3px;background:var(--b1);border-radius:2px;overflow:hidden">

            <div style="height:100%;width:${o.bos>0?Math.min(100,Math.round(o.kullan/o.bos*100)):100}%;background:var(--acc);border-radius:2px"></div>

          </div>

          <div style="font-size:10px;color:var(--t3);font-family:var(--mono);margin-top:2px">

            Limitin ${o.bos>0?Math.min(100,Math.round(o.kullan/o.bos*100)):100}% kullanılıyor

          </div>

        </div>`).join('')}

    </div>`;

    if(d.kalan_karsilanamiyan>0){

      onHtml+=`<div style="background:rgba(248,113,113,.1);border:1px solid rgba(248,113,113,.25);border-radius:var(--rad);padding:12px 14px;font-size:12.5px;color:var(--r);margin-top:6px">

        ⚠ Mevcut kredi limitleri ₺${fmtN(d.kalan_karsilanamiyan)} tutarındaki açığı karşılamaya yetmiyor.

        Ek finansman kaynağı gerekli.

      </div>`;

    } else {

      onHtml+=`<div style="background:rgba(52,211,153,.08);border:1px solid rgba(52,211,153,.2);border-radius:var(--rad);padding:10px 14px;font-size:12px;color:var(--g-d);margin-top:6px">

        ✓ Yukarıdaki kullanım planı ile finansman ihtiyacı tam karşılanıyor.

      </div>`;

    }

  } else {

    onHtml=`<div style="background:rgba(248,113,113,.1);border:1px solid rgba(248,113,113,.25);border-radius:var(--rad);padding:14px;font-size:13px;color:var(--r)">

      ⚠ Kullanılabilir kredi limiti bulunamadı. Lütfen kredi limitlerinizi güncelleyin.

    </div>`;

  }

  onDiv.innerHTML=onHtml;



  // Tüm kullanılabilir krediler tablosu

  const allR=await api('/api/finansman/oneri?'+params); const allD=await allR.json();

  const ktDiv=document.getElementById('fn-kredi-tablo');

  const kredilerSiralı=(allD.oneriler||[]);

  if(kredilerSiralı.length>0){

    ktDiv.innerHTML=`

      <div style="font-size:12px;font-weight:600;color:var(--t2);margin-bottom:7px">

        Tüm Kullanılabilir Kredi Limitleri — Düşük Faizden Yükseğe

      </div>

      <div class="tc">

        <table style="font-size:12px">

          <thead><tr>

            <th>Banka</th><th>Firma</th><th>Tür</th>

            <th class="num">Boş Limit ₺</th>

            <th class="num">Faiz % / ay</th>

            <th class="num">Tahmini Maliyet ₺</th>

            <th style="width:40px"></th>

          </tr></thead>

          <tbody>

            ${krediData.filter(k=>(k.limit||0)-(k.kullanilan||0)>0)

              .sort((a,b)=>(a.faiz||0)-(b.faiz||0))

              .map(k=>{

                const idx=krediData.indexOf(k);

                const bos=(k.limit||0)-(k.kullanilan||0);

                const mal=Math.round(bos*(k.faiz||0)/100);

                return `<tr>

                  <td><strong>${k.banka}</strong></td>

                  <td><span class="bdg ${FC[k.firma]||'bbi'}">${k.firma}</span></td>

                  <td style="font-size:11px;color:var(--t2)">${k.tur}</td>

                  <td class="num" style="color:var(--acc);font-family:var(--mono)">${fmtN(bos)}</td>

                  <td class="num" style="font-family:var(--mono)">${k.faiz||0}%</td>

                  <td class="num" style="color:var(--am);font-family:var(--mono)">${fmtN(mal)}</td>

                  <td style="text-align:center"><button title="Sil" style="background:none;border:1px solid var(--r);border-radius:4px;color:var(--r);cursor:pointer;font-size:13px;padding:2px 7px" onclick="if(confirm('${k.banka} silinsin mi?')){krediData.splice(${idx},1);saveKredi();renderFinansman();}">🗑</button></td>

                </tr>`;}).join('')}

          </tbody>

        </table>

      </div>`;

  }

}





// ── CARİ HESAPLAR ─────────────────────────────────────────────────────────────

let cariEditIdx = null;

let cariPg = 1; const CPPP = 30;



function cariTab(tab){

  ['ozet','detay'].forEach(t=>{

    document.getElementById('ctab-'+t+'-view').style.display = t===tab?'':'none';

    const btn=document.getElementById('ctab-'+t);

    if(btn){ btn.style.borderBottom=t===tab?'2px solid var(--acc)':'none'; btn.style.color=t===tab?'var(--acc)':''; }

  });

  if(tab==='ozet')  renderCariOzet();

  if(tab==='detay') renderCariDetay();

}



async function renderCari(){

  // Özet metrikleri yükle

  const r = await api('/api/cari'); const d = await r.json();

  const fm = n=>{n=n||0;if(n>=1e6)return'₺'+(n/1e6).toFixed(2)+' M';if(n>=1e3)return'₺'+Math.round(n/1e3)+'K';return'₺'+fmtN(n);};

  document.getElementById('cari-mets').innerHTML=`

    <div class="met"><div class="met-stripe" style="background:var(--g)"></div>

      <div class="mlb">Toplam Alacak</div>

      <div class="mvl" style="color:var(--g);font-size:17px">${fm(d.toplam_alacak)}</div>

      <div class="msb">Bize borçlu firmalar</div></div>

    <div class="met"><div class="met-stripe" style="background:var(--r)"></div>

      <div class="mlb">Toplam Borç</div>

      <div class="mvl" style="color:var(--r);font-size:17px">${fm(d.toplam_borc)}</div>

      <div class="msb">Biz borçlu firmalar</div></div>

    <div class="met"><div class="met-stripe" style="background:${d.net>=0?'var(--g)':'var(--r)'}"></div>

      <div class="mlb">Net Pozisyon</div>

      <div class="mvl" style="color:${d.net>=0?'var(--g)':'var(--r)'};font-size:17px">${d.net>=0?'+':''}${fm(d.net)}</div>

      <div class="msb">${d.net>=0?'Net alacaklı durumu':'Net borçlu durumu'}</div></div>

    <div class="met"><div class="met-stripe" style="background:var(--acc)"></div>

      <div class="mlb">Toplam Firma</div>

      <div class="mvl" style="color:var(--acc);font-size:17px">${d.firmalar.length}</div>

      <div class="msb">${d.total} hareket kaydı</div></div>

  `;

  // Firma datalist güncelle

  const dl=document.getElementById('cari-firma-list');

  if(dl){ dl.innerHTML=''; d.firmalar.forEach(f=>{const o=document.createElement('option');o.value=f;dl.appendChild(o);}); }

  // Firma filtresi güncelle

  const fs=document.getElementById('cari-firma-f');

  if(fs){ const cur=fs.value; fs.innerHTML='<option value="">Tüm Firmalar</option>';

    d.firmalar.forEach(f=>{const o=document.createElement('option');o.value=f;o.text=f;fs.appendChild(o);}); fs.value=cur; }

  renderCariOzet();

}



async function renderCariOzet(){

  const r = await api('/api/cari/ozet'); const d = await r.json();

  let ozet = d.ozet || [];

  const q = (document.getElementById('cari-ozet-q')?.value||'').toLowerCase();

  const tur = document.getElementById('cari-ozet-tur')?.value||'';

  if(q) ozet = ozet.filter(o=>o.cari_firma.toLowerCase().includes(q));

  if(tur==='ALACAK') ozet = ozet.filter(o=>o.alacak > 0);

  if(tur==='BORC')   ozet = ozet.filter(o=>o.borc > 0);



  const tbody = document.getElementById('cari-ozet-tbody');

  if(!ozet.length){

    tbody.innerHTML='<tr><td colspan="8" style="padding:24px;text-align:center;color:var(--t3)">Henüz cari kayıt yok — sağ üstten hareket ekleyin</td></tr>';

    document.getElementById('cari-ozet-tfoot').innerHTML='';

    return;

  }

  tbody.innerHTML = ozet.map(o=>{

    const netCls = o.net>0?'color:var(--g)':o.net<0?'color:var(--r)':'color:var(--t3)';

    const netPfx = o.net>0?'+':'';

    const durum  = o.net>0

      ? '<span class="bdg bgi">Alacaklı</span>'

      : o.net<0

        ? '<span class="bdg bri">Borçlu</span>'

        : '<span class="bdg bbi">Sıfır</span>';

    return `<tr style="cursor:pointer" onclick="filterDetayByFirma('${o.cari_firma}')">

      <td style="font-weight:500">${o.cari_firma}</td>

      <td><span class="bdg ${FC[o.grup_firma]||'bbi'}" style="font-size:10px">${o.grup_firma||'—'}</span></td>

      <td class="num" style="color:var(--g)">${o.alacak?fmtN(o.alacak):'—'}</td>

      <td class="num" style="color:var(--r)">${o.borc?fmtN(o.borc):'—'}</td>

      <td class="num" style="${netCls};font-weight:600">${netPfx}${fmtN(o.net)}</td>

      <td>${durum}</td>

      <td style="text-align:center;font-size:11px;color:var(--t3)">${o.hareket_sayisi}</td>

      <td><button class="btn bgh" onclick="event.stopPropagation();filterDetayByFirma('${o.cari_firma}')" title="Hareketleri gör" style="font-size:11px">≡</button></td>

    </tr>`;

  }).join('');



  // Footer

  const totA=ozet.reduce((s,o)=>s+o.alacak,0);

  const totB=ozet.reduce((s,o)=>s+o.borc,0);

  const totN=totA-totB;

  document.getElementById('cari-ozet-tfoot').innerHTML=`

    <tr style="border-top:2px solid var(--b2);font-weight:600">

      <td colspan="2" style="padding:8px 10px;font-family:var(--mono);font-size:11px;color:var(--t2)">TOPLAM (${ozet.length} firma)</td>

      <td class="num" style="color:var(--g)">${fmtN(totA)}</td>

      <td class="num" style="color:var(--r)">${fmtN(totB)}</td>

      <td class="num" style="color:${totN>=0?'var(--g)':'var(--r)'};font-weight:700">${totN>=0?'+':''}${fmtN(totN)}</td>

      <td colspan="3"></td>

    </tr>`;

}



function filterDetayByFirma(firma){

  const fs=document.getElementById('cari-firma-f');

  if(fs) fs.value=firma;

  cariTab('detay');

}



async function renderCariDetay(){

  const q=(document.getElementById('cari-q')?.value||'');

  const firma=document.getElementById('cari-firma-f')?.value||'';

  const tur=document.getElementById('cari-tur-f')?.value||'';

  const durum=document.getElementById('cari-durum-f')?.value||'';

  const params=new URLSearchParams();

  if(q) params.set('q',q); if(firma) params.set('firma',firma);

  if(tur) params.set('tur',tur); if(durum) params.set('durum',durum);

  const r=await api('/api/cari?'+params); const d=await r.json();

  const rows=d.records;



  // Tarihe göre sırala - en yeni önce

  rows.sort((a,b)=>(b.tarih||'').localeCompare(a.tarih||''));



  const start=(cariPg-1)*CPPP, end=start+CPPP;

  const page=rows.slice(start,end);



  document.getElementById('cari-detay-tbody').innerHTML=page.map(rec=>{

    const idx=rec._idx;

    const turCls=rec.tur==='ALACAK'?'bgi':'bri';

    const durumCls=rec.durum==='KAPALI'?'bgi':rec.durum==='VADEDE'?'bami':'bri';

    const vadeWarn = rec.vade && rec.durum==='AÇIK' && new Date(rec.vade.split('.').reverse().join('-')) < new Date()

      ? ' style="color:var(--r)"' : '';

    return `<tr>

      <td style="font-size:12px;font-weight:500">${rec.cari_firma||'—'}</td>

      <td><span class="bdg ${FC[rec.grup_firma]||'bbi'}" style="font-size:10px">${rec.grup_firma||'—'}</span></td>

      <td><span class="bdg ${turCls}">${rec.tur||'—'}</span></td>

      <td style="font-family:var(--mono);font-size:11px;color:var(--t2)">${rec.belge_no||'—'}</td>

      <td style="font-family:var(--mono);font-size:11px">${rec.tarih||'—'}</td>

      <td style="font-family:var(--mono);font-size:11px"${vadeWarn}>${rec.vade||'—'}${vadeWarn?' ⚠':''}</td>

      <td class="num" style="font-weight:600;color:${rec.tur==='ALACAK'?'var(--g)':'var(--r)'}">${fmtN(rec.tutar)}</td>

      <td><span class="bdg ${durumCls} cst" ondblclick="toggleCari(${idx})" title="Çift tıkla: durum değiştir">${rec.durum||'AÇIK'}</span></td>

      <td style="font-size:11px;color:var(--t2)" class="clip" title="${rec.aciklama||''}">${rec.aciklama?rec.aciklama.slice(0,20)+(rec.aciklama.length>20?'…':''):'—'}</td>

      <td style="white-space:nowrap">

        <button class="btn bgh" onclick="openCariEdit(${idx})" title="Düzenle">✏</button>

        <button class="btn bgh" style="color:var(--r)" onclick="delCari(${idx})" title="Sil">✕</button>

      </td>

    </tr>`;

  }).join('');



  // Pagination

  const tot=Math.ceil(rows.length/CPPP)||1;

  let ph=`<span class="pgi">${rows.length} hareket</span><div class="pgsp"></div>`;

  if(cariPg>1) ph+=`<button class="pbn" onclick="cariPg--;renderCariDetay()">‹</button>`;

  for(let i=Math.max(1,cariPg-2);i<=Math.min(tot,cariPg+2);i++)

    ph+=`<button class="pbn${i===cariPg?' on':''}" onclick="cariPg=${i};renderCariDetay()">${i}</button>`;

  if(cariPg<tot) ph+=`<button class="pbn" onclick="cariPg++;renderCariDetay()">›</button>`;

  document.getElementById('cari-pgn').innerHTML=ph;

}



// MODAL

function openCariAdd(){

  cariEditIdx=null;

  document.getElementById('cari-mtit').textContent='Yeni Cari Hareket';

  document.getElementById('cm-save').textContent='Kaydet';

  ['cm-firma','cm-belge','cm-aciklama'].forEach(id=>document.getElementById(id).value='');

  document.getElementById('cm-grup').value='ULUSAL';

  document.getElementById('cm-tur').value='ALACAK';

  document.getElementById('cm-durum').value='AÇIK';

  document.getElementById('cm-tutar').value='';

  document.getElementById('cm-tarih').value=new Date().toISOString().slice(0,10);

  document.getElementById('cm-vade').value='';

  document.getElementById('cari-mod').classList.add('on');

}

function openCariEdit(idx){

  cariEditIdx=idx;

  // Veriyi API'den al

  api('/api/cari').then(r=>r.json()).then(d=>{

    const rec=d.records.find(r=>r._idx===idx);

    if(!rec) return;

    document.getElementById('cari-mtit').textContent='Cari Hareketi Düzenle';

    document.getElementById('cm-save').textContent='Güncelle';

    document.getElementById('cm-firma').value=rec.cari_firma||'';

    document.getElementById('cm-grup').value=rec.grup_firma||'ULUSAL';

    document.getElementById('cm-tur').value=rec.tur||'ALACAK';

    document.getElementById('cm-belge').value=rec.belge_no||'';

    document.getElementById('cm-aciklama').value=rec.aciklama||'';

    document.getElementById('cm-durum').value=rec.durum||'AÇIK';

    document.getElementById('cm-tutar').value=rec.tutar||'';

    if(rec.tarih&&rec.tarih.length===10){

      const p=rec.tarih.split('.');

      document.getElementById('cm-tarih').value=p[2]+'-'+p[1]+'-'+p[0];

    }

    if(rec.vade&&rec.vade.length===10){

      const p=rec.vade.split('.');

      document.getElementById('cm-vade').value=p[2]+'-'+p[1]+'-'+p[0];

    }

    document.getElementById('cari-mod').classList.add('on');

  });

}

function closeCariMod(){ document.getElementById('cari-mod').classList.remove('on'); }



function fmtDate(v){ if(!v) return ''; const d=new Date(v); return d.toLocaleDateString('tr-TR',{day:'2-digit',month:'2-digit',year:'numeric'}); }



async function saveCari(){

  const rec={

    cari_firma: document.getElementById('cm-firma').value.trim(),

    grup_firma: document.getElementById('cm-grup').value,

    tur:        document.getElementById('cm-tur').value,

    belge_no:   document.getElementById('cm-belge').value.trim(),

    tarih:      fmtDate(document.getElementById('cm-tarih').value),

    vade:       fmtDate(document.getElementById('cm-vade').value),

    tutar:      parseFloat(document.getElementById('cm-tutar').value)||0,

    durum:      document.getElementById('cm-durum').value,

    aciklama:   document.getElementById('cm-aciklama').value.trim(),

  };

  if(!rec.cari_firma){ toast('Firma adı zorunlu','e'); return; }

  if(!rec.tutar)      { toast('Tutar zorunlu','e'); return; }



  const url=cariEditIdx!==null?'/api/cari/edit':'/api/cari/add';

  const body=cariEditIdx!==null?JSON.stringify({idx:cariEditIdx,record:rec}):JSON.stringify(rec);

  const r=await api(url,{method:'POST',headers:{'Content-Type':'application/json'},body});

  if((await r.json()).ok){

    toast(cariEditIdx!==null?'Hareket güncellendi':'Hareket eklendi','s');

    closeCariMod(); renderCari();

  }

}

async function delCari(idx){

  if(!confirm('Bu hareketi silmek istediğinizden emin misiniz?')) return;

  const r=await api('/api/cari/delete',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({idx})});

  if((await r.json()).ok){ toast('Hareket silindi','i'); renderCari(); }

}

async function toggleCari(idx){

  const r=await api('/api/cari/toggle',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({idx})});

  const d=await r.json();

  if(d.ok){ toast(d.durum==='KAPALI'?'✓ Kapandı':'Açıldı', d.durum==='KAPALI'?'s':'i'); renderCariDetay(); renderCariOzet(); }

}



async function expCariExcel(){

  const r=await api('/api/cari/ozet'); const d=await r.json();

  const ozet=d.ozet||[];

  if(!ozet.length){ toast('Veri yok','e'); return; }

  const hdrs=['Cari Firma','Grup Firma','Alacak','Borç','Net Bakiye','Son Hareket','Hareket Sayısı'];

  const csv=[hdrs,...ozet.map(o=>[o.cari_firma,o.grup_firma,o.alacak,o.borc,o.net,o.son_hareket,o.hareket_sayisi])]

    .map(r=>r.map(v=>'"'+String(v||0).replace(/"/g,'""')+'"').join(',')).join(String.fromCharCode(10));

  const a=document.createElement('a');

  a.href='data:text/csv;charset=utf-8,\uFEFF'+encodeURIComponent(csv);

  a.download='cari_hesaplar.csv'; a.click();

  toast('CSV indirildi','s');

}





// ── BANKA API & EKSTRE ────────────────────────────────────────────────────────

let _ekstreData = null;

let _apiEditIdx = null;



function baTab(tab){

  ['ekstre','api'].forEach(t=>{

    document.getElementById('batab-'+t+'-view').style.display = t===tab?'':'none';

    const btn=document.getElementById('batab-'+t);

    if(btn){ btn.style.borderBottom=t===tab?'2px solid var(--acc)':'none'; btn.style.color=t===tab?'var(--acc)':''; }

  });

  if(tab==='api') renderApiList();

  if(tab==='ekstre') renderEkstreGecmis();

}



async function renderBankaApi(){

  renderEkstreGecmis();

}



// ── EKSTRE ──────────────────────────────────────────────────────────────────

function handleEkstreFile(file){

  if(!file) return;

  const zone = document.getElementById('ekstre-drop-zone');

  zone.innerHTML = '<div style="font-size:14px;color:var(--acc)">⏳ '+file.name+' okunuyor...</div>';



  const formData = new FormData();

  formData.append('file', file, file.name);



  fetch('/api/banka/import_ekstre', {method:'POST', body: formData})

    .then(r=>r.json())

    .then(d=>{

      zone.innerHTML = '<div style="font-size:13px;color:var(--t2)">📂 Dosya seç veya sürükle</div><input type="file" id="ekstre-file-input" style="display:none" accept=".xlsx,.xls,.csv,.txt" onchange="handleEkstreFile(this.files[0])">';

      if(!d.ok){ toast('Hata: '+(d.msg||'Okunamadı'),'e'); return; }

      _ekstreData = d;

      showEkstrePreview(d);

    })

    .catch(e=>{ toast('Hata: '+e.message,'e'); });

}



function showEkstrePreview(d){

  const resDiv = document.getElementById('ekstre-result');

  resDiv.style.display = '';



  const infoDiv = document.getElementById('ekstre-info');

  const statusClr = d.banka!=='BILINMIYOR'?'var(--g)':'var(--am)';

  infoDiv.innerHTML = `

    <span>🏦 <strong style="color:${statusClr}">${d.banka}</strong></span>

    <span>📊 ${d.count} hareket</span>

    ${d.bakiye_son!=null?`<span>💰 Son Bakiye: <strong style="color:var(--acc)">₺${fmtN(d.bakiye_son)}</strong></span>`:''}

  `;



  const tbody = document.getElementById('ekstre-tbody');

  const rows = (d.records||[]).slice(0,50);

  tbody.innerHTML = rows.map(r=>`<tr>

    <td style="font-family:var(--mono);font-size:11px">${r.tarih||'—'}</td>

    <td style="font-size:11px" class="clip" title="${r.aciklama||''}">${r.aciklama?r.aciklama.slice(0,45)+(r.aciklama.length>45?'…':''):'—'}</td>

    <td class="num" style="color:var(--r);font-family:var(--mono);font-size:11px">${r.borc?fmtN(r.borc):'—'}</td>

    <td class="num" style="color:var(--g);font-family:var(--mono);font-size:11px">${r.alacak?fmtN(r.alacak):'—'}</td>

    <td class="num" style="font-family:var(--mono);font-size:11px">${r.bakiye!=null?fmtN(r.bakiye):'—'}</td>

  </tr>`).join('');

  if(d.count>50) tbody.innerHTML += `<tr><td colspan="5" style="text-align:center;color:var(--t3);font-size:11px;padding:8px">... ${d.count-50} satır daha</td></tr>`;

}



async function saveEkstreBakiye(){

  if(!_ekstreData || _ekstreData.bakiye_son == null){

    toast('Son bakiye bilgisi bulunamadı','e'); return;

  }

  // Bakiye listesini güncelle - aynı banka varsa üzerine yaz

  const r0 = await api('/api/banka'); const d0 = await r0.json();

  let bankaList = d0.records || [];

  const today = new Date().toISOString().slice(0,10);

  const existing = bankaList.findIndex(b=>b.banka===_ekstreData.banka);

  const newEntry = {

    firma: 'ULUSAL', banka: _ekstreData.banka,

    hesap_turu: 'Vadesiz TL', bakiye: _ekstreData.bakiye_son,

    tarih: today, kaynak: 'ekstre'

  };

  if(existing>=0) bankaList[existing] = {...bankaList[existing], ...newEntry};

  else bankaList.push(newEntry);



  const r = await api('/api/banka/save',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(bankaList)});

  if((await r.json()).ok){

    toast(`✓ ${_ekstreData.banka} bakiyesi güncellendi: ₺${fmtN(_ekstreData.bakiye_son)}`,'s');

    resetEkstre(); renderEkstreGecmis();

    // Likidite sayfası açıksa yenile

    if(document.getElementById('vw-liki')?.classList.contains('on')) renderLikidite();

  }

}



function resetEkstre(){

  _ekstreData = null;

  document.getElementById('ekstre-result').style.display='none';

  document.getElementById('ekstre-drop-zone').innerHTML=`

    <div style="font-size:28px;margin-bottom:8px">📂</div>

    <div style="font-size:13px;color:var(--t2)">Ekstre dosyasını buraya sürükleyin veya tıklayın</div>

    <div style="font-size:11px;color:var(--t3);margin-top:4px">Excel (.xlsx, .xls) veya CSV (.csv)</div>

    <input type="file" id="ekstre-file-input" style="display:none" accept=".xlsx,.xls,.csv,.txt" onchange="handleEkstreFile(this.files[0])">`;

  document.getElementById('ekstre-drop-zone').onclick=()=>document.getElementById('ekstre-file-input').click();

}



async function renderEkstreGecmis(){

  const r = await api('/api/banka'); const d = await r.json();

  const list = (d.records||[]).filter(b=>b.kaynak==='ekstre');

  const el = document.getElementById('ekstre-gecmis');

  if(!el) return;

  if(!list.length){ el.innerHTML='Henüz ekstre yüklenmedi'; return; }

  el.innerHTML=`<div class="tc"><table><thead><tr>

    <th>Banka</th><th>Firma</th><th class="num">Bakiye ₺</th><th>Güncelleme</th>

  </tr></thead><tbody>${list.map(b=>`<tr>

    <td>${b.banka}</td>

    <td><span class="bdg ${FC[b.firma]||'bbi'}">${b.firma}</span></td>

    <td class="num" style="font-family:var(--mono)">${fmtN(b.bakiye)}</td>

    <td style="font-size:11px;color:var(--t2)">${b.tarih||'—'}</td>

  </tr>`).join('')}</tbody></table></div>`;

}



// ── API AYARLARI ─────────────────────────────────────────────────────────────

let apiList = [];



async function renderApiList(){

  const r = await api('/api/banka_api/config'); const d = await r.json();

  apiList = d.config || [];

  const el = document.getElementById('api-list');

  if(!apiList.length){

    el.innerHTML='<div style="color:var(--t3);font-size:12px;padding:12px 0">Henüz API bağlantısı eklenmedi. Bankadan API bilgilerinizi aldıktan sonra buraya girin.</div>';

    return;

  }

  el.innerHTML = apiList.map((a,i)=>`

    <div style="background:var(--s2);border:1px solid ${a.aktif?'var(--g-d)':'var(--b1)'};border-radius:var(--rad);padding:14px 16px;display:flex;align-items:center;gap:12px;flex-wrap:wrap">

      <div style="flex:1;min-width:200px">

        <div style="font-weight:600;font-size:13px">${a.banka_adi||a.banka} <span style="font-size:11px;color:var(--t3)">${a.tanim||''}</span></div>

        <div style="font-size:11px;color:var(--t3);margin-top:3px;font-family:var(--mono)">${a.hesap||''} · ${a.firma||''}</div>

        ${a.not?`<div style="font-size:11px;color:var(--am);margin-top:3px">📝 ${a.not}</div>`:''}

      </div>

      <div style="display:flex;align-items:center;gap:8px">

        ${a.aktif

          ? '<span class="bdg bgi">✓ Aktif</span>'

          : '<span class="bdg bami">⏳ Beklemede</span>'

        }

        <button class="btn bo bsm" onclick="editApiEntry(${i})">✏ Düzenle</button>

        <button class="btn bgh bsm" style="color:var(--r)" onclick="delApiEntry(${i})">✕</button>

      </div>

    </div>`).join('');

}



function addApiEntry(){

  _apiEditIdx = null;

  document.getElementById('api-edit-title').textContent = 'Yeni API Bağlantısı';

  ['api-tanim','api-url','api-key','api-secret','api-hesap','api-not'].forEach(id=>document.getElementById(id).value='');

  document.getElementById('api-banka').value='VAKIFBANK';

  document.getElementById('api-firma').value='ULUSAL';

  document.getElementById('api-aktif').checked=false;

  document.getElementById('api-edit-area').style.display='';

  document.getElementById('api-test-result').style.display='none';

  document.getElementById('api-edit-area').scrollIntoView({behavior:'smooth'});

}



function editApiEntry(i){

  _apiEditIdx = i;

  const a = apiList[i];

  document.getElementById('api-edit-title').textContent = 'API Bağlantısını Düzenle';

  document.getElementById('api-banka').value   = a.banka||'VAKIFBANK';

  document.getElementById('api-tanim').value   = a.tanim||'';

  document.getElementById('api-url').value     = a.url||'';

  document.getElementById('api-key').value     = a.api_key||'';

  document.getElementById('api-secret').value  = a.api_secret||'';

  document.getElementById('api-hesap').value   = a.hesap||'';

  document.getElementById('api-firma').value   = a.firma||'ULUSAL';

  document.getElementById('api-aktif').checked = a.aktif||false;

  document.getElementById('api-not').value     = a.not||'';

  document.getElementById('api-edit-area').style.display='';

  document.getElementById('api-test-result').style.display='none';

}



async function saveApiEntry(){

  const entry = {

    id:         `${document.getElementById('api-banka').value}_${Date.now()}`,

    banka:      document.getElementById('api-banka').value,

    banka_adi:  document.getElementById('api-banka').value,

    tanim:      document.getElementById('api-tanim').value,

    url:        document.getElementById('api-url').value,

    api_key:    document.getElementById('api-key').value,

    api_secret: document.getElementById('api-secret').value,

    hesap:      document.getElementById('api-hesap').value,

    firma:      document.getElementById('api-firma').value,

    aktif:      document.getElementById('api-aktif').checked,

    not:        document.getElementById('api-not').value,

    eklenme:    new Date().toLocaleDateString('tr-TR'),

  };

  if(_apiEditIdx !== null) apiList[_apiEditIdx] = entry;

  else apiList.push(entry);



  const r = await api('/api/banka_api/save',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(apiList)});

  if((await r.json()).ok){

    toast('API ayarı kaydedildi','s');

    document.getElementById('api-edit-area').style.display='none';

    renderApiList();

  }

}



async function delApiEntry(i){

  if(!confirm('Bu API bağlantısını silmek istediğinizden emin misiniz?')) return;

  apiList.splice(i,1);

  const r = await api('/api/banka_api/save',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(apiList)});

  if((await r.json()).ok){ toast('Silindi','i'); renderApiList(); }

}



async function testApiConn(){

  const btn = document.getElementById('api-test-btn');

  const res = document.getElementById('api-test-result');

  btn.textContent='⏳ Test ediliyor...'; btn.disabled=true;

  // Kayıtlı değilse önce geçici test

  const entry = {

    url:    document.getElementById('api-url').value,

    api_key:document.getElementById('api-key').value,

    banka:  document.getElementById('api-banka').value,

  };

  // Şimdilik: sadece URL dolu mu kontrol et

  await new Promise(r=>setTimeout(r,800));

  btn.textContent='⚡ Bağlantı Test Et'; btn.disabled=false;

  res.style.display='';

  if(!entry.url || !entry.api_key){

    res.innerHTML='<span style="color:var(--am)">⚠ API URL ve Key girilmeden test yapılamaz. Bankadan bilgileri aldıktan sonra girin.</span>';

  } else {

    res.innerHTML=`<span style="color:var(--g)">✓ Ayarlar kaydedildi. Banka API erişimi aktifleştiğinde gerçek bağlantı test edilecek.</span><br>

      <span style="color:var(--t3);font-size:11px">URL: ${entry.url}</span>`;

  }

}





// ── FİNANSMAN YÖNETİMİ - EXCEL İMPORT ──────────────────────────────────────



function showBankaExcelImport(){

  const el = document.getElementById('banka-excel-import');

  el.style.display = el.style.display==='none' ? '' : 'none';

}



async function importBankaExcel(){

  const fileInput = document.getElementById('banka-excel-file');

  const file = fileInput.files[0];

  if(!file){ toast('Dosya seçin','e'); return; }



  const resEl = document.getElementById('banka-import-result');

  resEl.style.display=''; resEl.style.color='var(--t2)';

  resEl.textContent='⏳ Okunuyor...';



  const formData = new FormData();

  formData.append('file', file, file.name);



  try {

    const r = await fetch('/api/banka/import_excel', {method:'POST', body: formData});

    const d = await r.json();



    if(!d.ok){

      resEl.style.color='var(--r)';

      resEl.innerHTML='Hata: '+d.msg;

      return;

    }



    // Mevcut listeye merge et (aynı banka varsa üzerine yaz)

    const r2 = await api('/api/banka'); const d2 = await r2.json();

    let existing = d2.records || [];



    d.records.forEach(newRec=>{

      const idx = existing.findIndex(e=>

        e.banka?.toUpperCase()===newRec.banka?.toUpperCase() &&

        e.firma===newRec.firma

      );

      if(idx>=0) existing[idx] = {...existing[idx], ...newRec};

      else existing.push(newRec);

    });



    const r3 = await api('/api/banka/save',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(existing)});

    if((await r3.json()).ok){

      resEl.style.color='var(--g)';

      resEl.innerHTML=`✓ ${d.count} banka bakiyesi yüklendi ve kaydedildi`;

      toast(`✓ ${d.count} banka bakiyesi yüklendi`,'s');



      // Finansman sayfasındaki listeyi yenile

      if(typeof loadFinansmanData==='function') loadFinansmanData();

      // Likidite sayfası açıksa yenile

      if(document.getElementById('vw-liki')?.classList.contains('on')) renderLikidite();



      setTimeout(()=>{ document.getElementById('banka-excel-import').style.display='none'; },1500);

    }

  } catch(e) {

    resEl.style.color='var(--r)';

    resEl.innerHTML='Hata: '+e.message;

  }

}



function downloadBankaTemplate(){

  // Örnek banka bakiye şablonu CSV olarak indir

  const csv = [

    ['Banka Adı','Firma','Bakiye','Hesap Türü','Tarih'],

    ['VAKIFBANK','ULUSAL','1250000','Vadesiz TL','2026-03-14'],

    ['HALKBANK','ULUSAL','850000','Vadesiz TL','2026-03-14'],

    ['ZİRAAT BANKASI','ABC','320000','Vadesiz TL','2026-03-14'],

    ['GARANTİ BBVA','ULUSAL','2100000','Vadesiz TL','2026-03-14'],

  ].map(r=>r.map(v=>'"'+v+'"').join(',')).join(String.fromCharCode(10));

  const a=document.createElement('a');

  a.href='data:text/csv;charset=utf-8,\uFEFF'+encodeURIComponent(csv);

  a.download='banka_bakiye_sablon.csv'; a.click();

  toast('Şablon indirildi','i');

}



async function saveBankaFinans(){

  // Finansman sayfasındaki banka listesini kaydet

  if(typeof bankaFinansList !== 'undefined'){

    const r = await api('/api/banka/save',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(bankaFinansList)});

    if((await r.json()).ok){ toast('Banka bakiyeleri kaydedildi','s'); }

  } else if(typeof bankaList !== 'undefined'){

    const r = await api('/api/banka/save',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(bankaList)});

    if((await r.json()).ok){ toast('Banka bakiyeleri kaydedildi','s'); }

  }

}



async function importKrediExcelFinans(){

  const btn = document.getElementById('kredi-finans-import-btn') ||

              document.querySelector('[onclick="importKrediExcelFinans()"]');

  if(btn){ btn.textContent='⏳ Okunuyor...'; btn.disabled=true; }

  try {

    const r = await api('/api/kredi/import_excel',{method:'POST',headers:{'Content-Type':'application/json'},body:'{}'});

    const d = await r.json();

    if(btn){ btn.textContent="⬆ Kaynak Excel'den Yükle"; btn.disabled=false; }

    if(d.ok){

      krediList = d.records||[];

      if(typeof drawKrediTbl==='function') drawKrediTbl();

      if(typeof renderKrediTbl==='function') renderKrediTbl();

      toast('✓ '+d.count+' kredi limiti yüklendi','s');

      await saveKrediFinans();

    } else {

      toast('Hata: '+(d.msg||'Kaynak dosya bulunamadı — Kaynak Dosya sayfasından dosyayı tanımlayın'),'e');

    }

  } catch(e) {

    if(btn){ btn.textContent="⬆ Kaynak Excel'den Yükle"; btn.disabled=false; }

    toast('Hata: '+e.message,'e');

  }

}



async function saveKrediFinans(){

  if(typeof krediList !== 'undefined'){

    krediList.forEach(k=>{ if(!k.faiz_aylik&&k.faiz_yillik) k.faiz_aylik=k.faiz_yillik/12; });

    const r = await api('/api/kredi/save',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(krediList)});

    if((await r.json()).ok){ toast('Kredi limitleri kaydedildi','s'); }

  }

}





// ── DÖVİZ KURLARI (TCMB) ─────────────────────────────────────────────────────

let _kurGecmis = []; // Sorgulanan tarihler

const _ONEMLI_KURLAR = ['USD','EUR','GBP','CHF','JPY','SAR','AED','DKK','SEK','NOK'];

const _KUR_RENK = {USD:'#3b82f6',EUR:'#10b981',GBP:'#8b5cf6',CHF:'#f59e0b',JPY:'#ef4444',SAR:'#06b6d4',AED:'#6366f1'};



function sonIsGunu(date){

  // Hafta sonu veya pazar ise önceki cuma/cumartesiye geç

  const d = new Date(date);

  const gun = d.getDay(); // 0=Pazar, 6=Cumartesi

  if(gun === 0) d.setDate(d.getDate()-2); // Pazar -> Cuma

  else if(gun === 6) d.setDate(d.getDate()-1); // Cumartesi -> Cuma

  return d;

}



function tarihToInput(d){

  return d.getFullYear()+'-'+String(d.getMonth()+1).padStart(2,'0')+'-'+String(d.getDate()).padStart(2,'0');

}



function renderKur(){

  const d = sonIsGunu(new Date());

  document.getElementById('kur-tarih').value = tarihToInput(d);

  kurCek();

}



function kurBugun(){

  const d = sonIsGunu(new Date());

  document.getElementById('kur-tarih').value = tarihToInput(d);

  kurCek();

}



async function kurCek(){

  const dateVal = document.getElementById('kur-tarih').value;

  let tarihTR = '';

  if(dateVal){

    const [y,m,d] = dateVal.split('-');

    tarihTR = d+'.'+m+'.'+y;

  }



  const infoEl = document.getElementById('kur-info');

  const kartlarEl = document.getElementById('kur-ozet-kartlar');

  const tblWrap = document.getElementById('kur-tablo-wrap');



  infoEl.style.display='flex';

  infoEl.innerHTML='<span style="color:var(--acc)">⏳ TCMB sorgulanıyor...</span>';

  kartlarEl.innerHTML='';



  try {

    const url = tarihTR ? '/api/kur?tarih='+encodeURIComponent(tarihTR) : '/api/kur';

    const r = await api(url);

    const d = await r.json();



    if(!d.ok){

      infoEl.innerHTML=`<span style="color:var(--r)">✕ ${d.hata||'Kur çekilemedi'}</span>

        <span style="color:var(--t3);font-size:11px">Hafta sonu veya tatil günleri için bir önceki iş gününü seçin</span>`;

      return;

    }



    const kurlar = d.kurlar || {};

    const tarih  = d.tarih || tarihTR;



    // Info bar

    infoEl.innerHTML=`

      <span style="color:var(--g)">✓ TCMB</span>

      <span style="color:var(--t2)">${tarih} tarihi kurları</span>

      <span style="color:var(--t3)">${Object.keys(kurlar).length} para birimi</span>

      ${d.from_cache?'<span style="color:var(--am);font-size:10px">📦 Önbellekten</span>':''}

    `;



    // Özet kartları (önemli kurlar)

    kartlarEl.innerHTML = _ONEMLI_KURLAR

      .filter(k => kurlar[k])

      .map(k => {

        const kur = kurlar[k];

        const renk = _KUR_RENK[k] || 'var(--acc)';

        const satis = kur.satis || kur.alis || 0;

        const alis  = kur.alis  || kur.satis || 0;

        return `<div class="met" style="cursor:default">

          <div class="met-stripe" style="background:${renk}"></div>

          <div style="font-family:var(--mono);font-size:11px;font-weight:700;color:${renk};margin-bottom:6px">${k}

            ${kur.birim>1?`<span style="font-size:9px;color:var(--t3)">(${kur.birim} birim)</span>`:''}

          </div>

          <div style="font-size:10px;color:var(--t3);margin-bottom:8px">${kur.isim||''}</div>

          <div style="font-family:var(--mono);font-size:14px;font-weight:600;color:var(--t1)">

            ₺ ${satis?satis.toLocaleString('tr-TR',{minimumFractionDigits:4}):'—'}

          </div>

          <div style="font-size:10px;color:var(--t3);margin-top:3px;font-family:var(--mono)">

            Alış: ₺ ${alis?alis.toLocaleString('tr-TR',{minimumFractionDigits:4}):'—'}

          </div>

        </div>`;

      }).join('');



    // Tüm kurlar tablosu

    tblWrap.style.display='';

    const sirali = Object.values(kurlar).sort((a,b)=>{

      const ia = _ONEMLI_KURLAR.indexOf(a.kod);

      const ib = _ONEMLI_KURLAR.indexOf(b.kod);

      if(ia>=0 && ib>=0) return ia-ib;

      if(ia>=0) return -1;

      if(ib>=0) return 1;

      return a.kod.localeCompare(b.kod);

    });



    document.getElementById('kur-tbody').innerHTML = sirali.map(k => {

      const onemli = _ONEMLI_KURLAR.includes(k.kod);

      return `<tr style="${onemli?'background:rgba(255,255,255,.02)':''}">

        <td style="font-family:var(--mono);font-weight:${onemli?'700':'400'};color:${_KUR_RENK[k.kod]||'var(--t1)'}">${k.kod}</td>

        <td style="font-family:var(--mono);font-size:11px;color:var(--t3);text-align:center">${k.birim||1}</td>

        <td style="font-size:12px">${k.isim||'—'}</td>

        <td class="num" style="font-family:var(--mono);font-size:12px;color:var(--g)">${k.alis?k.alis.toLocaleString('tr-TR',{minimumFractionDigits:4}):'—'}</td>

        <td class="num" style="font-family:var(--mono);font-size:12px;color:var(--r)">${k.satis?k.satis.toLocaleString('tr-TR',{minimumFractionDigits:4}):'—'}</td>

        <td class="num" style="font-family:var(--mono);font-size:11px;color:var(--t2)">${k.efektif_alis?k.efektif_alis.toLocaleString('tr-TR',{minimumFractionDigits:4}):'—'}</td>

        <td class="num" style="font-family:var(--mono);font-size:11px;color:var(--t2)">${k.efektif_satis?k.efektif_satis.toLocaleString('tr-TR',{minimumFractionDigits:4}):'—'}</td>

      </tr>`;

    }).join('');



    // Geçmiş liste

    if(tarih && !_kurGecmis.includes(tarih)){

      _kurGecmis.unshift(tarih);

      if(_kurGecmis.length > 15) _kurGecmis.pop();

    }

    const gecmisWrap = document.getElementById('kur-gecmis-wrap');

    if(gecmisWrap) gecmisWrap.style.display='';

    document.getElementById('kur-gecmis').innerHTML = _kurGecmis.map(t=>`

      <button class="btn bo bsm" style="font-family:var(--mono);font-size:11px"

        onclick="kurGecmisSec('${t}')">${t}</button>

    `).join('');



    // Window'a kaydet (diğer sayfalar kullanabilsin)

    window._sonKurlar = {tarih, kurlar};



  } catch(e) {

    infoEl.innerHTML=`<span style="color:var(--r)">✕ Bağlantı hatası: ${e.message}</span>`;

    console.error('Kur hatası:', e);

  }

}



function kurGecmisSec(tarih){

  // TR formatı DD.MM.YYYY → input value YYYY-MM-DD

  const [d,m,y] = tarih.split('.');

  document.getElementById('kur-tarih').value = y+'-'+m+'-'+d;

  kurCek();

}



async function kurExport(){

  const url = '/api/kur' + (document.getElementById('kur-tarih').value

    ? '?tarih='+encodeURIComponent(document.getElementById('kur-tarih').value.split('-').reverse().join('.'))

    : '');

  const r = await api(url); const d = await r.json();

  if(!d.ok){ toast('Kur verisi yok','e'); return; }

  const kurlar = Object.values(d.kurlar||{});

  const hdrs = ['Kod','Birim','Para Birimi','Döviz Alış','Döviz Satış','Efektif Alış','Efektif Satış'];

  const rows = [hdrs, ...kurlar.map(k=>[k.kod,k.birim||1,k.isim,k.alis||'',k.satis||'',k.efektif_alis||'',k.efektif_satis||''])];

  const csv = rows.map(r=>r.map(v=>'"'+String(v).replace(/"/g,'""')+'"').join(',')).join(String.fromCharCode(10));

  const a = document.createElement('a');

  a.href = 'data:text/csv;charset=utf-8,\uFEFF'+encodeURIComponent(csv);

  a.download = 'tcmb_kurlar_'+( d.tarih||'bugun').replace(/\./g,'')+'.csv';

  a.click();

  toast('Kurlar CSV olarak indirildi','s');

}





// ── LOGO TIGER FATURALAR ──────────────────────────────────────────────────────

let _logoTip = 'satis';

let _logoFaturalar = [];

let _logoPg = 1; const LOGO_PPG = 50;



async function renderLogo(){

  // Tarih aralığı: son 3 ay

  const today = new Date();

  const bitVal = today.toISOString().slice(0,10);

  const bas = new Date(today); bas.setMonth(bas.getMonth()-3);

  const basVal = bas.toISOString().slice(0,10);

  if(!document.getElementById('logo-bas').value)

    document.getElementById('logo-bas').value = basVal;

  if(!document.getElementById('logo-bit').value)

    document.getElementById('logo-bit').value = bitVal;



  // Bağlantı durumu

  await logoConnDurum();

}



async function logoConnDurum(){

  const dot  = document.getElementById('logo-dot');

  const msg  = document.getElementById('logo-conn-msg');



  // Config var mı?

  const rc = await api('/api/logo/config'); const dc = await rc.json();

  if(!dc.config?.server){

    dot.className='src-dot none';

    msg.textContent='Logo bağlantısı ayarlanmamış — ⚙ Bağlantı Ayarları butonuna tıklayın';

    return;

  }



  dot.className='src-dot none';

  msg.textContent='Bağlanıyor...';



  const r = await api('/api/logo/test'); const d = await r.json();

  if(d.ok){

    dot.className='src-dot ok';

    msg.innerHTML=`✓ Logo Tiger bağlı — <span style="color:var(--t3)">${dc.config.server} / ${dc.config.database}</span> · <span style="color:var(--acc)">${(dc.config.firmalar||[]).length} firma</span>`;

    // Firma seçimlerini güncelle

    if(dc.config.firmalar?.length){

      logoFirmaSecimGuncelle(dc.config.firmalar);

      // Stok ve cari sayfaları için de güncelle

      logoFirmaSecimGuncelle2('stok-firma-secim', dc.config.firmalar);

      logoFirmaSecimGuncelle2('cari-firma-secim', dc.config.firmalar);

      logoYilSecimGuncelle2('stok-yil-secim');

      logoYilSecimGuncelle2('cari-yil-secim');

    }

    // Otomatik yükle

    logoFaturaYukle();

  } else {

    dot.className='src-dot err';

    msg.innerHTML=`✕ Bağlantı hatası: <span style="color:var(--r)">${d.msg||'Bilinmiyor'}</span>`;

  }

}



let _logoFirmalar = []; // Ayarlardaki firma listesi



// ── LOGO ORTAK YARDIMCI FONKSİYONLAR ────────────────────────────────────────

function _logoAyarFormu(serverEl, portEl, dbEl, userEl, firmaListEl){

  api('/api/logo/config').then(r=>r.json()).then(d=>{

    const c = d.config||{};

    document.getElementById(serverEl).value = c.server||'';

    document.getElementById(portEl).value   = c.port||1433;

    document.getElementById(dbEl).value     = c.database||'';

    document.getElementById(userEl).value   = c.username||'';

    const firmalar = c.firmalar||[{no:c.firma_no||'001',ad:'Firma 1'}];

    if(firmaListEl.includes('stok')) { _stokLogoFirmalar=[...firmalar]; _renderLogoFirmaForm(firmaListEl,_stokLogoFirmalar,'stokLogoFirmaEkle'); }

    else if(firmaListEl.includes('cari')) { _cariLogoFirmalar=[...firmalar]; _renderLogoFirmaForm(firmaListEl,_cariLogoFirmalar,'cariLogoFirmaEkle'); }

    else { _logoFirmalar=[...firmalar]; _renderLogoFirmaForm(firmaListEl,_logoFirmalar,'logoFirmaEkle'); }

  });

}



function _renderLogoFirmaForm(elId, list, ekleFunc){

  const el = document.getElementById(elId); if(!el) return;

  el.innerHTML = list.map((f,i)=>`

    <div style="display:flex;gap:6px;align-items:center">

      <input class="fi2" type="text" value="${f.no||''}" placeholder="001"

        style="width:70px;font-family:var(--mono)"

        onchange="${elId.split('-')[0]==='logo'?'_logoFirmalar':'_'+elId.split('-')[0]+'LogoFirmalar'}[${i}].no=this.value">

      <input class="fi2" type="text" value="${f.ad||''}" placeholder="Firma Adı"

        style="flex:1"

        onchange="${elId.split('-')[0]==='logo'?'_logoFirmalar':'_'+elId.split('-')[0]+'LogoFirmalar'}[${i}].ad=this.value">

      <button class="btn bgh" onclick="${elId.split('-')[0]==='logo'?'_logoFirmalar':'_'+elId.split('-')[0]+'LogoFirmalar'}.splice(${i},1);${ekleFunc}()" style="color:var(--r);font-size:12px;padding:4px 8px">✕</button>

    </div>`).join('');

}



async function _logoAyarKaydetOrtak(serverEl, portEl, dbEl, userEl, passEl, firmaList){

  const cfg = {

    server:   document.getElementById(serverEl).value.trim(),

    port:     parseInt(document.getElementById(portEl).value)||1433,

    database: document.getElementById(dbEl).value.trim(),

    username: document.getElementById(userEl).value.trim(),

    password: document.getElementById(passEl)?.value||'',

    firmalar: firmaList.filter(f=>f.no),

    firma_no: (firmaList[0]?.no||'001'),

  };

  if(!cfg.server||!cfg.database){ toast('Sunucu ve veritabanı adı zorunlu','e'); return false; }

  if(!cfg.firmalar.length){ toast('En az bir firma ekleyin','e'); return false; }

  const r = await api('/api/logo/config/save',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(cfg)});

  if((await r.json()).ok){

    toast(`Logo ayarları kaydedildi — ${cfg.firmalar.length} firma`,'s');

    // Tüm Logo sayfalarındaki firma seçimlerini güncelle

    logoFirmaSecimGuncelle(cfg.firmalar);

    logoFirmaSecimGuncelle2('stok-firma-secim', cfg.firmalar);

    logoFirmaSecimGuncelle2('cari-firma-secim', cfg.firmalar);

    logoFirmaSecimGuncelle2('sa-firma-secim',   cfg.firmalar);

    logoYilSecimGuncelle2('stok-yil-secim');

    logoYilSecimGuncelle2('cari-yil-secim');

    logoYilSecimGuncelle2('sa-yil-secim');

    return true;

  }

  return false;

}



function showLogoAyarlar(){

  api('/api/logo/config').then(r=>r.json()).then(d=>{

    const c = d.config||{};

    document.getElementById('logo-server').value = c.server||'';

    document.getElementById('logo-port').value   = c.port||1433;

    document.getElementById('logo-db').value     = c.database||'';

    document.getElementById('logo-user').value   = c.username||'';

    document.getElementById('logo-pass').value   = '';

    _logoFirmalar = c.firmalar || [{no:'001', ad:'Firma 1'}];

    renderLogoFirmaListesi();

  });

  document.getElementById('logo-ayarlar').style.display='';

}



function renderLogoFirmaListesi(){

  const el = document.getElementById('logo-firma-listesi');

  if(!el) return;

  el.innerHTML = _logoFirmalar.map((f,i)=>`

    <div style="display:flex;gap:6px;align-items:center">

      <input class="fi2" type="text" value="${f.no||''}" placeholder="001"

        style="width:70px;font-family:var(--mono)"

        onchange="_logoFirmalar[${i}].no=this.value">

      <input class="fi2" type="text" value="${f.ad||''}" placeholder="Firma Adı"

        style="flex:1"

        onchange="_logoFirmalar[${i}].ad=this.value">

      <button class="btn bgh" onclick="_logoFirmalar.splice(${i},1);renderLogoFirmaListesi()"

        style="color:var(--r);font-size:12px;padding:4px 8px">✕</button>

    </div>`).join('');

}



function logoFirmaEkle(){

  _logoFirmalar.push({no:'', ad:''});

  renderLogoFirmaListesi();

}



async function logoAyarKaydet(){

  const ok = await _logoAyarKaydetOrtak('logo-server','logo-port','logo-db','logo-user','logo-pass',_logoFirmalar);

  if(ok){

    document.getElementById('logo-ayarlar').style.display='none';

    logoConnDurum();

  }

}



function logoFirmaSecimGuncelle(firmalar){

  // Fatura, stok, cari sayfalarındaki firma seçimlerini güncelle

  ['logo-firma-secim','stok-firma-secim','cari-firma-secim'].forEach(elId=>{

    const el = document.getElementById(elId); if(!el) return;

    el.innerHTML = firmalar.map(f=>`

      <label style="display:flex;align-items:center;gap:4px;background:var(--s3);border:1px solid var(--b2);border-radius:4px;padding:3px 8px;cursor:pointer;font-size:11px">

        <input type="checkbox" checked value="${f.no}" style="margin:0">

        <span>${f.ad||f.no}</span>

      </label>`).join('');

  });

  // Yıl seçimini de güncelle (son 5 yıl)

  const yilEl = document.getElementById('logo-yil-secim');

  if(yilEl){

    const buYil = new Date().getFullYear();

    yilEl.innerHTML = Array.from({length:5},(_,i)=>buYil-i).map(y=>`

      <label style="display:flex;align-items:center;gap:4px;background:var(--s3);border:1px solid var(--b2);border-radius:4px;padding:3px 8px;cursor:pointer;font-size:11px">

        <input type="checkbox" ${y>=buYil-1?'checked':''} value="${y}" style="margin:0">

        <span>${y}</span>

      </label>`).join('');

  }

}



function logoSeciliFirmalar(elId){

  const el = document.getElementById(elId||'logo-firma-secim');

  if(!el) return null;

  const checks = el.querySelectorAll('input[type=checkbox]:checked');

  if(!checks.length) return null;

  return Array.from(checks).map(c=>c.value).join(',');

}



function logoSeciliYillar(){

  const el = document.getElementById('logo-yil-secim');

  if(!el) return null;

  const checks = el.querySelectorAll('input[type=checkbox]:checked');

  if(!checks.length) return null;

  return Array.from(checks).map(c=>c.value).join(',');

}



async function logoTestEt(){

  await logoAyarKaydet();

  const btn = document.getElementById('logo-test-btn');

  const res = document.getElementById('logo-test-result');

  btn.textContent='⏳ Test ediliyor...'; btn.disabled=true;

  const r = await api('/api/logo/test'); const d = await r.json();

  btn.textContent='⚡ Bağlantı Test Et'; btn.disabled=false;

  res.style.display='';

  if(d.ok){

    res.innerHTML=`<span style="color:var(--g)">✓ Bağlantı başarılı!</span>

      <span style="color:var(--t3);margin-left:8px;font-size:11px">Veritabanı: ${d.database}</span>`;

  } else {

    res.innerHTML=`<span style="color:var(--r)">✕ ${d.msg}</span>

      <div style="margin-top:6px;font-size:11px;color:var(--t3)">

        IT'e şunu sorun: SQL Server'da TCP/IP bağlantısı açık mı? 1433 portu firewall'da açık mı?

      </div>`;

  }

}



function logoTab(tip){

  _logoTip = tip;

  ['satis','alis'].forEach(t=>{

    const btn = document.getElementById('logo-tab-'+t);

    if(t===tip){ btn.style.background='var(--acc)'; btn.style.color='#fff'; }

    else { btn.style.background='transparent'; btn.style.color='var(--t2)'; }

  });

  logoFaturaYukle();

}



async function logoFaturaYukle(){

  const bas       = document.getElementById('logo-bas').value;

  const bit       = document.getElementById('logo-bit').value;

  const cari      = document.getElementById('logo-cari-f').value;

  const minTutar  = document.getElementById('logo-min-tutar')?.value||'';

  const maxTutar  = document.getElementById('logo-max-tutar')?.value||'';

  const firmalar  = logoSeciliFirmalar('logo-firma-secim');

  const yillar    = logoSeciliYillar();



  const tbody = document.getElementById('logo-tbody');

  tbody.innerHTML='<tr><td colspan="9" style="padding:20px;text-align:center;color:var(--t3)">⏳ Logo veritabanından yükleniyor...</td></tr>';



  const params = new URLSearchParams({tip:_logoTip, limit:500});

  if(bas)       params.set('baslangic', bas);

  if(bit)       params.set('bitis', bit);

  if(cari)      params.set('cari', cari);

  if(minTutar)  params.set('min_tutar', minTutar);

  if(maxTutar)  params.set('max_tutar', maxTutar);

  if(firmalar)  params.set('firmalar', firmalar);

  if(yillar)    params.set('yillar', yillar);



  try {

    const r = await api('/api/logo/fatura?'+params); const d = await r.json();

    if(!d.ok){

      tbody.innerHTML=`<tr><td colspan="9" style="padding:20px;text-align:center;color:var(--r)">${d.msg||'Veri çekilemedi'}</td></tr>`;

      return;

    }

    _logoFaturalar = d.records||[];

    _logoPg = 1;

    renderLogoTablo();

    // Sonuçlardan yıl listesini güncelle

    if(d.yillar?.length){

      const yilEl=document.getElementById('logo-yil-secim');

      if(yilEl && yilEl.children.length<=1){

        const mevcut=logoSeciliYillar();

        yilEl.innerHTML=d.yillar.map(y=>`

          <label style="display:flex;align-items:center;gap:4px;background:var(--s3);border:1px solid var(--b2);border-radius:4px;padding:3px 8px;cursor:pointer;font-size:11px">

            <input type="checkbox" checked value="${y}" style="margin:0"><span>${y}</span>

          </label>`).join('');

      }

    }

  } catch(e) {

    tbody.innerHTML=`<tr><td colspan="9" style="padding:20px;text-align:center;color:var(--r)">Hata: ${e.message}</td></tr>`;

  }

}



function renderLogoTablo(){

  const rows = _logoFaturalar;

  const totNet   = rows.reduce((s,r)=>s+(parseFloat(r.net_toplam)||0),0);

  const totKdv   = rows.reduce((s,r)=>s+(parseFloat(r.toplam_kdv)||0),0);

  const totBrut  = rows.reduce((s,r)=>s+(parseFloat(r.brut_toplam)||0),0);

  const iadeAdet = rows.filter(r=>r.tip_kodu=='8'||r.tip_kodu=='2').length;



  // Metrikler

  // Firma bazlı özet

  const firmaOzet = {};

  rows.forEach(r=>{ const f=r.logo_firma_adi||r.logo_firma_no||'?'; firmaOzet[f]=(firmaOzet[f]||0)+parseFloat(r.net_toplam||0); });

  const firmaOzetHTML = Object.entries(firmaOzet).map(([f,v])=>

    `<span style="font-size:10px;color:var(--t3)">${f}: <strong>₺${fmtN(v)}</strong></span>`).join(' · ');



  document.getElementById('logo-mets').innerHTML=`

    <div class="met"><div class="met-stripe" style="background:var(--acc)"></div>

      <div class="mlb">Toplam Fatura</div>

      <div class="mvl" style="color:var(--acc)">${rows.length}</div>

      <div class="msb">${iadeAdet} iade dahil</div></div>

    <div class="met"><div class="met-stripe" style="background:var(--g)"></div>

      <div class="mlb">Net Toplam ₺</div>

      <div class="mvl" style="color:var(--g);font-size:16px">₺${fmtN(totNet)}</div></div>

    <div class="met"><div class="met-stripe" style="background:var(--am)"></div>

      <div class="mlb">Toplam KDV ₺</div>

      <div class="mvl" style="color:var(--am);font-size:16px">₺${fmtN(totKdv)}</div></div>

    <div class="met"><div class="met-stripe" style="background:var(--r)"></div>

      <div class="mlb">Brüt Toplam ₺</div>

      <div class="mvl" style="color:var(--r);font-size:16px">₺${fmtN(totBrut)}</div></div>

  `;



  // Tablo

  const start=(_logoPg-1)*LOGO_PPG, end=start+LOGO_PPG;

  const page = rows.slice(start, end);



  document.getElementById('logo-tbody').innerHTML = page.map((r,i)=>{

    const isIade = r.tip_kodu=='8'||r.tip_kodu=='2';

    const net  = parseFloat(r.net_toplam)||0;

    const kdv  = parseFloat(r.toplam_kdv)||0;

    const brut = parseFloat(r.brut_toplam)||0;

    return `<tr style="${isIade?'opacity:.7':''}">

      <td style="text-align:center;font-size:10px;color:var(--t3);cursor:pointer" onclick="logoDetay('${r.id}','${r.fatura_no}')" title="Kalemleri gör">▶</td>

      <td style="font-family:var(--mono);font-size:12px;font-weight:500">${r.fatura_no||'—'}</td>

      <td style="font-family:var(--mono);font-size:11px">${r.tarih||'—'}</td>

      <td style="font-family:var(--mono);font-size:11px;color:var(--acc);text-align:center">${r.yil||'—'}</td>

      <td style="font-size:10px;color:var(--t3)" class="clip" title="${r.logo_firma_adi||''}">${r.logo_firma_adi||r.logo_firma_no||'—'}</td>

      <td style="font-size:12px" class="clip" title="${r.cari_unvan||''}">${r.cari_unvan||'—'}</td>

      <td style="font-family:var(--mono);font-size:11px;color:var(--t3)">${r.cari_kod||'—'}</td>

      <td class="num" style="font-family:var(--mono);font-size:12px;color:var(--g);font-weight:600">${fmtN(net)}</td>

      <td class="num" style="font-family:var(--mono);font-size:11px;color:var(--am)">${fmtN(kdv)}</td>

      <td class="num" style="font-family:var(--mono);font-size:12px">${fmtN(brut)}</td>

      <td><span class="bdg ${isIade?'bri':'bgi'}" style="font-size:10px">${r.tip_adi||'—'}</span></td>

    </tr>

    <tr id="logo-det-${r.id}" style="display:none">

      <td colspan="9" style="padding:0;background:var(--s2)">

        <div id="logo-det-inner-${r.id}" style="padding:12px"></div>

      </td>

    </tr>`;

  }).join('');



  // Footer

  document.getElementById('logo-tfoot').innerHTML=`

    <tr style="border-top:2px solid var(--b2);font-weight:600">

      <td colspan="5" style="padding:8px 10px;font-family:var(--mono);font-size:11px;color:var(--t2)">${rows.length} fatura</td>

      <td class="num" style="color:var(--g)">${fmtN(totNet)}</td>

      <td class="num" style="color:var(--am)">${fmtN(totKdv)}</td>

      <td class="num">${fmtN(totBrut)}</td>

      <td></td>

    </tr>`;



  // Pagination

  const tot=Math.ceil(rows.length/LOGO_PPG)||1;

  let ph=`<span class="pgi">${rows.length} fatura · Sayfa ${_logoPg}/${tot}</span><div class="pgsp"></div>`;

  if(_logoPg>1) ph+=`<button class="pbn" onclick="_logoPg--;renderLogoTablo()">‹</button>`;

  for(let i=Math.max(1,_logoPg-2);i<=Math.min(tot,_logoPg+2);i++)

    ph+=`<button class="pbn${i===_logoPg?' on':''}" onclick="_logoPg=${i};renderLogoTablo()">${i}</button>`;

  if(_logoPg<tot) ph+=`<button class="pbn" onclick="_logoPg++;renderLogoTablo()">›</button>`;

  document.getElementById('logo-pgn').innerHTML=ph;

}



async function logoDetay(id, faturaNo){

  const detRow   = document.getElementById('logo-det-'+id);

  const detInner = document.getElementById('logo-det-inner-'+id);

  if(!detRow) return;

  if(detRow.style.display!=='none'){ detRow.style.display='none'; return; }

  detRow.style.display='';

  detInner.innerHTML='<span style="color:var(--t3);font-size:12px">⏳ Kalemler yükleniyor...</span>';



  const r = await api('/api/logo/fatura/detay?id='+id); const d = await r.json();

  if(!d.ok){ detInner.innerHTML=`<span style="color:var(--r)">${d.msg}</span>`; return; }

  const satirlar = d.satirlar||[];

  if(!satirlar.length){ detInner.innerHTML='<span style="color:var(--t3)">Kalem bulunamadı</span>'; return; }



  detInner.innerHTML=`

    <div style="font-family:var(--mono);font-size:10px;color:var(--acc);margin-bottom:8px">📋 ${faturaNo} — Fatura Kalemleri</div>

    <table style="width:100%;border-collapse:collapse;font-size:11.5px">

      <thead><tr style="background:var(--s3)">

        <th style="padding:4px 8px;text-align:left;color:var(--t3);font-size:10px">Stok Kodu</th>

        <th style="padding:4px 8px;text-align:left;color:var(--t3);font-size:10px">Stok Adı</th>

        <th style="padding:4px 8px;text-align:right;color:var(--t3);font-size:10px">Miktar</th>

        <th style="padding:4px 8px;text-align:left;color:var(--t3);font-size:10px">Birim</th>

        <th style="padding:4px 8px;text-align:right;color:var(--t3);font-size:10px">Birim Fiyat ₺</th>

        <th style="padding:4px 8px;text-align:right;color:var(--t3);font-size:10px">KDV %</th>

        <th style="padding:4px 8px;text-align:right;color:var(--t3);font-size:10px">Net Tutar ₺</th>

      </tr></thead>

      <tbody>${satirlar.map(s=>`<tr style="border-bottom:1px solid var(--b1)">

        <td style="padding:4px 8px;font-family:var(--mono);color:var(--t2)">${s.stok_kodu||'—'}</td>

        <td style="padding:4px 8px">${s.stok_adi||'—'}</td>

        <td style="padding:4px 8px;text-align:right;font-family:var(--mono)">${parseFloat(s.miktar||0).toLocaleString('tr-TR',{maximumFractionDigits:2})}</td>

        <td style="padding:4px 8px;color:var(--t3)">${s.birim||'—'}</td>

        <td style="padding:4px 8px;text-align:right;font-family:var(--mono)">${parseFloat(s.birim_fiyat||0).toLocaleString('tr-TR',{minimumFractionDigits:2})}</td>

        <td style="padding:4px 8px;text-align:right;font-family:var(--mono);color:var(--am)">${s.kdv_orani||0}%</td>

        <td style="padding:4px 8px;text-align:right;font-family:var(--mono);font-weight:600;color:var(--g)">${fmtN(parseFloat(s.net_tutar||0))}</td>

      </tr>`).join('')}</tbody>

    </table>`;

}



async function logoExport(){

  if(!_logoFaturalar.length){ toast('Veri yok','e'); return; }

  const hdrs=['Fatura No','Tarih','Cari Unvan','Cari Kod','Net Toplam','KDV','Brüt Toplam','Tür'];

  const csv=[hdrs,..._logoFaturalar.map(r=>[r.fatura_no,r.tarih,r.cari_unvan,r.cari_kod,

    r.net_toplam,r.toplam_kdv,r.brut_toplam,r.tip_adi])]

    .map(r=>r.map(v=>'"'+String(v||'').replace(/"/g,'""')+'"').join(',')).join(String.fromCharCode(10));

  const a=document.createElement('a');

  a.href='data:text/csv;charset=utf-8,\uFEFF'+encodeURIComponent(csv);

  a.download=`logo_${_logoTip}_faturalar.csv`; a.click();

  toast('Faturalar CSV olarak indirildi','s');

}





// ── STOK LİSTESİ ──────────────────────────────────────────────────────────────

let _stokData=[], _stokPg=1; const STOK_PPG=50;



async function renderStok(){

  await stokConnDurum();

}



async function stokConnDurum(){

  const rc = await api('/api/logo/config'); const dc = await rc.json();

  const dot = document.getElementById('stok-conn-dot');

  const msg = document.getElementById('stok-conn-msg');



  if(!dc.config?.server){

    dot.className='src-dot none';

    msg.textContent='Logo bağlantısı ayarlanmamış — ⚙ Bağlantı Ayarları butonuna tıklayın';

    return;

  }



  const r = await api('/api/logo/test'); const d = await r.json();

  if(d.ok){

    dot.className='src-dot ok';

    const firmalar = dc.config.firmalar||[{no:dc.config.firma_no||'001',ad:'Firma 1'}];

    msg.innerHTML=`✓ Logo bağlı — <span style="color:var(--t3)">${dc.config.server} / ${dc.config.database}</span> · <span style="color:var(--acc)">${firmalar.length} firma</span>`;

    logoFirmaSecimGuncelle2('stok-firma-secim', firmalar);

    logoYilSecimGuncelle2('stok-yil-secim');

    stokYukle();

  } else {

    dot.className='src-dot err';

    msg.innerHTML=`✕ Bağlantı hatası: <span style="color:var(--r)">${d.msg||'Bilinmiyor'}</span>`;

  }

}



function showStokAyarlar(){

  _logoAyarFormu('stok-logo-server','stok-logo-port','stok-logo-db','stok-logo-user','stok-logo-firma-listesi');

  document.getElementById('stok-ayarlar').style.display='';

}



let _stokLogoFirmalar = [];

function stokLogoFirmaEkle(){ _stokLogoFirmalar.push({no:'',ad:''}); _renderLogoFirmaForm('stok-logo-firma-listesi',_stokLogoFirmalar,'stokLogoFirmaEkle'); }



async function stokLogoAyarKaydet(){

  await _logoAyarKaydetOrtak('stok-logo-server','stok-logo-port','stok-logo-db','stok-logo-user','stok-logo-pass',_stokLogoFirmalar);

  document.getElementById('stok-ayarlar').style.display='none';

  stokConnDurum();

}



async function stokLogoTestEt(){

  await _logoAyarKaydetOrtak('stok-logo-server','stok-logo-port','stok-logo-db','stok-logo-user','stok-logo-pass',_stokLogoFirmalar);

  const btn=document.querySelector('[onclick="stokLogoTestEt()"]');

  const res=document.getElementById('stok-test-result');

  if(btn){btn.textContent='⏳';btn.disabled=true;}

  const r=await api('/api/logo/test'); const d=await r.json();

  if(btn){btn.textContent='⚡ Test Et';btn.disabled=false;}

  res.style.display='';

  res.innerHTML=d.ok?`<span style="color:var(--g)">✓ Bağlantı başarılı! Veritabanı: ${d.database}</span>`:`<span style="color:var(--r)">✕ ${d.msg}</span>`;

}



async function stokYukle(){

  const q = document.getElementById('stok-q').value;

  document.getElementById('stok-tbody').innerHTML='<tr><td colspan="5" style="padding:20px;text-align:center;color:var(--t3)">⏳ Stoklar yükleniyor...</td></tr>';

  const firmaSec  = logoSeciliFirmalar('stok-firma-secim');

  const yilSec    = Array.from(document.querySelectorAll('#stok-yil-secim input:checked')).map(c=>c.value).join(',') || null;

  const grupSec   = document.getElementById('stok-grup-f')?.value||'';

  const aktifSec  = document.getElementById('stok-aktif-hareket')?.checked ? '1' : '0';

  const basSec    = document.getElementById('stok-bas')?.value||'';

  const bitSec    = document.getElementById('stok-bit')?.value||'';

  const params = new URLSearchParams({q, limit:500});

  if(firmaSec)     params.set('firmalar', firmaSec);

  if(yilSec)       params.set('yillar', yilSec);

  if(grupSec)      params.set('grup', grupSec);

  if(aktifSec==='1') params.set('aktif_hareket','1');

  if(basSec)       params.set('baslangic', basSec);

  if(bitSec)       params.set('bitis', bitSec);

  const r = await api('/api/logo/stok?'+params);

  const d = await r.json();

  if(!d.ok){

    document.getElementById('stok-tbody').innerHTML=`<tr><td colspan="5" style="padding:20px;text-align:center;color:var(--r)">${d.msg}</td></tr>`;

    return;

  }

  _stokData = d.records||[]; _stokPg=1;

  document.getElementById('stok-count').textContent=`${_stokData.length} stok kartı`;

  // Grup filtresini güncelle

  const grupSel = document.getElementById('stok-grup-f');

  if(grupSel && d.gruplar?.length){

    const curG = grupSel.value;

    grupSel.innerHTML='<option value="">Tüm Gruplar</option>'+d.gruplar.map(g=>`<option value="${g}">${g}</option>`).join('');

    grupSel.value = curG;

  }

  renderStokTablo();

}



function renderStokTablo(){

  const start=(_stokPg-1)*STOK_PPG, page=_stokData.slice(start,start+STOK_PPG);

  document.getElementById('stok-tbody').innerHTML=page.map(s=>`<tr>

    <td style="font-family:var(--mono);font-size:11px;font-weight:600;color:var(--acc)">${s.stok_kodu}</td>

    <td style="font-size:12px">${s.stok_adi||'—'}</td>

    <td style="font-size:10px;color:var(--t3)">${s.logo_firma_adi||s.logo_firma_no||'—'}</td>

    <td style="font-size:11px;color:var(--t2)">${s.stok_grubu||'—'}</td>

    <td style="font-family:var(--mono);font-size:11px">${s.birim||'—'}</td>

    <td class="num" style="font-family:var(--mono);font-size:11px;color:${parseFloat(s.stok_bakiye||0)<0?'var(--r)':parseFloat(s.stok_bakiye||0)===0?'var(--am)':'var(--g)'};font-weight:600">${s.stok_bakiye!==''?parseFloat(s.stok_bakiye||0).toLocaleString('tr-TR',{maximumFractionDigits:2}):'—'}</td>

    <td class="num" style="font-family:var(--mono);font-size:11px">${s.donem_satis_adet!==''?parseFloat(s.donem_satis_adet||0).toLocaleString('tr-TR',{maximumFractionDigits:1}):'—'}</td>

    <td class="num" style="font-family:var(--mono);font-size:11px;color:var(--g)">${s.donem_satis_tutar!==''&&parseFloat(s.donem_satis_tutar||0)>0?'₺'+fmtN(parseFloat(s.donem_satis_tutar||0)):'—'}</td>

    <td class="num" style="font-family:var(--mono);font-size:11px">${s.son_alis_fiyati!==''&&parseFloat(s.son_alis_fiyati||0)>0?'₺'+fmtN(parseFloat(s.son_alis_fiyati||0)):'—'}</td>

    <td class="num" style="font-family:var(--mono);font-size:11px">${s.son_satis_fiyati!==''&&parseFloat(s.son_satis_fiyati||0)>0?'₺'+fmtN(parseFloat(s.son_satis_fiyati||0)):'—'}</td>

  </tr>`).join('');

  const tot=Math.ceil(_stokData.length/STOK_PPG)||1;

  let ph=`<span class="pgi">${_stokData.length} kayıt</span><div class="pgsp"></div>`;

  if(_stokPg>1) ph+=`<button class="pbn" onclick="_stokPg--;renderStokTablo()">‹</button>`;

  for(let i=Math.max(1,_stokPg-2);i<=Math.min(tot,_stokPg+2);i++)

    ph+=`<button class="pbn${i===_stokPg?' on':''}" onclick="_stokPg=${i};renderStokTablo()">${i}</button>`;

  if(_stokPg<tot) ph+=`<button class="pbn" onclick="_stokPg++;renderStokTablo()">›</button>`;

  document.getElementById('stok-pgn').innerHTML=ph;

}



async function stokExport(){

  if(!_stokData.length){toast('Veri yok','e');return;}

  const sep=',', nl='\r\n';

  let csv='Stok Kodu,Stok Adi,Grup,Birim,Ozel Kod'+nl;

  _stokData.forEach(s=>{

    csv+='"'+(s.stok_kodu||'').replace(/"/g,'""')+'",';

    csv+='"'+(s.stok_adi||'').replace(/"/g,'""')+'",';

    csv+='"'+(s.stok_grubu||'').replace(/"/g,'""')+'",';

    csv+='"'+(s.birim||'').replace(/"/g,'""')+'",';

    csv+='"'+(s.ozel_kod||'').replace(/"/g,'""')+'"'+nl;

  });

  const a=document.createElement('a'); a.href='data:text/csv;charset=utf-8,\uFEFF'+encodeURIComponent(csv);

  a.download='logo_stok_listesi.csv'; a.click(); toast('CSV indirildi','s');

}



// ── CARİ LİSTESİ (LOGO) ───────────────────────────────────────────────────────

let _cariLogoData=[], _cariLogoPg=1; const CARI_LOGO_PPG=50;



async function renderCariLogo(){

  await cariLogoConnDurum();

}



async function cariLogoConnDurum(){

  const rc = await api('/api/logo/config'); const dc = await rc.json();

  const dot = document.getElementById('cari-logo-conn-dot');

  const msg = document.getElementById('cari-logo-conn-msg');



  if(!dc.config?.server){

    dot.className='src-dot none';

    msg.textContent='Logo bağlantısı ayarlanmamış — ⚙ Bağlantı Ayarları butonuna tıklayın';

    return;

  }



  const r = await api('/api/logo/test'); const d = await r.json();

  if(d.ok){

    dot.className='src-dot ok';

    const firmalar = dc.config.firmalar||[{no:dc.config.firma_no||'001',ad:'Firma 1'}];

    msg.innerHTML=`✓ Logo bağlı — <span style="color:var(--t3)">${dc.config.server} / ${dc.config.database}</span> · <span style="color:var(--acc)">${firmalar.length} firma</span>`;

    logoFirmaSecimGuncelle2('cari-firma-secim', firmalar);

    logoYilSecimGuncelle2('cari-yil-secim');

    cariLogoYukle();

  } else {

    dot.className='src-dot err';

    msg.innerHTML=`✕ Bağlantı hatası: <span style="color:var(--r)">${d.msg||'Bilinmiyor'}</span>`;

  }

}



function showCariLogoAyarlar(){

  _logoAyarFormu('cari-logo-server','cari-logo-port','cari-logo-db','cari-logo-user','cari-logo-firma-listesi');

  document.getElementById('cari-logo-ayarlar').style.display='';

}



let _cariLogoFirmalar = [];

function cariLogoFirmaEkle(){ _cariLogoFirmalar.push({no:'',ad:''}); _renderLogoFirmaForm('cari-logo-firma-listesi',_cariLogoFirmalar,'cariLogoFirmaEkle'); }



async function cariLogoAyarKaydet(){

  await _logoAyarKaydetOrtak('cari-logo-server','cari-logo-port','cari-logo-db','cari-logo-user','cari-logo-pass',_cariLogoFirmalar);

  document.getElementById('cari-logo-ayarlar').style.display='none';

  cariLogoConnDurum();

}



async function cariLogoTestEt(){

  await _logoAyarKaydetOrtak('cari-logo-server','cari-logo-port','cari-logo-db','cari-logo-user','cari-logo-pass',_cariLogoFirmalar);

  const btn=document.querySelector('[onclick="cariLogoTestEt()"]');

  const res=document.getElementById('cari-logo-test-result');

  if(btn){btn.textContent='⏳';btn.disabled=true;}

  const r=await api('/api/logo/test'); const d=await r.json();

  if(btn){btn.textContent='⚡ Test Et';btn.disabled=false;}

  res.style.display='';

  res.innerHTML=d.ok?`<span style="color:var(--g)">✓ Bağlantı başarılı! Veritabanı: ${d.database}</span>`:`<span style="color:var(--r)">✕ ${d.msg}</span>`;

}



async function cariLogoYukle(){

  const q   = document.getElementById('cari-logo-q').value;

  const tip = document.getElementById('cari-logo-tip').value;

  document.getElementById('cari-logo-tbody').innerHTML='<tr><td colspan="7" style="padding:20px;text-align:center;color:var(--t3)">⏳ Cari listesi yükleniyor...</td></tr>';

  const params = new URLSearchParams({limit:500});

  if(q) params.set('q',q); if(tip) params.set('tip',tip);

  const r = await api('/api/logo/cari?'+params); const d = await r.json();

  if(!d.ok){

    document.getElementById('cari-logo-tbody').innerHTML=`<tr><td colspan="7" style="padding:20px;text-align:center;color:var(--r)">${d.msg}</td></tr>`;

    return;

  }

  let cariRecords = d.records||[];

  const siralaC = document.getElementById('cari-sirala')?.value||'kod';

  if(siralaC==='ciro_desc') cariRecords.sort((a,b)=>parseFloat(b.donem_satis_ciro||0)-parseFloat(a.donem_satis_ciro||0));

  else if(siralaC==='bakiye_desc') cariRecords.sort((a,b)=>Math.abs(parseFloat(b.cari_bakiye||0))-Math.abs(parseFloat(a.cari_bakiye||0)));

  _cariLogoData = cariRecords; _cariLogoPg=1;

  document.getElementById('cari-logo-count').textContent=`${_cariLogoData.length} cari`;

  renderCariLogoTablo();

}



function renderCariLogoTablo(){

  const tipRenk={'Müşteri':'bgi','Tedarikçi':'bri','Hem Müşteri Hem Tedarikçi':'bami'};

  const start=(_cariLogoPg-1)*CARI_LOGO_PPG, page=_cariLogoData.slice(start,start+CARI_LOGO_PPG);

  document.getElementById('cari-logo-tbody').innerHTML=page.map(c=>`<tr>

    <td style="font-family:var(--mono);font-size:11px;font-weight:600;color:var(--acc)">${c.cari_kodu}</td>

    <td style="font-size:12px;font-weight:500">${c.cari_adi||'—'}</td>

    <td><span class="bdg ${tipRenk[c.cari_tipi]||'bbi'}" style="font-size:10px">${(c.cari_tipi||'—').slice(0,8)}</span></td>

    <td style="font-size:10px;color:var(--t3)">${c.logo_firma_adi||'—'}</td>

    <td class="num" style="font-family:var(--mono);font-size:11px;color:var(--g)">${parseFloat(c.donem_satis_ciro||0)>0?'₺'+fmtN(parseFloat(c.donem_satis_ciro)):'—'}</td>

    <td class="num" style="font-family:var(--mono);font-size:11px;color:var(--am)">${parseFloat(c.donem_alis_ciro||0)>0?'₺'+fmtN(parseFloat(c.donem_alis_ciro)):'—'}</td>

    <td class="num" style="font-family:var(--mono);font-size:11px;font-weight:600;color:${parseFloat(c.cari_bakiye||0)>0?'var(--g)':parseFloat(c.cari_bakiye||0)<0?'var(--r)':'var(--t3)'}">${c.cari_bakiye!==''?'₺'+fmtN(Math.abs(parseFloat(c.cari_bakiye||0)))+(parseFloat(c.cari_bakiye||0)<0?' ⬇':parseFloat(c.cari_bakiye||0)>0?' ⬆':''):'—'}</td>

    <td style="font-family:var(--mono);font-size:11px">${c.vergi_no||'—'}</td>

    <td style="font-size:11px;color:var(--t2)">${c.sehir||'—'}</td>

    <td style="font-family:var(--mono);font-size:11px">${c.telefon||'—'}</td>

  </tr>`).join('');

  const tot=Math.ceil(_cariLogoData.length/CARI_LOGO_PPG)||1;

  let ph=`<span class="pgi">${_cariLogoData.length} cari</span><div class="pgsp"></div>`;

  if(_cariLogoPg>1) ph+=`<button class="pbn" onclick="_cariLogoPg--;renderCariLogoTablo()">‹</button>`;

  for(let i=Math.max(1,_cariLogoPg-2);i<=Math.min(tot,_cariLogoPg+2);i++)

    ph+=`<button class="pbn${i===_cariLogoPg?' on':''}" onclick="_cariLogoPg=${i};renderCariLogoTablo()">${i}</button>`;

  if(_cariLogoPg<tot) ph+=`<button class="pbn" onclick="_cariLogoPg++;renderCariLogoTablo()">›</button>`;

  document.getElementById('cari-logo-pgn').innerHTML=ph;

}



async function cariLogoExport(){

  if(!_cariLogoData.length){toast('Veri yok','e');return;}

  const nl='\r\n';

  let csv='Cari Kodu,Cari Adi,Tur,Vergi No,Vergi Dairesi,Sehir,Telefon,Email'+nl;

  _cariLogoData.forEach(c=>{

    csv+='"'+(c.cari_kodu||'').replace(/"/g,'""')+'",';

    csv+='"'+(c.cari_adi||'').replace(/"/g,'""')+'",';

    csv+='"'+(c.cari_tipi||'').replace(/"/g,'""')+'",';

    csv+='"'+(c.vergi_no||'').replace(/"/g,'""')+'",';

    csv+='"'+(c.vergi_dairesi||'').replace(/"/g,'""')+'",';

    csv+='"'+(c.sehir||'').replace(/"/g,'""')+'",';

    csv+='"'+(c.telefon||'').replace(/"/g,'""')+'",';

    csv+='"'+(c.email||'').replace(/"/g,'""')+'"'+nl;

  });

  const a=document.createElement('a'); a.href='data:text/csv;charset=utf-8,\uFEFF'+encodeURIComponent(csv);

  a.download='logo_cari_listesi.csv'; a.click(); toast('CSV indirildi','s');

}



// ── GMAIL ─────────────────────────────────────────────────────────────────────

let _gmailMails=[], _aktifMailId=null, _cevapKime='', _cevapKonu='';



async function renderGmail(){

  const r=await api('/api/gmail/config'); const d=await r.json();

  const dot=document.getElementById('gmail-dot');

  const msg=document.getElementById('gmail-conn-msg');

  if(!d.config?.client_id){

    dot.className='src-dot none';

    msg.textContent='Gmail API ayarlanmamış — ⚙ API Ayarları butonuna tıklayın';

    return;

  }

  if(!d.connected){

    dot.className='src-dot err';

    msg.innerHTML='Gmail bağlı değil — <button class="btn bp bsm" onclick="gmailOAuthBaslat()">🔐 Google ile Bağlan</button>';

    return;

  }

  dot.className='src-dot ok';

  msg.innerHTML=`✓ Gmail bağlı — <span style="color:var(--t3)">${d.config.email||'sahinerikin@gmail.com'}</span>`;

  gmailYukle();

}



function showGmailAyarlar(){

  api('/api/gmail/config').then(r=>r.json()).then(d=>{

    const c=d.config||{};

    document.getElementById('gmail-client-id').value=c.client_id||'';

    document.getElementById('gmail-client-secret').value='';

    document.getElementById('gmail-email').value=c.email||'sahinerikin@gmail.com';

  });

  document.getElementById('gmail-ayarlar').style.display='';

}



async function gmailAyarKaydet(){

  const cfg={

    client_id:     document.getElementById('gmail-client-id').value.trim(),

    client_secret: document.getElementById('gmail-client-secret').value.trim(),

    email:         document.getElementById('gmail-email').value.trim(),

  };

  if(!cfg.client_id||!cfg.client_secret){toast('Client ID ve Secret zorunlu','e');return;}

  const r=await api('/api/gmail/config/save',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(cfg)});

  if((await r.json()).ok){ toast('Gmail ayarları kaydedildi','s'); }

}



async function gmailOAuthBaslat(){

  await gmailAyarKaydet();

  const r=await api('/api/gmail/oauth/url'); const d=await r.json();

  if(!d.ok){

    const res=document.getElementById('gmail-ayar-result');

    res.style.display=''; res.style.color='var(--r)';

    res.textContent='Hata: '+d.msg; return;

  }

  window.open(d.url,'_blank','width=600,height=700');

  toast('Google onay sayfası açıldı. Onaylayıp geri dönün.','i');

  // 5 sn sonra bağlantıyı kontrol et

  setTimeout(()=>renderGmail(), 5000);

}



async function gmailYukle(){

  const q=document.getElementById('gmail-q')?.value||'';

  document.getElementById('gmail-tbody').innerHTML='<tr><td colspan="4" style="padding:20px;text-align:center;color:var(--t3)">⏳ Mailler yükleniyor...</td></tr>';

  const r=await api('/api/gmail/mails?max=50&q='+encodeURIComponent(q)); const d=await r.json();

  if(!d.ok){

    document.getElementById('gmail-tbody').innerHTML=`<tr><td colspan="4" style="padding:20px;text-align:center;color:var(--r)">${d.msg}</td></tr>`;

    return;

  }

  _gmailMails = d.mails||[];

  document.getElementById('gmail-tbody').innerHTML=_gmailMails.map(m=>`

    <tr style="cursor:pointer;${m.okunmadi?'font-weight:600':''}" onclick="gmailOku('${m.id}')">

      <td style="text-align:center;width:20px">${m.okunmadi?'<span style="color:var(--acc);font-size:8px">●</span>':''}</td>

      <td style="font-size:12px;max-width:220px" class="clip" title="${m.kimden}">${m.kimden}</td>

      <td style="font-size:12px">

        <div style="font-weight:${m.okunmadi?700:400}">${m.konu||'(Konusuz)'}</div>

        <div style="font-size:11px;color:var(--t3);margin-top:2px">${m.snippet||''}</div>

      </td>

      <td style="font-size:11px;color:var(--t3);white-space:nowrap">${m.tarih||''}</td>

    </tr>`).join('');

}



async function gmailOku(id){

  const detay=document.getElementById('gmail-detay');

  detay.style.display='';

  document.getElementById('gd-konu').textContent='⏳ Yükleniyor...';

  document.getElementById('gd-body').textContent='';

  const r=await api('/api/gmail/mail?id='+id); const d=await r.json();

  if(!d.ok){ document.getElementById('gd-konu').textContent='Hata: '+d.msg; return; }

  _aktifMailId=id; _cevapKime=d.kimden; _cevapKonu='Re: '+d.konu;

  document.getElementById('gd-konu').textContent=d.konu||'(Konusuz)';

  document.getElementById('gd-kimden').textContent='Kimden: '+d.kimden+' → Kime: '+d.kime;

  document.getElementById('gd-tarih').textContent=d.tarih;

  document.getElementById('gd-body').textContent=d.body;

  detay.scrollIntoView({behavior:'smooth'});

  // Okundu güncelle

  _gmailMails = _gmailMails.map(m=>m.id===id?{...m,okunmadi:false}:m);

  gmailYukle();

}



function openMailYaz(){ _cevapKime=''; _cevapKonu=''; _openMailModal('Yeni Mail','',''); }

function openMailCevapla(){ _openMailModal('Cevapla — '+_cevapKonu, _cevapKime, _cevapKonu); }

function _openMailModal(baslik, kime, konu){

  document.getElementById('mail-mtit').textContent=baslik;

  document.getElementById('mail-kime').value=kime;

  document.getElementById('mail-konu').value=konu;

  document.getElementById('mail-body').value='';

  document.getElementById('mail-mod').classList.add('on');

}

function closeMailMod(){ document.getElementById('mail-mod').classList.remove('on'); }



async function mailGonder(){

  const kime  = document.getElementById('mail-kime').value.trim();

  const konu  = document.getElementById('mail-konu').value.trim();

  const body  = document.getElementById('mail-body').value.trim();

  if(!kime||!body){ toast('Alıcı ve mesaj zorunlu','e'); return; }

  const btn=document.getElementById('mail-mod').querySelector('.btn.bp');

  btn.textContent='⏳ Gönderiliyor...'; btn.disabled=true;

  const r=await api('/api/gmail/send',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({kime,konu,body})});

  const d=await r.json();

  btn.textContent='📤 Gönder'; btn.disabled=false;

  if(d.ok){ toast('✓ Mail gönderildi','s'); closeMailMod(); }

  else toast('Hata: '+d.msg,'e');

}



// ── NOTLAR ────────────────────────────────────────────────────────────────────

let _notlar = [];

const _notRenkler = {

  default: {bg:'var(--s2)',border:'var(--b1)',label:''},

  mavi:    {bg:'rgba(59,130,246,.08)',border:'rgba(59,130,246,.3)',label:'🟦'},

  yesil:   {bg:'rgba(16,185,129,.08)',border:'rgba(16,185,129,.3)',label:'🟩'},

  sari:    {bg:'rgba(245,158,11,.08)',border:'rgba(245,158,11,.3)',label:'🟨'},

  kirmizi: {bg:'rgba(239,68,68,.08)', border:'rgba(239,68,68,.3)', label:'🟥'},

};



async function renderNotlar(){

  const r=await api('/api/notlar'); const d=await r.json();

  _notlar = d.notlar||[];

  _renderNotlar();

}



function _renderNotlar(){

  const q=(document.getElementById('not-q')?.value||'').toLowerCase();

  const rf=document.getElementById('not-renk-f')?.value||'';

  let filtered=_notlar;

  if(q) filtered=filtered.filter(n=>(n.baslik||'').toLowerCase().includes(q)||(n.metin||'').toLowerCase().includes(q));

  if(rf) filtered=filtered.filter(n=>(n.renk||'default')===rf);

  document.getElementById('not-count').textContent=`${filtered.length} not`;

  document.getElementById('not-listesi').innerHTML=filtered.map(n=>{

    const r=_notRenkler[n.renk||'default']||_notRenkler.default;

    const guncTag = n.guncellendi

      ? `<span style="color:var(--am);font-size:10px" title="Güncellendi: ${n.guncellendi}">✎</span>`

      : '';

    return `<div style="background:${r.bg};border:1px solid ${r.border};border-radius:var(--radl);padding:14px;position:relative;display:flex;flex-direction:column;gap:0;transition:box-shadow .15s" onmouseenter="this.style.boxShadow='0 4px 16px rgba(0,0,0,.25)'" onmouseleave="this.style.boxShadow='none'">

      <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:8px">

        <div style="font-size:10px;color:var(--t3);font-family:var(--mono)">${n.tarih||''} ${guncTag}</div>

        <div style="display:flex;gap:4px">

          <button onclick="openNotEdit(${n.id})"

            style="background:rgba(255,255,255,.06);border:1px solid var(--b2);color:var(--t2);cursor:pointer;font-size:11px;padding:2px 7px;border-radius:4px;line-height:1.4"

            title="Düzenle">✏</button>

          <button onclick="notSil(${n.id})"

            style="background:none;border:none;color:var(--t3);cursor:pointer;font-size:13px;padding:2px 4px;line-height:1"

            title="Sil">✕</button>

        </div>

      </div>

      ${n.baslik?`<div style="font-weight:700;font-size:13px;margin-bottom:6px;color:var(--t1)">${n.baslik}</div>`:''}

      <div style="font-size:13px;color:var(--t1);white-space:pre-wrap;word-break:break-word;flex:1;cursor:pointer" onclick="openNotEdit(${n.id})" title="Düzenlemek için tıklayın">${n.metin||''}</div>

    </div>`;

  }).join('');

  if(!filtered.length) document.getElementById('not-listesi').innerHTML='<div style="color:var(--t3);font-size:12px;padding:20px">Henüz not yok — yukarıdan ekleyin</div>';

}



let _notEditId = null;



function openNotEdit(id){

  const n = _notlar.find(x=>x.id===id);

  if(!n) return;

  _notEditId = id;

  document.getElementById('not-edit-baslik').value = n.baslik||'';

  document.getElementById('not-edit-metin').value  = n.metin||'';

  document.getElementById('not-edit-renk').value   = n.renk||'default';

  const tarihEl = document.getElementById('not-edit-tarih');

  tarihEl.innerHTML = 'Oluşturuldu: '+n.tarih+(n.guncellendi?' · Güncellendi: '+n.guncellendi:'');

  document.getElementById('not-edit-mod').classList.add('on');

  setTimeout(()=>document.getElementById('not-edit-metin').focus(), 100);

}



function closeNotEditMod(){

  document.getElementById('not-edit-mod').classList.remove('on');

  _notEditId = null;

}



async function notEditKaydet(){

  if(_notEditId === null) return;

  const baslik = document.getElementById('not-edit-baslik').value.trim();

  const metin  = document.getElementById('not-edit-metin').value.trim();

  const renk   = document.getElementById('not-edit-renk').value;

  if(!metin){ toast('Not metni boş olamaz','e'); return; }



  const r = await api('/api/notlar/edit',{

    method:'POST',

    headers:{'Content-Type':'application/json'},

    body: JSON.stringify({id:_notEditId, baslik, metin, renk})

  });

  if((await r.json()).ok){

    // Yerel listeyi güncelle

    const idx = _notlar.findIndex(n=>n.id===_notEditId);

    if(idx>=0){

      _notlar[idx].baslik      = baslik;

      _notlar[idx].metin       = metin;

      _notlar[idx].renk        = renk;

      _notlar[idx].guncellendi = new Date().toLocaleString('tr-TR',{day:'2-digit',month:'2-digit',year:'numeric',hour:'2-digit',minute:'2-digit'});

    }

    closeNotEditMod();

    _renderNotlar();

    toast('Not güncellendi','s');

  }

}



async function notEkle(){

  const baslik = document.getElementById('not-baslik').value.trim();

  const metin  = document.getElementById('not-metin').value.trim();

  const renk   = document.getElementById('not-renk').value;

  if(!metin){ toast('Not metni boş olamaz','e'); return; }

  const r=await api('/api/notlar/add',{method:'POST',headers:{'Content-Type':'application/json'},

    body:JSON.stringify({baslik,metin,renk})});

  const d=await r.json();

  if(d.ok){

    _notlar.unshift(d.not);

    document.getElementById('not-baslik').value='';

    document.getElementById('not-metin').value='';

    document.getElementById('not-renk').value='default';

    _renderNotlar();

    toast('Not eklendi','s');

  }

}



async function notSil(id){

  if(!confirm('Bu notu silmek istediğinizden emin misiniz?')) return;

  const r=await api('/api/notlar/delete',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({id})});

  if((await r.json()).ok){

    _notlar=_notlar.filter(n=>n.id!==id);

    _renderNotlar();

    toast('Not silindi','i');

  }

}





// ── AI ASİSTAN ────────────────────────────────────────────────────────────────

let _aiMsgs = [];  // Konuşma geçmişi

let _aiTyping = false;



async function renderAI(){

  // API key durumu

  const r = await api('/api/ai/config'); const d = await r.json();

  const statusEl = document.getElementById('ai-conn-status');

  if(!statusEl) return;

  if(d.config.has_key){

    statusEl.innerHTML=`<span style="color:var(--g)">✓ API bağlı</span>

      <span style="color:var(--t3);margin-left:6px">${d.config.model||'gpt-4o-mini'}</span>`;

  } else {

    statusEl.innerHTML=`<span style="color:var(--am)">⚠ API key girilmemiş</span>

      <button class="btn bp bsm" style="margin-left:8px" onclick="document.getElementById('ai-ayar-bar').style.display=''">⚙ Ayarla</button>`;

  }

  // Model seçeneğini set et

  const modelSel = document.getElementById('ai-model');

  if(modelSel && d.config.model) modelSel.value = d.config.model;

}



async function aiAyarKaydet(){

  const key   = document.getElementById('ai-api-key').value.trim();

  const model = document.getElementById('ai-model').value;

  if(!key){ toast('API key zorunlu','e'); return; }

  const cfg = { api_key: key, model };

  const r = await api('/api/ai/config/save',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(cfg)});

  if((await r.json()).ok){

    toast('AI ayarları kaydedildi','s');

    document.getElementById('ai-ayar-bar').style.display='none';

    document.getElementById('ai-api-key').value='';

    renderAI();

  }

}



function aiTemizle(){

  _aiMsgs = [];

  const msgs = document.getElementById('ai-msgs');

  msgs.innerHTML=`<div id="ai-welcome" style="text-align:center;padding:30px 20px">

    <div style="font-size:32px;margin-bottom:12px">🤖</div>

    <div style="font-size:14px;font-weight:600;margin-bottom:8px">Konuşma temizlendi</div>

    <div style="font-size:12px;color:var(--t3)">Yeni bir soru yazabilirsiniz.</div>

  </div>`;

}



function aiHizliSor(soru){

  document.getElementById('ai-input').value = soru;

  aiGonder();

}



function _aiRenderMsg(role, text){

  const msgs = document.getElementById('ai-msgs');

  const welcome = document.getElementById('ai-welcome');

  if(welcome) welcome.remove();



  const isUser = role === 'user';

  const div = document.createElement('div');

  div.style.cssText = `display:flex;gap:10px;align-items:flex-start;${isUser?'flex-direction:row-reverse':''}`;



  // Avatar

  const av = document.createElement('div');

  av.style.cssText = `width:32px;height:32px;border-radius:50%;display:flex;align-items:center;justify-content:center;font-size:14px;flex-shrink:0;${isUser?'background:var(--acc)':'background:var(--s3);border:1px solid var(--b2)'}`;

  av.textContent = isUser ? '👤' : '🤖';



  // Bubble

  const bubble = document.createElement('div');

  bubble.style.cssText = `max-width:78%;padding:12px 15px;border-radius:${isUser?'16px 4px 16px 16px':'4px 16px 16px 16px'};font-size:13px;line-height:1.6;${isUser?'background:var(--acc);color:#fff':'background:var(--s3);border:1px solid var(--b1);color:var(--t1)'}`;



  // Markdown benzeri format

  let formatted = text

    .replace(/\*\*(.+?)\*\*/g, '<strong>$1</strong>')

    .replace(/\*(.+?)\*/g, '<em>$1</em>')

    .replace(/`(.+?)`/g, '<code style="background:rgba(0,0,0,.2);padding:1px 5px;border-radius:3px;font-family:var(--mono)">$1</code>')

    .replace(/^#{1,3}\s+(.+)$/gm, '<div style="font-weight:700;margin:6px 0 3px">$1</div>')

    .replace(/^[-•]\s+(.+)$/gm, '<div style="padding-left:14px">• $1</div>')

    .replace(/\\n\\n/g, '<br><br>')

    .replace(/\\n/g, '<br>');



  bubble.innerHTML = formatted;

  div.appendChild(av);

  div.appendChild(bubble);

  msgs.appendChild(div);

  msgs.scrollTop = msgs.scrollHeight;

}



function _aiRenderTyping(){

  const msgs = document.getElementById('ai-msgs');

  const div = document.createElement('div');

  div.id = 'ai-typing-indicator';

  div.style.cssText = 'display:flex;gap:10px;align-items:flex-start';

  div.innerHTML=`

    <div style="width:32px;height:32px;border-radius:50%;background:var(--s3);border:1px solid var(--b2);display:flex;align-items:center;justify-content:center;font-size:14px">🤖</div>

    <div style="background:var(--s3);border:1px solid var(--b1);border-radius:4px 16px 16px 16px;padding:12px 16px">

      <div style="display:flex;gap:4px;align-items:center">

        <div class="typing-dot"></div><div class="typing-dot"></div><div class="typing-dot"></div>

      </div>

    </div>`;

  msgs.appendChild(div);

  msgs.scrollTop = msgs.scrollHeight;

}



async function aiGonder(){

  if(_aiTyping) return;

  const input = document.getElementById('ai-input');

  const soru = input.value.trim();

  if(!soru) return;



  input.value = '';

  input.style.height = '44px';



  // Kullanıcı mesajı ekle

  _aiMsgs.push({role:'user', content: soru});

  _aiRenderMsg('user', soru);

  _aiTyping = true;



  const btn = document.getElementById('ai-send-btn');

  btn.textContent = '⏳'; btn.disabled = true;

  _aiRenderTyping();



  try {

    const r = await fetch('/api/ai/chat', {

      method: 'POST',

      headers: {'Content-Type': 'application/json'},

      body: JSON.stringify({messages: _aiMsgs})

    });

    const d = await r.json();



    const typingEl = document.getElementById('ai-typing-indicator');

    if(typingEl) typingEl.remove();



    if(d.ok){

      _aiMsgs.push({role:'assistant', content: d.content});

      _aiRenderMsg('assistant', d.content);

    } else {

      _aiRenderMsg('assistant', '❌ Hata: ' + (d.msg||'Bilinmiyor'));

      if(d.msg && d.msg.includes('API key')) {

        document.getElementById('ai-ayar-bar').style.display='';

      }

    }

  } catch(e) {

    const typingEl = document.getElementById('ai-typing-indicator');

    if(typingEl) typingEl.remove();

    _aiRenderMsg('assistant', '❌ Bağlantı hatası: ' + e.message);

  } finally {

    _aiTyping = false;

    btn.textContent = '➤ Gönder'; btn.disabled = false;

    document.getElementById('ai-input').focus();

  }

}





// ── VERİTABANI YÖNETİMİ ──────────────────────────────────────────────────────

let _dbSorguRows = [];



async function renderDB(){ await dbDurumYukle(); }



async function dbDurumYukle(){

  const r = await api('/api/db/status'); const d = await r.json();

  if(!d.ok){

    document.getElementById('db-status-grid').innerHTML=`<div style="color:var(--r);font-size:13px;grid-column:1/-1">${d.msg}</div>`;

    return;

  }

  const info = d.info;



  // Özet kartlar

  const toplam = Object.entries(info)

    .filter(([k])=>['giderler','gelirler','banka_bakiyeleri','kredi_limitleri','cari_hareketler','notlar'].includes(k))

    .reduce((s,[,v])=>s+v, 0);



  document.getElementById('db-status-grid').innerHTML=`

    <div class="met"><div class="met-stripe" style="background:var(--acc)"></div>

      <div class="mlb">Toplam Kayıt</div>

      <div class="mvl" style="color:var(--acc);font-size:20px">${toplam.toLocaleString('tr-TR')}</div>

      <div class="msb">Tüm tablolarda</div></div>

    <div class="met"><div class="met-stripe" style="background:var(--g)"></div>

      <div class="mlb">Veritabanı Boyutu</div>

      <div class="mvl" style="color:var(--g);font-size:20px">${info.db_boyut_kb} KB</div>

      <div class="msb">nakit_akis.db</div></div>

    <div class="met"><div class="met-stripe" style="background:var(--am)"></div>

      <div class="mlb">Son Senkronizasyon</div>

      <div class="mvl" style="font-size:12px;color:var(--am)">${info.son_sync?.created_at||'—'}</div>

      <div class="msb">${info.son_sync?.kayit_sayisi||0} kayıt</div></div>

  `;



  // Tablo detayları

  const tblMap = {

    giderler: {label:'Giderler', icon:'💸', renk:'var(--r)'},

    gelirler: {label:'Gelirler', icon:'💰', renk:'var(--g)'},

    banka_bakiyeleri: {label:'Banka Bakiyeleri', icon:'🏦', renk:'var(--acc)'},

    kredi_limitleri: {label:'Kredi Limitleri', icon:'💳', renk:'var(--am)'},

    cari_hareketler: {label:'Cari Hareketler', icon:'⇄', renk:'var(--acc2)'},

    notlar: {label:'Notlar', icon:'📝', renk:'var(--t2)'},

  };

  document.getElementById('db-tablo-detay').innerHTML = Object.entries(tblMap).map(([k,v])=>`

    <div style="background:var(--s2);border:1px solid var(--b1);border-radius:var(--rad);padding:12px 14px;display:flex;align-items:center;gap:12px">

      <div style="font-size:20px">${v.icon}</div>

      <div style="flex:1">

        <div style="font-size:12px;font-weight:600">${v.label}</div>

        <div style="font-family:var(--mono);font-size:18px;color:${v.renk};font-weight:700">${(info[k]||0).toLocaleString('tr-TR')}</div>

      </div>

      <button class="btn bgh bsm" onclick="dbHazirSorgu('SELECT * FROM ${k} LIMIT 50')" style="font-size:10px">Gör</button>

    </div>`).join('');

}



async function dbSync(){

  const msg = document.getElementById('db-sync-msg');

  msg.textContent = '⏳ Senkronize ediliyor...';

  const r = await api('/api/db/sync',{method:'POST',headers:{'Content-Type':'application/json'},body:'{}'});

  const d = await r.json();

  if(d.ok){

    const total = Object.values(d.results).filter(v=>typeof v==='number').reduce((s,v)=>s+v,0);

    msg.textContent = `✓ ${total} kayıt senkronize edildi`;

    msg.style.color = 'var(--g)';

    toast('Veritabanı güncellendi','s');

    dbDurumYukle();

  } else {

    msg.textContent = '✕ Hata: ' + d.msg;

    msg.style.color = 'var(--r)';

  }

  setTimeout(()=>{ msg.textContent=''; msg.style.color=''; }, 5000);

}



function dbHazirSorgu(sql){

  document.getElementById('db-sql-input').value = sql;

  dbSorguCalistir();

}



async function dbSorguCalistir(){

  const sql = document.getElementById('db-sql-input').value.trim();

  if(!sql){ toast('SQL yazın','e'); return; }



  const resDiv  = document.getElementById('db-sorgu-result');

  const hataDiv = document.getElementById('db-sorgu-hata');

  resDiv.style.display='none'; hataDiv.style.display='none';



  const r = await api('/api/db/sorgu?sql='+encodeURIComponent(sql));

  const d = await r.json();



  if(!d.ok){

    hataDiv.style.display=''; hataDiv.textContent='Hata: '+d.msg;

    return;

  }



  _dbSorguRows = d.rows||[];

  resDiv.style.display='';

  document.getElementById('db-sorgu-info').textContent =

    `${d.count} kayıt bulundu${d.count>500?' (ilk 500 gösteriliyor)':''}`;



  if(!_dbSorguRows.length){

    document.getElementById('db-sorgu-table').innerHTML='<tr><td style="padding:16px;color:var(--t3)">Sonuç yok</td></tr>';

    return;

  }



  const cols = Object.keys(_dbSorguRows[0]);

  const thead = '<thead><tr>'+cols.map(c=>`<th style="padding:6px 10px;font-size:11px;text-align:left;white-space:nowrap">${c}</th>`).join('')+'</tr></thead>';

  const tbody = '<tbody>'+_dbSorguRows.map(row=>

    '<tr>'+cols.map(c=>{

      const v = row[c];

      const isNum = typeof v==='number';

      const isId  = c==='id';

      return `<td style="padding:5px 10px;font-size:11.5px;font-family:${isNum||isId?'var(--mono)':'inherit'};text-align:${isNum?'right':'left'};max-width:200px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap" title="${v||''}">${v===null||v===undefined?'<span style="color:var(--t3)">NULL</span>':v}</td>`;

    }).join('')+'</tr>'

  ).join('')+'</tbody>';

  document.getElementById('db-sorgu-table').innerHTML = thead+tbody;

}



async function dbSorguExport(){

  if(!_dbSorguRows.length){ toast('Sonuç yok','e'); return; }

  const cols = Object.keys(_dbSorguRows[0]);

  const nl='\r\n';

  let csv = cols.join(',')+nl;

  _dbSorguRows.forEach(row=>{

    csv += cols.map(c=>'"'+String(row[c]===null||row[c]===undefined?'':row[c]).replace(/"/g,'""')+'"').join(',')+nl;

  });

  const a=document.createElement('a');

  a.href='data:text/csv;charset=utf-8,\uFEFF'+encodeURIComponent(csv);

  a.download='sorgu_sonucu.csv'; a.click();

  toast('CSV indirildi','s');

}





// ── OTEL YÖNETİMİ ────────────────────────────────────────────────────────────

let _aktifOtel = 'otel1';

let _otelRezervler = [];

let _otelPg = 1; const OTEL_PPG = 30;

let _otelEditId = null;



async function renderOtel(){

  await otelAyarlarYukle();

  await otelOzetYukle();

}



async function otelAyarlarYukle(){

  const r = await api('/api/otel/config'); const d = await r.json();

  const cfg = d.config || {};

  // Tab adlarını güncelle

  const o1 = cfg.otel1?.ad || 'Otel 1';

  const o2 = cfg.otel2?.ad || 'Otel 2';

  document.getElementById('otel-tab-1').textContent = '🏨 ' + o1;

  document.getElementById('otel-tab-2').textContent = '🏨 ' + o2;

}



function otelTab(otel){

  _aktifOtel = otel;

  ['otel1','otel2'].forEach(o=>{

    const btn = document.getElementById('otel-tab-'+o.slice(-1));

    if(btn){ btn.style.background = o===otel?'var(--acc)':'transparent'; btn.style.color = o===otel?'#fff':''; }

  });

  otelSekme(_aktifSekme||'ozet');

}



let _aktifSekme = 'ozet';

function otelSekme(sekme){

  _aktifSekme = sekme;

  ['ozet','takvim','odalar','rezerv','gelir'].forEach(s=>{

    document.getElementById('ostab-'+s+'-view').style.display = s===sekme?'':'none';

    const btn = document.getElementById('ostab-'+s);

    if(btn){ btn.style.borderBottom = s===sekme?'2px solid var(--acc)':'none'; btn.style.color = s===sekme?'var(--acc)':''; }

  });

  if(sekme==='ozet')   otelOzetYukle();

  if(sekme==='takvim') otelTakvimYukle();

  if(sekme==='odalar') otelOdalarYukle();

  if(sekme==='rezerv') otelRezervYukle();

  if(sekme==='gelir')  otelGelirYukle();

}



async function otelOzetYukle(){

  const r = await api('/api/otel/istatistik?otel='+_aktifOtel); const d = await r.json();

  const cfg = (await (await api('/api/otel/config')).json()).config || {};

  const otelAd = cfg[_aktifOtel]?.ad || (_aktifOtel==='otel1'?'Otel 1':'Otel 2');



  // Metrikler

  document.getElementById('otel-mets').innerHTML=`

    <div class="met"><div class="met-stripe" style="background:var(--acc)"></div>

      <div class="mlb">Doluluk Oranı</div>

      <div class="mvl" style="color:${d.doluluk_oran>=70?'var(--g)':d.doluluk_oran>=40?'var(--am)':'var(--r)'};font-size:22px;font-weight:700">%${d.doluluk_oran}</div>

      <div class="msb">${d.aktif_konak} / ${d.oda_sayisi} oda dolu</div></div>

    <div class="met"><div class="met-stripe" style="background:var(--g)"></div>

      <div class="mlb">Toplam Gelir</div>

      <div class="mvl" style="color:var(--g);font-size:16px">₺${fmtN(d.toplam_gelir)}</div>

      <div class="msb">${d.toplam_gece} gece · ${d.toplam_rezerv} rezerv</div></div>

    <div class="met"><div class="met-stripe" style="background:var(--am)"></div>

      <div class="mlb">Bugün Check-in</div>

      <div class="mvl" style="color:var(--am);font-size:22px">${d.bugun_checkin}</div>

      <div class="msb">misafir</div></div>

    <div class="met"><div class="met-stripe" style="background:var(--r)"></div>

      <div class="mlb">Bugün Check-out</div>

      <div class="mvl" style="color:var(--r);font-size:22px">${d.bugun_checkout}</div>

      <div class="msb">misafir</div></div>`;



  // Bugün check-in listesi

  document.getElementById('otel-bugun-checkin-card').innerHTML=`

    <div style="font-weight:600;font-size:13px;margin-bottom:10px;color:var(--am)">⬇ Bugün Check-in (${d.bugun_checkin})</div>

    ${d.bugun_checkin_list.length

      ? d.bugun_checkin_list.map(r=>`<div style="display:flex;justify-content:space-between;padding:6px 0;border-bottom:1px solid var(--b1);font-size:12px">

          <div><strong>${r.misafir_adi||'—'}</strong> <span style="color:var(--t3);font-size:10px">Oda ${r.oda||'?'}</span></div>

          <div style="display:flex;gap:6px;align-items:center">

            <span style="font-size:10px;color:var(--t3)">${r.kanal||''}</span>

            ${r.telefon?`<button onclick="waHizliAc('${r.telefon}','${r.misafir_adi}')" class="btn bgh" style="font-size:10px;padding:2px 6px">💬</button>`:''}

          </div>

        </div>`).join('')

      : '<div style="color:var(--t3);font-size:12px;padding:8px 0">Bugün check-in yok</div>'}`;



  // Bugün check-out listesi

  document.getElementById('otel-bugun-checkout-card').innerHTML=`

    <div style="font-weight:600;font-size:13px;margin-bottom:10px;color:var(--r)">⬆ Bugün Check-out (${d.bugun_checkout})</div>

    ${d.bugun_checkout_list.length

      ? d.bugun_checkout_list.map(r=>`<div style="display:flex;justify-content:space-between;padding:6px 0;border-bottom:1px solid var(--b1);font-size:12px">

          <div><strong>${r.misafir_adi||'—'}</strong> <span style="color:var(--t3);font-size:10px">Oda ${r.oda||'?'}</span></div>

          <div style="display:flex;gap:6px;align-items:center">

            <span style="font-size:10px;color:var(--t3)">${r.kanal||''}</span>

            ${r.telefon?`<button onclick="waHizliAc('${r.telefon}','${r.misafir_adi}')" class="btn bgh" style="font-size:10px;padding:2px 6px">💬</button>`:''}

          </div>

        </div>`).join('')

      : '<div style="color:var(--t3);font-size:12px;padding:8px 0">Bugün check-out yok</div>'}`;



  // Kanal dağılımı SVG bar

  const kanallar = Object.entries(d.kanal_dagilim||{}).sort((a,b)=>b[1]-a[1]);

  if(kanallar.length){

    const maxK = Math.max(...kanallar.map(([,v])=>v));

    const clrs = ['#3b82f6','#10b981','#f59e0b','#ef4444','#8b5cf6','#06b6d4'];

    document.getElementById('otel-kanal-chart').innerHTML=`

      <div style="font-weight:600;font-size:12px;color:var(--t2);margin-bottom:8px">Kanal Dağılımı</div>

      ${kanallar.map(([k,v],i)=>`

        <div style="display:flex;align-items:center;gap:10px;margin-bottom:6px">

          <div style="width:100px;font-size:12px;color:var(--t2);text-align:right;flex-shrink:0">${k}</div>

          <div style="flex:1;background:var(--s2);border-radius:3px;height:18px;overflow:hidden">

            <div style="width:${Math.round(v/maxK*100)}%;background:${clrs[i%clrs.length]};height:100%;border-radius:3px;transition:width .4s"></div>

          </div>

          <div style="width:80px;font-family:var(--mono);font-size:12px;color:var(--t1)">₺${fmtN(v)}</div>

        </div>`).join('')}`;

  }

}



async function otelTakvimYukle(){

  const r = await api('/api/otel/istatistik?otel='+_aktifOtel); const d = await r.json();

  const takvim = d.takvim || [];

  const odaSayisi = d.oda_sayisi || 1;

  const bugun = new Date().toISOString().slice(0,10);



  let html = '<div style="display:flex;flex-wrap:wrap;gap:4px;min-width:700px">';

  const aylar = {};

  takvim.forEach(g=>{

    const ay = g.tarih.slice(0,7);

    if(!aylar[ay]) aylar[ay]=[];

    aylar[ay].push(g);

  });



  Object.entries(aylar).forEach(([ay, gunler])=>{

    const [yil,ayNo] = ay.split('-');

    const ayAdlari=['','Ocak','Şubat','Mart','Nisan','Mayıs','Haziran','Temmuz','Ağustos','Eylül','Ekim','Kasım','Aralık'];

    html += `<div style="margin-bottom:12px;width:100%">

      <div style="font-size:12px;font-weight:600;color:var(--t2);margin-bottom:6px">${ayAdlari[parseInt(ayNo)]} ${yil}</div>

      <div style="display:flex;flex-wrap:wrap;gap:3px">`;

    gunler.forEach(g=>{

      const oran = g.oran;

      const renk = oran>=80?'#10b981':oran>=50?'#f59e0b':oran>=20?'#3b82f6':'var(--s3)';

      const isBugun = g.tarih===bugun;

      const gunNo = parseInt(g.tarih.slice(8,10));

      html += `<div title="${g.tarih}: %${oran} dolu (${g.dolu}/${odaSayisi} oda)"

        style="width:36px;height:36px;border-radius:6px;background:${renk};opacity:${oran>0?0.85:0.4};

               display:flex;flex-direction:column;align-items:center;justify-content:center;cursor:default;

               border:${isBugun?'2px solid #fff':'1px solid rgba(255,255,255,.1)'}">

        <div style="font-size:10px;font-weight:600;color:#fff;line-height:1">${gunNo}</div>

        <div style="font-size:8px;color:rgba(255,255,255,.8);line-height:1">${oran>0?'%'+oran:''}</div>

      </div>`;

    });

    html += '</div></div>';

  });



  html += '</div>';

  html += `<div style="display:flex;gap:16px;margin-top:12px;font-size:11px;flex-wrap:wrap">

    <span><span style="display:inline-block;width:12px;height:12px;background:#10b981;border-radius:2px"></span> %80+ Yüksek</span>

    <span><span style="display:inline-block;width:12px;height:12px;background:#f59e0b;border-radius:2px"></span> %50-80 Orta</span>

    <span><span style="display:inline-block;width:12px;height:12px;background:#3b82f6;border-radius:2px"></span> %20-50 Düşük</span>

    <span><span style="display:inline-block;width:12px;height:12px;background:var(--s3);border-radius:2px;border:1px solid var(--b1)"></span> Boş</span>

  </div>`;



  document.getElementById('otel-takvim-grid').innerHTML = html;

}



async function otelRezervYukle(){

  const q     = document.getElementById('otel-q')?.value||'';

  const durum = document.getElementById('otel-durum-f')?.value||'';

  const kanal = document.getElementById('otel-kanal-f')?.value||'';

  const bas   = document.getElementById('otel-bas-f')?.value||'';

  const bit   = document.getElementById('otel-bit-f')?.value||'';

  const params = new URLSearchParams({otel:_aktifOtel});

  if(q) params.set('q',q); if(durum) params.set('durum',durum);

  if(kanal) params.set('kanal',kanal); if(bas) params.set('baslangic',bas); if(bit) params.set('bitis',bit);

  const r = await api('/api/otel/rezervasyonlar?'+params); const d = await r.json();

  _otelRezervler = d.records||[];



  // Kanal filtresi doldur

  const kanalSel = document.getElementById('otel-kanal-f');

  const curKanal = kanalSel.value;

  kanalSel.innerHTML='<option value="">Tüm Kanallar</option>';

  (d.kanallar||[]).forEach(k=>{ const o=document.createElement('option'); o.value=k; o.text=k; kanalSel.appendChild(o); });

  kanalSel.value=curKanal;



  _otelPg=1; renderOtelTablo();

}



function renderOtelTablo(){

  const durumRenk = {'Rezerve':'bami','Check-in':'bgi','Konaklamakta':'bgi','Check-out':'bbi','İptal':'bri','No-show':'bri'};

  const start=(_otelPg-1)*OTEL_PPG, page=_otelRezervler.slice(start,start+OTEL_PPG);

  let totTutar=0, totGece=0;

  _otelRezervler.forEach(r=>{

    totTutar+=parseFloat(r.tutar||0);

    try{ const g=new Date(r.giris_tarihi),c=new Date(r.cikis_tarihi); totGece+=Math.max(0,(c-g)/86400000); }catch{}

  });



  document.getElementById('otel-tbody').innerHTML=page.map(r=>{

    let gece=0;

    try{ const g=new Date(r.giris_tarihi),c=new Date(r.cikis_tarihi); gece=Math.max(0,Math.round((c-g)/86400000)); }catch{}

    return `<tr>

      <td style="font-family:var(--mono);font-size:11px;color:var(--acc)">${r.rezervasyon_no||'—'}</td>

      <td style="font-size:12px;font-weight:500">${r.misafir_adi||'—'}</td>

      <td style="font-family:var(--mono);font-size:11px">${r.giris_tarihi||'—'}</td>

      <td style="font-family:var(--mono);font-size:11px">${r.cikis_tarihi||'—'}</td>

      <td style="text-align:center;font-family:var(--mono);font-size:11px">${gece}</td>

      <td style="font-size:11px;color:var(--t2)">${r.oda||'—'}</td>

      <td style="font-size:11px"><span class="bdg bbi" style="font-size:9px">${r.kanal||'—'}</span></td>

      <td class="num" style="font-family:var(--mono);font-size:12px;color:var(--g)">${fmtN(parseFloat(r.tutar||0))}</td>

      <td><span class="bdg ${durumRenk[r.durum]||'bbi'}" style="font-size:10px;cursor:pointer" ondblclick="otelDurumToggle(${r.id})">${r.durum||'—'}</span></td>

      <td style="white-space:nowrap">

        ${r.telefon?`<button class="btn bgh" onclick="waHizliAc('${r.telefon}','${r.misafir_adi||''}')" title="WhatsApp" style="font-size:11px;padding:2px 6px">💬</button>`:''}

        <button class="btn bgh" onclick="openOtelRezervEdit(${r.id})" style="font-size:11px;padding:2px 6px">✏</button>

        <button class="btn bgh" style="color:var(--r);font-size:11px;padding:2px 6px" onclick="otelRezervSil(${r.id})">✕</button>

      </td>

    </tr>`;

  }).join('');



  document.getElementById('otel-tfoot').innerHTML=`

    <tr style="border-top:2px solid var(--b2);font-weight:600">

      <td colspan="7" style="padding:8px 10px;font-size:11px;color:var(--t2)">${_otelRezervler.length} rezervasyon · ${Math.round(totGece)} toplam gece</td>

      <td class="num" style="color:var(--g)">${fmtN(totTutar)}</td>

      <td colspan="2"></td>

    </tr>`;



  const tot=Math.ceil(_otelRezervler.length/OTEL_PPG)||1;

  let ph=`<span class="pgi">${_otelRezervler.length} rezervasyon</span><div class="pgsp"></div>`;

  if(_otelPg>1) ph+=`<button class="pbn" onclick="_otelPg--;renderOtelTablo()">‹</button>`;

  for(let i=Math.max(1,_otelPg-2);i<=Math.min(tot,_otelPg+2);i++)

    ph+=`<button class="pbn${i===_otelPg?' on':''}" onclick="_otelPg=${i};renderOtelTablo()">${i}</button>`;

  if(_otelPg<tot) ph+=`<button class="pbn" onclick="_otelPg++;renderOtelTablo()">›</button>`;

  document.getElementById('otel-pgn').innerHTML=ph;

}



async function otelGelirYukle(){

  const r = await api('/api/otel/istatistik?otel='+_aktifOtel); const d = await r.json();

  const rezervler = _otelRezervler.length ? _otelRezervler : (await (await api('/api/otel/rezervasyonlar?otel='+_aktifOtel)).json()).records||[];



  // Aylık gelir

  const ayGelir = {};

  rezervler.forEach(r=>{

    if((r.durum||'').toLowerCase().includes('iptal')) return;

    const ay = (r.giris_tarihi||'').slice(0,7);

    if(!ay) return;

    ayGelir[ay] = (ayGelir[ay]||0) + parseFloat(r.tutar||0);

  });

  const aylar = Object.entries(ayGelir).sort(([a],[b])=>a.localeCompare(b));

  const maxG = Math.max(...aylar.map(([,v])=>v), 1);



  document.getElementById('otel-gelir-content').innerHTML=`

    <div style="display:grid;grid-template-columns:1fr 1fr;gap:16px;margin-bottom:16px">

      <div class="met"><div class="met-stripe" style="background:var(--g)"></div>

        <div class="mlb">Toplam Gelir</div>

        <div class="mvl" style="color:var(--g);font-size:18px">₺${fmtN(d.toplam_gelir)}</div>

        <div class="msb">${d.toplam_rezerv} rezervasyon · ${d.toplam_gece} gece</div></div>

      <div class="met"><div class="met-stripe" style="background:var(--acc)"></div>

        <div class="mlb">Ortalama Geceleme Geliri</div>

        <div class="mvl" style="color:var(--acc);font-size:18px">₺${d.toplam_gece?fmtN(Math.round(d.toplam_gelir/d.toplam_gece)):0}</div>

        <div class="msb">gece başına</div></div>

    </div>

    <div style="font-weight:600;font-size:12px;color:var(--t2);margin-bottom:10px">Aylık Gelir</div>

    ${aylar.map(([ay,val])=>{

      const [yil,ayNo]=ay.split('-');

      const ayAdlari=['','Oca','Şub','Mar','Nis','May','Haz','Tem','Ağu','Eyl','Eki','Kas','Ara'];

      return `<div style="display:flex;align-items:center;gap:10px;margin-bottom:5px">

        <div style="width:70px;font-size:11px;color:var(--t2)">${ayAdlari[parseInt(ayNo)]} ${yil}</div>

        <div style="flex:1;background:var(--s2);border-radius:3px;height:20px">

          <div style="width:${Math.round(val/maxG*100)}%;background:var(--g);height:100%;border-radius:3px"></div>

        </div>

        <div style="width:90px;font-family:var(--mono);font-size:12px;text-align:right">₺${fmtN(val)}</div>

      </div>`;

    }).join('')}

    <div style="margin-top:16px">

      <div style="font-weight:600;font-size:12px;color:var(--t2);margin-bottom:10px">Kanal Bazlı Gelir</div>

      ${Object.entries(d.kanal_dagilim||{}).sort((a,b)=>b[1]-a[1]).map(([k,v])=>`

        <div style="display:flex;justify-content:space-between;padding:6px 0;border-bottom:1px solid var(--b1);font-size:12px">

          <span>${k}</span><span style="font-family:var(--mono);font-weight:600">₺${fmtN(v)}</span>

        </div>`).join('')}

    </div>`;

}



// Rezervasyon CRUD

function openOtelRezervAdd(){

  _otelEditId=null;

  document.getElementById('otel-mod-title').textContent='Yeni Rezervasyon';

  ['orm-no','orm-misafir','orm-oda','orm-tel','orm-aciklama'].forEach(id=>document.getElementById(id).value='');

  document.getElementById('orm-tutar').value='';

  document.getElementById('orm-kanal').value='Direkt';

  document.getElementById('orm-durum').value='Rezerve';

  document.getElementById('orm-giris').value=new Date().toISOString().slice(0,10);

  const bit=new Date(); bit.setDate(bit.getDate()+1);

  document.getElementById('orm-cikis').value=bit.toISOString().slice(0,10);

  document.getElementById('otel-rezerv-mod').classList.add('on');

}



function openOtelRezervEdit(id){

  const r=_otelRezervler.find(x=>x.id===id); if(!r) return;

  _otelEditId=id;

  document.getElementById('otel-mod-title').textContent='Rezervasyon Düzenle';

  document.getElementById('orm-no').value=r.rezervasyon_no||'';

  document.getElementById('orm-misafir').value=r.misafir_adi||'';

  document.getElementById('orm-giris').value=r.giris_tarihi||'';

  document.getElementById('orm-cikis').value=r.cikis_tarihi||'';

  document.getElementById('orm-oda').value=r.oda||'';

  document.getElementById('orm-kanal').value=r.kanal||'Direkt';

  document.getElementById('orm-tutar').value=r.tutar||'';

  document.getElementById('orm-durum').value=r.durum||'Rezerve';

  document.getElementById('orm-tel').value=r.telefon||'';

  document.getElementById('orm-aciklama').value=r.aciklama||'';

  document.getElementById('otel-rezerv-mod').classList.add('on');

}



function closeOtelMod(){ document.getElementById('otel-rezerv-mod').classList.remove('on'); }



async function otelRezervKaydet(){

  const rezerv={

    rezervasyon_no: document.getElementById('orm-no').value.trim(),

    misafir_adi:    document.getElementById('orm-misafir').value.trim(),

    giris_tarihi:   document.getElementById('orm-giris').value,

    cikis_tarihi:   document.getElementById('orm-cikis').value,

    oda:            document.getElementById('orm-oda').value.trim(),

    kanal:          document.getElementById('orm-kanal').value,

    tutar:          parseFloat(document.getElementById('orm-tutar').value)||0,

    durum:          document.getElementById('orm-durum').value,

    telefon:        document.getElementById('orm-tel').value.trim(),

    aciklama:       document.getElementById('orm-aciklama').value.trim(),

  };

  if(!rezerv.misafir_adi||!rezerv.giris_tarihi||!rezerv.cikis_tarihi){

    toast('Misafir adı ve tarihler zorunlu','e'); return;

  }

  const url = _otelEditId!==null ? '/api/otel/rezervasyon/edit' : '/api/otel/rezervasyon/add';

  const body = _otelEditId!==null

    ? JSON.stringify({otel_id:_aktifOtel, id:_otelEditId, rezervasyon:rezerv})

    : JSON.stringify({otel_id:_aktifOtel, rezervasyon:rezerv});

  const r=await api(url,{method:'POST',headers:{'Content-Type':'application/json'},body});

  if((await r.json()).ok){

    toast(_otelEditId!==null?'Rezervasyon güncellendi':'Rezervasyon eklendi','s');

    closeOtelMod(); otelRezervYukle(); otelOzetYukle();

  }

}



async function otelRezervSil(id){

  if(!confirm('Bu rezervasyonu silmek istediğinizden emin misiniz?')) return;

  const r=await api('/api/otel/rezervasyon/delete',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({otel_id:_aktifOtel,id})});

  if((await r.json()).ok){ toast('Rezervasyon silindi','i'); otelRezervYukle(); otelOzetYukle(); }

}



async function otelDurumToggle(id){

  const r=_otelRezervler.find(x=>x.id===id); if(!r) return;

  const sira=['Rezerve','Check-in','Konaklamakta','Check-out'];

  const cur=sira.indexOf(r.durum);

  const yeni=sira[(cur+1)%sira.length];

  r.durum=yeni;

  const rr=await api('/api/otel/rezervasyon/edit',{method:'POST',headers:{'Content-Type':'application/json'},

    body:JSON.stringify({otel_id:_aktifOtel,id,rezervasyon:r})});

  if((await rr.json()).ok){ toast(yeni,'s'); renderOtelTablo(); }

}



function showOtelAyarlar(){

  api('/api/otel/config').then(r=>r.json()).then(d=>{

    const c=d.config||{};

    ['otel1','otel2'].forEach(o=>{

      document.getElementById(o+'-ad').value=c[o]?.ad||'';

      document.getElementById(o+'-oda').value=c[o]?.oda_sayisi||'';

      document.getElementById(o+'-konum').value=c[o]?.konum||'';

      document.getElementById(o+'-pms').value=c[o]?.pms||'';

    });

  });

  document.getElementById('otel-ayarlar').style.display='';

}



async function otelAyarKaydet(){

  const cfg={

    otel1:{ ad:document.getElementById('otel1-ad').value, oda_sayisi:parseInt(document.getElementById('otel1-oda').value)||50, konum:document.getElementById('otel1-konum').value, pms:document.getElementById('otel1-pms').value },

    otel2:{ ad:document.getElementById('otel2-ad').value, oda_sayisi:parseInt(document.getElementById('otel2-oda').value)||50, konum:document.getElementById('otel2-konum').value, pms:document.getElementById('otel2-pms').value },

  };

  const r=await api('/api/otel/config/save',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(cfg)});

  if((await r.json()).ok){ toast('Otel ayarları kaydedildi','s'); document.getElementById('otel-ayarlar').style.display='none'; otelAyarlarYukle(); }

}



function showOtelExcelImport(){ document.getElementById('otel-excel-import').style.display=''; }



async function otelExcelImport(){

  const file = document.getElementById('otel-excel-file').files[0];

  if(!file){ toast('Dosya seçin','e'); return; }

  const resEl = document.getElementById('otel-import-result');

  resEl.style.display=''; resEl.textContent='⏳ Okunuyor...';



  const formData = new FormData(); formData.append('file', file, file.name);

  const r = await fetch('/api/banka/import_ekstre',{method:'POST',body:formData});

  const d = await r.json();



  if(!d.ok){ resEl.style.color='var(--r)'; resEl.textContent='Hata: '+d.msg; return; }



  // Ekstre formatını rezervasyon formatına çevir

  const rows = (d.records||[]).map((r,i)=>({

    rezervasyon_no: 'IMP-'+(i+1),

    misafir_adi: r.aciklama||'Misafir '+(i+1),

    giris_tarihi: r.tarih||'',

    cikis_tarihi: r.tarih||'',

    tutar: r.alacak||r.borc||0,

    kanal: 'Excel Import',

    durum: 'Rezerve',

  }));



  const rr = await api('/api/otel/import_excel',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({otel_id:_aktifOtel,rows})});

  const dd = await rr.json();

  if(dd.ok){

    resEl.style.color='var(--g)'; resEl.textContent=`✓ ${dd.count} rezervasyon yüklendi`;

    setTimeout(()=>{ document.getElementById('otel-excel-import').style.display='none'; otelRezervYukle(); },1500);

  }

}



// ── ODA BAZLI ANALİZ ─────────────────────────────────────────────────────────

async function otelOdalarYukle(){

  const r = await api('/api/otel/rezervasyonlar?otel='+_aktifOtel+'&limit=1000');

  const d = await r.json();

  const cfg = (await (await api('/api/otel/config')).json()).config||{};

  const otelCfg = cfg[_aktifOtel]||{};

  const odaSayisi = parseInt(otelCfg.oda_sayisi)||20;

  const odaListesi = otelCfg.odalar || [];



  // Rezervasyonlardan oda listesi oluştur

  const odaMap = {};

  // Önceden tanımlanmış odalar

  for(let i=1;i<=odaSayisi;i++){

    const no = String(i).padStart(3,'0');

    const cfg2 = odaListesi.find(o=>o.no===no)||{};

    odaMap[no] = {no, tip: cfg2.tip||'Standart', kat: cfg2.kat||Math.ceil(i/10), rezervasyonlar:[], gelir:0};

  }

  // Rezervasyonları odalara dağıt

  (d.records||[]).forEach(r=>{

    const oda = (r.oda||'').replace(/[^0-9a-zA-Z]/g,'').padStart(3,'0');

    if(!odaMap[oda]) odaMap[oda]={no:oda,tip:'Standart',kat:'?',rezervasyonlar:[],gelir:0};

    odaMap[oda].rezervasyonlar.push(r);

    if(!['iptal','İptal'].includes(r.durum||''))

      odaMap[oda].gelir += parseFloat(r.tutar||0);

  });



  const bugun = new Date().toISOString().slice(0,10);

  window._odaMap = odaMap;

  otelOdalarRender();

}



function otelOdalarRender(){

  const q = (document.getElementById('oda-no-f')?.value||'').toLowerCase();

  const durumF = document.getElementById('oda-durum-f')?.value||'';

  const bugun = new Date().toISOString().slice(0,10);

  const odaMap = window._odaMap||{};



  const odalar = Object.values(odaMap).filter(o=>{

    if(q && !o.no.toLowerCase().includes(q)) return false;

    const aktif = o.rezervasyonlar.find(r=>

      r.giris_tarihi<=bugun && r.cikis_tarihi>bugun &&

      !['iptal','İptal'].includes(r.durum||'')

    );

    if(durumF==='dolu' && !aktif) return false;

    if(durumF==='bos'  &&  aktif) return false;

    return true;

  }).sort((a,b)=>a.no.localeCompare(b.no));



  document.getElementById('oda-grid').innerHTML = odalar.map(o=>{

    const bugun2 = new Date().toISOString().slice(0,10);

    const aktif = o.rezervasyonlar.find(r=>

      r.giris_tarihi<=bugun2 && r.cikis_tarihi>bugun2 &&

      !['iptal','İptal'].includes(r.durum||'')

    );

    const gelecek = o.rezervasyonlar.filter(r=>r.giris_tarihi>bugun2 && !['iptal','İptal'].includes(r.durum||'')).length;

    const gecmis  = o.rezervasyonlar.filter(r=>r.cikis_tarihi<=bugun2).length;

    const doluluk = o.rezervasyonlar.length

      ? Math.round(o.rezervasyonlar.filter(r=>!['iptal','İptal'].includes(r.durum||'')).length / Math.max(1,o.rezervasyonlar.length)*100)

      : 0;



    const renk = aktif ? 'var(--r)' : gelecek ? 'var(--am)' : 'var(--s3)';

    const border = aktif ? 'var(--r)' : gelecek ? 'var(--am)' : 'var(--b1)';



    return `<div style="background:var(--s2);border:1px solid ${border};border-radius:var(--radl);padding:12px;cursor:pointer"

      onclick="otelOdaDetay('${o.no}')">

      <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:8px">

        <div style="font-size:18px;font-weight:700;font-family:var(--mono);color:var(--acc)">Oda ${o.no}</div>

        <div style="width:12px;height:12px;border-radius:50%;background:${renk}"></div>

      </div>

      <div style="font-size:11px;color:var(--t3);margin-bottom:8px">${o.tip} · Kat ${o.kat}</div>

      ${aktif ? `<div style="background:rgba(239,68,68,.12);border:1px solid rgba(239,68,68,.3);border-radius:6px;padding:6px 8px;font-size:11px;margin-bottom:6px">

        <div style="font-weight:600;color:var(--r)">🔴 Dolu</div>

        <div style="color:var(--t2)">${aktif.misafir_adi||'—'}</div>

        <div style="color:var(--t3);font-family:var(--mono)">${aktif.giris_tarihi} → ${aktif.cikis_tarihi}</div>

      </div>` : `<div style="background:rgba(16,185,129,.08);border:1px solid rgba(16,185,129,.2);border-radius:6px;padding:6px 8px;font-size:11px;margin-bottom:6px">

        <div style="font-weight:600;color:var(--g)">🟢 Boş</div>

        ${gelecek?`<div style="color:var(--am)">${gelecek} yaklaşan rezervasyon</div>`:'<div style="color:var(--t3)">Rezervasyon yok</div>'}

      </div>`}

      <div style="display:flex;justify-content:space-between;font-size:11px;color:var(--t3);margin-top:4px">

        <span>Geçmiş: ${gecmis}</span>

        <span style="color:var(--g);font-family:var(--mono)">₺${fmtN(o.gelir)}</span>

      </div>

    </div>`;

  }).join('');



  if(!odalar.length) document.getElementById('oda-grid').innerHTML='<div style="color:var(--t3);font-size:12px;padding:20px">Oda bulunamadı</div>';

}



async function otelOdaDetay(odaNo){

  const oda = (window._odaMap||{})[odaNo];

  if(!oda) return;

  const bugun = new Date().toISOString().slice(0,10);



  const gecmis  = oda.rezervasyonlar.filter(r=>r.cikis_tarihi<=bugun).sort((a,b)=>b.giris_tarihi.localeCompare(a.giris_tarihi));

  const aktif   = oda.rezervasyonlar.filter(r=>r.giris_tarihi<=bugun&&r.cikis_tarihi>bugun&&!['iptal','İptal'].includes(r.durum||''));

  const gelecek = oda.rezervasyonlar.filter(r=>r.giris_tarihi>bugun&&!['iptal','İptal'].includes(r.durum||'')).sort((a,b)=>a.giris_tarihi.localeCompare(b.giris_tarihi));

  const iptal   = oda.rezervasyonlar.filter(r=>['iptal','İptal'].includes(r.durum||''));



  const toplamGelir = oda.gelir;

  const toplamGece  = oda.rezervasyonlar.reduce((s,r)=>{

    if(['iptal','İptal'].includes(r.durum||'')) return s;

    try{ return s+Math.max(0,Math.round((new Date(r.cikis_tarihi)-new Date(r.giris_tarihi))/86400000)); }catch{ return s; }

  },0);



  const rezervHTML = (list, baslik, renk) => list.length ? `

    <div style="font-size:12px;font-weight:600;color:${renk};margin:12px 0 6px">${baslik} (${list.length})</div>

    ${list.map(r=>`<div style="background:var(--s3);border-radius:6px;padding:8px 10px;margin-bottom:4px;font-size:12px">

      <div style="display:flex;justify-content:space-between;margin-bottom:2px">

        <strong>${r.misafir_adi||'—'}</strong>

        <span style="font-family:var(--mono);color:var(--g)">₺${fmtN(parseFloat(r.tutar||0))}</span>

      </div>

      <div style="color:var(--t3);font-family:var(--mono);font-size:11px">${r.giris_tarihi} → ${r.cikis_tarihi}</div>

      <div style="color:var(--t3);font-size:11px">${r.kanal||''} ${r.rezervasyon_no?'· '+r.rezervasyon_no:''}</div>

    </div>`).join('')}` : '';



  // Modal oluştur

  const modal = document.createElement('div');

  modal.className='ov on';

  modal.id='oda-detay-modal';

  modal.innerHTML=`<div class="mod" style="width:560px;max-width:95vw;max-height:85vh;overflow-y:auto">

    <div class="mtit">

      <span>🚪 Oda ${odaNo} — ${oda.tip} · Kat ${oda.kat}</span>

      <button class="mclose" onclick="document.getElementById('oda-detay-modal').remove()">✕</button>

    </div>

    <div style="display:grid;grid-template-columns:repeat(3,1fr);gap:8px;margin-bottom:12px">

      <div class="met" style="padding:10px"><div class="mlb">Toplam Gelir</div><div class="mvl" style="color:var(--g);font-size:16px">₺${fmtN(toplamGelir)}</div></div>

      <div class="met" style="padding:10px"><div class="mlb">Toplam Gece</div><div class="mvl" style="color:var(--acc);font-size:16px">${toplamGece}</div></div>

      <div class="met" style="padding:10px"><div class="mlb">Toplam Rezerv</div><div class="mvl" style="color:var(--am);font-size:16px">${oda.rezervasyonlar.length}</div></div>

    </div>

    ${rezervHTML(aktif,'🔴 Şu An Konaklamakta','var(--r)')}

    ${rezervHTML(gelecek,'🟡 Gelecek Rezervasyonlar','var(--am)')}

    ${rezervHTML(gecmis,'⚫ Geçmiş Rezervasyonlar','var(--t2)')}

    ${rezervHTML(iptal,'❌ İptal Edilenler','var(--r)')}

    ${!oda.rezervasyonlar.length?'<div style="color:var(--t3);padding:20px;text-align:center">Bu oda için rezervasyon kaydı yok</div>':''}

    <div style="margin-top:14px;text-align:center">

      <button class="btn bp" onclick="document.getElementById('oda-detay-modal').remove();openOtelRezervAdd()" style="font-size:12px">+ Bu Odaya Rezervasyon Ekle</button>

    </div>

  </div>`;

  document.body.appendChild(modal);

}



function otelOdaEkle(){ toast('Odalar otel ayarlarından oda sayısı girilerek otomatik oluşturulur','i'); }



// ── WHATSAPP ──────────────────────────────────────────────────────────────────

let _waSablonlar = (()=>{ try{ return JSON.parse(localStorage.getItem('wa_sablonlar')||'[]'); }catch(e){ return []; } })();

let _waKisiler   = (()=>{ try{ return JSON.parse(localStorage.getItem('wa_kisiler')||'[]');   }catch(e){ return []; } })();

let _waSablonEditId = null, _waKisiEditId = null;



function renderWA(){

  if(_waSablonlar.length===0) _waSablonlar = [..._waVarsayilanSablonlar];

  waTab('gonder');

}



const _waVarsayilanSablonlar = [

  {id:1, ad:'Merhaba / Selamlama', kategori:'Genel', metin:'Merhaba,\n\nUmarım iyisinizdir. Size birkaç önemli konuyu iletmek istedim.\n\nİyi günler diliyorum.'},

  {id:2, ad:'Toplantı Hatırlatması', kategori:'Hatırlatma', metin:'Merhaba,\n\nYarın saat {SAAT} toplantımız olduğunu hatırlatmak istedim.\nKonu: {KONU}\n\nGörüşmek üzere.'},

  {id:3, ad:'Ödeme Hatırlatması', kategori:'Ödeme', metin:'Merhaba,\n\nVadesi yaklaşan {TUTAR} tutarındaki ödemenizi hatırlatmak istedim.\nVade: {TARIH}\n\nTeşekkürler.'},

  {id:4, ad:'Bilgi Talebi', kategori:'Genel', metin:'Merhaba,\n\n{KONU} hakkında bilgi alabilir miyim?\n\nTeşekkürler.'},

  {id:5, ad:'Randevu Onayı', kategori:'Genel', metin:'Merhaba,\n\n{TARIH} tarihindeki randevumuzu onaylıyorum.\nSaat: {SAAT}\n\nGörüşmek üzere.'},

  {id:6, ad:'Teşekkür', kategori:'Genel', metin:'Merhaba,\n\nGöstermiş olduğunuz ilgi ve destek için çok teşekkür ederim.\n\nİyi günler.'},

];



function waTab(tab){

  ['gonder','sablon','liste'].forEach(t=>{

    document.getElementById('watab-'+t+'-view').style.display=t===tab?'':'none';

    const btn=document.getElementById('watab-'+t);

    if(btn){ btn.style.borderBottom=t===tab?'2px solid var(--acc)':'none'; btn.style.color=t===tab?'var(--acc)':''; }

  });

  if(tab==='gonder') waGonderRender();

  if(tab==='sablon') waSablonListeRender();

  if(tab==='liste')  waKisiListeRender();

}



function waGonderRender(){

  const sel = document.getElementById('wa-sablon-sec');

  sel.innerHTML='<option value="">— Şablon seç —</option>';

  _waSablonlar.forEach(s=>{ const o=document.createElement('option'); o.value=s.id; o.text=s.ad; sel.appendChild(o); });



  document.getElementById('wa-hizli-sablon').innerHTML=_waSablonlar.slice(0,6).map(s=>`

    <button class="btn bgh" onclick="waSablonHizliUygula(${s.id})"

      style="text-align:left;padding:8px 12px;font-size:12px;justify-content:flex-start">

      <span style="font-size:10px;color:var(--t3);margin-right:6px">${s.kategori}</span>${s.ad}

    </button>`).join('');

}



function waSablonUygula(){

  const id = parseInt(document.getElementById('wa-sablon-sec').value);

  const s = _waSablonlar.find(x=>x.id===id);

  if(s) document.getElementById('wa-mesaj').value=s.metin;

}



function waSablonHizliUygula(id){

  const s = _waSablonlar.find(x=>x.id===id);

  if(s){ document.getElementById('wa-mesaj').value=s.metin; waTab('gonder'); }

}



function waHizliAc(tel, ad){

  showView('wa');

  document.getElementById('wa-tel').value = tel.replace(/[^+0-9]/g,'');

  const s = _waSablonlar.find(x=>x.ad.includes('Check-in')||x.ad.includes('Rezervasyon'));

  if(s) document.getElementById('wa-mesaj').value = s.metin.replace('{MISAFIR_ADI}',ad||'');

  waTab('gonder');

}



function waGonder(){

  let tel   = waTelFormat(document.getElementById('wa-tel').value);

  const msg = document.getElementById('wa-mesaj').value.trim();

  if(!tel){ toast('Telefon numarası girin','e'); return; }

  const url = 'https://wa.me/'+tel.replace('+','')+(msg?'?text='+encodeURIComponent(msg):'');

  window.open(url,'_blank');

  toast('WhatsApp açıldı','s');

}



function waKisiSec(){

  if(!_waKisiler.length){ toast('Kişi listesi boş — önce kişi ekleyin','e'); return; }

  const kisi = _waKisiler.find(k=>true);

  if(kisi) document.getElementById('wa-tel').value=kisi.telefon;

}



function waSablonListeRender(){

  document.getElementById('wa-sablon-listesi').innerHTML=_waSablonlar.map(s=>`

    <div style="background:var(--s2);border:1px solid var(--b1);border-radius:var(--radl);padding:12px">

      <div style="display:flex;justify-content:space-between;margin-bottom:6px">

        <div>

          <span style="font-weight:600;font-size:12px">${s.ad}</span>

          <span class="bdg bbi" style="font-size:9px;margin-left:6px">${s.kategori}</span>

        </div>

        <div style="display:flex;gap:4px">

          <button class="btn bgh" onclick="waSablonKullan(${s.id})" style="font-size:10px;padding:2px 6px">Kullan</button>

          <button class="btn bgh" onclick="waSablonDuzenle(${s.id})" style="font-size:10px;padding:2px 6px">✏</button>

          <button class="btn bgh" onclick="waSablonSil(${s.id})" style="font-size:10px;padding:2px 6px;color:var(--r)">✕</button>

        </div>

      </div>

      <div style="font-size:11px;color:var(--t3);white-space:pre-wrap;max-height:60px;overflow:hidden">${s.metin||''}</div>

    </div>`).join('');

}



function waKisiListeRender(){

  const q=(document.getElementById('wa-kisi-q')?.value||'').toLowerCase();

  const filtered=_waKisiler.filter(k=>!q||(k.ad||'').toLowerCase().includes(q)||(k.telefon||'').includes(q));

  document.getElementById('wa-kisi-listesi').innerHTML=filtered.map(k=>`

    <div style="background:var(--s2);border:1px solid var(--b1);border-radius:var(--rad);padding:10px 12px;display:flex;align-items:center;gap:10px">

      <div style="width:36px;height:36px;border-radius:50%;background:var(--acc);display:flex;align-items:center;justify-content:center;font-weight:700;color:#fff;font-size:14px;flex-shrink:0">${(k.ad||'?')[0].toUpperCase()}</div>

      <div style="flex:1;min-width:0">

        <div style="font-weight:600;font-size:12px">${k.ad||'—'}</div>

        <div style="font-family:var(--mono);font-size:11px;color:var(--t3)">${k.telefon||'—'}</div>

        ${k.grup?`<span class="bdg bbi" style="font-size:9px">${k.grup}</span>`:''}

      </div>

      <div style="display:flex;gap:4px;flex-shrink:0">

        <button class="btn bp bsm" onclick="waKisiMesajYaz('${k.telefon}')" style="font-size:11px;padding:3px 8px">💬</button>

        <button class="btn bgh" onclick="waKisiDuzenle('${k.telefon}')" style="font-size:10px;padding:3px 6px">✏</button>

        <button class="btn bgh" onclick="waKisiSilFn('${k.telefon}')" style="font-size:10px;padding:3px 6px;color:var(--r)">✕</button>

      </div>

    </div>`).join('');

  if(!filtered.length) document.getElementById('wa-kisi-listesi').innerHTML='<div style="color:var(--t3);font-size:12px;padding:16px">Kişi yok</div>';

}



function waKisiMesajYaz(tel){ document.getElementById('wa-tel').value=tel; waTab('gonder'); }



function openWaSablonEkle(){ _waSablonEditId=null; document.getElementById('wa-sablon-mod-title').textContent='Yeni Şablon'; ['wasm-ad','wasm-metin'].forEach(id=>document.getElementById(id).value=''); document.getElementById('wasm-kategori').value='Genel'; document.getElementById('wa-sablon-mod').classList.add('on'); }



function waSablonDuzenle(id){ _waSablonEditId=id; const s=_waSablonlar.find(x=>x.id===id); if(!s) return; document.getElementById('wa-sablon-mod-title').textContent='Şablonu Düzenle'; document.getElementById('wasm-ad').value=s.ad||''; document.getElementById('wasm-metin').value=s.metin||''; document.getElementById('wasm-kategori').value=s.kategori||'Genel'; document.getElementById('wa-sablon-mod').classList.add('on'); }



function closeWaSablonMod(){ document.getElementById('wa-sablon-mod').classList.remove('on'); }



function waSablonKullan(id){ const s=_waSablonlar.find(x=>x.id===id); if(s){ document.getElementById('wa-mesaj').value=s.metin; waTab('gonder'); } }



function waSablonKaydet(){

  const ad=document.getElementById('wasm-ad').value.trim(); const metin=document.getElementById('wasm-metin').value.trim();

  if(!ad||!metin){ toast('Ad ve metin zorunlu','e'); return; }

  if(_waSablonEditId!==null){

    const idx=_waSablonlar.findIndex(x=>x.id===_waSablonEditId);

    if(idx>=0){ _waSablonlar[idx].ad=ad; _waSablonlar[idx].metin=metin; _waSablonlar[idx].kategori=document.getElementById('wasm-kategori').value; }

  } else {

    _waSablonlar.push({id:Date.now(),ad,metin,kategori:document.getElementById('wasm-kategori').value});

  }

  localStorage.setItem('wa_sablonlar',JSON.stringify(_waSablonlar));

  closeWaSablonMod(); waSablonListeRender(); toast('Şablon kaydedildi','s');

}



function waSablonSil(id){ if(!confirm('Şablonu sil?')) return; _waSablonlar=_waSablonlar.filter(x=>x.id!==id); localStorage.setItem('wa_sablonlar',JSON.stringify(_waSablonlar)); waSablonListeRender(); toast('Silindi','i'); }



function openWaKisiEkle(){ _waKisiEditId=null; ['wakm-ad','wakm-tel','wakm-not'].forEach(id=>document.getElementById(id).value=''); document.getElementById('wakm-grup').value='Misafir'; document.getElementById('wa-kisi-mod').classList.add('on'); }



function waKisiDuzenle(tel){ const k=_waKisiler.find(x=>x.telefon===tel); if(!k) return; _waKisiEditId=tel; document.getElementById('wakm-ad').value=k.ad||''; document.getElementById('wakm-tel').value=k.telefon||''; document.getElementById('wakm-grup').value=k.grup||'Misafir'; document.getElementById('wakm-not').value=k.not||''; document.getElementById('wa-kisi-mod').classList.add('on'); }



function closeWaKisiMod(){ document.getElementById('wa-kisi-mod').classList.remove('on'); }



function waTelFormat(raw){

  let t=raw.trim().replace(/[^+0-9]/g,'');

  if(t.startsWith('+')) return t;

  if(t.startsWith('90') && t.length>=11) return '+'+t;

  if(t.startsWith('05') && t.length>=10) return '+9'+t;

  if(t.startsWith('5') && t.length>=9)  return '+90'+t;

  if(t.length>=7) return '+90'+t;

  return t;

}

function waKisiKaydet(){

  const ad=document.getElementById('wakm-ad').value.trim();

  const tel=waTelFormat(document.getElementById('wakm-tel').value);

  if(!ad||!tel){ toast('Ad ve telefon zorunlu','e'); return; }

  if(_waKisiEditId){ _waKisiler=_waKisiler.filter(x=>x.telefon!==_waKisiEditId); }

  _waKisiler.push({ad,telefon:tel,grup:document.getElementById('wakm-grup').value,not:document.getElementById('wakm-not').value.trim()});

  localStorage.setItem('wa_kisiler',JSON.stringify(_waKisiler));

  closeWaKisiMod(); waKisiListeRender(); toast('Kişi kaydedildi','s');

}



function waKisiSilFn(tel){ if(!confirm('Bu kişiyi sil?')) return; _waKisiler=_waKisiler.filter(x=>x.telefon!==tel); localStorage.setItem('wa_kisiler',JSON.stringify(_waKisiler)); waKisiListeRender(); toast('Silindi','i'); }



// ── WHAPI.CLOUD ENTEGRASYONu ──────────────────────────────────────────────────

let _whapiToken = '';

let _aktifChatId = '';

let _aktifChatAd = '';

let _waOtomatikYenile = null;

let _waKayitliSohbetler = {};



async function whapiConfigYukle(){

  try {

    const r = await api('/api/whapi/config'); const d = await r.json();

    _whapiToken = d.config?.token || '';

    const inp = document.getElementById('whapi-token');

    if(inp) inp.value = _whapiToken;

    const dur = document.getElementById('whapi-durum');

    if(dur) dur.textContent = _whapiToken ? '✓ Token kayıtlı — bağlantıyı test etmek için butona tıklayın' : 'Token girilmedi';

    if(dur) dur.style.color = _whapiToken ? 'var(--g)' : 'var(--t3)';

  } catch(e){}

}



async function whapiKaydet(){

  const token = document.getElementById('whapi-token').value.trim();

  if(!token){ toast('Token boş olamaz','e'); return; }

  const r = await api('/api/whapi/config/save',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({token})});

  const d = await r.json();

  if(d.ok){ _whapiToken=token; toast('Token kaydedildi ✓','s');

    document.getElementById('whapi-durum').textContent='✓ Token kayıtlı';

    document.getElementById('whapi-durum').style.color='var(--g)';

  } else toast('Kayıt hatası: '+d.msg,'e');

}



async function whapiTest(){

  const token = document.getElementById('whapi-token').value.trim();

  if(!token){ toast('Önce token girin','e'); return; }

  const dur = document.getElementById('whapi-durum');

  dur.textContent = '⏳ Test ediliyor...'; dur.style.color='var(--am)';

  try {

    const resp = await fetch('https://gate.whapi.cloud/health', {headers:{'Authorization':'Bearer '+token}});

    const d = await resp.json();

    if(resp.ok){

      dur.textContent = '🟢 Bağlantı başarılı! Durum: '+(d.status||'active');

      dur.style.color = 'var(--g)'; toast('Whapi bağlantısı tamam ✓','s');

    } else {

      dur.textContent = '🔴 Hata: '+(d.message||d.error||'Geçersiz token');

      dur.style.color = 'var(--r)'; toast('Bağlantı başarısız','e');

    }

  } catch(e){

    dur.textContent = '🔴 Bağlantı hatası: '+e.message;

    dur.style.color = 'var(--r)';

  }

}



// Kayıtlı sohbetleri önce göster, sonra Whapi'den güncelle

// ── WHATSAPP KALICI SOHBET SİSTEMİ ───────────────────────────────────────────

let _waSohbetler = {};

let _waPollingTimer = null;

let _waSohbetAraQ = '';



async function waGelenYukle(){

  await waSohbetListeYukle(false);

  const bilgi = document.getElementById('wa-webhook-bilgi');

  if(bilgi) bilgi.innerHTML = `📡 Webhook URL: <strong style="color:var(--acc)">${location.origin}/api/whapi/webhook</strong> — Whapi panelinde bu URL'yi webhook olarak tanimlayin`;

  if(_waPollingTimer) clearInterval(_waPollingTimer);

  _waPollingTimer = setInterval(()=>{ waSohbetListeYukle(false); if(_aktifChatId) waMesajlarYenile(false); }, 30000);

}



async function waSohbetListeYukle(force=true){

  const listEl = document.getElementById('wa-sohbet-listesi');

  if(!listEl) return;

  try {

    const r = await api('/api/whapi/kayitli_sohbetler'); const d = await r.json();

    _waSohbetler = d.sohbetler || {};

    if(_whapiToken && force){

      try { await api('/api/whapi/chats',{method:'POST',headers:{'Content-Type':'application/json'},body:'{}'}); } catch(e){}

    }

    waSohbetListeRender();

  } catch(e){

    if(listEl) listEl.innerHTML='<div style="font-size:12px;color:var(--r);padding:16px">Hata: '+e.message+'</div>';

  }

}



function waSohbetFiltrele(){

  _waSohbetAraQ = (document.getElementById('wa-sohbet-ara')?.value||'').toLowerCase();

  waSohbetListeRender();

}



function waSohbetListeRender(){

  const listEl = document.getElementById('wa-sohbet-listesi');

  if(!listEl) return;

  let girdi = Object.entries(_waSohbetler);

  if(_waSohbetAraQ) girdi = girdi.filter(([id,v])=>(v.ad||id).toLowerCase().includes(_waSohbetAraQ));

  girdi.sort((a,b)=>(b[1].son_zaman||0)-(a[1].son_zaman||0));

  const toplamOkunmadi = girdi.reduce((s,[,v])=>s+(v.okunmadi||0),0);

  const badge = document.getElementById('wa-toplam-okunmadi');

  if(badge){ badge.textContent=toplamOkunmadi||''; badge.style.display=toplamOkunmadi?'':'none'; }

  if(!girdi.length){

    listEl.innerHTML='<div style="font-size:12px;color:var(--t3);padding:24px;text-align:center">'+(Object.keys(_waSohbetler).length===0?'💬 Henüz mesaj yok<br><span style="font-size:10px;margin-top:6px;display:block">Bağlantı sekmesinden token girip mesajlaşmaya başlayın</span>':'🔍 Arama sonucu bulunamadı')+'</div>';

    return;

  }

  listEl.innerHTML = girdi.map(([chatId, v])=>{

    const ad = v.ad || chatId;

    const son = v.son_mesaj || '';

    const zaman = v.son_zaman ? (()=>{ const now=Date.now()/1000,diff=now-v.son_zaman; if(diff<86400) return new Date(v.son_zaman*1000).toLocaleTimeString("tr-TR",{hour:"2-digit",minute:"2-digit"}); return new Date(v.son_zaman*1000).toLocaleDateString("tr-TR",{day:"2-digit",month:"2-digit"}); })() : "";

    const aktif = _aktifChatId===chatId;

    const okunmadi = v.okunmadi||0;

    const escQ = s=>(s||'').replace(/['\\']/g,'').replace(/"/g,'');

    return `<div onclick="waSohbetAc('${escQ(chatId)}','${escQ(ad)}')" style="padding:10px 12px;cursor:pointer;border-bottom:1px solid var(--b1);background:${aktif?'var(--s4)':'transparent'};border-left:${aktif?'3px solid var(--acc)':'3px solid transparent'}">

      <div style="display:flex;align-items:flex-start;gap:8px">

        <div style="width:38px;height:38px;border-radius:50%;background:var(--acc-d);display:flex;align-items:center;justify-content:center;font-size:16px;font-weight:700;color:#fff;flex-shrink:0">${ad.charAt(0).toUpperCase()}</div>

        <div style="flex:1;min-width:0">

          <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:3px">

            <div style="font-size:13px;font-weight:${okunmadi?700:600};color:var(--t1);overflow:hidden;text-overflow:ellipsis;white-space:nowrap;max-width:160px">${ad}</div>

            <div style="font-size:10px;color:${okunmadi?'var(--g)':'var(--t3)'};flex-shrink:0;margin-left:4px">${zaman}</div>

          </div>

          <div style="display:flex;justify-content:space-between;align-items:center">

            <div style="font-size:11px;color:var(--t3);overflow:hidden;text-overflow:ellipsis;white-space:nowrap;max-width:190px">${son}</div>

            ${okunmadi?`<span style="background:var(--g);color:#000;border-radius:10px;font-size:10px;padding:1px 6px;font-weight:700;flex-shrink:0">${okunmadi}</span>`:''}

          </div>

        </div>

      </div>

    </div>`;

  }).join('');

}



async function waSohbetAc(chatId, chatAd){

  _aktifChatId = chatId; _aktifChatAd = chatAd;

  document.getElementById('wa-baslik-ad').textContent = chatAd;

  document.getElementById('wa-baslik-alt').textContent = 'Yükleniyor...';

  const yBtn=document.getElementById('wa-yenile-btn'); if(yBtn) yBtn.style.display='';

  const sBtn=document.getElementById('wa-sil-btn'); if(sBtn) sBtn.style.display='';

  const listEl = document.getElementById('wa-mesaj-listesi');

  listEl.innerHTML='<div style="text-align:center;padding:30px;color:var(--t3);font-size:12px">⏳ Yükleniyor...</div>';

  waSohbetListeRender();

  try {

    const r = await api('/api/whapi/kayitli_mesajlar?chat_id='+encodeURIComponent(chatId));

    const d = await r.json();

    waMesajlarRender(d.mesajlar||[], chatAd);

    if(_waSohbetler[chatId]) _waSohbetler[chatId].okunmadi=0;

    waSohbetListeRender();

    if(_whapiToken) setTimeout(()=>waMesajlarYenile(false), 400);

  } catch(e){

    listEl.innerHTML='<div style="color:var(--r);font-size:12px;padding:16px">Hata: '+e.message+'</div>';

  }

}



async function waMesajlarYenile(showLoading=true){

  if(!_aktifChatId||!_whapiToken) return;

  try {

    const r = await api('/api/whapi/messages',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({chat_id:_aktifChatId,chat_ad:_aktifChatAd,limit:60})});

    const d = await r.json();

    if(d.ok){

      waMesajlarRender(d.messages||[], _aktifChatAd);

      const alt=document.getElementById('wa-baslik-alt');

      if(alt) alt.textContent=(d.yeni>0?d.yeni+' yeni mesaj · ':'')+'Son: '+new Date().toLocaleTimeString('tr-TR');

      waSohbetListeYukle(false);

    }

  } catch(e){}

}



function waMesajlarRender(msgs, baslik){

  const listEl = document.getElementById('wa-mesaj-listesi');

  if(!listEl) return;

  const alt=document.getElementById('wa-baslik-alt');

  if(alt) alt.textContent=msgs.length+' mesaj · '+new Date().toLocaleTimeString('tr-TR');

  if(!msgs.length){ listEl.innerHTML='<div style="text-align:center;padding:40px;color:var(--t3);font-size:12px">💬 Henüz mesaj yok</div>'; return; }

  const sorted=[...msgs].sort((a,b)=>(a.timestamp||0)-(b.timestamp||0));

  let lastDate='';

  listEl.innerHTML=sorted.map(m=>{

    const benden=m.from_me;

    const metin=m.text?.body||m.caption||(m.type==='image'?'📷 Fotoğraf':m.type==='document'?'📎 Dosya':m.type==='audio'?'🎵 Ses':'['+m.type+']');

    const ts=m.timestamp?new Date(m.timestamp*1000):null;

    const zaman=ts?ts.toLocaleTimeString('tr-TR',{hour:'2-digit',minute:'2-digit'}):'';

    const tarih=ts?ts.toLocaleDateString('tr-TR',{day:'2-digit',month:'long',year:'numeric'}):'';

    let dateSep='';

    if(tarih&&tarih!==lastDate){ lastDate=tarih; dateSep=`<div style="text-align:center;margin:10px 0"><span style="background:var(--s3);color:var(--t3);font-size:10px;padding:3px 12px;border-radius:10px">${tarih}</span></div>`; }

    return dateSep+`<div style="display:flex;justify-content:${benden?'flex-end':'flex-start'}">

      <div style="max-width:72%;background:${benden?'#1a4a8a':'var(--s3)'};border-radius:${benden?'14px 14px 4px 14px':'14px 14px 14px 4px'};padding:9px 13px;box-shadow:0 1px 3px rgba(0,0,0,.2)">

        ${!benden?'<div style="font-size:10px;color:var(--acc);font-weight:600;margin-bottom:3px">'+(m.from_name||baslik)+'</div>':''}

        <div style="font-size:13px;color:var(--t1);word-break:break-word;line-height:1.5">${metin}</div>

        <div style="font-size:10px;color:${benden?'rgba(255,255,255,.5)':'var(--t3)'};margin-top:4px;text-align:right">${zaman}${benden?' ✓✓':''}</div>

      </div>

    </div>`;

  }).join('');

  listEl.scrollTop=listEl.scrollHeight;

}



async function waSohbetiSil(){

  if(!_aktifChatId) return;

  if(!confirm(_aktifChatAd+' sohbetini kalıcı olarak sil?')) return;

  await api('/api/whapi/mesaj_sil',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({chat_id:_aktifChatId})});

  _aktifChatId=''; _aktifChatAd='';

  document.getElementById('wa-baslik-ad').textContent='← Sol taraftan sohbet seçin';

  document.getElementById('wa-baslik-alt').textContent='Kalıcı mesaj geçmişi';

  document.getElementById('wa-mesaj-listesi').innerHTML='<div style="font-size:12px;color:var(--t3);text-align:center;margin-top:40px">Sol taraftan bir sohbet seçin</div>';

  const yBtn=document.getElementById('wa-yenile-btn'); if(yBtn) yBtn.style.display='none';

  const sBtn=document.getElementById('wa-sil-btn'); if(sBtn) sBtn.style.display='none';

  waSohbetListeYukle(false); toast('Sohbet silindi','i');

}



async function waYanitGonder(){

  if(!_aktifChatId){ toast('Önce bir sohbet seçin','e'); return; }

  const mesajEl=document.getElementById('wa-yanit-mesaj');

  const mesaj=mesajEl.value.trim();

  if(!mesaj) return;

  mesajEl.value='';

  try {

    const r=await api('/api/whapi/send',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({to:_aktifChatId,chat_ad:_aktifChatAd,message:mesaj})});

    const d=await r.json();

    if(d.ok){ setTimeout(()=>waMesajlarYenile(false),500); waSohbetListeYukle(false); }

    else { mesajEl.value=mesaj; toast('Hata: '+d.msg,'e'); }

  } catch(e){ mesajEl.value=mesaj; toast('Gönderim hatası','e'); }

}



// ── HATIRLATICI ───────────────────────────────────────────────────────────────

let _hatFiltre = 'tumu';

let _hatEditId = null;



async function renderHatirlatma(){

  const q = (document.getElementById('hat-q')?.value||'').toLowerCase();

  const r = await api('/api/hatirlatma?filtre='+_hatFiltre);

  const d = await r.json();

  let hatlar = d.records||[];

  if(q) hatlar = hatlar.filter(h=>

    (h.baslik||'').toLowerCase().includes(q)||(h.not||'').toLowerCase().includes(q));



  // Header badge güncelle (header'da bildirim sayısı)

  const nvHat = document.getElementById('nv-hat');

  if(nvHat){

    if(d.acil>0){

      nvHat.innerHTML=`🔔 HATIRLATICI <span style="background:var(--r);color:#fff;border-radius:10px;padding:1px 6px;font-size:10px;margin-left:4px">${d.acil}</span>`;

    } else {

      nvHat.innerHTML=`🔔 HATIRLATICI`;

    }

  }



  // Özet

  const ozEl = document.getElementById('hat-ozet');

  if(ozEl) ozEl.innerHTML = `<span style="color:var(--am)">${d.bugun_bekleyen||0} bugün</span> · <span style="color:var(--r)">${d.acil||0} bekleyen</span>`;



  const bugun = new Date().toISOString().slice(0,10);

  const listEl = document.getElementById('hat-listesi');



  if(!hatlar.length){

    listEl.innerHTML='<div style="color:var(--t3);font-size:12px;padding:24px;text-align:center">'+

      (_hatFiltre==='tamamlandi'?'Tamamlanan hatırlatma yok':'Hatırlatma yok — yukarıdan ekleyin')+'</div>';

    return;

  }



  // Gruplara ayır: geciken, bugün, yarın, bu hafta, diğer

  const grupla = (h)=>{

    const t = h.tarih||'';

    if(!t) return 'Tarifsiz';

    if(t < bugun) return '⚠ Geciken';

    if(t === bugun) return '📅 Bugün';

    const yarin = new Date(); yarin.setDate(yarin.getDate()+1);

    if(t === yarin.toISOString().slice(0,10)) return '➡ Yarın';

    const haftaSonu = new Date(); haftaSonu.setDate(haftaSonu.getDate()+7);

    if(t <= haftaSonu.toISOString().slice(0,10)) return '📆 Bu Hafta';

    return '🗓 Sonraki';

  };



  const grupMap = {};

  hatlar.forEach(h=>{

    if(h.tamamlandi){ if(!grupMap['✅ Tamamlanan']) grupMap['✅ Tamamlanan']=[]; grupMap['✅ Tamamlanan'].push(h); return; }

    const g = grupla(h);

    if(!grupMap[g]) grupMap[g]=[];

    grupMap[g].push(h);

  });



  const siralama = ['⚠ Geciken','📅 Bugün','➡ Yarın','📆 Bu Hafta','🗓 Sonraki','Tarifsiz','✅ Tamamlanan'];

  let html = '';

  siralama.forEach(grup=>{

    if(!grupMap[grup]?.length) return;

    const isGeciken   = grup.includes('Geciken');

    const isBugun     = grup.includes('Bugün');

    const isTamamlandi= grup.includes('Tamamlanan');

    const grupRenk = isGeciken?'var(--r)':isBugun?'var(--am)':isTamamlandi?'var(--t3)':'var(--t2)';

    html += `<div style="font-size:11px;font-weight:600;color:${grupRenk};text-transform:uppercase;letter-spacing:.06em;margin:14px 0 6px;padding-bottom:4px;border-bottom:1px solid var(--b1)">${grup} (${grupMap[grup].length})</div>`;

    html += grupMap[grup].map(h=>{

      const oncelikRenk = h.oncelik==='yuksek'?'var(--r)':h.oncelik==='dusuk'?'var(--t3)':'var(--acc)';

      const oncelikIcon = h.oncelik==='yuksek'?'🔴':h.oncelik==='dusuk'?'⬇':'●';

      const satir = `

        <div style="display:flex;align-items:center;gap:10px;background:var(--s2);border:1px solid ${h.gecti&&!h.tamamlandi?'rgba(239,68,68,.3)':'var(--b1)'};border-radius:var(--rad);padding:10px 12px;margin-bottom:6px;${h.tamamlandi?'opacity:.55':''}">

          <!-- Checkbox -->

          <div onclick="hatToggle(${h.id})" style="width:20px;height:20px;border-radius:50%;border:2px solid ${h.tamamlandi?'var(--g)':'var(--b2)'};background:${h.tamamlandi?'var(--g)':'transparent'};cursor:pointer;flex-shrink:0;display:flex;align-items:center;justify-content:center;font-size:12px">

            ${h.tamamlandi?'✓':''}

          </div>

          <!-- İçerik -->

          <div style="flex:1;min-width:0">

            <div style="display:flex;align-items:center;gap:6px;margin-bottom:2px">

              <span style="color:${oncelikRenk};font-size:10px">${oncelikIcon}</span>

              <span style="font-size:13px;font-weight:${h.oncelik==='yuksek'?600:400};${h.tamamlandi?'text-decoration:line-through':''}">${h.baslik||'—'}</span>

            </div>

            ${h.not?`<div style="font-size:11px;color:var(--t3);margin-top:1px">${h.not}</div>`:''}

          </div>

          <!-- Tarih/Saat -->

          <div style="text-align:right;flex-shrink:0;font-size:11px">

            <div style="font-family:var(--mono);color:${h.gecti&&!h.tamamlandi?'var(--r)':'var(--t2)'};font-weight:${h.gecti&&!h.tamamlandi?600:400}">${h.tarih||'—'}</div>

            ${h.saat?`<div style="font-family:var(--mono);color:var(--t3)">${h.saat}</div>`:''}

          </div>

          <!-- Butonlar -->

          <div style="display:flex;gap:4px;flex-shrink:0">

            <button class="btn bgh" onclick="hatDuzenle(${h.id})" style="font-size:10px;padding:3px 7px">✏</button>

            <button class="btn bgh" onclick="hatSil(${h.id})" style="font-size:10px;padding:3px 7px;color:var(--r)">✕</button>

          </div>

        </div>`;

      return satir;

    }).join('');

  });



  listEl.innerHTML = html;

}



function hatFiltre(f){

  _hatFiltre = f;

  ['tumu','bugun','bekleyen','tamamlandi'].forEach(t=>{

    const btn = document.getElementById('hf-'+t);

    if(btn){ btn.style.background=t===f?'var(--acc)':'transparent'; btn.style.color=t===f?'#fff':''; }

  });

  renderHatirlatma();

}



async function hatEkle(){

  const baslik = document.getElementById('hat-baslik').value.trim();

  const tarih  = document.getElementById('hat-tarih').value;

  const saat   = document.getElementById('hat-saat').value;

  const oncelik= document.getElementById('hat-oncelik').value;

  const not    = document.getElementById('hat-not').value.trim();

  if(!baslik){ toast('Hatırlatma metni boş olamaz','e'); return; }



  const r = await api('/api/hatirlatma/add',{method:'POST',

    headers:{'Content-Type':'application/json'},

    body:JSON.stringify({baslik,tarih,saat,oncelik,not})});

  const d = await r.json();

  if(d.ok){

    document.getElementById('hat-baslik').value='';

    document.getElementById('hat-saat').value='';

    document.getElementById('hat-not').value='';

    document.getElementById('hat-oncelik').value='normal';

    renderHatirlatma();

    toast('Hatırlatma eklendi','s');

  }

}



async function hatToggle(id){

  const r=await api('/api/hatirlatma/toggle',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({id})});

  if((await r.json()).ok) renderHatirlatma();

}



function hatDuzenle(id){

  _hatEditId=id;

  api('/api/hatirlatma?filtre=tumu').then(r=>r.json()).then(d=>{

    const h=d.records.find(x=>x.id===id); if(!h) return;

    document.getElementById('hem-baslik').value=h.baslik||'';

    document.getElementById('hem-tarih').value=h.tarih||'';

    document.getElementById('hem-saat').value=h.saat||'';

    document.getElementById('hem-oncelik').value=h.oncelik||'normal';

    document.getElementById('hem-not').value=h.not||'';

    document.getElementById('hat-edit-mod').classList.add('on');

  });

}



function closeHatMod(){ document.getElementById('hat-edit-mod').classList.remove('on'); }



async function hatEditKaydet(){

  const data={

    baslik:  document.getElementById('hem-baslik').value.trim(),

    tarih:   document.getElementById('hem-tarih').value,

    saat:    document.getElementById('hem-saat').value,

    oncelik: document.getElementById('hem-oncelik').value,

    not:     document.getElementById('hem-not').value.trim(),

    tamamlandi: false,

  };

  if(!data.baslik){ toast('Başlık boş olamaz','e'); return; }

  const r=await api('/api/hatirlatma/edit',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({id:_hatEditId,data})});

  if((await r.json()).ok){ closeHatMod(); renderHatirlatma(); toast('Güncellendi','s'); }

}



async function hatSil(id){

  if(!confirm('Bu hatırlatmayı silmek istediğinizden emin misiniz?')) return;

  const r=await api('/api/hatirlatma/delete',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({id})});

  if((await r.json()).ok){ renderHatirlatma(); toast('Silindi','i'); }

}



// Init: bugünün tarihini hatırlatıcı forma otomatik set et



// ── SATIŞ ANALİZİ ────────────────────────────────────────────────────────────

let _saRows    = [];   // Ham satış kalem verileri

let _saStokRows= [];   // Stok bakiye verileri

let _saAktifSekme = 'ozet';

let _saUrunPg = 1, _saMarjPg = 1, _saStokPg = 1;

const SA_PPG = 50;

let _ssFiltre = 'tumu';



async function renderSatisAnaliz(){

  // Firma ve yıl seçimlerini güncelle

  const rc = await api('/api/logo/config'); const dc = await rc.json();

  const firmalar = dc.config?.firmalar || [{no: dc.config?.firma_no||'001', ad:'Firma'}];

  logoFirmaSecimGuncelle2('sa-firma-secim', firmalar);

  logoYilSecimGuncelle2('sa-yil-secim');



  // Tarih aralığını bu yıl olarak ayarla

  const y = new Date().getFullYear();

  if(!document.getElementById('sa-bas').value)

    document.getElementById('sa-bas').value = y+'-01-01';

  if(!document.getElementById('sa-bit').value)

    document.getElementById('sa-bit').value = y+'-12-31';

}



function logoFirmaSecimGuncelle2(elId, firmalar){

  const el = document.getElementById(elId); if(!el) return;

  el.innerHTML = firmalar.map(f=>`

    <label style="display:flex;align-items:center;gap:4px;background:var(--s3);border:1px solid var(--b2);border-radius:4px;padding:3px 8px;cursor:pointer;font-size:11px">

      <input type="checkbox" checked value="${f.no}" style="margin:0">

      <span>${f.ad||f.no}</span>

    </label>`).join('');

}



function logoYilSecimGuncelle2(elId){

  const el = document.getElementById(elId); if(!el) return;

  const buYil = new Date().getFullYear();

  el.innerHTML = Array.from({length:5},(_,i)=>buYil-i).map(y=>`

    <label style="display:flex;align-items:center;gap:4px;background:var(--s3);border:1px solid var(--b2);border-radius:4px;padding:3px 8px;cursor:pointer;font-size:11px">

      <input type="checkbox" ${y>=buYil-1?'checked':''} value="${y}" style="margin:0">

      <span>${y}</span>

    </label>`).join('');

}



function saGetSeciliFirmalar(){

  return Array.from(document.querySelectorAll('#sa-firma-secim input:checked')).map(c=>c.value).join(',') || null;

}

function saGetSeciliYillar(){

  return Array.from(document.querySelectorAll('#sa-yil-secim input:checked')).map(c=>c.value).join(',') || null;

}



async function satisAnalizYukle(){

  const loading = document.getElementById('sa-loading');

  loading.style.display = '';



  const params = new URLSearchParams();

  const bas = document.getElementById('sa-bas').value;

  const bit = document.getElementById('sa-bit').value;

  const stokQ = document.getElementById('sa-stok-q').value;

  const firmalar = saGetSeciliFirmalar();

  const yillar   = saGetSeciliYillar();



  if(bas)     params.set('baslangic', bas);

  if(bit)     params.set('bitis', bit);

  if(stokQ)   params.set('stok_q', stokQ);

  if(firmalar) params.set('firmalar', firmalar);

  if(yillar)  params.set('yillar', yillar);



  try {

    const [r1, r2] = await Promise.all([

      api('/api/logo/satis_analiz?'+params),

      api('/api/logo/stok_bakiye'+(firmalar?'?firmalar='+encodeURIComponent(firmalar):''))

    ]);

    const d1 = await r1.json();

    const d2 = await r2.json();



    if(!d1.ok){ toast('Hata: '+(d1.msg||'Veri çekilemedi'),'e'); loading.style.display='none'; return; }



    _saRows     = d1.rows||[];

    _saStokRows = d2.rows||[];



    loading.style.display = 'none';

    toast(`✓ ${_saRows.length} satır analiz edildi`,'s');



    // Dönem yıl filtresini güncelle

    const yillar2 = [...new Set(_saRows.map(r=>r.yil).filter(Boolean))].sort().reverse();

    const yilSel = document.getElementById('sa-donem-yil-f');

    if(yilSel){

      yilSel.innerHTML = '<option value="">Tüm Yıllar</option>' +

        yillar2.map(y=>`<option value="${y}">${y}</option>`).join('');

    }



    saTab(_saAktifSekme || 'ozet');

  } catch(e){

    loading.style.display='none';

    toast('Bağlantı hatası: '+e.message,'e');

  }

}



function saTab(sekme){

  _saAktifSekme = sekme;

  ['ozet','urun','donem','marj','stok'].forEach(s=>{

    document.getElementById('satab-'+s+'-view').style.display = s===sekme?'':'none';

    const btn = document.getElementById('satab-'+s);

    if(btn){ btn.style.borderBottom = s===sekme?'2px solid var(--acc)':'none'; btn.style.color = s===sekme?'var(--acc)':''; }

  });

  if(sekme==='ozet')  saOzetRender();

  if(sekme==='urun')  saUrunRender();

  if(sekme==='donem') saDonemRender();

  if(sekme==='marj')  saMarjRender();

  if(sekme==='stok')  saStokRender();

}



// ── GENEL ÖZET ──────────────────────────────────────────────────────────────

function saOzetRender(){

  if(!_saRows.length){

    document.getElementById('sa-mets').innerHTML='<div style="color:var(--t3);font-size:12px;grid-column:1/-1;padding:20px">Veri yok — "Analiz Et" butonuna tıklayın</div>';

    return;

  }

  const topTutar  = _saRows.reduce((s,r)=>s+parseFloat(r.net_tutar||0),0);

  const topAdet   = _saRows.reduce((s,r)=>s+parseFloat(r.miktar||0),0);

  const urunSet   = new Set(_saRows.map(r=>r.stok_kodu).filter(Boolean));

  const faturaSay = new Set(_saRows.map(r=>r.fatura_no).filter(Boolean)).size;

  const ortFatura = faturaSay ? topTutar/faturaSay : 0;



  document.getElementById('sa-mets').innerHTML=`

    <div class="met"><div class="met-stripe" style="background:var(--g)"></div>

      <div class="mlb">Toplam Satış ₺</div>

      <div class="mvl" style="color:var(--g);font-size:18px">₺${fmtN(topTutar)}</div>

      <div class="msb">${_saRows.length} kalem</div></div>

    <div class="met"><div class="met-stripe" style="background:var(--acc)"></div>

      <div class="mlb">Toplam Satış Adedi</div>

      <div class="mvl" style="color:var(--acc);font-size:18px">${Math.round(topAdet).toLocaleString('tr-TR')}</div>

      <div class="msb">${urunSet.size} farklı ürün</div></div>

    <div class="met"><div class="met-stripe" style="background:var(--am)"></div>

      <div class="mlb">Fatura Sayısı</div>

      <div class="mvl" style="color:var(--am);font-size:18px">${faturaSay.toLocaleString('tr-TR')}</div>

      <div class="msb">Ort. ₺${fmtN(ortFatura)}/fatura</div></div>

    <div class="met"><div class="met-stripe" style="background:var(--r)"></div>

      <div class="mlb">Negatif Stok</div>

      <div class="mvl" style="color:var(--r);font-size:18px">${_saStokRows.filter(r=>parseFloat(r.bakiye||0)<0).length}</div>

      <div class="msb">ürün eksiye düşmüş</div></div>`;



  // Aylık satış SVG grafiği

  const ayMap = {};

  _saRows.forEach(r=>{

    const k = `${r.yil}-${String(r.ay||'').padStart(2,'0')}`;

    if(!k.includes('undefined') && k.length>=7){

      ayMap[k] = (ayMap[k]||0) + parseFloat(r.net_tutar||0);

    }

  });

  const aylar = Object.entries(ayMap).sort(([a],[b])=>a.localeCompare(b)).slice(-24);

  const maxA = Math.max(...aylar.map(([,v])=>v),1);

  const W=500, H=160, PL=44, PR=8, PT=10, PB=32, cW=W-PL-PR, cH=H-PT-PB;

  const bW = Math.max(4, Math.floor(cW/aylar.length)-1);

  let svgA = `<svg width="${W}" height="${H}" xmlns="http://www.w3.org/2000/svg" style="display:block">`;

  [0,.25,.5,.75,1].forEach(p=>{

    const y=PT+cH*(1-p);

    svgA+=`<line x1="${PL}" y1="${y.toFixed(1)}" x2="${W-PR}" y2="${y.toFixed(1)}" stroke="#1e2535" stroke-width="1"/>`;

    if(p>0) svgA+=`<text x="${PL-3}" y="${(y+4).toFixed(1)}" text-anchor="end" font-size="8" fill="#3d4f6b">${(maxA*p/1e6).toFixed(1)}M</text>`;

  });

  aylar.forEach(([ay,val],i)=>{

    const bH=Math.max(2,Math.round(val/maxA*cH));

    const bX=PL+i*(bW+1); const bY=PT+cH-bH;

    svgA+=`<rect x="${bX}" y="${bY}" width="${bW}" height="${bH}" fill="rgba(16,185,129,.8)" rx="1"/>`;

    if(i%3===0){

      const lbl=ay.slice(2);

      svgA+=`<text x="${bX+bW/2}" y="${H-3}" text-anchor="middle" font-size="7" fill="#4a5568">${lbl}</text>`;

    }

  });

  svgA+='</svg>';

  document.getElementById('sa-aylik-grafik').innerHTML=`<div style="font-weight:600;font-size:12px;margin-bottom:8px;color:var(--t2)">Aylık Satış ₺</div>${svgA}`;



  // Firma bazlı dağılım

  const firmaMap={};

  _saRows.forEach(r=>{ const f=r.logo_firma_adi||r.logo_firma_no||'?'; firmaMap[f]=(firmaMap[f]||0)+parseFloat(r.net_tutar||0); });

  const firmaList=Object.entries(firmaMap).sort((a,b)=>b[1]-a[1]);

  const maxF=Math.max(...firmaList.map(([,v])=>v),1);

  const clrs=['#3b82f6','#10b981','#f59e0b','#ef4444','#8b5cf6'];

  document.getElementById('sa-firma-grafik').innerHTML=`

    <div style="font-weight:600;font-size:12px;margin-bottom:8px;color:var(--t2)">Firma Bazlı Satış</div>

    ${firmaList.map(([f,v],i)=>`

      <div style="display:flex;align-items:center;gap:8px;margin-bottom:5px">

        <div style="width:90px;font-size:11px;color:var(--t2);text-align:right;flex-shrink:0">${f}</div>

        <div style="flex:1;background:var(--s3);border-radius:2px;height:18px">

          <div style="width:${Math.round(v/maxF*100)}%;background:${clrs[i%clrs.length]};height:100%;border-radius:2px"></div>

        </div>

        <div style="width:80px;font-family:var(--mono);font-size:11px">₺${fmtN(v)}</div>

      </div>`).join('')}`;

}



// ── ÜRÜN BAZLI ──────────────────────────────────────────────────────────────

function saUrunRender(){

  const q      = (document.getElementById('sa-urun-q')?.value||'').toLowerCase();

  const sirala = document.getElementById('sa-urun-sirala')?.value||'tutar';



  const urunMap = {};

  _saRows.forEach(r=>{

    const k = r.stok_kodu||'?';

    if(!urunMap[k]) urunMap[k]={ stok_kodu:k, stok_adi:r.stok_adi||'—', grup:r.stok_grubu||'', firma:r.logo_firma_adi||r.logo_firma_no||'?', adet:0, tutar:0, fatura_sayisi:new Set() };

    urunMap[k].adet   += parseFloat(r.miktar||0);

    urunMap[k].tutar  += parseFloat(r.net_tutar||0);

    urunMap[k].fatura_sayisi.add(r.fatura_no);

  });



  let urunler = Object.values(urunMap).map(u=>({...u, fatura_sayisi: u.fatura_sayisi.size}));

  const topTutar = urunler.reduce((s,u)=>s+u.tutar,0);



  if(q) urunler = urunler.filter(u=>

    (u.stok_kodu||'').toLowerCase().includes(q)||(u.stok_adi||'').toLowerCase().includes(q));



  urunler.sort((a,b)=> sirala==='adet'?b.adet-a.adet : sirala==='fatura'?b.fatura_sayisi-a.fatura_sayisi : b.tutar-a.tutar);



  document.getElementById('sa-urun-count').textContent = `${urunler.length} ürün`;



  const start=(_saUrunPg-1)*SA_PPG, page=urunler.slice(start,start+SA_PPG);

  document.getElementById('sa-urun-tbody').innerHTML = page.map((u,i)=>{

    const pay = topTutar>0 ? (u.tutar/topTutar*100).toFixed(1) : 0;

    const ortFiyat = u.adet>0 ? u.tutar/u.adet : 0;

    return `<tr>

      <td style="text-align:center;font-size:11px;color:var(--t3)">${start+i+1}</td>

      <td style="font-family:var(--mono);font-size:11px;color:var(--acc)">${u.stok_kodu}</td>

      <td style="font-size:12px">${u.stok_adi}</td>

      <td style="font-size:10px;color:var(--t3)">${u.grup||'—'}</td>

      <td style="font-size:10px;color:var(--t3)">${u.firma}</td>

      <td class="num" style="font-family:var(--mono)">${Math.round(u.adet).toLocaleString('tr-TR')}</td>

      <td class="num" style="font-family:var(--mono)">${u.fatura_sayisi}</td>

      <td class="num" style="font-family:var(--mono);color:var(--g);font-weight:600">₺${fmtN(u.tutar)}</td>

      <td class="num" style="font-family:var(--mono);font-size:11px">₺${fmtN(ortFiyat)}</td>

      <td class="num">

        <div style="display:flex;align-items:center;gap:4px">

          <div style="flex:1;background:var(--s3);border-radius:2px;height:8px">

            <div style="width:${Math.min(100,parseFloat(pay)*5)}%;background:var(--acc);height:100%;border-radius:2px"></div>

          </div>

          <span style="font-size:10px;font-family:var(--mono);width:32px;text-align:right">%${pay}</span>

        </div>

      </td>

    </tr>`;

  }).join('');



  const tot=Math.ceil(urunler.length/SA_PPG)||1;

  let ph=`<span class="pgi">${urunler.length} ürün</span><div class="pgsp"></div>`;

  if(_saUrunPg>1) ph+=`<button class="pbn" onclick="_saUrunPg--;saUrunRender()">‹</button>`;

  for(let i=Math.max(1,_saUrunPg-2);i<=Math.min(tot,_saUrunPg+2);i++)

    ph+=`<button class="pbn${i===_saUrunPg?' on':''}" onclick="_saUrunPg=${i};saUrunRender()">${i}</button>`;

  if(_saUrunPg<tot) ph+=`<button class="pbn" onclick="_saUrunPg++;saUrunRender()">›</button>`;

  document.getElementById('sa-urun-pgn').innerHTML=ph;

}



// ── DÖNEM ANALİZİ ────────────────────────────────────────────────────────────

function saDonemRender(){

  const tip  = document.getElementById('sa-donem-tip')?.value||'aylik';

  const yilF = document.getElementById('sa-donem-yil-f')?.value||'';



  const donemMap = {};

  _saRows.forEach(r=>{

    if(yilF && r.yil != yilF) return;

    let k;

    if(tip==='yillik')    k = String(r.yil||'?');

    else if(tip==='aylik') k = `${r.yil}-${String(r.ay||'').padStart(2,'0')}`;

    else { const d=new Date(r.tarih?.split('.').reverse().join('-')||r.tarih); k=`${r.yil}-H${Math.ceil((d.getMonth()+1)/3)}`; }

    if(!donemMap[k]) donemMap[k]={donem:k,adet:0,tutar:0,faturalar:new Set(),urunler:new Set()};

    donemMap[k].adet+=parseFloat(r.miktar||0);

    donemMap[k].tutar+=parseFloat(r.net_tutar||0);

    donemMap[k].faturalar.add(r.fatura_no);

    donemMap[k].urunler.add(r.stok_kodu);

  });



  const donemler = Object.values(donemMap)

    .map(d=>({...d,fatura_sayisi:d.faturalar.size,urun_sayisi:d.urunler.size}))

    .sort((a,b)=>a.donem.localeCompare(b.donem));



  // SVG grafik

  const maxD = Math.max(...donemler.map(d=>d.tutar),1);

  const W=600, H=160, PL=44, PR=8, PT=10, PB=32, cW=W-PL-PR, cH=H-PT-PB;

  const bW = Math.max(6,Math.floor(cW/Math.max(donemler.length,1))-1);

  let svgD = `<svg width="${W}" height="${H}" xmlns="http://www.w3.org/2000/svg" style="display:block;overflow:visible">`;

  [0,.25,.5,.75,1].forEach(p=>{

    const y=PT+cH*(1-p);

    svgD+=`<line x1="${PL}" y1="${y.toFixed(1)}" x2="${W-PR}" y2="${y.toFixed(1)}" stroke="#1e2535" stroke-width="1"/>`;

    if(p>0) svgD+=`<text x="${PL-3}" y="${(y+4).toFixed(1)}" text-anchor="end" font-size="8" fill="#3d4f6b">${(maxD*p>=1e6?(maxD*p/1e6).toFixed(1)+'M':Math.round(maxD*p/1e3)+'K')}</text>`;

  });

  donemler.forEach((d,i)=>{

    const bH=Math.max(2,Math.round(d.tutar/maxD*cH));

    const bX=PL+i*(bW+1); const bY=PT+cH-bH;

    svgD+=`<rect x="${bX}" y="${bY}" width="${bW}" height="${bH}" fill="rgba(79,156,249,.8)" rx="1" title="₺${Math.round(d.tutar).toLocaleString()}"/>`;

    if(tip!=='aylik'||donemler.length<=18)

      svgD+=`<text x="${bX+bW/2}" y="${H-3}" text-anchor="middle" font-size="7" fill="#4a5568">${d.donem.slice(-5)}</text>`;

  });

  svgD+='</svg>';

  document.getElementById('sa-donem-grafik').innerHTML=svgD;



  // Tablo

  const rows = [...donemler].reverse();

  document.getElementById('sa-donem-tbody').innerHTML = rows.map((d,i)=>{

    const prev = rows[i+1];

    const degisim = prev && prev.tutar>0 ? ((d.tutar-prev.tutar)/prev.tutar*100).toFixed(1) : null;

    const degisimHTML = degisim!==null

      ? `<div style="display:flex;align-items:center;gap:4px">

          <span style="color:${parseFloat(degisim)>=0?'var(--g)':'var(--r)'};font-family:var(--mono);font-size:11px">${parseFloat(degisim)>=0?'▲':'▼'} %${Math.abs(parseFloat(degisim))}</span>

         </div>`

      : '—';

    return `<tr>

      <td style="font-family:var(--mono);font-weight:600">${d.donem}</td>

      <td class="num" style="font-family:var(--mono)">${Math.round(d.adet).toLocaleString('tr-TR')}</td>

      <td class="num" style="font-family:var(--mono);color:var(--g);font-weight:600">₺${fmtN(d.tutar)}</td>

      <td class="num" style="font-family:var(--mono)">₺${fmtN(d.fatura_sayisi>0?d.tutar/d.fatura_sayisi:0)}</td>

      <td class="num" style="font-family:var(--mono)">${d.fatura_sayisi}</td>

      <td class="num" style="font-family:var(--mono)">${d.urun_sayisi}</td>

      <td>${degisimHTML}</td>

    </tr>`;

  }).join('');

}



// ── ALIŞ-SATIŞ MARJI ────────────────────────────────────────────────────────

function saMarjRender(){

  const q     = (document.getElementById('sa-marj-q')?.value||'').toLowerCase();

  const sirala= document.getElementById('sa-marj-sirala')?.value||'marj_desc';



  // Ürün bazlı tutar ve adet

  const urunTutar = {};

  _saRows.forEach(r=>{

    const k=r.stok_kodu||'?';

    urunTutar[k]=(urunTutar[k]||0)+parseFloat(r.net_tutar||0);

  });



  let marjList = _saStokRows

    .filter(r=>r.son_alis_fiyati||r.son_satis_fiyati)

    .map(r=>{

      const alis  = parseFloat(r.son_alis_fiyati||0);

      const satis = parseFloat(r.son_satis_fiyati||0);

      const fark  = satis - alis;

      const marj  = satis>0 ? (fark/satis*100) : 0;

      const totSatis = urunTutar[r.stok_kodu]||0;

      const karEst= totSatis>0 && satis>0 ? totSatis * (marj/100) : 0;

      return {...r, alis, satis, fark, marj: parseFloat(marj.toFixed(2)), kar_est: karEst};

    });



  if(q) marjList=marjList.filter(r=>

    (r.stok_kodu||'').toLowerCase().includes(q)||(r.stok_adi||'').toLowerCase().includes(q));

  if(sirala==='negatif') marjList=marjList.filter(r=>r.marj<0);



  marjList.sort((a,b)=>{

    if(sirala==='marj_desc') return b.marj-a.marj;

    if(sirala==='marj_asc')  return a.marj-b.marj;

    if(sirala==='tutar')     return b.kar_est-a.kar_est;

    return a.marj-b.marj;

  });



  const start=(_saMarjPg-1)*SA_PPG, page=marjList.slice(start,start+SA_PPG);

  document.getElementById('sa-marj-tbody').innerHTML=page.map(r=>{

    const marjRenk = r.marj>30?'var(--g)':r.marj>10?'var(--am)':r.marj>0?'var(--r)':'#ef4444';

    const marjBg   = r.marj<0?'rgba(239,68,68,.1)':'';

    return `<tr style="background:${marjBg}">

      <td style="font-family:var(--mono);font-size:11px;color:var(--acc)">${r.stok_kodu}</td>

      <td style="font-size:12px">${r.stok_adi||'—'}</td>

      <td style="font-size:10px;color:var(--t3)">${r.logo_firma_adi||'—'}</td>

      <td class="num" style="font-family:var(--mono);font-size:12px">${r.alis?'₺'+fmtN(r.alis):'—'}</td>

      <td class="num" style="font-family:var(--mono);font-size:12px">${r.satis?'₺'+fmtN(r.satis):'—'}</td>

      <td class="num" style="font-family:var(--mono);font-size:12px;color:${r.fark>=0?'var(--g)':'var(--r)'};font-weight:600">${r.fark?'₺'+fmtN(r.fark):'—'}</td>

      <td class="num" style="font-family:var(--mono);color:${marjRenk};font-weight:700">%${r.marj.toFixed(1)}</td>

      <td class="num" style="font-family:var(--mono);font-size:11px;color:var(--g)">${r.kar_est?'₺'+fmtN(r.kar_est):'—'}</td>

    </tr>`;

  }).join('');



  const tot=Math.ceil(marjList.length/SA_PPG)||1;

  let ph=`<span class="pgi">${marjList.length} ürün</span><div class="pgsp"></div>`;

  if(_saMarjPg>1) ph+=`<button class="pbn" onclick="_saMarjPg--;saMarjRender()">‹</button>`;

  for(let i=Math.max(1,_saMarjPg-2);i<=Math.min(tot,_saMarjPg+2);i++)

    ph+=`<button class="pbn${i===_saMarjPg?' on':''}" onclick="_saMarjPg=${i};saMarjRender()">${i}</button>`;

  if(_saMarjPg<tot) ph+=`<button class="pbn" onclick="_saMarjPg++;saMarjRender()">›</button>`;

  document.getElementById('sa-marj-pgn').innerHTML=ph;

}



// ── STOK DURUMU ──────────────────────────────────────────────────────────────

function saStokFiltre(f){

  _ssFiltre=f;

  ['tumu','negatif','sifir','kritik'].forEach(t=>{

    const b=document.getElementById('ss-'+t);

    if(b){ b.style.background=t===f?'var(--acc)':''; b.style.color=t===f?'#fff':''; }

  });

  saStokRender();

}



function saStokRender(){

  let list = _saStokRows;

  if(_ssFiltre==='negatif') list=list.filter(r=>parseFloat(r.bakiye||0)<0);

  if(_ssFiltre==='sifir')   list=list.filter(r=>parseFloat(r.bakiye||0)===0);

  if(_ssFiltre==='kritik')  list=list.filter(r=>{const b=parseFloat(r.bakiye||0);return b>0&&b<10;});



  list.sort((a,b)=>parseFloat(a.bakiye||0)-parseFloat(b.bakiye||0));

  document.getElementById('ss-count').textContent=`${list.length} ürün`;



  const start=(_saStokPg-1)*SA_PPG, page=list.slice(start,start+SA_PPG);

  document.getElementById('sa-stok-tbody').innerHTML=page.map(r=>{

    const bak = parseFloat(r.bakiye||0);

    const durumCls = bak<0?'bri':bak===0?'bami':bak<10?'bami':'bgi';

    const durumTxt = bak<0?'⚠ Negatif':bak===0?'Tükenmiş':bak<10?'Kritik':'Normal';

    return `<tr style="${bak<0?'background:rgba(239,68,68,.06)':''}">

      <td style="font-family:var(--mono);font-size:11px;color:var(--acc)">${r.stok_kodu}</td>

      <td style="font-size:12px">${r.stok_adi||'—'}</td>

      <td style="font-size:10px;color:var(--t3)">${r.stok_grubu||'—'}</td>

      <td style="font-size:10px;color:var(--t3)">${r.logo_firma_adi||'—'}</td>

      <td class="num" style="font-family:var(--mono);font-weight:700;color:${bak<0?'var(--r)':bak<10?'var(--am)':'var(--g)'}">${bak.toLocaleString('tr-TR',{maximumFractionDigits:2})}</td>

      <td class="num" style="font-family:var(--mono);font-size:11px">${r.son_alis_fiyati?'₺'+fmtN(parseFloat(r.son_alis_fiyati)):'—'}</td>

      <td class="num" style="font-family:var(--mono);font-size:11px">${r.son_satis_fiyati?'₺'+fmtN(parseFloat(r.son_satis_fiyati)):'—'}</td>

      <td><span class="bdg ${durumCls}" style="font-size:10px">${durumTxt}</span></td>

    </tr>`;

  }).join('');



  const tot=Math.ceil(list.length/SA_PPG)||1;

  let ph=`<span class="pgi">${list.length} ürün</span><div class="pgsp"></div>`;

  if(_saStokPg>1) ph+=`<button class="pbn" onclick="_saStokPg--;saStokRender()">‹</button>`;

  for(let i=Math.max(1,_saStokPg-2);i<=Math.min(tot,_saStokPg+2);i++)

    ph+=`<button class="pbn${i===_saStokPg?' on':''}" onclick="_saStokPg=${i};saStokRender()">${i}</button>`;

  if(_saStokPg<tot) ph+=`<button class="pbn" onclick="_saStokPg++;saStokRender()">›</button>`;

  document.getElementById('sa-stok-pgn').innerHTML=ph;

}



async function satisExport(){

  if(!_saRows.length){ toast('Önce analiz çalıştırın','e'); return; }

  const nl='\r\n';

  let csv='Stok Kodu,Stok Adı,Grup,Firma,Tarih,Yıl,Ay,Miktar,Birim,Satış Fiyatı,Net Tutar,Fatura No,Cari'+nl;

  _saRows.forEach(r=>{

    csv+=`"${r.stok_kodu||''}","${r.stok_adi||''}","${r.stok_grubu||''}","${r.logo_firma_adi||''}",`;

    csv+=`"${r.tarih||''}","${r.yil||''}","${r.ay||''}","${r.miktar||''}","${r.birim||''}",`;

    csv+=`"${r.satis_fiyati||''}","${r.net_tutar||''}","${r.fatura_no||''}","${r.cari_adi||''}"${nl}`;

  });

  const a=document.createElement('a');

  a.href='data:text/csv;charset=utf-8,\uFEFF'+encodeURIComponent(csv);

  a.download='satis_analiz.csv'; a.click();

  toast('Satış analizi CSV indirildi','s');

}



document.addEventListener('keydown',e=>{if(e.key==='Escape'){closeMod();closeGelirMod();closeCariMod();closeMailMod();closeNotEditMod();closeOtelMod();closeWaSablonMod();closeWaKisiMod();closeHatMod();}});





// ── DOKÜMAN YÖNETİM SİSTEMİ ─────────────────────────────────────────────────

let _docsVerisi = [];

let _docsEditId = null;



async function renderDocs(){

  try { await fetch('/api/docs/temizle_duplicate',{method:'POST',headers:{'Content-Type':'application/json'},body:'{}'}); } catch(e){}

  await docsListeYukle();

}



async function docsListeYukle(){

  const q   = (document.getElementById('docs-ara')?.value||'').trim();

  const kat = document.getElementById('docs-kat-filtre')?.value||'';

  try {

    const r = await fetch('/api/docs/list?q='+encodeURIComponent(q)+'&kategori='+encodeURIComponent(kat));

    if(!r.ok) throw new Error('HTTP '+r.status);

    const d = await r.json();

    _docsVerisi = d.docs||[];

    docsKategorileriGuncelle(d.kategoriler||[]);

    docsStatsGuncelle(d.toplam||0, _docsVerisi);

    docsGridRender(_docsVerisi);

  } catch(e){

    const g = document.getElementById('docs-grid');

    if(g) g.innerHTML = '<div style="color:var(--r);padding:20px">Yükleme hatası: '+e.message+'</div>';

  }

}



function docsFiltrele(){ docsListeYukle(); }



function docsKategorileriGuncelle(kategoriler){

  const sel = document.getElementById('docs-kat-filtre');

  if(!sel) return;

  const mevcut = sel.value;

  sel.innerHTML = '<option value="">Tüm Kategoriler</option>';

  kategoriler.forEach(k=>{ const o=document.createElement('option'); o.value=k; o.textContent=k; sel.appendChild(o); });

  sel.value = mevcut;

  const editSel = document.getElementById('docs-edit-kat');

  if(editSel){

    const sabit = ['Genel','Fatura','Sözleşme','Ekstre','Rapor','Diğer'];

    editSel.innerHTML = [...new Set([...sabit,...kategoriler])].map(k=>'<option>'+k+'</option>').join('');

  }

}



function docsStatsGuncelle(toplam, docs){

  const el = document.getElementById('docs-stats');

  if(!el) return;

  const toplamBoyut = docs.reduce((s,d)=>s+(d.boyut||0),0);

  const fmt = b=>b>1048576?(b/1048576).toFixed(1)+' MB':b>1024?(b/1024).toFixed(0)+' KB':b+' B';

  const kategoriler = [...new Set(docs.map(d=>d.kategori||'Genel'))].length;

  el.innerHTML = [

    ['📄','Toplam Belge',toplam+' adet','var(--acc)'],

    ['💾','Toplam Boyut',fmt(toplamBoyut),'var(--g)'],

    ['📂','Kategori',kategoriler+' adet','var(--am)'],

  ].map(([ic,lb,vl,cl])=>'<div style="background:var(--s2);border:1px solid var(--b1);border-radius:var(--rad);padding:8px 14px;display:flex;align-items:center;gap:10px"><span style="font-size:20px">'+ic+'</span><div><div style="font-size:10px;color:var(--t3)">'+lb+'</div><div style="font-size:14px;font-weight:700;color:'+cl+'">'+vl+'</div></div></div>').join('');

}



function docsGridRender(docs){

  const grid = document.getElementById('docs-grid');

  const bos  = document.getElementById('docs-bos');

  if(!grid) return;

  if(!docs||!docs.length){ grid.innerHTML=''; if(bos) bos.style.display='block'; return; }

  if(bos) bos.style.display='none';

  const extIcon = e=>{ e=(e||'').toLowerCase(); if(e==='pdf') return '📕'; if(['jpg','jpeg','png','gif','webp','bmp','svg'].includes(e)) return '🖼'; if(['doc','docx'].includes(e)) return '📘'; if(['xls','xlsx','csv'].includes(e)) return '📗'; if(['ppt','pptx'].includes(e)) return '📙'; if(['zip','rar','7z'].includes(e)) return '📦'; if(['mp4','avi','mov','webm'].includes(e)) return '🎬'; if(['mp3','wav','m4a','ogg','aac','flac'].includes(e)) return '🎵'; if(['txt','md','log'].includes(e)) return '📄'; return '📎'; };

  const fmtB = b=>{ b=b||0; if(b>1048576) return (b/1048576).toFixed(1)+' MB'; if(b>1024) return (b/1024).toFixed(0)+' KB'; return b+' B'; };

  window._docsMap = {};

  docs.forEach(d=>{ window._docsMap[d.id]=d; });

  grid.innerHTML = '';

  docs.forEach(d=>{

    const div = document.createElement('div');

    div.style.cssText = 'background:var(--s2);border:1px solid var(--b1);border-radius:var(--radl);overflow:hidden;display:flex;flex-direction:column';

    const ad = (d.ad||d.orijinal_ad||'Belge').replace(/</g,'&lt;');

    div.innerHTML =

      '<div style="padding:16px;flex:1;text-align:center;cursor:pointer;border-bottom:1px solid var(--b1)">'

      +'<div style="font-size:38px;margin-bottom:8px">'+extIcon(d.ext)+'</div>'

      +'<div style="font-size:12px;font-weight:600;color:var(--t1);word-break:break-word;line-height:1.4">'+ad+'</div>'

      +'<div style="margin-top:6px"><span style="background:var(--s4);color:var(--acc);font-size:10px;padding:2px 8px;border-radius:10px">'+(d.kategori||'Genel')+'</span></div>'

      +(d.not?'<div style="font-size:10px;color:var(--t3);margin-top:4px;font-style:italic">'+d.not.slice(0,40)+'</div>':'')

      +'<div style="font-size:10px;color:var(--t3);margin-top:6px">'+fmtB(d.boyut)+' · '+(d.tarih||'').slice(0,10)+'</div>'

      +'</div>'

      +'<div style="display:flex;gap:3px;padding:8px;background:var(--s3)">'

      +'<button style="flex:1;padding:6px 2px;background:var(--acc-d);border:none;border-radius:var(--rad);color:#fff;cursor:pointer;font-size:11px;font-weight:600">👁 Görüntüle</button>'

      +'<button style="flex:1;padding:6px 2px;background:var(--s4);border:1px solid var(--b2);border-radius:var(--rad);color:var(--t2);cursor:pointer;font-size:11px">✏ Düzenle</button>'

      +'<button style="padding:6px 8px;background:var(--s4);border:1px solid var(--b2);border-radius:var(--rad);color:var(--acc);cursor:pointer;font-size:11px">⬇</button>'

      +'<button style="padding:6px 8px;background:rgba(248,113,113,.15);border:1px solid var(--r);border-radius:var(--rad);color:var(--r);cursor:pointer;font-size:11px">🗑</button>'

      +'</div>';

    const btns = div.querySelectorAll('button');

    btns[0].onclick = ()=>docsOnizle(d.id);

    btns[1].onclick = ()=>docsEdit(d.id);

    btns[2].onclick = ()=>docsIndir(d.id);

    btns[3].onclick = ()=>docsSil(d.id, ad);

    div.querySelector('div').onclick = ()=>docsOnizle(d.id);

    grid.appendChild(div);

  });

}



// ── YÜKLEME ──────────────────────────────────────────────────────────────────

async function docsYukle(files){

  if(!files||!files.length) return;

  const kategori = prompt('Kategori (Genel/Fatura/Sözleşme/Ekstre/Rapor/Diğer):', 'Genel')||'Genel';

  const progText = document.getElementById('docs-progress-text');

  const progBar  = document.getElementById('docs-progress-bar');

  const logEl    = document.getElementById('docs-yukle-log');

  if(logEl) logEl.innerHTML = '';

  let yuklenen=0, hata=0;

  for(let i=0;i<files.length;i++){

    const file = files[i];

    if(progText) progText.textContent = '⏳ ('+( i+1)+'/'+files.length+'): '+file.name;

    const fd = new FormData();

    fd.append('dosya', file, file.name);

    fd.append('kategori', kategori);

    fd.append('not', '');

    try {

      const tokenCookie = document.cookie.split(';').map(c=>c.trim()).find(c=>c.startsWith('na_token='));

      const tokenVal = tokenCookie ? tokenCookie.split('=').slice(1).join('=') : '';

      if(tokenVal) fd.append('_token', tokenVal);

      const r = await fetch('/api/docs/yukle',{method:'POST',body:fd,credentials:'include'});

      const d = await r.json();

      if(d.ok){ yuklenen++; if(logEl) logEl.innerHTML+='<div style="color:var(--g)">✓ '+file.name+'</div>'; }

      else { hata++; if(logEl) logEl.innerHTML+='<div style="color:var(--r)">✗ '+file.name+': '+d.msg+'</div>'; }

    } catch(e){ hata++; if(logEl) logEl.innerHTML+='<div style="color:var(--r)">✗ '+file.name+'</div>'; }

    if(progBar) progBar.style.width=((i+1)/files.length*100)+'%';

  }

  if(progText) progText.textContent = '✓ '+yuklenen+' yüklendi'+(hata?' · '+hata+' hata':'');

  if(progBar) progBar.style.background = hata?'var(--am)':'var(--g)';

  const inp = document.getElementById('docs-file-input');

  if(inp) inp.value='';

  if(yuklenen>0){ toast(yuklenen+' belge yüklendi ✓','s'); await docsListeYukle(); }

}



// ── ÖNİZLEME ─────────────────────────────────────────────────────────────────

function docsOnizle(docId){

  const doc = (window._docsMap||{})[docId]||_docsVerisi.find(d=>d.id===docId);

  const mod = document.getElementById('docs-onizle-mod');

  const baslik = document.getElementById('docs-mod-baslik');

  const icerik = document.getElementById('docs-mod-icerik');

  if(!mod||!icerik) return;

  mod.style.cssText = 'display:flex;flex-direction:column;position:fixed;inset:0;background:rgba(0,0,0,.9);z-index:9999';

  if(baslik) baslik.textContent = doc?(doc.ad||doc.orijinal_ad||'Belge'):'Belge';

  const indirBtn = document.getElementById('docs-mod-indir');

  if(indirBtn) indirBtn.onclick = ()=>docsIndir(docId);

  const waBtn = document.getElementById('docs-mod-wa');

  if(waBtn) waBtn.onclick = ()=>docsBelgeWAGonder(docId, doc);

  const mailBtn = document.getElementById('docs-mod-mail');

  if(mailBtn) mailBtn.onclick = ()=>docsBelgeMailGonder(docId, doc);

  if(!doc){ icerik.innerHTML='<div style="color:var(--t3);text-align:center;padding:40px">⏳ Yükleniyor...</div>'; fetch('/api/docs/list').then(r=>r.json()).then(d=>{ const f=(d.docs||[]).find(x=>x.id===docId); if(f) docsOnizleRender(f,docId,icerik); }); return; }

  docsOnizleRender(doc,docId,icerik);

}



function docsOnizleRender(doc, docId, icerik){

  const ext = (doc.ext||'').toLowerCase();

  const url = '/api/docs/onizle/'+docId;

  const ad  = (doc.ad||doc.orijinal_ad||'Belge').replace(/</g,'&lt;');



  if(['jpg','jpeg','png','gif','webp','bmp','svg'].includes(ext)){

    const img = document.createElement('img');

    img.src=url; img.style.cssText='max-width:100%;max-height:88vh;border-radius:8px;display:block;margin:auto';

    img.onerror=()=>{ icerik.innerHTML='<div style="color:var(--r);padding:40px;text-align:center">Resim yüklenemedi</div>'; };

    icerik.innerHTML=''; icerik.appendChild(img);



  } else if(ext==='pdf'){

    icerik.innerHTML='<div style="width:100%;height:88vh"><embed src="'+url+'" type="application/pdf" width="100%" height="100%" style="border:none"><p style="text-align:center;margin-top:8px"><a href="'+url+'" target="_blank" style="color:#4f9cf9;font-size:12px">↗ Yeni sekmede aç</a> &nbsp; <button class="btn bgh bsm" onclick="docsIndir(&quot;"+docId+"&quot;)">⬇ İndir</button></p></div>';



  } else if(['doc','docx'].includes(ext)){

    icerik.innerHTML='<div style="text-align:center;padding:60px"><div style="font-size:64px;margin-bottom:16px">📘</div><div style="font-size:15px;color:var(--t1);margin-bottom:8px">'+ad+'</div><div style="font-size:12px;color:var(--t3);margin-bottom:20px">Word dosyaları tarayıcıda önizlenemez — indirerek açın</div><button class="btn bp" onclick="docsIndir(\"'+docId+'\")">⬇ Word\'de Aç</button></div>';



  } else if(['txt','md','csv','log','json','xml'].includes(ext)){

    icerik.innerHTML='<div style="color:var(--t3);padding:20px;text-align:center">⏳</div>';

    fetch(url).then(r=>r.text()).then(t=>{ const pre=document.createElement('pre'); pre.style.cssText='font-family:var(--mono);font-size:13px;color:var(--t1);white-space:pre-wrap;padding:20px;background:var(--s2);border-radius:8px;width:100%;max-height:88vh;overflow:auto;box-sizing:border-box'; pre.textContent=t; icerik.innerHTML=''; icerik.appendChild(pre); }).catch(e=>{ icerik.innerHTML='<div style="color:var(--r);padding:20px">Okunamadı: '+e.message+'</div>'; });



  } else if(['mp4','webm','mov'].includes(ext)){

    icerik.innerHTML='<div style="text-align:center;padding:20px"><video controls style="max-width:100%;max-height:80vh;border-radius:8px;background:#000"><source src="'+url+'"><source src="'+url+'" type="video/'+ext+'"></video></div>';



  } else if(['mp3','wav','m4a','ogg','aac','flac','wma','opus'].includes(ext)){

    const wrap=document.createElement('div');

    wrap.style.cssText='text-align:center;padding:60px';

    const icon=document.createElement('div'); icon.style.cssText='font-size:64px;margin-bottom:16px'; icon.textContent='🎵'; wrap.appendChild(icon);

    const adEl=document.createElement('div'); adEl.style.cssText='font-size:15px;color:var(--t1);margin-bottom:24px;font-weight:600'; adEl.textContent=doc.ad||doc.orijinal_ad||''; wrap.appendChild(adEl);

    const audio=document.createElement('audio'); audio.controls=true; audio.preload='metadata'; audio.style.cssText='width:100%;max-width:500px;display:block;margin:0 auto';

    const src=document.createElement('source'); src.src=url; audio.appendChild(src);

    wrap.appendChild(audio);

    const ib=document.createElement('div'); ib.style.cssText='margin-top:16px;font-size:12px;color:var(--t3)'; ib.textContent='Calmazsa: ';

    const ib2=document.createElement('button'); ib2.className='btn bgh bsm'; ib2.textContent='Indir'; ib2.onclick=()=>docsIndir(docId); ib.appendChild(ib2); wrap.appendChild(ib);

    icerik.innerHTML=''; icerik.appendChild(wrap);

    setTimeout(()=>audio.play().catch(()=>{}),400);



  } else {

    icerik.innerHTML='<div style="text-align:center;padding:60px"><div style="font-size:64px;margin-bottom:16px">📎</div><div style="font-size:15px;color:var(--t1);margin-bottom:8px">'+ad+'</div><div style="font-size:12px;color:var(--t3);margin-bottom:24px">.'+(ext||'?')+' dosyası önizlenemez</div><button class="btn bp" onclick="docsIndir(&quot;"+docId+"&quot;)">⬇ Dosyayı İndir</button></div>';

  }

}



function docsOnizleKapat(){

  const mod=document.getElementById('docs-onizle-mod'); if(mod) mod.style.display='none';

  const ic=document.getElementById('docs-mod-icerik'); if(ic) ic.innerHTML='';

}



// ── BELGE GÖNDER ─────────────────────────────────────────────────────────────

function docsBelgeWAGonder(docId, doc){

  const ad = doc ? (doc.ad||doc.orijinal_ad||'Belge') : 'Belge';

  const url = window.location.origin + '/api/docs/indir/' + docId;

  const mesaj = ad + ' belgesini indirmek icin: ' + url;

  // Kişi listesinden seç veya numara gir

  const tel = prompt('WhatsApp numarası (+90 ile başlayın):', '+90');

  if(!tel || tel === '+90') return;

  const temiz = tel.replace(/[^0-9]/g,'');

  const waUrl = 'https://wa.me/' + temiz + '?text=' + encodeURIComponent(mesaj);

  window.open(waUrl, '_blank');

  toast('WhatsApp açıldı', 's');

}



function docsBelgeMailGonder(docId, doc){

  const ad = doc ? (doc.ad||doc.orijinal_ad||'Belge') : 'Belge';

  const url = window.location.origin + '/api/docs/indir/' + docId;

  const konu = encodeURIComponent('Belge: ' + ad);

  const govde = encodeURIComponent(

    'Merhaba,\n\n' + ad + ' belgesini paylaşıyorum.\n\n' +

    'İndirme linki: ' + url + '\n\n' +

    'Saygılarımla'

  );

  window.location.href = 'mailto:?subject=' + konu + '&body=' + govde;

  toast('Mail uygulaması açıldı', 's');

}



// ── İNDİR ────────────────────────────────────────────────────────────────────

function docsIndir(docId){

  const a = document.createElement('a');

  a.style.display = 'none';

  a.href = '/api/docs/indir/'+docId;

  document.body.appendChild(a);

  a.click();

  setTimeout(()=>{ try{document.body.removeChild(a);}catch(e){} }, 2000);

}



// ── DÜZENLE ───────────────────────────────────────────────────────────────────

function docsEdit(docId){

  const doc=(window._docsMap||{})[docId]||_docsVerisi.find(d=>d.id===docId);

  if(!doc){ toast('Belge bulunamadı','e'); return; }

  _docsEditId=docId;

  const ad=document.getElementById('docs-edit-ad'); if(ad) ad.value=doc.ad||doc.orijinal_ad||'';

  const nt=document.getElementById('docs-edit-not'); if(nt) nt.value=doc.not||'';

  const kat=document.getElementById('docs-edit-kat'); if(kat) kat.value=doc.kategori||'Genel';

  const katyeni=document.getElementById('docs-edit-kat-yeni'); if(katyeni) katyeni.value='';

  document.getElementById('docs-edit-mod').style.display='flex';

}

function docsEditKapat(){ document.getElementById('docs-edit-mod').style.display='none'; _docsEditId=null; }

async function docsEditKaydet(){

  if(!_docsEditId) return;

  const ad=(document.getElementById('docs-edit-ad').value||'').trim();

  const nt=(document.getElementById('docs-edit-not').value||'').trim();

  const katYeni=(document.getElementById('docs-edit-kat-yeni').value||'').trim();

  const kat=katYeni||document.getElementById('docs-edit-kat').value;

  try {

    const r=await fetch('/api/docs/not_guncelle',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({id:_docsEditId,ad,not:nt,kategori:kat})});

    const d=await r.json();

    if(d.ok){ toast('Kaydedildi ✓','s'); docsEditKapat(); docsListeYukle(); }

    else toast('Hata: '+d.msg,'e');

  } catch(e){ toast('Hata: '+e.message,'e'); }

}



// ── SİL ───────────────────────────────────────────────────────────────────────

async function docsSil(docId, ad){

  if(!confirm((ad||'Bu belge')+' silinsin mi?')) return;

  try {

    const r=await fetch('/api/docs/sil',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({id:docId})});

    const d=await r.json();

    if(d.ok){ toast('Belge silindi','i'); docsListeYukle(); }

    else toast('Silinemedi: '+d.msg,'e');

  } catch(e){ toast('Hata: '+e.message,'e'); }

}



// Eski isim uyumluluğu

function docsOnizleById(id){ docsOnizle(id); }

function docsEditById(id)  { docsEdit(id);   }

function docsIndirById(id) { docsIndir(id);  }

function docsSilById(id)   { docsSil(id,''); }



// INIT — XMLHttpRequest ile senkron yükleme (async sorunlarını bypass eder)

async function initApp(){

  try {

    const fm=function(n){n=n||0;return'\u20ba'+n.toLocaleString('tr-TR',{minimumFractionDigits:2,maximumFractionDigits:2});};

    const [rs,rd] = await Promise.all([fetch('/api/stats').then(r=>r.json()), fetch('/api/data').then(r=>r.json())]);

    const s=rs; const d=rd;

    allData=d.records||[]; filtered=[...allData]; doSort(); renderTbl();

    document.getElementById('st-tot').textContent=fm(s.total||0);

    document.getElementById('st-tc').textContent=(s.total_count||0)+' kayıt';

    document.getElementById('st-paid').textContent=fm(s.odendi||0);

    document.getElementById('st-pc').textContent=(s.odendi_count||0)+' kayıt';

    document.getElementById('st-pend').textContent=fm(s.bekleyen||0);

    document.getElementById('st-bc').textContent=(s.bekleyen_count||0)+' kayıt';

    document.getElementById('st-soon').textContent=fm(s.soon||0);

    document.getElementById('st-sc').textContent=(s.soon_count||0)+' kayıt · 30 gün';

    document.getElementById('db-sub').textContent=(s.total_count||0)>0

      ? (s.total_count)+' kayıt yüklü'

      : 'Veri yok — sol menüden Veri İçe Aktar kullanın';

    populateFilters(s.firmas||[],s.turler||[],s.aylar||[],s.yillar||[]);

    loadSourceStatus();

    var urlPage=new URLSearchParams(window.location.search).get('page');

    if(urlPage) showView(urlPage);

    _chartsLoaded=false;

    setTimeout(function(){ loadCharts(); },200);

    setTimeout(loadCurrentUser,300);

  } catch(e) {

    console.error('Init hatasi:',e);

    var sub=document.getElementById('db-sub');

    if(sub) sub.textContent='Yuklenemedi - sayfayi yenileyin';

    ['ch-mo','ch-pie','ch-st'].forEach(function(id){

      var el=document.getElementById(id);

      if(el) el.innerHTML='<div style="color:#f87171;font-size:11px;padding:20px;text-align:center">Yuklenemedi</div>';

    });

  }

}





</script>



<script>

// Script yüklenince hemen çalıştır

initApp();

// ── BANKA KARTLARI ────────────────────────────────────────────────────────────

var _bkData=[], _bkFiltered=[], _bkEditId=null, _bkHesapSayaci=0;

const BK_TURLER = ['taksitli','rotatif','kmh','kredi_karti','teminat','diger'];

const BK_TUR_ADLAR = {'taksitli':'Taksitli Kredi','rotatif':'Rotatif Kredi','kmh':'KMH','kredi_karti':'Kredi Kartı','teminat':'Teminat Mektubu','diger':'Diğer'};

const HESAP_TURLERI = ['Vadesiz TL','Vadeli TL','Vadesiz USD','Vadesiz EUR','Mevduat','DBS','Diğer'];



async function renderBankaKart(){

  try{

    const r=await api('/api/banka_kart'); const d=await r.json();

    _bkData=d.records||[]; _bkFiltered=[..._bkData];

    bkOzetGuncelle(); bkListeRender();

  }catch(e){console.error('BankaKart:',e);}

}



function bkOzetGuncelle(){

  const fm=n=>'₺'+(n||0).toLocaleString('tr-TR',{minimumFractionDigits:0});

  document.getElementById('bk-toplam-banka').textContent=_bkFiltered.length;

  const totL=_bkFiltered.reduce((s,b)=>s+bkToplam(b,'limit'),0);

  const totK=_bkFiltered.reduce((s,b)=>s+bkToplam(b,'kullanim'),0);

  document.getElementById('bk-toplam-limit').textContent=fm(totL);

  document.getElementById('bk-toplam-kullanim').textContent=fm(totK);

  document.getElementById('bk-toplam-kalan').textContent=fm(totL-totK);

}



function bkToplam(banka, tip){

  return BK_TURLER.reduce((s,t)=>s+(parseFloat(banka[tip+'_'+t])||0),0);

}



function bankaKartFiltrele(){

  const q=(document.getElementById('bk-ara').value||'').toLowerCase();

  _bkFiltered=q?_bkData.filter(b=>(b.banka||'').toLowerCase().includes(q)||(b.sube||'').toLowerCase().includes(q)):([..._bkData]);

  bkOzetGuncelle(); bkListeRender();

}



function bkListeRender(){

  const container=document.getElementById('bk-liste');

  if(!container)return;

  if(!_bkFiltered.length){

    container.innerHTML='<div style="color:var(--t3);padding:30px;text-align:center">Henüz banka kartı eklenmedi — sağ üstten ekleyin</div>'; return;

  }

  const fm=n=>n?(parseFloat(n)||0).toLocaleString('tr-TR',{minimumFractionDigits:0}):'0';

  const fmtl=n=>'₺'+fm(n);

  container.innerHTML=_bkFiltered.map(b=>{

    const totLimit=bkToplam(b,'limit'), totKull=bkToplam(b,'kullanim'), totKalan=totLimit-totKull;

    const kullOran=totLimit>0?Math.min(100,Math.round(totKull/totLimit*100)):0;

    const renk=kullOran>80?'var(--r)':kullOran>60?'var(--am)':'var(--acc)';

    const hesaplar=(b.hesaplar||[]);

    const ipotekSayisi=(b.ipotekler||[]).length;

    return `

    <div style="background:var(--s2);border:1px solid var(--b2);border-radius:12px;overflow:hidden">

      <!-- Banka Başlığı -->

      <div style="background:linear-gradient(135deg,#1e3a5f,#0f2744);padding:14px 18px;display:flex;align-items:center;gap:12px;flex-wrap:wrap">

        <div style="width:44px;height:44px;background:#fff;border-radius:10px;display:flex;align-items:center;justify-content:center;font-size:20px;flex-shrink:0">🏦</div>

        <div style="flex:1;min-width:0">

          <div style="font-size:16px;font-weight:700;color:#fff">${b.banka||'—'}</div>

          <div style="font-size:12px;color:rgba(255,255,255,.6)">${b.sube||''} ${b.firma?'· '+b.firma:''}</div>

        </div>

        <div style="display:flex;gap:8px;flex-shrink:0">

          <button onclick="bankaKartDuzenle('${b.id}')" style="background:rgba(255,255,255,.15);color:#fff;border:none;padding:6px 12px;border-radius:6px;cursor:pointer;font-size:12px">✏️ Düzenle</button>

          <button onclick="bkAktarKredi('${b.id}')" style="background:rgba(16,185,129,.3);color:#fff;border:none;padding:6px 12px;border-radius:6px;cursor:pointer;font-size:12px" title="Limitleri Kredi Limitleri bölümüne aktar">⬆ Kredi'ye Aktar</button>

          <button onclick="bkSil('${b.id}')" style="background:rgba(239,68,68,.3);color:#fff;border:none;padding:6px 10px;border-radius:6px;cursor:pointer;font-size:12px">🗑</button>

        </div>

      </div>

      <!-- İçerik Grid -->

      <div style="padding:14px 18px;display:grid;grid-template-columns:1fr 1fr 1fr;gap:14px">



        <!-- Limit Özeti -->

        <div style="grid-column:1/-1">

          <div style="display:flex;justify-content:space-between;margin-bottom:6px;font-size:11px;color:var(--t3)">

            <span>Kullanım: ${fmtl(totKull)} / ${fmtl(totLimit)}</span>

            <span style="font-weight:700;color:${renk}">${kullOran}%</span>

          </div>

          <div style="height:8px;background:var(--s3);border-radius:4px;overflow:hidden">

            <div style="height:100%;width:${kullOran}%;background:${renk};border-radius:4px;transition:.3s"></div>

          </div>

          <div style="display:flex;gap:12px;margin-top:8px;flex-wrap:wrap">

            ${BK_TURLER.map(t=>{

              const lim=parseFloat(b['limit_'+t])||0; if(!lim)return'';

              const kull=parseFloat(b['kullanim_'+t])||0; const kal=lim-kull;

              return `<span style="font-size:11px;background:var(--s3);padding:3px 8px;border-radius:6px">

                <span style="color:var(--t3)">\${BK_TUR_ADLAR[t]}:</span>

                <span style="font-family:var(--mono);font-weight:600;color:\${kal>0?'var(--g)':'var(--r)'}">\${fmtl(kal)}</span>

                <span style="color:var(--t3);font-size:10px"> kalan</span>

              </span>`;

            }).filter(Boolean).join('')}

          </div>

        </div>



        <!-- Müdür -->

        <div style="background:var(--s3);border-radius:8px;padding:10px">

          <div style="font-size:10px;color:var(--am);font-weight:700;margin-bottom:6px">ŞUBE MÜDÜRÜ</div>

          <div style="font-size:13px;font-weight:600;color:var(--t1)">${b.mudur_ad||'—'}</div>

          ${b.mudur_tel?`<div style="font-size:11px;color:var(--t3);margin-top:3px">📞 <a href="tel:${b.mudur_tel}" style="color:var(--acc)">${b.mudur_tel}</a></div>`:''}

          ${b.mudur_mail?`<div style="font-size:11px;color:var(--t3);margin-top:2px">✉ <a href="mailto:${b.mudur_mail}" style="color:var(--acc)">${b.mudur_mail}</a></div>`:''}

        </div>



        <!-- Personel -->

        <div style="background:var(--s3);border-radius:8px;padding:10px">

          <div style="font-size:10px;color:var(--g);font-weight:700;margin-bottom:6px">İLGİLİ PERSONEL</div>

          <div style="font-size:13px;font-weight:600;color:var(--t1)">${b.personel_ad||'—'}</div>

          ${b.personel_tel?`<div style="font-size:11px;color:var(--t3);margin-top:3px">📞 <a href="tel:${b.personel_tel}" style="color:var(--acc)">${b.personel_tel}</a></div>`:''}

          ${b.personel_mail?`<div style="font-size:11px;color:var(--t3);margin-top:2px">✉ <a href="mailto:${b.personel_mail}" style="color:var(--acc)">${b.personel_mail}</a></div>`:''}

        </div>



        <!-- Hesaplar + İpotekler -->

        <div style="background:var(--s3);border-radius:8px;padding:10px">

          <div style="font-size:10px;color:var(--acc);font-weight:700;margin-bottom:6px">HESAPLAR & İPOTEKLER</div>

          ${hesaplar.length?hesaplar.map(h=>`<div style="font-size:11px;color:var(--t2);margin-bottom:3px">

            <span style="background:var(--s2);padding:2px 6px;border-radius:4px;font-size:10px;color:var(--acc)">${h.tur||''}</span>

            <span style="font-family:var(--mono);font-size:11px;margin-left:4px">${h.no||''}</span>

          </div>`).join(''):'<span style="color:var(--t3);font-size:11px">Hesap eklenmedi</span>'}

          ${ipotekSayisi>0?`<div style="margin-top:6px;font-size:11px;color:var(--am)">🏠 ${ipotekSayisi} ipotek kayıtlı</div>`:''}

          <button onclick="bkIpotekYonet('${b.id}')" style="margin-top:8px;background:var(--s2);color:var(--t2);border:1px solid var(--b2);padding:4px 10px;border-radius:6px;cursor:pointer;font-size:11px;width:100%">🏠 İpotek Yönet (${ipotekSayisi})</button>

        </div>

      </div>

      ${b.notlar?`<div style="padding:0 18px 14px;font-size:12px;color:var(--t3);font-style:italic">📝 ${b.notlar}</div>`:''}

    </div>`;

  }).join('');

}



function bankaKartBankaSecildi(){

  // İleride otomatik logo vs. eklenebilir

}



function bankaKartHesapEkle(){

  const container=document.getElementById('bk-hesaplar');

  const idx=_bkHesapSayaci++;

  const div=document.createElement('div');

  div.style.cssText='display:grid;grid-template-columns:140px 1fr 30px;gap:6px;align-items:center';

  div.innerHTML=`

    <select class="fi2" id="bk-hesap-tur-${idx}" style="font-size:12px">

      ${HESAP_TURLERI.map(t=>`<option>${t}</option>`).join('')}

    </select>

    <input class="fi2" type="text" id="bk-hesap-no-${idx}" placeholder="IBAN / Hesap No" style="font-size:12px;font-family:var(--mono)">

    <button onclick="this.parentElement.remove()" style="background:var(--s3);color:var(--r);border:none;border-radius:6px;cursor:pointer;font-size:14px;padding:4px 8px">✕</button>`;

  container.appendChild(div);

}



function bkHesapla(){

  const fm=n=>'₺'+(parseFloat(n)||0).toLocaleString('tr-TR',{minimumFractionDigits:0});

  let totL=0,totK=0;

  BK_TURLER.forEach(t=>{

    const l=parseFloat(document.getElementById('bk-limit-'+t)?.value)||0;

    const k=parseFloat(document.getElementById('bk-kullanim-'+t)?.value)||0;

    const kal=l-k;

    const el=document.getElementById('bk-kalan-'+t);

    if(el){el.textContent=fm(kal);el.style.color=kal>0?'var(--g)':'var(--r)';}

    totL+=l; totK+=k;

  });

  const tlEl=document.getElementById('bk-toplam-limit-modal');

  const tkEl=document.getElementById('bk-toplam-kullanim-modal');

  const trEl=document.getElementById('bk-toplam-kalan-modal');

  if(tlEl)tlEl.textContent=fm(totL);

  if(tkEl)tkEl.textContent=fm(totK);

  if(trEl){trEl.textContent=fm(totL-totK);trEl.style.color=(totL-totK)>0?'var(--g)':'var(--r)';}

}



function bankaKartEkleAc(){

  _bkEditId=null;

  document.getElementById('bk-mod-baslik').textContent='Yeni Banka Kartı';

  // Tüm alanları temizle

  ['bk-banka','bk-sube','bk-firma','bk-mudur-ad','bk-mudur-tel','bk-mudur-mail',

   'bk-personel-ad','bk-personel-tel','bk-personel-mail','bk-notlar'].forEach(id=>{

    const el=document.getElementById(id); if(el)el.value='';

  });

  BK_TURLER.forEach(t=>{

    ['bk-limit-','bk-kullanim-','bk-faiz-'].forEach(pre=>{

      const el=document.getElementById(pre+t); if(el)el.value='';

    });

    const el=document.getElementById('bk-kalan-'+t); if(el)el.textContent='₺0';

  });

  document.getElementById('bk-hesaplar').innerHTML='';

  _bkHesapSayaci=0;

  bkHesapla();

  document.getElementById('bk-mod').style.display='flex';

}



function bankaKartDuzenle(id){

  const b=_bkData.find(x=>String(x.id)===String(id)); if(!b)return;

  _bkEditId=id;

  document.getElementById('bk-mod-baslik').textContent='Banka Kartını Düzenle';

  const f=(fid,val)=>{const el=document.getElementById(fid);if(el)el.value=val||'';};

  f('bk-banka',b.banka); f('bk-sube',b.sube); f('bk-firma',b.firma||'ULUSAL');

  f('bk-mudur-ad',b.mudur_ad); f('bk-mudur-tel',b.mudur_tel); f('bk-mudur-mail',b.mudur_mail);

  f('bk-personel-ad',b.personel_ad); f('bk-personel-tel',b.personel_tel); f('bk-personel-mail',b.personel_mail);

  f('bk-notlar',b.notlar);

  BK_TURLER.forEach(t=>{

    f('bk-limit-'+t, b['limit_'+t]);

    f('bk-kullanim-'+t, b['kullanim_'+t]);

    f('bk-faiz-'+t, b['faiz_'+t]?Math.round(b['faiz_'+t]*1000)/10:'');

  });

  // Hesaplar

  const container=document.getElementById('bk-hesaplar');

  container.innerHTML=''; _bkHesapSayaci=0;

  (b.hesaplar||[]).forEach(h=>{ bankaKartHesapEkle(); const idx=_bkHesapSayaci-1;

    const turEl=document.getElementById('bk-hesap-tur-'+idx);

    const noEl=document.getElementById('bk-hesap-no-'+idx);

    if(turEl)turEl.value=h.tur||'Vadesiz TL';

    if(noEl)noEl.value=h.no||'';

  });

  bkHesapla();

  document.getElementById('bk-mod').style.display='flex';

}



function bankaKartKapat(){ const m=document.getElementById('bk-mod'); m.classList.remove('on'); m.style.display=''; _bkEditId=null; }



async function bankaKartKaydet(){
  console.log('[BK] Kaydet başladi');
  const gv=id=>{const el=document.getElementById(id);return el?el.value.trim():'';};

  const gn=id=>parseFloat(document.getElementById(id)?.value)||0;

  const bkBankaEl = document.getElementById('bk-banka');
  const bkBankaVal = bkBankaEl ? bkBankaEl.value.trim() : '';
  if(!bkBankaVal){ toast('Lütfen bir banka seçin','e'); bkBankaEl?.focus(); return; }

  // Hesapları topla

  const hesaplar=[];

  document.querySelectorAll('#bk-hesaplar > div').forEach((div,i)=>{

    const tur=div.querySelector('[id^="bk-hesap-tur-"]')?.value;

    const no=div.querySelector('[id^="bk-hesap-no-"]')?.value?.trim();

    if(no) hesaplar.push({tur,no});

  });

  const kayit={

    id: _bkEditId||String(Date.now()),

    banka:gv('bk-banka'), sube:gv('bk-sube'), firma:gv('bk-firma'),

    mudur_ad:gv('bk-mudur-ad'), mudur_tel:gv('bk-mudur-tel'), mudur_mail:gv('bk-mudur-mail'),

    personel_ad:gv('bk-personel-ad'), personel_tel:gv('bk-personel-tel'), personel_mail:gv('bk-personel-mail'),

    notlar:gv('bk-notlar'), hesaplar,

    ipotekler: _bkEditId ? (_bkData.find(x=>String(x.id)===String(_bkEditId))?.ipotekler||[]) : [],

  };

  BK_TURLER.forEach(t=>{

    kayit['limit_'+t]=gn('bk-limit-'+t);

    kayit['kullanim_'+t]=gn('bk-kullanim-'+t);

    const faizYuzde=parseFloat(document.getElementById('bk-faiz-'+t)?.value)||0;

    kayit['faiz_'+t]=faizYuzde/100;

  });

  console.log('[BK] POST gonderiliyor, banka:', kayit.banka);
  try{
    const r=await api('/api/banka_kart/kaydet',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(kayit)});
    console.log('[BK] Response status:', r.status);
    if(r.status===401){ toast('Oturum sorunu - yeniden giriş yapın','e'); setTimeout(()=>window.location.href='/login',1500); return; }
    if(!r.ok){ toast('Sunucu hatası: '+r.status,'e'); return; }
    let d; try{ d=await r.json(); }catch(e){ console.error('[BK] JSON hatası:',e); toast('Yanıt okunamadı','e'); return; }
    console.log('[BK] Yanıt:', JSON.stringify(d));
    if(d.ok){ toast('Banka kartı kaydedildi ✓','s'); bankaKartKapat(); renderBankaKart(); }
    else toast('Hata: '+(d.msg||'Bilinmeyen hata'),'e');

  }catch(e){ toast('Hata: '+e.message,'e'); }

}



async function bkSil(id){

  if(!confirm('Bu banka kartı silinsin mi?'))return;

  const r=await api('/api/banka_kart/sil',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({id})});

  const d=await r.json();

  if(d.ok){ toast('Silindi','i'); renderBankaKart(); } else toast('Hata','e');

}



async function bkAktarKredi(id){

  if(!confirm('Bu bankanın limitleri Kredi Limitleri bölümüne aktarılsın mı?\n(Daha önce aktarılmış kayıtlar güncellenecek)'))return;

  const r=await api('/api/banka_kart/aktar_kredi',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({id})});

  const d=await r.json();

  if(d.ok){ toast(d.msg||'Aktarıldı','s'); }

  else toast('Hata: '+(d.msg||''),'e');

}



async function bkIpotekYonet(id){

  const b=_bkData.find(x=>String(x.id)===String(id)); if(!b)return;

  // İpotek listesini çek ve modal aç

  const r=await api('/api/banka_kart/ipotek_listesi?banka='+encodeURIComponent(b.banka));

  const d=await r.json();

  const mevcutIpotekIds=(b.ipotekler||[]).map(x=>String(x));

  const allIpotekler=d.records||[];

  const modal=document.createElement('div');

  modal.className='ov'; modal.style.display='flex';

  modal.innerHTML=`<div class="mod" style="max-width:700px;width:94%;max-height:85vh;overflow-y:auto">

    <div class="mtit"><span>🏠 ${b.banka} — İpotek Yönetimi</span><button class="mclose" onclick="this.closest('.ov').remove()">✕</button></div>

    <div style="padding:16px">

      <p style="font-size:13px;color:var(--t3);margin-bottom:12px">İpotek listesindeki ${b.banka} kayıtları. İşaretlenenler bu bankaya bağlanır.</p>

      ${allIpotekler.length?allIpotekler.map(ip=>`

        <label style="display:flex;align-items:center;gap:10px;padding:8px 10px;background:var(--s2);border-radius:8px;margin-bottom:6px;cursor:pointer">

          <input type="checkbox" value="${ip.id}" ${mevcutIpotekIds.includes(String(ip.id))?'checked':''} style="accent-color:var(--acc)">

          <div style="flex:1">

            <span style="font-weight:600">${ip.sahip||'—'}</span>

            <span style="color:var(--t3);font-size:12px;margin-left:8px">${ip.il||''} ${ip.ilce||''} · ${ip.nitelik||''}</span>

          </div>

          <span style="font-family:var(--mono);font-size:12px;color:var(--am)">₺${(ip.tutar||0).toLocaleString('tr-TR')}</span>

        </label>`).join(''):'<p style="color:var(--t3);text-align:center;padding:20px">Bu bankaya ait ipotek kaydı bulunamadı.<br><small>İpotek Takibi bölümünde banka adını eşleştirin.</small></p>'}

    </div>

    <div style="padding:0 16px 16px;display:flex;gap:10px">

      <button class="btn bp" style="flex:1" onclick="bkIpotekKaydet('${id}',this.closest('.ov'))">💾 Kaydet</button>

      <button class="btn bo" onclick="this.closest('.ov').remove()">İptal</button>

    </div>

  </div>`;

  document.body.appendChild(modal);

}



async function bkIpotekKaydet(bkId, modal){

  const checkboxes=modal.querySelectorAll('input[type=checkbox]');

  const secilen=[...checkboxes].filter(cb=>cb.checked).map(cb=>cb.value);

  const b=_bkData.find(x=>String(x.id)===String(bkId)); if(!b)return;

  b.ipotekler=secilen;

  const r=await api('/api/banka_kart/kaydet',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(b)});

  const d=await r.json();

  if(d.ok){ toast(secilen.length+' ipotek bağlandı','s'); modal.remove(); renderBankaKart(); }

  else toast('Hata','e');

}



// ── İPOTEK TAKİBİ ────────────────────────────────────────────────────────────

var _ipotekData=[],_ipotekFiltered=[],_ipotekEditId=null;

async function renderIpotek(){

  try{const r=await fetch('/api/ipotek');const d=await r.json();_ipotekData=d.records||[];_ipotekFiltered=[..._ipotekData];ipotekFiltreleriDoldur();ipotekTabloRender();ipotekOzetGuncelle();}catch(e){console.error('Ipotek:',e);}

}

function ipotekFiltreleriDoldur(){

  const fm=(id,vals)=>{const el=document.getElementById(id);if(!el)return;const cur=el.value;el.innerHTML=el.options[0].outerHTML+vals.map(v=>`<option ${v===cur?'selected':''}>${v}</option>`).join('');};

  fm('ipt-banka-f',[...new Set(_ipotekData.map(r=>r.banka).filter(Boolean))].sort());

  fm('ipt-il-f',[...new Set(_ipotekData.map(r=>r.il).filter(Boolean))].sort());

  fm('ipt-sahip-f',[...new Set(_ipotekData.map(r=>r.sahip).filter(Boolean))].sort());

}

function ipotekFiltrele(){

  const ara=(document.getElementById('ipt-ara').value||'').toLowerCase();

  const banka=document.getElementById('ipt-banka-f').value;

  const il=document.getElementById('ipt-il-f').value;

  const sahip=document.getElementById('ipt-sahip-f').value;

  _ipotekFiltered=_ipotekData.filter(r=>{

    if(banka&&r.banka!==banka)return false;if(il&&r.il!==il)return false;if(sahip&&r.sahip!==sahip)return false;

    if(ara){const hay=Object.values(r).join(' ').toLowerCase();if(!hay.includes(ara))return false;}return true;});

  ipotekTabloRender();ipotekOzetGuncelle();

}

function ipotekOzetGuncelle(){

  const fm=n=>'\u20ba'+(n||0).toLocaleString('tr-TR',{minimumFractionDigits:0,maximumFractionDigits:0});

  document.getElementById('ipt-toplam-sayi').textContent=_ipotekFiltered.length;

  const totExp=_ipotekFiltered.reduce((s,r)=>s+(r.exp_deger||0),0);

  const totIptek=_ipotekFiltered.reduce((s,r)=>s+(r.tutar||0),0);

  const fark=totExp-totIptek;

  document.getElementById('ipt-toplam-exp').textContent=fm(totExp);

  document.getElementById('ipt-toplam-limit').textContent=fm(totIptek);

  const farkEl=document.getElementById('ipt-limit-fark');

  const lblEl=document.getElementById('ipt-limit-fark-lbl');

  if(farkEl){farkEl.textContent=fm(Math.abs(fark));farkEl.style.color=fark>=0?'#10b981':'#ef4444';}

  if(lblEl)lblEl.textContent=fark>=0?'Ekspertiz Fazlası':'İpotek Fazlası';

}

function ipotekTabloRender(){

  const fm=n=>n?'\u20ba'+(n||0).toLocaleString('tr-TR',{minimumFractionDigits:0}):'—';

  const tbody=document.getElementById('ipt-tbody');

  if(!tbody)return;

  if(!_ipotekFiltered.length){tbody.innerHTML='<tr><td colspan="23" style="text-align:center;padding:30px;color:var(--t3)">Kayıt bulunamadı</td></tr>';return;}

  tbody.innerHTML=_ipotekFiltered.map((r,i)=>{

    const bg=i%2===0?'background:var(--bg2)':'';

    const td='padding:6px 6px;vertical-align:middle;border-bottom:1px solid var(--b2)';

    return `<tr style="${bg}">

      <td style="${td};text-align:center">${r.sno||i+1}</td>

      <td style="${td};font-size:10px">${r.kod||'—'}</td>

      <td style="${td};white-space:nowrap"><strong>${r.sahip||'—'}</strong></td>

      <td style="${td}">${r.il||'—'}</td><td style="${td}">${r.ilce||'—'}</td>

      <td style="${td};max-width:120px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap" title="${r.adres||''}">${r.adres||'—'}</td>

      <td style="${td};text-align:center">${r.ada||'—'}</td><td style="${td};text-align:center">${r.parsel||'—'}</td>

      <td style="${td};text-align:center">${r.bb||'—'}</td><td style="${td};text-align:center">${r.kat||'—'}</td><td style="${td};text-align:center">${r.blok||'—'}</td>

      <td style="${td};max-width:90px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap">${r.nitelik||'—'}</td>

      <td style="${td}"><strong style="color:#f59e0b">${r.banka||'—'}</strong></td>

      <td style="${td};text-align:right;font-family:var(--mono);color:#f59e0b">${fm(r.tutar)}</td>

      <td style="${td};text-align:center;font-size:10px">${r.ipotek_tarih||'—'}</td>

      <td style="${td};text-align:right;font-family:var(--mono);color:#0ea5e9"><strong>${fm(r.exp_deger)}</strong></td>

      <td style="${td};text-align:center;font-size:10px">${r.exp_tarih||'—'}</td>

      <td style="${td};font-size:10px">${r.tbl_tip||'—'}</td>

      <td style="${td};font-size:10px">${r.dask_no||'—'}</td><td style="${td};text-align:center;font-size:10px">${r.dask_vade||'—'}</td>

      <td style="${td};font-size:10px">${r.konut_no||'—'}</td><td style="${td};text-align:center;font-size:10px">${r.konut_vade||'—'}</td>

      <td style="${td};text-align:center;white-space:nowrap">

        <button onclick="ipotekDuzenle('${r.id}')" style="background:#e0f2fe;color:#0369a1;border:none;padding:3px 7px;border-radius:4px;cursor:pointer;font-size:10px">✏️</button>

        <button onclick="ipotekSil('${r.id}')" style="background:#fee2e2;color:#dc2626;border:none;padding:3px 7px;border-radius:4px;cursor:pointer;font-size:10px">🗑️</button>

      </td></tr>`;}).join('');

}

function ipotekOtoCap(){

  const exp=parseFloat(document.getElementById('ipt-exp-deger').value)||0;

  const tutar=parseFloat(document.getElementById('ipt-tutar').value)||0;

  const fark=exp-tutar;

  const farkEl=document.getElementById('ipt-fark-preview');

  const durumEl=document.getElementById('ipt-durum-preview');

  if(farkEl){farkEl.textContent=(fark>=0?'+':'')+('\u20ba'+Math.abs(fark).toLocaleString('tr-TR'));farkEl.style.color=fark>=0?'#10b981':'#ef4444';}

  if(durumEl){durumEl.textContent=fark>=0?'Ekspertiz fazla':'İpotek fazla';durumEl.style.color=fark>=0?'#166534':'#dc2626';}

}

function ipotekEkleAc(){

  _ipotekEditId=null;

  document.getElementById('ipt-mod-baslik').textContent='Yeni Taşınmaz Ekle';

  ['ipt-sno','ipt-kod','ipt-sahip','ipt-il','ipt-ilce','ipt-adres','ipt-ada','ipt-parsel','ipt-bb','ipt-kat','ipt-blok','ipt-nitelik','ipt-banka','ipt-tasınmaz-tip','ipt-dask-no','ipt-dask-vade','ipt-konut-no','ipt-konut-vade'].forEach(id=>{const el=document.getElementById(id);if(el)el.value='';});

  ['ipt-tutar','ipt-exp-deger'].forEach(id=>{const el=document.getElementById(id);if(el)el.value='';});

  ['ipt-ipotek-tarih','ipt-exp-tarih'].forEach(id=>{const el=document.getElementById(id);if(el)el.value='';});

  document.getElementById('ipt-mod').style.display='flex';

}

function ipotekDuzenle(id){

  const r=_ipotekData.find(x=>String(x.id)===String(id));if(!r)return;

  _ipotekEditId=id;document.getElementById('ipt-mod-baslik').textContent='Taşınmaz Düzenle';

  const f=(fid,val)=>{const el=document.getElementById(fid);if(el)el.value=val||'';};

  f('ipt-sno',r.sno);f('ipt-kod',r.kod);f('ipt-sahip',r.sahip);f('ipt-il',r.il);f('ipt-ilce',r.ilce);f('ipt-adres',r.adres);

  f('ipt-ada',r.ada);f('ipt-parsel',r.parsel);f('ipt-bb',r.bb);f('ipt-kat',r.kat);f('ipt-blok',r.blok);f('ipt-nitelik',r.nitelik);

  f('ipt-banka',r.banka);f('ipt-tutar',r.tutar);f('ipt-ipotek-tarih',r.ipotek_tarih);

  f('ipt-exp-deger',r.exp_deger);f('ipt-exp-tarih',r.exp_tarih);f('ipt-tasınmaz-tip',r.tbl_tip);

  f('ipt-dask-no',r.dask_no);f('ipt-dask-vade',r.dask_vade);f('ipt-konut-no',r.konut_no);f('ipt-konut-vade',r.konut_vade);

  ipotekOtoCap();document.getElementById('ipt-mod').style.display='flex';

}

function ipotekKapat(){document.getElementById('ipt-mod').style.display='none';_ipotekEditId=null;}

async function ipotekKaydet(){

  const gv=id=>{const el=document.getElementById(id);return el?el.value.trim():'';};

  const gn=id=>parseFloat(gv(id))||0;

  const kayit={id:_ipotekEditId||String(Date.now()),sno:gv('ipt-sno'),kod:gv('ipt-kod'),sahip:gv('ipt-sahip'),il:gv('ipt-il'),ilce:gv('ipt-ilce'),adres:gv('ipt-adres'),ada:gv('ipt-ada'),parsel:gv('ipt-parsel'),bb:gv('ipt-bb'),kat:gv('ipt-kat'),blok:gv('ipt-blok'),nitelik:gv('ipt-nitelik'),banka:gv('ipt-banka'),tutar:gn('ipt-tutar'),ipotek_tarih:gv('ipt-ipotek-tarih'),exp_deger:gn('ipt-exp-deger'),exp_tarih:gv('ipt-exp-tarih'),'tbl_tip':gv('ipt-tasınmaz-tip'),dask_no:gv('ipt-dask-no'),dask_vade:gv('ipt-dask-vade'),konut_no:gv('ipt-konut-no'),konut_vade:gv('ipt-konut-vade')};

  if(!kayit.il){toast('İl zorunlu!','e');return;}

  try{const r=await fetch('/api/ipotek/kaydet',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(kayit)});const d=await r.json();if(d.ok){toast('Kaydedildi','s');ipotekKapat();renderIpotek();}else toast('Hata: '+(d.msg||''),'e');}catch(e){toast('Hata: '+e.message,'e');}

}

async function ipotekSil(id){

  if(!confirm('Silinsin mi?'))return;

  try{const r=await fetch('/api/ipotek/sil',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({id})});const d=await r.json();if(d.ok){toast('Silindi','i');renderIpotek();}else toast('Hata','e');}catch(e){toast('Hata','e');}

}

async function ipotekExcel(){window.location.href='/api/ipotek/excel';}

async function ipotekSablonIndir(){window.location.href='/api/ipotek/sablon';}

async function ipotekExcelYukle(input){

  const file=input.files[0];if(!file)return;

  const formData=new FormData();formData.append('file',file);

  toast('Excel yükleniyor...','i');

  try{const r=await fetch('/api/ipotek/excel_yukle',{method:'POST',body:formData});const d=await r.json();if(d.ok){toast(d.msg||'Yuklendi','s');renderIpotek();}else toast('Hata: '+(d.msg||''),'e');}catch(e){toast('Hata: '+e.message,'e');}

  input.value='';

}



// ── KULLANICI YÖNETİMİ ────────────────────────────────────────────────────────

async function cikisYap(){if(!confirm('Çıkış yapılsın mı?'))return;await fetch('/api/logout');window.location.href='/login';}

async function loadCurrentUser(){

  try{const r=await fetch('/api/me');const d=await r.json();if(d.ok){const lbl=document.getElementById('nav-user-label');if(lbl)lbl.textContent='👤 '+d.ad+' ('+d.username+')';}}catch(e){}

}

async function loadKullanicilar(){

  const container=document.getElementById('kullanici-liste');if(!container)return;

  try{const r=await fetch('/api/users/list',{method:'POST',headers:{'Content-Type':'application/json'},body:'{}'});const d=await r.json();

  if(!d.ok){container.innerHTML='<p style="color:#9ca3af">'+(d.msg||'Yetki yok')+'</p>';return;}

  container.innerHTML=d.users.map(u=>`<div style="display:flex;align-items:center;gap:10px;padding:10px 14px;background:#f9fafb;border-radius:8px;margin-bottom:8px"><div style="flex:1"><strong>${u.username}</strong> <span style="color:#6b7280;font-size:12px">${u.ad}</span> <span style="background:${u.rol==='admin'?'#fef3c7':'#e0f2fe'};color:${u.rol==='admin'?'#92400e':'#0369a1'};padding:2px 8px;border-radius:12px;font-size:11px">${u.rol}</span></div><span style="font-size:11px;color:#9ca3af">${u.created_at}</span>${u.username!=='admin'?`<button onclick="kullaniciSil('${u.username}')" style="background:#fee2e2;color:#dc2626;border:none;padding:4px 10px;border-radius:6px;cursor:pointer;font-size:12px">Sil</button>`:''}</div>`).join('');}catch(e){container.innerHTML='<p style="color:#ef4444">Hata</p>';}

}

async function kullaniciEkle(){

  const uname=document.getElementById('yeni-uname').value.trim();

  const ad=document.getElementById('yeni-ad').value.trim();

  const pw=document.getElementById('yeni-pw').value;

  const rol=document.getElementById('yeni-rol').value;

  if(!uname||!pw){toast('Kullanıcı adı ve şifre zorunlu','e');return;}

  const r=await fetch('/api/users/add',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({username:uname,ad,password:pw,rol})});

  const d=await r.json();if(d.ok){toast('Eklendi','s');document.getElementById('yeni-uname').value='';document.getElementById('yeni-pw').value='';loadKullanicilar();}else toast('Hata: '+(d.msg||''),'e');

}

async function kullaniciSil(uname){

  if(!confirm(uname+' silinsin mi?'))return;

  const r=await fetch('/api/users/delete',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({username:uname})});

  const d=await r.json();if(d.ok){toast('Silindi','i');loadKullanicilar();}

}

async function sifreDegistir(){

  const yeni=document.getElementById('yeni-sifre').value;const tekrar=document.getElementById('yeni-sifre2').value;

  if(yeni!==tekrar){toast('Şifreler eşleşmiyor','e');return;}if(yeni.length<6){toast('En az 6 karakter','e');return;}

  const r=await fetch('/api/users/change_password',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({new_password:yeni})});

  const d=await r.json();if(d.ok){toast('Şifre değiştirildi','s');document.getElementById('yeni-sifre').value='';document.getElementById('yeni-sifre2').value='';}else toast('Hata','e');

}



// ── JSON İÇE AKTARMA ─────────────────────────────────────────────────────────

async function jsonYukle(){

  const tur=document.getElementById('imp-json-tur').value;const file=document.getElementById('imp-json-file').files[0];

  const msg=document.getElementById('imp-json-msg');if(!file){msg.textContent='Dosya seçin';msg.style.color='#ef4444';return;}

  msg.textContent='Yükleniyor...';msg.style.color='#6b7280';

  try{const text=await file.text();const data=JSON.parse(text);const r=await fetch('/api/import/json',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({tur,data})});const d=await r.json();if(d.ok){msg.textContent='✓ '+d.msg;msg.style.color='#166534';initApp();}else{msg.textContent='Hata: '+d.msg;msg.style.color='#ef4444';}}catch(e){msg.textContent='Hata: '+e.message;msg.style.color='#ef4444';}

}

async function topluYukle(){

  const files=document.getElementById('imp-toplu-files').files;const msg=document.getElementById('imp-toplu-msg');const detail=document.getElementById('imp-toplu-detail');

  if(!files.length){msg.textContent='Dosya seçin';msg.style.color='#ef4444';return;}

  msg.textContent='Yükleniyor...';msg.style.color='#6b7280';detail.innerHTML='';let results=[];

  for(const file of files){

    try{const text=await file.text();const data=JSON.parse(text);const name=file.name.toLowerCase();

    let tur='gider';if(name.includes('gelir'))tur='gelir';else if(name.includes('banka'))tur='banka';else if(name.includes('kredi'))tur='kredi';else if(name.includes('cari'))tur='cari';else if(name.includes('notlar'))tur='notlar';else if(name.includes('otel'))tur='otel';else if(name.includes('hatirlatma'))tur='hatirlatma';

    const r=await fetch('/api/import/json',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({tur,data})});const d=await r.json();results.push({file:file.name,ok:d.ok,msg:d.msg});}

    catch(e){results.push({file:file.name,ok:false,msg:e.message});}

  }

  msg.textContent='Tamamlandı';msg.style.color='#166534';

  detail.innerHTML=results.map(r=>`<div style="font-size:12px;padding:4px 0;color:${r.ok?'#166534':'#ef4444'}">${r.ok?'✓':'✗'} ${r.file} → ${r.msg}</div>`).join('');

  initApp();

}



async function githubYedekAl(){

  const btn=event.target;btn.textContent='⏳ Yedekleniyor...';btn.disabled=true;

  try{const r=await fetch('/api/yedek/github');const d=await r.json();if(d.ok){toast('GitHub yedegi tamamlandi','s');}else toast('Hata: '+(d.msg||''),'e');}catch(e){toast('Hata','e');}

  btn.textContent='☁️ GitHub Yedek';btn.disabled=false;

}





</script>

</body>

</html>'''





# ── LAUNCHER ──────────────────────────────────────────────────────────────────

PORT = 5678



def find_port(start=5678):

    import socket

    for p in range(start, start+20):

        with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:

            if s.connect_ex(('localhost', p)) != 0:

                return p

    return start



def kill_old_instances():

    """5678-5698 arasındaki tüm eski uygulamaları kapat"""

    import socket, urllib.request, time

    print("  Eski süreçler kontrol ediliyor...")

    for p in range(5678, 5698):

        try:

            with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:

                if s.connect_ex(('localhost', p)) == 0:

                    # Bu portta bir şey çalışıyor - kapat

                    try:

                        urllib.request.urlopen(f'http://localhost:{p}/api/shutdown', timeout=2)

                    except:

                        pass

            time.sleep(0.3)

        except:

            pass



def main():

    global PORT

    load_config()

    global SOURCE_SHEET

    SOURCE_SHEET = None

    try:

        if os.path.exists(CONFIG_FILE):

            import json as _json

            with open(CONFIG_FILE,'r',encoding='utf-8') as _f:

                _cfg = _json.load(_f)

            _cfg['source_sheet'] = None

            with open(CONFIG_FILE,'w',encoding='utf-8') as _f:

                _json.dump(_cfg,_f,ensure_ascii=False,indent=2)

    except: pass



    # Eski uygulamaları kapat

    kill_old_instances()

    import time; time.sleep(0.5)



    load_data()

    load_ipotek()

    load_banka_kart()

    load_users()



    # GitHub yedeginden yukle (arka planda)

    IS_RENDER = os.environ.get('RENDER') or os.environ.get('PORT')

    if GITHUB_TOKEN and len(DATA) == 0:

        def _restore():

            import time as _t; _t.sleep(3)

            github_yedekten_yukle()

        threading.Thread(target=_restore, daemon=True).start()



    if SOURCE_FILE:

        reload_from_source(force=True)



    PORT = int(os.environ.get('PORT', 5678))

    HOST = '0.0.0.0'

    from socketserver import TCPServer

    TCPServer.allow_reuse_address = True

    server = HTTPServer((HOST, PORT), AppHandler)

    url = f'http://localhost:{PORT}'



    print("=" * 55)

    print("  NAKİT AKIŞ YÖNETİM SİSTEMİ  v3.0")

    print("=" * 55)

    print(f"  Adres  : {url}")

    print(f"  Veriler: {SAVE_FILE}")

    if SOURCE_FILE:

        print(f"  Kaynak : {SOURCE_FILE}")

        print(f"  Durum  : {LAST_SYNC_MSG}")

    print("  Kapatmak için bu pencereyi kapatın (Ctrl+C)")

    print("=" * 55)



    # Arka planda kaynak dosya izleyici başlat

    threading.Thread(target=source_watcher, daemon=True).start()

    # Gece yedek zamanlayıcısı

    if GITHUB_TOKEN:

        threading.Thread(target=yedek_zamanlayici, daemon=True).start()



    if not os.environ.get('RENDER') and not os.environ.get('PORT'):

        def open_browser():

            import time; time.sleep(0.9)

            webbrowser.open(url)

        threading.Thread(target=open_browser, daemon=True).start()



    try:

        server.serve_forever()

    except KeyboardInterrupt:

        print("\n  Kapatılıyor...")

        server.shutdown()



if __name__ == '__main__':

    main()
